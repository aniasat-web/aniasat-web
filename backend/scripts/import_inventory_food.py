#!/usr/bin/env python3
"""Import storage inventory from Google Sheet workbook into inventory_items.

Default mapping matches the "Inventory - Food" tab format:
- Ingredient name: column B
- Unit: column C
- Storage quantity: column D
- Kitchen pantry quantity in column E is intentionally ignored.

Usage:
  cd backend
  . .venv/bin/activate
  python scripts/import_inventory_food.py --xlsx /tmp/spring_2026_inventory_file.xlsx
  python scripts/import_inventory_food.py --xlsx /tmp/spring_2026_inventory_file.xlsx --apply
"""

from __future__ import annotations

import argparse
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DB_PATH = SCRIPT_DIR.parent / "data" / "retreat_ops.db"

MASS_TO_G = {
    "g": 1.0,
    "kg": 1000.0,
    "lb": 453.59237,
    "oz": 28.349523125,
}

VOLUME_TO_ML = {
    "ml": 1.0,
    "l": 1000.0,
    "cup": 240.0,
    "tbsp": 14.7868,
    "tsp": 4.92892,
}

UNIT_ALIASES = {
    "gms": "g",
    "gram": "g",
    "grams": "g",
    "kgs": "kg",
    "kilogram": "kg",
    "kilograms": "kg",
    "kilo": "kg",
    "kilos": "kg",
    "lbs": "lb",
    "pound": "lb",
    "pounds": "lb",
    "litre": "l",
    "litres": "l",
    "liter": "l",
    "liters": "l",
    "ltr": "l",
    "ltrs": "l",
    "lt": "l",
    "lts": "l",
    "milliliter": "ml",
    "milliliters": "ml",
    "millilitre": "ml",
    "millilitres": "ml",
    "cups": "cup",
    "tablespoon": "tbsp",
    "tablespoons": "tbsp",
    "tbs": "tbsp",
    "teaspoon": "tsp",
    "teaspoons": "tsp",
    "pieces": "piece",
    "packets": "packet",
    "cans": "can",
    "bunches": "bunch",
    "loaves": "loaf",
    "sprigs": "sprig",
    "leaves": "leaf",
    "bags": "bag",
    "pinches": "pinch",
}


@dataclass
class ParsedRow:
    ingredient_name: str
    quantity: float
    unit: str
    source_row: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import storage inventory from workbook Inventory - Food tab.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Path to source XLSX file")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH, help=f"SQLite DB path (default: {DEFAULT_DB_PATH})")
    parser.add_argument("--sheet", default="Inventory - Food", help="Worksheet name (default: Inventory - Food)")
    parser.add_argument("--name-col", default="B", help="Ingredient column letter (default: B)")
    parser.add_argument("--unit-col", default="C", help="Unit column letter (default: C)")
    parser.add_argument("--storage-col", default="D", help="Storage quantity column letter (default: D)")
    parser.add_argument("--start-row", type=int, default=7, help="First row to scan (default: 7)")
    parser.add_argument("--source", default="Inventory - Food: Storage", help="inventory_items.source value")
    parser.add_argument("--apply", action="store_true", help="Write changes (default is dry-run)")
    parser.add_argument("--no-backup", action="store_true", help="Skip backup when --apply is used")
    return parser.parse_args()


def col_letter_to_index(col: str) -> int:
    col = col.strip().upper()
    index = 0
    for ch in col:
        if ch < "A" or ch > "Z":
            raise ValueError(f"Invalid column letter: {col!r}")
        index = index * 26 + (ord(ch) - ord("A") + 1)
    return index


def normalize_unit(raw_unit: str | None) -> str:
    value = str(raw_unit or "").strip().lower()
    if not value:
        return ""
    return UNIT_ALIASES.get(value, value)


def parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_canonical(quantity: float, unit: str) -> tuple[float, str]:
    if unit in MASS_TO_G:
        return quantity * MASS_TO_G[unit], "g"
    if unit in VOLUME_TO_ML:
        return quantity * VOLUME_TO_ML[unit], "ml"
    return quantity, unit


def backup_db(db_path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = db_path.with_name(f"{db_path.name}.bak-{ts}")
    backup_path.write_bytes(db_path.read_bytes())
    return backup_path


def read_rows(
    xlsx_path: Path,
    sheet_name: str,
    start_row: int,
    name_col: int,
    unit_col: int,
    storage_col: int,
) -> list[ParsedRow]:
    workbook = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found. Available: {', '.join(workbook.sheetnames)}")
    ws = workbook[sheet_name]

    parsed: list[ParsedRow] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        def value_at(col_idx: int) -> object | None:
            if col_idx <= 0:
                return None
            return row[col_idx - 1] if col_idx - 1 < len(row) else None

        ingredient_name = str(value_at(name_col) or "").strip()
        if not ingredient_name:
            continue

        storage_qty = parse_float(value_at(storage_col))
        if storage_qty is None or storage_qty <= 0:
            continue

        unit = normalize_unit(value_at(unit_col))
        if not unit:
            continue

        parsed.append(
            ParsedRow(
                ingredient_name=ingredient_name,
                quantity=storage_qty,
                unit=unit,
                source_row=row_idx,
            )
        )

    return parsed


def get_or_create_ingredient_id(cur: sqlite3.Cursor, name: str) -> int:
    existing = cur.execute(
        "SELECT id FROM ingredients WHERE lower(name) = lower(?)",
        (name,),
    ).fetchone()
    if existing:
        return int(existing["id"])
    created = cur.execute(
        "INSERT INTO ingredients(name) VALUES (?) RETURNING id",
        (name,),
    ).fetchone()
    return int(created["id"])


def apply_inventory(
    cur: sqlite3.Cursor,
    rows: Iterable[ParsedRow],
    source: str,
) -> tuple[int, int, int]:
    # Replace previous imported storage snapshot for this source.
    removed_existing = cur.execute("DELETE FROM inventory_items WHERE source = ?", (source,)).rowcount

    # Aggregate by ingredient and canonical unit.
    aggregate: dict[tuple[str, str], float] = {}
    display_name_by_key: dict[str, str] = {}
    for row in rows:
        canonical_qty, canonical_unit = to_canonical(row.quantity, row.unit)
        key_name = row.ingredient_name.strip().lower()
        key = (key_name, canonical_unit)
        aggregate[key] = aggregate.get(key, 0.0) + canonical_qty
        display_name_by_key[key_name] = row.ingredient_name.strip()

    inserted = 0
    canonical_updates = 0
    for (ingredient_key, unit), total_qty in aggregate.items():
        ingredient_name = display_name_by_key[ingredient_key]
        ingredient_id = get_or_create_ingredient_id(cur, ingredient_name)

        # Fill canonical_unit if missing.
        existing = cur.execute(
            "SELECT canonical_unit FROM ingredients WHERE id = ?",
            (ingredient_id,),
        ).fetchone()
        current_canonical = str(existing["canonical_unit"] or "").strip().lower() if existing else ""
        if not current_canonical and unit:
            cur.execute("UPDATE ingredients SET canonical_unit = ? WHERE id = ?", (unit, ingredient_id))
            canonical_updates += cur.rowcount

        cur.execute(
            """
            INSERT INTO inventory_items(ingredient_id, quantity, unit, source, updated_at)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (ingredient_id, round(total_qty, 4), unit, source),
        )
        inserted += 1

    return removed_existing, inserted, canonical_updates


def main() -> None:
    args = parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(f"XLSX not found: {args.xlsx}")
    if not args.db.exists():
        raise FileNotFoundError(f"DB not found: {args.db}")

    if args.apply and not args.no_backup:
        backup_path = backup_db(args.db)
        print(f"Backup created: {backup_path}")

    name_col = col_letter_to_index(args.name_col)
    unit_col = col_letter_to_index(args.unit_col)
    storage_col = col_letter_to_index(args.storage_col)

    parsed_rows = read_rows(
        xlsx_path=args.xlsx,
        sheet_name=args.sheet,
        start_row=args.start_row,
        name_col=name_col,
        unit_col=unit_col,
        storage_col=storage_col,
    )

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("PRAGMA foreign_keys = ON")
    cur.execute("BEGIN")

    removed_existing, inserted, canonical_updates = apply_inventory(cur, parsed_rows, args.source)

    if args.apply:
        conn.commit()
        mode = "APPLY"
    else:
        conn.rollback()
        mode = "DRY RUN"
    conn.close()

    print(f"\n{mode} summary")
    print(f"- Source workbook: {args.xlsx}")
    print(f"- Sheet: {args.sheet}")
    print(f"- Parsed storage rows: {len(parsed_rows)}")
    print(f"- Existing inventory rows removed for source '{args.source}': {removed_existing}")
    print(f"- Inventory rows inserted (aggregated): {inserted}")
    print(f"- Ingredient canonical_unit fields updated: {canonical_updates}")


if __name__ == "__main__":
    main()

