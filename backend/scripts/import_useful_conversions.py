from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

# Allow `python scripts/import_useful_conversions.py` from backend/.
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.db import get_connection, init_db


UNIT_ALIASES = {
    "gms": "g",
    "gram": "g",
    "grams": "g",
    "kilogram": "kg",
    "kilograms": "kg",
    "ml": "ml",
    "milliliter": "ml",
    "milliliters": "ml",
    "liter": "l",
    "liters": "l",
    "litre": "l",
    "litres": "l",
    "cup": "cup",
    "cups": "cup",
    "tbsp": "tbsp",
    "tbs": "tbsp",
    "tablespoon": "tbsp",
    "tablespoons": "tbsp",
    "tsp": "tsp",
    "tsb": "tsp",
    "teaspoon": "tsp",
    "teaspoons": "tsp",
    "oz": "oz",
    "ounce": "oz",
    "ounces": "oz",
    "lb": "lb",
    "lbs": "lb",
    "pound": "lb",
    "pounds": "lb",
}


@dataclass
class ConversionRow:
    item_name: str | None
    quantity_from: float
    unit_from: str
    quantity_to: float
    unit_to: str
    context: str
    source_sheet: str
    source_row: int
    notes: str | None = None


def as_text(value: object) -> str | None:
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return None


def as_float(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def normalize_unit(unit: str) -> str:
    value = unit.strip().lower()
    return UNIT_ALIASES.get(value, value)


def normalize_name(name: str) -> str:
    return " ".join(name.strip().split())


def parse_sheet(xlsx: Path, sheet_name: str) -> list[ConversionRow]:
    wb = load_workbook(xlsx, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")

    ws = wb[sheet_name]
    out: list[ConversionRow] = []

    # Generic conversions (solids and liquids): columns B-H
    for r in range(1, ws.max_row + 1):
        qty_from = as_float(ws.cell(r, 2).value)  # B
        unit_from = as_text(ws.cell(r, 3).value)  # C
        solid_to = as_float(ws.cell(r, 4).value)  # D
        solid_unit_to = as_text(ws.cell(r, 5).value)  # E
        liquid_to = as_float(ws.cell(r, 7).value)  # G
        liquid_unit_to = as_text(ws.cell(r, 8).value)  # H

        if qty_from is not None and unit_from and solid_to is not None and solid_unit_to:
            out.append(
                ConversionRow(
                    item_name=None,
                    quantity_from=qty_from,
                    unit_from=normalize_unit(unit_from),
                    quantity_to=solid_to,
                    unit_to=normalize_unit(solid_unit_to),
                    context="generic_solid",
                    source_sheet=sheet_name,
                    source_row=r,
                )
            )

        if qty_from is not None and unit_from and liquid_to is not None and liquid_unit_to:
            out.append(
                ConversionRow(
                    item_name=None,
                    quantity_from=qty_from,
                    unit_from=normalize_unit(unit_from),
                    quantity_to=liquid_to,
                    unit_to=normalize_unit(liquid_unit_to),
                    context="generic_liquid",
                    source_sheet=sheet_name,
                    source_row=r,
                )
            )

    # Ingredient-specific conversions: columns F-J
    for r in range(1, ws.max_row + 1):
        item_name = as_text(ws.cell(r, 6).value)  # F
        qty_from = as_float(ws.cell(r, 7).value)  # G
        unit_from = as_text(ws.cell(r, 8).value)  # H
        qty_to = as_float(ws.cell(r, 9).value)  # I
        unit_to = as_text(ws.cell(r, 10).value)  # J

        if not (item_name and qty_from is not None and unit_from and qty_to is not None and unit_to):
            continue

        low_name = item_name.lower()
        if low_name == "food":
            continue
        if low_name.startswith("unit"):
            continue

        out.append(
            ConversionRow(
                item_name=normalize_name(item_name),
                quantity_from=qty_from,
                unit_from=normalize_unit(unit_from),
                quantity_to=qty_to,
                unit_to=normalize_unit(unit_to),
                context="ingredient_specific",
                source_sheet=sheet_name,
                source_row=r,
            )
        )

    # De-duplicate exact repeats.
    deduped: list[ConversionRow] = []
    seen: set[tuple[object, ...]] = set()
    for row in out:
        key = (
            (row.item_name.lower() if row.item_name else None),
            round(row.quantity_from, 6),
            row.unit_from,
            round(row.quantity_to, 6),
            row.unit_to,
            row.context,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def upsert_ingredient(conn, ingredient_name: str) -> int:
    row = conn.execute(
        "SELECT id FROM ingredients WHERE lower(name) = lower(?)",
        (ingredient_name,),
    ).fetchone()
    if row:
        return int(row["id"])

    row = conn.execute(
        "INSERT INTO ingredients(name) VALUES (?) RETURNING id",
        (ingredient_name,),
    ).fetchone()
    return int(row["id"])


def maybe_set_grams_per_cup(
    conn,
    ingredient_name: str,
    qty_from: float,
    unit_from: str,
    qty_to: float,
    unit_to: str,
) -> bool:
    if qty_from <= 0:
        return False

    if unit_from != "cup":
        return False

    grams_per_cup: float
    if unit_to == "g":
        grams_per_cup = qty_to / qty_from
    elif unit_to == "kg":
        grams_per_cup = (qty_to * 1000.0) / qty_from
    else:
        return False

    ingredient_id = upsert_ingredient(conn, ingredient_name)

    row = conn.execute(
        "SELECT grams_per_cup, canonical_unit FROM ingredients WHERE id = ?",
        (ingredient_id,),
    ).fetchone()

    existing = row["grams_per_cup"]
    canonical = row["canonical_unit"]

    if existing is None:
        conn.execute(
            "UPDATE ingredients SET grams_per_cup = ?, canonical_unit = COALESCE(?, canonical_unit, 'g') WHERE id = ?",
            (grams_per_cup, canonical, ingredient_id),
        )
        return True

    # Update only if significantly different, assuming latest sheet value is authoritative.
    if abs(float(existing) - grams_per_cup) > 1e-6:
        conn.execute(
            "UPDATE ingredients SET grams_per_cup = ? WHERE id = ?",
            (grams_per_cup, ingredient_id),
        )
        return True

    return False


def import_rows(rows: list[ConversionRow], dry_run: bool, replace_existing: bool, sheet_name: str) -> None:
    init_db()

    inserted = 0
    updated_grams = 0

    with get_connection() as conn:
        if replace_existing and not dry_run:
            conn.execute("DELETE FROM unit_conversions WHERE source_sheet = ?", (sheet_name,))

        for row in rows:
            if not dry_run:
                conn.execute(
                    """
                    INSERT INTO unit_conversions(
                        item_name, quantity_from, unit_from, quantity_to, unit_to,
                        context, source_sheet, source_row, notes
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row.item_name,
                        row.quantity_from,
                        row.unit_from,
                        row.quantity_to,
                        row.unit_to,
                        row.context,
                        row.source_sheet,
                        row.source_row,
                        row.notes,
                    ),
                )

                if row.context == "ingredient_specific" and row.item_name:
                    if maybe_set_grams_per_cup(
                        conn,
                        ingredient_name=row.item_name,
                        qty_from=row.quantity_from,
                        unit_from=row.unit_from,
                        qty_to=row.quantity_to,
                        unit_to=row.unit_to,
                    ):
                        updated_grams += 1

            inserted += 1

        if not dry_run:
            conn.commit()

    by_context: dict[str, int] = {}
    for row in rows:
        by_context[row.context] = by_context.get(row.context, 0) + 1

    print("Import summary")
    print(f"- rows parsed: {len(rows)}")
    print(f"- rows inserted: {inserted if not dry_run else 0}")
    print(f"- grams_per_cup updates: {updated_grams if not dry_run else 0}")
    print(f"- by context: {by_context}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import useful conversions from workbook into unit_conversions table."
    )
    parser.add_argument("--xlsx", type=Path, required=True, help="Path to source .xlsx workbook")
    parser.add_argument(
        "--sheet",
        default="useful conversions",
        help="Sheet name (default: useful conversions)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse/report only; no DB writes")
    parser.add_argument(
        "--replace-existing",
        action="store_true",
        help="Delete prior rows from this sheet before inserting",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Workbook not found: {args.xlsx}")

    rows = parse_sheet(args.xlsx, args.sheet)

    print("Parsed conversion rows")
    for row in rows[:25]:
        print(
            f"- {row.context} | item={row.item_name} | "
            f"{row.quantity_from:g} {row.unit_from} -> {row.quantity_to:g} {row.unit_to} "
            f"(row {row.source_row})"
        )
    if len(rows) > 25:
        print(f"... ({len(rows) - 25} more rows)")

    import_rows(rows, dry_run=args.dry_run, replace_existing=args.replace_existing, sheet_name=args.sheet)


if __name__ == "__main__":
    main()
