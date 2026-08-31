from __future__ import annotations

import csv
import io
import json
import os
import re
import secrets
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Annotated
from typing import Any
from typing import Literal
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from .auth import (
    BOOTSTRAP_ADMIN_PASSWORD_ENV,
    BOOTSTRAP_ADMIN_USERNAME_ENV,
    GUEST_SCOPE_KITCHEN_RETREAT_VIEW,
    GUEST_SCOPE_KITCHEN_TESTING,
    ROLE_ADMIN,
    ROLE_PLANNER,
    ROLE_VIEWER,
    SESSION_COOKIE_NAME,
    AuthUser,
    authenticate_credentials,
    authenticate_guest_session_token,
    authenticate_session_token,
    cookie_secure_enabled,
    create_session,
    create_guest_session,
    create_user,
    default_route_for_role,
    delete_guest_session,
    delete_guest_sessions_for_scope,
    delete_session,
    ensure_bootstrap_admin,
    guest_session_cookie_name,
    guest_session_hours,
    hash_session_token,
    has_any_users,
    list_users,
    normalize_role,
    update_user,
)
from .db import get_connection, init_db, table_columns, table_exists
from .usda import populate_ingredient_conversions

app = FastAPI(title="Blossom Foundation Volunteering API", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR = Path(__file__).resolve().parents[2] / "frontend"


def using_sqlite_backend() -> bool:
    return not str(os.getenv("DATABASE_URL", "") or "").strip()

MASS_TO_G = {
    "g": 1.0,
    "kg": 1000.0,
    "lb": 453.59237,
    "lbs": 453.59237,
    "oz": 28.349523125,
}

VOLUME_TO_ML = {
    "ml": 1.0,
    "l": 1000.0,
    "fl oz": 29.5735295625,
    "qt": 946.352946,
    "gal": 3785.411784,
    "cup": 240.0,
    "cups": 240.0,
    "tbsp": 14.7868,
    "tsp": 4.92892,
}

COUNT_UNITS = {
    "each",
    "piece",
    "pieces",
    "pack",
    "packs",
    "packet",
    "packets",
    "can",
    "cans",
    "bunch",
    "bunches",
    "loaf",
    "loaves",
    "sprig",
    "sprigs",
    "leaf",
    "leaves",
    "pinch",
    "pinches",
    "bag",
    "bags",
    "box",
    "boxes",
    "case",
    "cases",
    "bottle",
    "bottles",
    "jug",
    "jugs",
    "jar",
    "jars",
    "carton",
    "cartons",
    "tub",
    "tubs",
    "package",
    "packages",
}

DAL_RICE_TOKEN_RE = re.compile(r"\b(rice|dal|moong|mung|toor|urad|masoor)\b")
STORAGE_GRID_LOCATION_RE = re.compile(r"^([A-Za-z])\s*-?\s*([1-9]|[1-9][0-9])$")
INVENTORY_BARCODE_RE = re.compile(r"^\d{8,14}$")
INFRA_CATEGORY_NAME = "Infra"
INFRA_CATEGORY_EXACT = {
    "infra",
    "infrastructure",
    "maintenance",
    "facility maintenance",
    "facilities maintenance",
    "janitorial",
    "housekeeping",
}
INFRA_CATEGORY_HINTS = (
    "cleaning",
    "maintenance",
    "janitorial",
    "housekeeping",
    "facility",
    "facilities",
)
INFRA_ITEM_HINTS = (
    "all purpose cleaner",
    "cleaner",
    "detergent",
    "dish soap",
    "dishwashing",
    "disinfect",
    "sanitizer",
    "sanitiser",
    "toilet cleaner",
    "trash bag",
    "garbage bag",
    "paper towel",
    "broom",
    "mop",
    "vacuum",
)

RECIPE_CATEGORIES = [
    "M's Recipes",
    "Breakfast",
    "Salads",
    "Vegetable Dishes",
    "Dals & Stews",
    "Khichdi & Kadhi",
    "Rice Dishes",
    "Desserts",
    "Chai & Coffee",
    "Pickles",
    "Ready to Serve",
]

OPENAI_API_KEY_ENV = "OPENAI_API_KEY"
OPENAI_API_BASE_ENV = "RETREAT_OPS_OPENAI_API_BASE"
OPENAI_INGREDIENT_DUP_MODEL_ENV = "RETREAT_OPS_INGREDIENT_DUP_MODEL"
DEFAULT_OPENAI_API_BASE = "https://api.openai.com/v1"
DEFAULT_INGREDIENT_DUP_MODEL = "gpt-5-mini"
INVENTORY_WITHDRAW_ACCESS_CODE_ENV = "RETREAT_OPS_INVENTORY_WITHDRAW_ACCESS_CODE"
APP_SETTING_INVENTORY_WITHDRAW_ACCESS_CODE = "inventory_withdraw_access_code"
APP_SETTING_KITCHEN_TESTING_ACCESS_CODE = "kitchen_testing_access_code"
APP_SETTING_KITCHEN_RETREAT_VIEW_ACCESS_CODE = "kitchen_retreat_view_access_code"
KITCHEN_ACCESS_SCOPE_TESTING = "testing"
KITCHEN_ACCESS_SCOPE_RETREAT_VIEW = "retreat-view"

PUBLIC_API_PATHS = {
    "/api/health",
    "/api/auth/login",
    "/api/auth/bootstrap-status",
    "/api/inventory/withdraw-search",
    "/api/inventory/withdraw-complete",
}

PUBLIC_API_PREFIXES = (
    "/api/kitchen-access/",
)

PUBLIC_API_GET_PREFIXES = (
)

PUBLIC_API_GET_PATHS = {
    "/api/inventory/withdraw-config",
}

KITCHEN_ACCESS_SCOPE_CONFIG = {
    KITCHEN_ACCESS_SCOPE_TESTING: {
        "setting_key": APP_SETTING_KITCHEN_TESTING_ACCESS_CODE,
        "guest_session_scope": GUEST_SCOPE_KITCHEN_TESTING,
        "label": "Kitchen Testing View",
    },
    KITCHEN_ACCESS_SCOPE_RETREAT_VIEW: {
        "setting_key": APP_SETTING_KITCHEN_RETREAT_VIEW_ACCESS_CODE,
        "guest_session_scope": GUEST_SCOPE_KITCHEN_RETREAT_VIEW,
        "label": "Kitchen Retreat View",
    },
}

HEADCOUNT_PROFILES = {"retreat", "test"}
DEFAULT_TEST_HEADCOUNT = 4.0
DEFAULT_SHOPPING_PROFILE = "retreat"
SHOPPING_PHASES = ["bulk", "fresh", "daily", "custom"]
MANUAL_INVENTORY_SOURCE = "Shopping Manual Override"
ORDER_PUTAWAY_INVENTORY_SOURCE = "Inventory Order Putaway"
OPEN_FOOD_FACTS_PRODUCT_ENDPOINT = "https://world.openfoodfacts.org/api/v2/product/{barcode}.json"
OPEN_PRODUCTS_FACTS_PRODUCT_ENDPOINT = "https://world.openproductsfacts.org/api/v2/product/{barcode}.json"
OPEN_BEAUTY_FACTS_PRODUCT_ENDPOINT = "https://world.openbeautyfacts.org/api/v2/product/{barcode}.json"
OPEN_FOOD_FACTS_SEARCH_ENDPOINT = "https://world.openfoodfacts.org/cgi/search.pl"
OPEN_PRODUCTS_FACTS_SEARCH_ENDPOINT = "https://world.openproductsfacts.org/cgi/search.pl"
OPEN_BEAUTY_FACTS_SEARCH_ENDPOINT = "https://world.openbeautyfacts.org/cgi/search.pl"
UPCITEMDB_LOOKUP_ENDPOINT = "https://api.upcitemdb.com/prod/trial/lookup?upc={barcode}"
UPCITEMDB_API_KEY_ENV = "UPCITEMDB_API_KEY"


class IngredientInput(BaseModel):
    name: str = Field(min_length=1)
    quantity: float = Field(gt=0)
    unit: str = Field(min_length=1)


class ScalePreviewRequest(BaseModel):
    base_servings: float = Field(gt=0)
    target_servings: float = Field(gt=0)
    ingredients: list[IngredientInput] = Field(min_length=1)


class ScaledIngredient(BaseModel):
    name: str
    input_quantity: float
    input_unit: str
    scaled_quantity: float
    scaled_unit: str
    canonical_quantity: float | None
    canonical_unit: str | None
    shopping_quantity: float | None
    shopping_unit: str | None
    note: str | None = None


class ScalePreviewResponse(BaseModel):
    scale_factor: float
    ingredients: list[ScaledIngredient]


class RecipeIngredientCreate(BaseModel):
    ingredient_name: str = Field(min_length=1)
    quantity: float = Field(gt=0)
    unit: str = Field(min_length=1)
    prep_notes: str | None = None


class RecipeCreate(BaseModel):
    name: str = Field(min_length=1)
    category: str | None = None
    base_servings: int = Field(gt=0)
    notes: str | None = None
    ingredients: list[RecipeIngredientCreate] = Field(min_length=1)
    steps: list[str] = Field(default_factory=list)


class RetreatPlanMeal(BaseModel):
    day: int = Field(ge=1)
    meal: str = Field(min_length=1)
    people: float = Field(ge=0)
    dishes: list[str] = Field(default_factory=list)


class RetreatPlanPayload(BaseModel):
    id: int | None = Field(default=None, ge=1)
    name: str = Field(min_length=1)
    startDate: str | None = None
    dayCount: int = Field(ge=1, le=10)
    defaultPeople: float = Field(gt=0)
    meals: list[RetreatPlanMeal] = Field(min_length=1)
    retreatDefaultPeople: float | None = Field(default=None, gt=0)
    testDefaultPeople: float = Field(default=4, gt=0)
    activeProfile: Literal["retreat", "test"] = "retreat"
    shoppingProfile: Literal["retreat", "test"] = "retreat"
    retreatMeals: list[RetreatPlanMeal] | None = None
    testMeals: list[RetreatPlanMeal] | None = None


class RetreatPlanDuplicatePayload(BaseModel):
    name: str | None = None


class AuthLoginPayload(BaseModel):
    username: str = Field(min_length=1)
    password: str = Field(min_length=1)


class AuthUserCreatePayload(BaseModel):
    username: str = Field(min_length=1)
    password: str = Field(min_length=1)
    role: str = Field(min_length=1)
    is_active: bool = True


class AuthUserUpdatePayload(BaseModel):
    password: str | None = None
    role: str | None = None
    is_active: bool | None = None


class AuthChangePasswordPayload(BaseModel):
    current_password: str = Field(min_length=1)
    new_password: str = Field(min_length=1)


class IngredientUpdatePayload(BaseModel):
    name: str = Field(min_length=1)
    canonical_unit: str | None = None
    grams_per_cup: float | None = None
    notes: str | None = None
    category: str | None = None
    purchase_tier: str | None = None


class IngredientDuplicateScanPayload(BaseModel):
    max_groups: int = Field(default=20, ge=1, le=60)
    model: str | None = None


class StandaloneInventoryCreate(BaseModel):
    item_name: str = Field(min_length=1)
    barcode: str = Field(min_length=8)
    barcodes: list[str] | None = None
    quantity: float = Field(ge=0, default=0)
    unit: str = Field(min_length=1)
    category: str | None = None
    location: str = Field(min_length=1)
    image_url: str = Field(min_length=1)
    notes: str | None = None


class StandaloneInventoryUpdate(BaseModel):
    item_name: str = Field(min_length=1)
    barcode: str = Field(min_length=8)
    barcodes: list[str] | None = None
    quantity: float = Field(ge=0, default=0)
    unit: str = Field(min_length=1)
    category: str | None = None
    location: str = Field(min_length=1)
    image_url: str = Field(min_length=1)
    notes: str | None = None


class StandaloneInventoryPatch(BaseModel):
    item_name: str | None = None
    barcode: str | None = None
    barcodes: list[str] | None = None
    quantity: float | None = Field(default=None, ge=0)
    unit: str | None = None
    category: str | None = None
    location: str | None = None
    image_url: str | None = None
    notes: str | None = None


class StandaloneInventoryCategoryPatch(BaseModel):
    category: str | None = None


class StandaloneInventoryNotesPatch(BaseModel):
    notes: str | None = None


class StandaloneInventoryNamePatch(BaseModel):
    item_name: str = Field(min_length=1)


class StandaloneInventoryBarcodeBindPayload(BaseModel):
    item_id: int = Field(gt=0)
    barcode: str = Field(min_length=8)


class StandaloneInventoryMergePayload(BaseModel):
    source_item_id: int = Field(gt=0)
    target_item_id: int = Field(gt=0)


class InventoryWithdrawSearchPayload(BaseModel):
    query: str = Field(min_length=1)
    accessCode: str | None = None
    limit: int = Field(default=12, ge=1, le=50)


class InventoryWithdrawItemInput(BaseModel):
    itemId: int = Field(gt=0)
    quantity: float = Field(gt=0)


class InventoryWithdrawCompletePayload(BaseModel):
    withdrawnBy: str | None = None
    accessCode: str | None = None
    reason: str | None = None
    items: list[InventoryWithdrawItemInput] = Field(default_factory=list)


class AdminInventoryWithdrawAccessPayload(BaseModel):
    accessCode: str | None = Field(default=None, max_length=128)


class AdminSharedAccessCodePayload(BaseModel):
    accessCode: str | None = Field(default=None, max_length=128)


class KitchenGuestAccessLoginPayload(BaseModel):
    accessCode: str = Field(min_length=1, max_length=128)


class InventoryOrderItemInput(BaseModel):
    itemType: Literal["INGREDIENT", "STANDALONE_INVENTORY"]
    itemId: int = Field(gt=0)
    requiredQuantity: float = Field(default=0, ge=0)
    orderedQuantity: float = Field(default=0, ge=0)
    receivedQuantity: float = Field(default=0, ge=0)
    appliedQuantity: float = Field(default=0, ge=0)
    unit: str | None = None
    purchaseUnit: str | None = None
    unitsPerPurchase: float = Field(default=1, gt=0)
    draftPurchaseUnit: str | None = None
    draftUnitsPerPurchase: float | None = Field(default=None, gt=0)
    draftOrderedPurchaseQuantity: float | None = Field(default=None, ge=0)
    orderedPurchaseQuantity: float | None = Field(default=None, ge=0)
    receivedPurchaseQuantity: float | None = Field(default=None, ge=0)
    sourceShoppingListItemId: int | None = Field(default=None, gt=0)
    orderUrlOverride: str | None = None
    notes: str | None = None


class InventoryOrderCreate(BaseModel):
    domain: Literal["FOOD", "NON_FOOD"]
    name: str | None = None
    sourceType: Literal["SHOPPING_LIST", "NON_FOOD_PLAN", "MANUAL", "LEGACY"] | None = None
    sourceId: int | None = Field(default=None, gt=0)
    workflowStage: Literal["PLANNING", "PURCHASING", "RECEIVING", "COMPLETE"] | None = None
    supplierName: str | None = None
    notes: str | None = None
    items: list[InventoryOrderItemInput] = Field(default_factory=list)


class InventoryOrderUpdate(BaseModel):
    name: str | None = None
    sourceType: Literal["SHOPPING_LIST", "NON_FOOD_PLAN", "MANUAL", "LEGACY"] | None = None
    sourceId: int | None = Field(default=None, gt=0)
    expectedWorkflowStage: Literal["PLANNING", "PURCHASING", "RECEIVING", "COMPLETE"] | None = None
    workflowStage: Literal["PLANNING", "PURCHASING", "RECEIVING", "COMPLETE"] | None = None
    supplierName: str | None = None
    notes: str | None = None
    items: list[InventoryOrderItemInput] | None = None
    preserveEmptyItems: bool = False


class InventoryOrderPutawayItemInput(BaseModel):
    orderItemId: int = Field(gt=0)
    quantity: float | None = Field(default=None, ge=0)
    location: str | None = None
    reason: str | None = None


class InventoryOrderPutawayPayload(BaseModel):
    items: list[InventoryOrderPutawayItemInput] = Field(default_factory=list)


class StandaloneInventoryOrderDraftItemCreate(BaseModel):
    itemName: str = Field(min_length=1)
    category: str | None = None
    unit: str | None = None
    barcode: str | None = None
    location: str | None = None
    imageUrl: str | None = None
    orderUrl: str | None = None
    notes: str | None = None


class RetreatInventoryCategoryCreate(BaseModel):
    name: str = Field(min_length=1)
    trackingMode: Literal["ITEM", "CATEGORY"] = "ITEM"
    imageUrl: str | None = None


class RetreatInventoryCategoryUpdate(BaseModel):
    name: str | None = None
    trackingMode: Literal["ITEM", "CATEGORY"] | None = None
    imageUrl: str | None = None
    active: bool | None = None


SHELF_LOCATION_PATTERN = re.compile(r"^[A-Za-z]\d+$")


def validate_shelf_location_name(name: str) -> str:
    clean = name.strip().upper()
    if not SHELF_LOCATION_PATTERN.match(clean):
        raise HTTPException(
            status_code=400,
            detail=f"Location name '{name}' must follow shelf-location pattern (letter + number, e.g. A1, B12).",
        )
    return clean


class RetreatInventoryLocationCreate(BaseModel):
    name: str = Field(min_length=1)
    description: str | None = None
    active: bool = True


class RetreatInventoryLocationUpdate(BaseModel):
    name: str | None = None
    description: str | None = None
    active: bool | None = None


class RetreatInventoryItemLocationInput(BaseModel):
    locationId: int = Field(gt=0)
    quantity: int = Field(default=0, ge=0)


class RetreatInventoryItemCreate(BaseModel):
    name: str = Field(min_length=1)
    barcode: str = Field(min_length=1)
    categoryId: int = Field(gt=0)
    # Deprecated legacy input, kept for backward compatibility.
    shelfLocation: str | None = None
    unit: str = Field(default="each", min_length=1)
    purchaseUrl: str | None = None
    brand: str | None = None
    description: str | None = None
    imageUrl: str | None = None
    active: bool = True
    locations: list[RetreatInventoryItemLocationInput] = Field(default_factory=list)


class RetreatInventoryItemUpdate(BaseModel):
    name: str | None = None
    barcode: str | None = None
    categoryId: int | None = Field(default=None, gt=0)
    # Deprecated legacy input, kept for backward compatibility.
    shelfLocation: str | None = None
    unit: str | None = None
    purchaseUrl: str | None = None
    brand: str | None = None
    description: str | None = None
    imageUrl: str | None = None
    active: bool | None = None
    locations: list[RetreatInventoryItemLocationInput] | None = None


class RetreatInventoryScanPayload(BaseModel):
    barcode: str = Field(min_length=1)
    transactionType: Literal["IN", "OUT", "ADJUSTMENT"] = "IN"
    quantity: int = 1
    reason: str | None = None
    locationId: int | None = Field(default=None, gt=0)


class PurchaseOrderCreate(BaseModel):
    supplierName: str | None = None
    status: Literal["DRAFT", "ORDERED", "PARTIAL", "RECEIVED"] = "DRAFT"
    expectedDate: str | None = None
    notes: str | None = None
    items: list["PurchaseOrderItemInput"] = Field(default_factory=list)


class PurchaseOrderUpdate(BaseModel):
    supplierName: str | None = None
    status: Literal["DRAFT", "ORDERED", "PARTIAL", "RECEIVED"] | None = None
    expectedDate: str | None = None
    notes: str | None = None
    items: list["PurchaseOrderItemInput"] | None = None


class PurchaseOrderItemInput(BaseModel):
    entityType: Literal["ITEM", "CATEGORY"] = "ITEM"
    entityId: int = Field(gt=0)
    orderedQuantity: int = Field(ge=0)
    receivedQuantity: int = Field(default=0, ge=0)


class ShoppingListGeneratePayload(BaseModel):
    retreatPlanId: int | None = Field(default=None, gt=0)
    retreatPlanIds: list[int] | None = None
    allRetreats: bool = False
    name: str | None = None
    phase: Literal["bulk", "fresh", "daily", "custom"] = "bulk"
    purchaseTiers: list[Literal["bulk", "fresh", "daily"]] | None = None
    profile: Literal["retreat", "test"] = "retreat"
    subtractInventory: bool = True
    includeZeroToBuy: bool = False


class ShoppingListItemVendorAllocationInput(BaseModel):
    id: int | None = Field(default=None, gt=0)
    vendorId: int | None = Field(default=None, ge=1)
    allocatedQty: float = Field(default=0, ge=0)
    allocatedUnit: str | None = None
    ordered: bool = False
    received: bool = False


class ShoppingListItemUpdatePayload(BaseModel):
    vendorId: int | None = Field(default=None, ge=1)
    inStockQty: float | None = Field(default=None, ge=0)
    orderedQty: float | None = Field(default=None, ge=0)
    orderedUnit: str | None = None
    ordered: bool | None = None
    received: bool | None = None
    notes: str | None = None
    vendorAllocations: list[ShoppingListItemVendorAllocationInput] | None = None


class ShoppingListUpdatePayload(BaseModel):
    name: str = Field(min_length=1)


class ShoppingListCarryForwardPayload(BaseModel):
    name: str | None = None
    phase: Literal["bulk", "fresh", "daily", "custom"] | None = None


class ShoppingPickupListCreatePayload(BaseModel):
    itemIds: list[int] = Field(min_length=1)
    name: str | None = None
    vendorId: int | None = Field(default=None, ge=1)
    assignee: str | None = None
    pickupDate: str | None = None
    notes: str | None = None


class ShoppingPickupListUpdatePayload(BaseModel):
    name: str | None = None
    vendorId: int | None = Field(default=None, ge=1)
    assignee: str | None = None
    pickupDate: str | None = None
    notes: str | None = None


class ServiceSnapshotIngredient(BaseModel):
    name: str = Field(min_length=1)
    scaledQty: str = Field(min_length=1)
    scaledUnit: str = Field(min_length=1)
    shopQty: str | None = None
    shopUnit: str | None = None


class ServiceSnapshotDish(BaseModel):
    name: str = Field(min_length=1)
    serves: float = Field(gt=0)
    baseServings: float = Field(gt=0)
    factor: float = Field(gt=0)
    ingredients: list[ServiceSnapshotIngredient] = Field(min_length=1)
    steps: list[str] = Field(min_length=1)


class ServiceSnapshotMeal(BaseModel):
    day: int = Field(ge=1)
    meal: str = Field(min_length=1)
    people: float = Field(gt=0)
    dishes: list[ServiceSnapshotDish] = Field(min_length=1)


class ServiceSnapshotPayload(BaseModel):
    version: int = Field(default=1, ge=1)
    retreatName: str = Field(min_length=1)
    generatedAt: str | None = None
    meals: list[ServiceSnapshotMeal] = Field(min_length=1)
    retreatPlanId: int | None = None
    profile: Literal["retreat", "test"] | None = None


@app.on_event("startup")
def on_startup() -> None:
    init_db()
    with get_connection() as conn:
        created = ensure_bootstrap_admin(conn)
        conn.commit()
    if created:
        print("Bootstrap admin user created from environment variables.")


def get_request_user(request: Request) -> AuthUser | None:
    candidate = getattr(request.state, "auth_user", None)
    if isinstance(candidate, AuthUser):
        return candidate
    return None


def require_authenticated_user(request: Request) -> AuthUser:
    user = get_request_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    return user


def require_roles(*allowed_roles: str):
    normalized_allowed = {normalize_role(role) for role in allowed_roles}

    def dependency(user: Annotated[AuthUser, Depends(require_authenticated_user)]) -> AuthUser:
        if user.role not in normalized_allowed:
            raise HTTPException(
                status_code=403,
                detail=f"Requires one of roles: {', '.join(sorted(normalized_allowed))}",
            )
        return user

    return dependency


def resolve_kitchen_access_scope(scope_slug: str) -> dict[str, str]:
    normalized = str(scope_slug or "").strip().lower()
    details = KITCHEN_ACCESS_SCOPE_CONFIG.get(normalized)
    if not details:
        raise HTTPException(status_code=404, detail="Kitchen access scope not found.")
    return {
        "scope": normalized,
        "setting_key": str(details["setting_key"]),
        "guest_session_scope": str(details["guest_session_scope"]),
        "label": str(details["label"]),
    }


def kitchen_guest_access_scope_for_request(method: str, path: str) -> str | None:
    if method not in {"GET", "HEAD"}:
        return None
    if path == "/api/retreat-plans" or path == "/api/recipes/full" or path.startswith("/api/retreat-plans/"):
        return KITCHEN_ACCESS_SCOPE_TESTING
    if (
        path in {"/api/service-snapshots/latest", "/api/ingredients", "/api/unit-conversions"}
        or path.startswith("/api/service-snapshots/by-plan/")
    ):
        return KITCHEN_ACCESS_SCOPE_RETREAT_VIEW
    return None


def load_optional_kitchen_guest_access(conn: Any, request: Request, scope_slug: str) -> bool:
    details = resolve_kitchen_access_scope(scope_slug)
    raw_token = request.cookies.get(guest_session_cookie_name(details["guest_session_scope"]))
    return authenticate_guest_session_token(conn, raw_token, scope=details["guest_session_scope"])


def set_kitchen_guest_access_cookie(response: Response, scope_slug: str, token: str) -> None:
    details = resolve_kitchen_access_scope(scope_slug)
    response.set_cookie(
        key=guest_session_cookie_name(details["guest_session_scope"]),
        value=token,
        httponly=True,
        secure=cookie_secure_enabled(),
        samesite="lax",
        path="/",
    )


def clear_kitchen_guest_access_cookie(response: Response, scope_slug: str) -> None:
    details = resolve_kitchen_access_scope(scope_slug)
    response.delete_cookie(guest_session_cookie_name(details["guest_session_scope"]), path="/")


@app.middleware("http")
async def api_auth_middleware(request: Request, call_next):
    path = request.url.path
    if not path.startswith("/api"):
        return await call_next(request)

    if (
        request.method == "OPTIONS"
        or path in PUBLIC_API_PATHS
        or any(path.startswith(prefix) for prefix in PUBLIC_API_PREFIXES)
    ):
        return await call_next(request)

    guest_scope_slug = kitchen_guest_access_scope_for_request(request.method, path)

    if request.method in {"GET", "HEAD"}:
        if path in PUBLIC_API_GET_PATHS or any(path.startswith(prefix) for prefix in PUBLIC_API_GET_PREFIXES):
            return await call_next(request)

    raw_token = request.cookies.get(SESSION_COOKIE_NAME)
    auth_detail = "Authentication required"

    with get_connection() as conn:
        if raw_token:
            user = authenticate_session_token(conn, raw_token)
            if user:
                conn.commit()
                request.state.auth_user = user
                return await call_next(request)
            auth_detail = "Session expired. Please log in again."

        if guest_scope_slug and load_optional_kitchen_guest_access(conn, request, guest_scope_slug):
            conn.commit()
            request.state.kitchen_guest_scope = guest_scope_slug
            return await call_next(request)

        conn.commit()

    return JSONResponse(status_code=401, content={"detail": auth_detail})


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/auth/bootstrap-status")
def auth_bootstrap_status() -> dict[str, Any]:
    with get_connection() as conn:
        configured = has_any_users(conn)
    return {
        "has_users": configured,
        "bootstrap_env_required": [BOOTSTRAP_ADMIN_USERNAME_ENV, BOOTSTRAP_ADMIN_PASSWORD_ENV],
    }


@app.post("/api/auth/login")
def login(payload: AuthLoginPayload, response: Response) -> dict[str, Any]:
    with get_connection() as conn:
        if not has_any_users(conn):
            raise HTTPException(
                status_code=503,
                detail=(
                    "No users configured. Set "
                    f"{BOOTSTRAP_ADMIN_USERNAME_ENV} and {BOOTSTRAP_ADMIN_PASSWORD_ENV}, "
                    "then restart the service."
                ),
            )
        user = authenticate_credentials(conn, payload.username, payload.password)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid username or password")
        token = create_session(conn, user.id)
        conn.commit()

    response.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=token,
        httponly=True,
        secure=cookie_secure_enabled(),
        samesite="lax",
        path="/",
    )

    return {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "default_path": default_route_for_role(user.role),
    }


@app.post("/api/auth/logout")
def logout(request: Request, response: Response) -> dict[str, str]:
    raw_token = request.cookies.get(SESSION_COOKIE_NAME)
    with get_connection() as conn:
        delete_session(conn, raw_token)
        conn.commit()

    response.delete_cookie(SESSION_COOKIE_NAME, path="/")
    return {"status": "ok"}


@app.get("/api/auth/me")
def auth_me(request: Request) -> dict[str, Any]:
    user = require_authenticated_user(request)
    return {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "is_active": user.is_active,
        "default_path": default_route_for_role(user.role),
    }


@app.get("/api/kitchen-access/{scope_slug}")
def kitchen_guest_access_status(
    scope_slug: str,
    request: Request,
    response: Response,
) -> dict[str, Any]:
    details = resolve_kitchen_access_scope(scope_slug)
    cookie_name = guest_session_cookie_name(details["guest_session_scope"])

    with get_connection() as conn:
        user = load_optional_session_user(conn, request)
        session_mode = "user" if user and user.role in {ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN} else "none"
        guest_authorized = False
        if session_mode == "none":
            raw_guest_token = request.cookies.get(cookie_name)
            guest_authorized = authenticate_guest_session_token(
                conn,
                raw_guest_token,
                scope=details["guest_session_scope"],
            )
            if raw_guest_token and not guest_authorized:
                clear_kitchen_guest_access_cookie(response, scope_slug)
            if guest_authorized:
                session_mode = "guest"
        state = resolve_kitchen_guest_access_state(conn, scope_slug)
        conn.commit()

    return {
        "scope": details["scope"],
        "label": details["label"],
        "authorized": session_mode in {"user", "guest"},
        "sessionMode": session_mode,
        "guestAccessEnabled": bool(state["guest_access_enabled"]),
        "guestSessionHours": guest_session_hours(),
    }


@app.post("/api/kitchen-access/{scope_slug}/login")
def kitchen_guest_access_login(
    scope_slug: str,
    payload: KitchenGuestAccessLoginPayload,
    request: Request,
    response: Response,
) -> dict[str, Any]:
    details = resolve_kitchen_access_scope(scope_slug)

    with get_connection() as conn:
        user = load_optional_session_user(conn, request)
        if user and user.role in {ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN}:
            conn.commit()
            return {
                "scope": details["scope"],
                "label": details["label"],
                "authorized": True,
                "sessionMode": "user",
                "guestSessionHours": guest_session_hours(),
            }

        configured_code = get_kitchen_guest_access_code(conn, scope_slug)
        if not configured_code:
            raise HTTPException(
                status_code=403,
                detail="Guest kitchen access is not configured. Ask an admin to set a shared code.",
            )

        provided_code = normalize_required_text(payload.accessCode, field_name="Access code")
        if not secrets.compare_digest(provided_code, configured_code):
            raise HTTPException(status_code=403, detail="Invalid access code.")

        token = create_guest_session(conn, details["guest_session_scope"])
        conn.commit()

    set_kitchen_guest_access_cookie(response, scope_slug, token)
    return {
        "scope": details["scope"],
        "label": details["label"],
        "authorized": True,
        "sessionMode": "guest",
        "guestSessionHours": guest_session_hours(),
    }


@app.post("/api/kitchen-access/{scope_slug}/logout")
def kitchen_guest_access_logout(scope_slug: str, request: Request, response: Response) -> dict[str, Any]:
    details = resolve_kitchen_access_scope(scope_slug)
    raw_token = request.cookies.get(guest_session_cookie_name(details["guest_session_scope"]))
    with get_connection() as conn:
        delete_guest_session(conn, raw_token, scope=details["guest_session_scope"])
        conn.commit()
    clear_kitchen_guest_access_cookie(response, scope_slug)
    return {"status": "ok", "scope": details["scope"]}


@app.post("/api/auth/change-password")
def auth_change_password(
    payload: AuthChangePasswordPayload,
    request: Request,
    user: Annotated[AuthUser, Depends(require_authenticated_user)],
) -> dict[str, str]:
    if payload.current_password == payload.new_password:
        raise HTTPException(status_code=400, detail="New password must be different from current password.")

    with get_connection() as conn:
        verified = authenticate_credentials(conn, user.username, payload.current_password)
        if not verified or verified.id != user.id:
            raise HTTPException(status_code=400, detail="Current password is incorrect.")

        try:
            update_user(conn, user.id, password=payload.new_password)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        # Keep the current browser session and revoke other sessions for this user.
        raw_token = request.cookies.get(SESSION_COOKIE_NAME)
        if raw_token:
            token_hash = hash_session_token(raw_token)
            conn.execute(
                "DELETE FROM auth_sessions WHERE user_id = ? AND token_hash != ?",
                (user.id, token_hash),
            )
        else:
            conn.execute("DELETE FROM auth_sessions WHERE user_id = ?", (user.id,))

        conn.commit()

    return {"status": "ok"}


@app.get("/api/auth/users")
def auth_list_users(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> list[dict[str, Any]]:
    with get_connection() as conn:
        return list_users(conn)


@app.post("/api/auth/users")
def auth_create_user(
    payload: AuthUserCreatePayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        try:
            created = create_user(
                conn,
                payload.username,
                payload.password,
                payload.role,
                is_active=payload.is_active,
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not create user: {exc}") from exc
        conn.commit()
    return {
        "id": created.id,
        "username": created.username,
        "role": created.role,
        "is_active": created.is_active,
    }


@app.put("/api/auth/users/{user_id}")
def auth_update_user(
    user_id: int,
    payload: AuthUserUpdatePayload,
    request: Request,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> Any:
    with get_connection() as conn:
        try:
            updated = update_user(
                conn,
                user_id,
                password=payload.password,
                role=payload.role,
                is_active=payload.is_active,
            )
            admins = conn.execute(
                "SELECT COUNT(*) AS admin_count FROM users WHERE role = 'admin' AND is_active = 1"
            ).fetchone()
            if admins and int(admins["admin_count"]) <= 0:
                conn.rollback()
                raise HTTPException(
                    status_code=400,
                    detail="Refusing to remove the last active admin user.",
                )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        conn.commit()

    current = get_request_user(request)
    if current and current.id == updated.id and not updated.is_active:
        response = JSONResponse(
            {
                "id": updated.id,
                "username": updated.username,
                "role": updated.role,
                "is_active": updated.is_active,
            }
        )
        response.delete_cookie(SESSION_COOKIE_NAME, path="/")
        return response

    return {
        "id": updated.id,
        "username": updated.username,
        "role": updated.role,
        "is_active": updated.is_active,
    }


@app.delete("/api/auth/users/{user_id}")
def auth_delete_user(
    user_id: int,
    request: Request,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> Any:
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id, username, role, is_active FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="User not found")

        if existing["role"] == ROLE_ADMIN and bool(existing["is_active"]):
            admins = conn.execute(
                "SELECT COUNT(*) AS admin_count FROM users WHERE role = 'admin' AND is_active = 1"
            ).fetchone()
            if admins and int(admins["admin_count"]) <= 1:
                raise HTTPException(
                    status_code=400,
                    detail="Refusing to remove the last active admin user.",
                )

        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()

    deleted_payload = {
        "id": int(existing["id"]),
        "username": existing["username"],
        "role": existing["role"],
        "is_active": bool(existing["is_active"]),
    }

    current = get_request_user(request)
    if current and current.id == int(existing["id"]):
        response = JSONResponse({**deleted_payload, "self_deleted": True})
        response.delete_cookie(SESSION_COOKIE_NAME, path="/")
        return response

    return {**deleted_payload, "self_deleted": False}


@app.get("/api/admin/inventory-withdraw-access")
def admin_get_inventory_withdraw_access(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        state = resolve_inventory_withdraw_access_state(conn)
        conn.commit()
    return {
        "accessCode": state.get("access_code"),
        "guestAccessEnabled": bool(state.get("guest_access_enabled")),
        "source": state.get("source"),
        "updatedAt": state.get("updated_at"),
        "updatedByUserId": state.get("updated_by_user_id"),
        "updatedByUsername": state.get("updated_by_username"),
    }


@app.put("/api/admin/inventory-withdraw-access")
@app.post("/api/admin/inventory-withdraw-access")
@app.patch("/api/admin/inventory-withdraw-access")
def admin_update_inventory_withdraw_access(
    payload: AdminInventoryWithdrawAccessPayload,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    normalized_code = normalize_optional_text(payload.accessCode)
    if normalized_code and len(normalized_code) < 4:
        raise HTTPException(status_code=400, detail="Access code must be at least 4 characters.")

    with get_connection() as conn:
        save_app_setting(
            conn,
            setting_key=APP_SETTING_INVENTORY_WITHDRAW_ACCESS_CODE,
            setting_value=normalized_code or "",
            updated_by_user_id=user.id,
        )
        conn.commit()
        state = resolve_inventory_withdraw_access_state(conn)

    return {
        "accessCode": state.get("access_code"),
        "guestAccessEnabled": bool(state.get("guest_access_enabled")),
        "source": state.get("source"),
        "updatedAt": state.get("updated_at"),
        "updatedByUserId": state.get("updated_by_user_id"),
        "updatedByUsername": state.get("updated_by_username"),
    }


@app.get("/api/admin/kitchen-access/{scope_slug}")
def admin_get_kitchen_access(
    scope_slug: str,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    details = resolve_kitchen_access_scope(scope_slug)
    with get_connection() as conn:
        state = resolve_kitchen_guest_access_state(conn, scope_slug)
        conn.commit()
    return {
        "scope": details["scope"],
        "label": details["label"],
        "accessCode": state.get("access_code"),
        "guestAccessEnabled": bool(state.get("guest_access_enabled")),
        "source": state.get("source"),
        "updatedAt": state.get("updated_at"),
        "updatedByUserId": state.get("updated_by_user_id"),
        "updatedByUsername": state.get("updated_by_username"),
        "guestSessionHours": guest_session_hours(),
    }


@app.put("/api/admin/kitchen-access/{scope_slug}")
@app.post("/api/admin/kitchen-access/{scope_slug}")
@app.patch("/api/admin/kitchen-access/{scope_slug}")
def admin_update_kitchen_access(
    scope_slug: str,
    payload: AdminSharedAccessCodePayload,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    details = resolve_kitchen_access_scope(scope_slug)
    normalized_code = normalize_optional_text(payload.accessCode)
    if normalized_code and len(normalized_code) < 4:
        raise HTTPException(status_code=400, detail="Access code must be at least 4 characters.")

    with get_connection() as conn:
        save_app_setting(
            conn,
            setting_key=details["setting_key"],
            setting_value=normalized_code or "",
            updated_by_user_id=user.id,
        )
        delete_guest_sessions_for_scope(conn, details["guest_session_scope"])
        conn.commit()
        state = resolve_kitchen_guest_access_state(conn, scope_slug)

    return {
        "scope": details["scope"],
        "label": details["label"],
        "accessCode": state.get("access_code"),
        "guestAccessEnabled": bool(state.get("guest_access_enabled")),
        "source": state.get("source"),
        "updatedAt": state.get("updated_at"),
        "updatedByUserId": state.get("updated_by_user_id"),
        "updatedByUsername": state.get("updated_by_username"),
        "guestSessionHours": guest_session_hours(),
    }


def resolve_openai_api_base() -> str:
    configured = str(os.getenv(OPENAI_API_BASE_ENV, DEFAULT_OPENAI_API_BASE) or "").strip()
    if not configured:
        return DEFAULT_OPENAI_API_BASE
    return configured.rstrip("/")


def read_project_env_value(key: str) -> str:
    env_path = Path(__file__).resolve().parents[2] / ".env"
    if not env_path.exists():
        return ""
    prefix = f"{key}="
    try:
        for raw_line in env_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or not line.startswith(prefix):
                continue
            return line.split("=", 1)[1].strip().strip('"').strip("'")
    except Exception:
        return ""
    return ""


def resolve_ingredient_duplicate_model(requested: str | None) -> str:
    candidate = str(requested or "").strip()
    if candidate:
        return candidate
    configured = str(os.getenv(OPENAI_INGREDIENT_DUP_MODEL_ENV, "") or "").strip()
    if not configured:
        configured = read_project_env_value(OPENAI_INGREDIENT_DUP_MODEL_ENV)
    if configured:
        return configured
    return DEFAULT_INGREDIENT_DUP_MODEL


def parse_float_confidence(value: object, default: float = 0.5) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        parsed = default
    if parsed < 0:
        return 0.0
    if parsed > 1:
        return 1.0
    return parsed


def normalize_duplicate_group_type(raw_type: object, confidence: float) -> str:
    value = str(raw_type or "").strip().lower().replace("-", "_").replace(" ", "_")
    strict_values = {"strict", "strict_duplicate", "exact_duplicate", "duplicate"}
    possible_values = {"possible", "possible_consolidation", "review", "candidate"}
    if value in strict_values:
        return "strict_duplicate"
    if value in possible_values:
        return "possible_consolidation"
    return "strict_duplicate" if confidence >= 0.82 else "possible_consolidation"


def normalize_llm_duplicate_groups(
    raw_groups: object,
    ingredient_by_id: dict[int, dict[str, Any]],
    max_groups: int,
) -> dict[str, list[dict[str, Any]]]:
    if not isinstance(raw_groups, list):
        return {
            "strict_duplicates": [],
            "possible_consolidations": [],
        }

    provisional: list[dict[str, Any]] = []
    seen_signatures: set[tuple[int, ...]] = set()

    for entry in raw_groups:
        if not isinstance(entry, dict):
            continue
        raw_member_ids = entry.get("member_ingredient_ids")
        if not isinstance(raw_member_ids, list):
            continue

        member_ids: list[int] = []
        for candidate in raw_member_ids:
            try:
                member_id = int(candidate)
            except (TypeError, ValueError):
                continue
            if member_id not in ingredient_by_id:
                continue
            if member_id not in member_ids:
                member_ids.append(member_id)
        if len(member_ids) < 2:
            continue

        canonical_raw = entry.get("canonical_ingredient_id")
        try:
            canonical_id = int(canonical_raw)
        except (TypeError, ValueError):
            canonical_id = member_ids[0]
        if canonical_id not in ingredient_by_id:
            canonical_id = member_ids[0]
        if canonical_id not in member_ids:
            member_ids = [canonical_id, *member_ids]
            member_ids = list(dict.fromkeys(member_ids))

        signature = tuple(sorted(member_ids))
        if signature in seen_signatures:
            continue
        seen_signatures.add(signature)

        reason_raw = entry.get("reason")
        reason = str(reason_raw).strip() if reason_raw is not None else ""
        confidence = parse_float_confidence(entry.get("confidence"), default=0.5)
        group_type = normalize_duplicate_group_type(entry.get("group_type"), confidence)
        provisional.append(
            {
                "canonical_id": canonical_id,
                "member_ids": member_ids,
                "confidence": confidence,
                "reason": reason,
                "group_type": group_type,
            }
        )

    strict_provisional = [
        group for group in provisional if group["group_type"] == "strict_duplicate"
    ]
    possible_provisional = [
        group for group in provisional if group["group_type"] == "possible_consolidation"
    ]
    strict_provisional.sort(key=lambda group: (-group["confidence"], -len(group["member_ids"]), group["canonical_id"]))
    possible_provisional.sort(key=lambda group: (-group["confidence"], -len(group["member_ids"]), group["canonical_id"]))

    used_member_ids: set[int] = set()
    strict_duplicates: list[dict[str, Any]] = []
    possible_consolidations: list[dict[str, Any]] = []

    def materialize(group: dict[str, Any], *, bucket: str) -> None:
        canonical = ingredient_by_id[group["canonical_id"]]
        members = [ingredient_by_id[member_id] for member_id in group["member_ids"]]
        payload = {
            "canonical": {
                "id": int(canonical["id"]),
                "name": canonical["name"],
            },
            "confidence": round(float(group["confidence"]), 3),
            "reason": group["reason"],
            "members": members,
            "group_type": bucket,
        }
        if bucket == "strict_duplicate":
            strict_duplicates.append(payload)
        else:
            possible_consolidations.append(payload)
        used_member_ids.update(group["member_ids"])

    for group in strict_provisional:
        if any(member_id in used_member_ids for member_id in group["member_ids"]):
            continue
        if len(strict_duplicates) + len(possible_consolidations) >= max_groups:
            break
        materialize(group, bucket="strict_duplicate")

    for group in possible_provisional:
        if any(member_id in used_member_ids for member_id in group["member_ids"]):
            continue
        if len(strict_duplicates) + len(possible_consolidations) >= max_groups:
            break
        materialize(group, bucket="possible_consolidation")

    return {
        "strict_duplicates": strict_duplicates,
        "possible_consolidations": possible_consolidations,
    }


def call_openai_for_ingredient_duplicates(
    *,
    model: str,
    ingredients: list[dict[str, Any]],
) -> dict[str, Any]:
    api_key = str(os.getenv(OPENAI_API_KEY_ENV, "") or "").strip()
    if not api_key:
        api_key = read_project_env_value(OPENAI_API_KEY_ENV)
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail=f"{OPENAI_API_KEY_ENV} is not configured on the backend.",
        )

    api_base = resolve_openai_api_base()
    endpoint = f"{api_base}/chat/completions"
    system_prompt = (
        "You are a strict ingredient-normalization assistant for a recipe database. "
        "Find likely duplicate ingredients using semantic understanding, including Hindi/English synonyms "
        "and transliteration variants (example: rajma vs kidney beans). "
        "Return two types of groups: strict duplicates and possible consolidations for manual review. "
        "A strict duplicate means the same ingredient with naming variants. "
        "A possible consolidation means likely same ingredient family/variant that may be merged with human review "
        "(example: spring onion vs green onion). "
        "Return JSON only with this exact shape: "
        "{"
        "\"groups\":["
        "{"
        "\"canonical_ingredient_id\":number,"
        "\"member_ingredient_ids\":[number,number],"
        "\"group_type\":\"strict_duplicate|possible_consolidation\","
        "\"confidence\":number,"
        "\"reason\":string"
        "}"
        "]"
        "}"
    )
    user_prompt = {
        "task": "Find duplicate ingredient groups.",
        "constraints": [
            "Each group must represent the same real-world ingredient.",
            "Each group must contain at least 2 IDs from the input list.",
            "Use the existing ingredient IDs only.",
            "Pick canonical_ingredient_id from member_ingredient_ids.",
            "Use group_type=strict_duplicate for clear exact synonyms/aliases only.",
            "Use group_type=possible_consolidation for likely merge candidates that need review.",
        ],
        "ingredients": ingredients,
    }
    payload = {
        "model": model,
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(user_prompt, ensure_ascii=True)},
        ],
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib_request.Request(
        endpoint,
        data=data,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib_request.urlopen(req, timeout=90) as response:
            raw = response.read().decode("utf-8")
    except urllib_error.HTTPError as exc:
        body = ""
        try:
            body = exc.read().decode("utf-8")
        except Exception:
            body = ""
        detail = f"LLM request failed ({exc.code})."
        if body:
            detail = f"{detail} {body[:500]}"
        raise HTTPException(status_code=502, detail=detail) from exc
    except urllib_error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"Could not reach LLM service: {exc.reason}") from exc

    try:
        parsed = json.loads(raw)
        content = parsed["choices"][0]["message"]["content"]
        if not isinstance(content, str) or not content.strip():
            raise ValueError("Missing model content")
        return json.loads(content)
    except Exception as exc:
        raise HTTPException(status_code=502, detail="Invalid LLM duplicate response format.") from exc


@app.get("/api/ingredients")
def list_ingredients() -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT id, name, category, purchase_tier, canonical_unit, grams_per_cup, notes FROM ingredients ORDER BY name"
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/ingredients/find-duplicates")
def find_ingredient_duplicates(
    payload: IngredientDuplicateScanPayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    model = resolve_ingredient_duplicate_model(payload.model)
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                i.id,
                i.name,
                i.category,
                i.purchase_tier,
                i.canonical_unit,
                COUNT(ri.id) AS recipe_usage_count
            FROM ingredients i
            LEFT JOIN recipe_ingredients ri ON ri.ingredient_id = i.id
            GROUP BY i.id
            ORDER BY lower(i.name), i.id
            """
        ).fetchall()

    ingredients = [
        {
            "id": int(row["id"]),
            "name": row["name"],
            "category": row["category"],
            "purchase_tier": row["purchase_tier"],
            "canonical_unit": row["canonical_unit"],
            "recipe_usage_count": int(row["recipe_usage_count"] or 0),
        }
        for row in rows
    ]
    if len(ingredients) < 2:
        return {
            "model": model,
            "ingredient_count": len(ingredients),
            "groups": [],
        }

    llm_response = call_openai_for_ingredient_duplicates(
        model=model,
        ingredients=ingredients,
    )
    ingredient_by_id = {int(item["id"]): item for item in ingredients}
    grouped = normalize_llm_duplicate_groups(
        llm_response.get("groups"),
        ingredient_by_id=ingredient_by_id,
        max_groups=payload.max_groups,
    )
    strict_duplicates = grouped["strict_duplicates"]
    possible_consolidations = grouped["possible_consolidations"]
    return {
        "model": model,
        "ingredient_count": len(ingredients),
        "strict_duplicate_count": len(strict_duplicates),
        "possible_consolidation_count": len(possible_consolidations),
        "group_count": len(strict_duplicates) + len(possible_consolidations),
        "strict_duplicates": strict_duplicates,
        "possible_consolidations": possible_consolidations,
        # Backward compatibility for older frontend handlers.
        "groups": [*strict_duplicates, *possible_consolidations],
    }


PURCHASE_TIERS = ["bulk", "fresh", "daily"]


@app.get("/api/purchase-tiers")
def list_purchase_tiers() -> list[str]:
    return PURCHASE_TIERS


@app.get("/api/shopping-phases")
def list_shopping_phases() -> list[str]:
    return SHOPPING_PHASES


@app.get("/api/vendors")
def list_vendors() -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT id, name, notes FROM vendors ORDER BY name"
        ).fetchall()
    return [dict(row) for row in rows]


@app.get("/api/shopping-lists")
def list_shopping_lists(
    retreatPlanId: int | None = Query(default=None, ge=1),
    phase: str | None = Query(default=None),
) -> list[dict[str, Any]]:
    filters: list[str] = []
    params: list[Any] = []
    if retreatPlanId is not None:
        filters.append("sl.retreat_plan_id = ?")
        params.append(retreatPlanId)
    if phase and phase.strip():
        filters.append("lower(sl.phase) = lower(?)")
        params.append(phase.strip())

    where_sql = f"WHERE {' AND '.join(filters)}" if filters else ""
    shopping_list_items_source = (
        "shopping_list_items sli NOT INDEXED"
        if using_sqlite_backend()
        else "shopping_list_items sli"
    )

    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT
                sl.id,
                sl.name,
                sl.phase,
                sl.status,
                sl.created_at,
                sl.retreat_plan_id,
                rp.name AS retreat_plan_name,
                COUNT(sli.id) AS item_count,
                COALESCE(SUM(CASE WHEN COALESCE(sli.ordered, 0) = 1 THEN 1 ELSE 0 END), 0) AS ordered_count,
                COALESCE(SUM(CASE WHEN COALESCE(sli.received, 0) = 1 THEN 1 ELSE 0 END), 0) AS received_count
            FROM shopping_lists sl
            LEFT JOIN retreat_plans rp ON rp.id = sl.retreat_plan_id
            LEFT JOIN {shopping_list_items_source} ON sli.shopping_list_id = sl.id
            {where_sql}
            GROUP BY sl.id
            ORDER BY sl.id DESC
            """,
            tuple(params),
        ).fetchall()

    return [
        {
            "id": int(row["id"]),
            "name": row["name"],
            "phase": row["phase"] or "bulk",
            "status": row["status"] or "draft",
            "created_at": row["created_at"],
            "retreat_plan_id": int(row["retreat_plan_id"]) if row["retreat_plan_id"] is not None else None,
            "retreat_plan_name": row["retreat_plan_name"],
            "item_count": int(row["item_count"] or 0),
            "ordered_count": int(row["ordered_count"] or 0),
            "received_count": int(row["received_count"] or 0),
        }
        for row in rows
    ]


@app.get("/api/shopping-lists/{shopping_list_id}")
def get_shopping_list(shopping_list_id: int) -> dict[str, Any]:
    with get_connection() as conn:
        detail = load_shopping_list_detail(conn, shopping_list_id)
    return detail


@app.get("/api/shopping-lists/{shopping_list_id}/pickup-lists")
def list_shopping_pickup_lists(shopping_list_id: int) -> list[dict[str, Any]]:
    with get_connection() as conn:
        ensure_shopping_list_exists(conn, shopping_list_id)
        return load_shopping_pickup_lists_for_shopping_list(conn, shopping_list_id)


@app.post("/api/shopping-lists/{shopping_list_id}/pickup-lists")
def create_shopping_pickup_list(
    shopping_list_id: int,
    payload: ShoppingPickupListCreatePayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    requested_name = normalize_optional_text(payload.name)
    assignee = normalize_optional_text(payload.assignee)
    pickup_date = normalize_optional_text(payload.pickupDate)
    notes = normalize_optional_text(payload.notes)
    with get_connection() as conn:
        detail = create_saved_shopping_pickup_list(
            conn,
            shopping_list_id=shopping_list_id,
            item_ids=payload.itemIds,
            requested_name=requested_name,
            vendor_id=payload.vendorId,
            assignee=assignee,
            pickup_date=pickup_date,
            notes=notes,
        )
        conn.commit()
    return detail


@app.get("/api/shopping-pickup-lists/{pickup_list_id}")
def get_shopping_pickup_list(pickup_list_id: int) -> dict[str, Any]:
    with get_connection() as conn:
        return load_shopping_pickup_list_detail(conn, pickup_list_id)


@app.patch("/api/shopping-pickup-lists/{pickup_list_id}")
def update_shopping_pickup_list(
    pickup_list_id: int,
    payload: ShoppingPickupListUpdatePayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    fields = payload.model_fields_set
    if not fields:
        raise HTTPException(status_code=400, detail="No fields supplied")

    with get_connection() as conn:
        pickup_list = ensure_shopping_pickup_list_exists(conn, pickup_list_id)
        updates: list[str] = []
        params: list[Any] = []

        if "name" in fields:
            requested_name = normalize_required_text(payload.name, field_name="name")
            next_name = unique_shopping_pickup_list_name(
                conn,
                int(pickup_list["shopping_list_id"]),
                requested_name,
                exclude_pickup_list_id=pickup_list_id,
            )
            updates.append("name = ?")
            params.append(next_name)

        if "vendorId" in fields:
            vendor_id = payload.vendorId
            if vendor_id is not None:
                ensure_vendor_exists(conn, vendor_id)
            updates.append("vendor_id = ?")
            params.append(vendor_id)

        if "assignee" in fields:
            updates.append("assignee = ?")
            params.append(normalize_optional_text(payload.assignee))

        if "pickupDate" in fields:
            updates.append("pickup_date = ?")
            params.append(normalize_optional_text(payload.pickupDate))

        if "notes" in fields:
            updates.append("notes = ?")
            params.append(normalize_optional_text(payload.notes))

        updates.append("updated_at = CURRENT_TIMESTAMP")
        params.append(pickup_list_id)
        conn.execute(
            f"""
            UPDATE shopping_pickup_lists
            SET {", ".join(updates)}
            WHERE id = ?
            """,
            tuple(params),
        )
        detail = load_shopping_pickup_list_detail(conn, pickup_list_id)
        conn.commit()
    return detail


@app.delete("/api/shopping-pickup-lists/{pickup_list_id}")
def delete_shopping_pickup_list(
    pickup_list_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        pickup_list = ensure_shopping_pickup_list_exists(conn, pickup_list_id)
        conn.execute("DELETE FROM shopping_pickup_lists WHERE id = ?", (pickup_list_id,))
        conn.commit()
    return {
        "id": int(pickup_list["id"]),
        "shopping_list_id": int(pickup_list["shopping_list_id"]),
        "name": pickup_list["name"],
        "status": "deleted",
    }


@app.patch("/api/shopping-lists/{shopping_list_id}")
def rename_shopping_list(
    shopping_list_id: int,
    payload: ShoppingListUpdatePayload,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    requested_name = " ".join(payload.name.strip().split())
    if not requested_name:
        raise HTTPException(status_code=400, detail="Shopping list name cannot be blank")

    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id, name
            FROM shopping_lists
            WHERE id = ?
            """,
            (shopping_list_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Shopping list not found")

        final_name = unique_shopping_list_name(
            conn,
            requested_name,
            exclude_shopping_list_id=shopping_list_id,
        )
        conn.execute(
            "UPDATE shopping_lists SET name = ? WHERE id = ?",
            (final_name, shopping_list_id),
        )
        sync_food_inventory_orders_for_shopping_list(conn, shopping_list_id, actor_user_id=user.id)
        detail = load_shopping_list_detail(conn, shopping_list_id)
        conn.commit()

    detail["renamed_from"] = existing["name"]
    return detail


@app.delete("/api/shopping-lists/{shopping_list_id}")
def delete_shopping_list(
    shopping_list_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id, name FROM shopping_lists WHERE id = ?",
            (shopping_list_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Shopping list not found")

        archive_food_inventory_orders_for_shopping_list(conn, shopping_list_id)
        conn.execute("DELETE FROM shopping_lists WHERE id = ?", (shopping_list_id,))
        conn.commit()

    return {
        "id": int(existing["id"]),
        "name": existing["name"],
        "status": "deleted",
    }


class ShoppingListCreatePayload(BaseModel):
    name: str = ""
    listDate: str = ""


@app.post("/api/shopping-lists")
def create_manual_shopping_list(
    payload: ShoppingListCreatePayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    list_date_text = (payload.listDate or "").strip()
    if list_date_text:
        try:
            list_date = date.fromisoformat(list_date_text).isoformat()
        except ValueError:
            raise HTTPException(status_code=400, detail="List date must be formatted YYYY-MM-DD.")
    else:
        list_date = date.today().isoformat()

    label = (payload.name or "").strip() or f"Manual List {list_date}"

    with get_connection() as conn:
        label = unique_shopping_list_name(conn, label)
        created = conn.execute(
            """
            INSERT INTO shopping_lists(name, phase, status)
            VALUES (?, 'custom', 'draft')
            RETURNING id
            """,
            (label,),
        ).fetchone()
        shopping_list_id = int(created["id"])
        conn.commit()
        detail = load_shopping_list_detail(conn, shopping_list_id)
    return detail


class ShoppingListItemAddPayload(BaseModel):
    ingredientName: str = Field(min_length=1)
    qty: float = Field(gt=0)
    unit: str = ""


@app.post("/api/shopping-lists/{shopping_list_id}/items")
def add_shopping_list_item(
    shopping_list_id: int,
    payload: ShoppingListItemAddPayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        list_row = conn.execute(
            "SELECT id FROM shopping_lists WHERE id = ?",
            (shopping_list_id,),
        ).fetchone()
        if not list_row:
            raise HTTPException(status_code=404, detail="Shopping list not found")

        ingredient_name = payload.ingredientName.strip()
        unit = normalize_unit((payload.unit or "").strip()) or "each"
        qty = round(float(payload.qty), 4)

        ingredient = conn.execute(
            "SELECT id, name FROM ingredients WHERE lower(name) = lower(?)",
            (ingredient_name,),
        ).fetchone()
        if ingredient is None:
            alias = conn.execute(
                "SELECT ingredient_name FROM ingredient_aliases WHERE lower(alias_name) = lower(?)",
                (ingredient_name,),
            ).fetchone()
            if alias:
                ingredient = conn.execute(
                    "SELECT id, name FROM ingredients WHERE lower(name) = lower(?)",
                    (str(alias["ingredient_name"]),),
                ).fetchone()
        if ingredient is None:
            if unit in MASS_TO_G:
                canonical_unit = "g"
            elif unit in VOLUME_TO_ML:
                canonical_unit = "ml"
            else:
                canonical_unit = unit
            ingredient = conn.execute(
                "INSERT INTO ingredients(name, canonical_unit) VALUES (?, ?) RETURNING id, name",
                (ingredient_name, canonical_unit),
            ).fetchone()

        ingredient_id = int(ingredient["id"])
        existing_item = conn.execute(
            """
            SELECT id, required_qty, required_unit, in_stock_qty
            FROM shopping_list_items
            WHERE shopping_list_id = ? AND ingredient_id = ? AND lower(required_unit) = lower(?)
            """,
            (shopping_list_id, ingredient_id, unit),
        ).fetchone()

        if existing_item:
            required_qty = round(float(existing_item["required_qty"] or 0.0) + qty, 4)
            in_stock_qty = float(existing_item["in_stock_qty"] or 0.0)
            to_buy_qty = round(max(required_qty - in_stock_qty, 0.0), 4)
            conn.execute(
                "UPDATE shopping_list_items SET required_qty = ?, to_buy_qty = ?, to_buy_unit = required_unit WHERE id = ?",
                (required_qty, to_buy_qty, int(existing_item["id"])),
            )
        else:
            conn.execute(
                """
                INSERT INTO shopping_list_items(
                    shopping_list_id, ingredient_id, required_qty, required_unit,
                    in_stock_qty, in_stock_unit, to_buy_qty, to_buy_unit,
                    ordered, received, status
                )
                VALUES (?, ?, ?, ?, 0, ?, ?, ?, 0, 0, 'open')
                """,
                (shopping_list_id, ingredient_id, qty, unit, unit, qty, unit),
            )

        refresh_shopping_list_status(conn, shopping_list_id)
        conn.commit()
        detail = load_shopping_list_detail(conn, shopping_list_id)
    return detail


@app.delete("/api/shopping-lists/{shopping_list_id}/items/{item_id}")
def delete_shopping_list_item(
    shopping_list_id: int,
    item_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        item_row = conn.execute(
            "SELECT id FROM shopping_list_items WHERE id = ? AND shopping_list_id = ?",
            (item_id, shopping_list_id),
        ).fetchone()
        if not item_row:
            raise HTTPException(status_code=404, detail="Shopping list item not found")
        conn.execute("DELETE FROM shopping_list_items WHERE id = ?", (item_id,))
        refresh_shopping_list_status(conn, shopping_list_id)
        conn.commit()
        detail = load_shopping_list_detail(conn, shopping_list_id)
    return detail


@app.post("/api/shopping-lists/generate")
def generate_shopping_list(
    payload: ShoppingListGeneratePayload,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        detail = materialize_shopping_list(
            conn,
            payload,
            fixed_name=payload.name,
            allow_empty_result=False,
        )
        sync_food_inventory_orders_for_shopping_list(conn, int(detail["id"]), actor_user_id=user.id)
        detail = load_shopping_list_detail(conn, int(detail["id"]))
        conn.commit()
    return detail


@app.post("/api/shopping-lists/{shopping_list_id}/refresh")
def refresh_shopping_list(
    shopping_list_id: int,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        list_row = conn.execute(
            """
            SELECT id, name, phase, generation_config_json
            FROM shopping_lists
            WHERE id = ?
            """,
            (shopping_list_id,),
        ).fetchone()
        if not list_row:
            raise HTTPException(status_code=404, detail="Shopping list not found")
        if str(list_row["phase"] or "").strip().lower() == "custom" and not list_row["generation_config_json"]:
            raise HTTPException(
                status_code=400,
                detail="This is a manual list with no retreat plan behind it; there is nothing to refresh from.",
            )

        payload = infer_shopping_list_generation_payload(conn, shopping_list_id)
        detail = materialize_shopping_list(
            conn,
            payload,
            shopping_list_id=shopping_list_id,
            fixed_name=str(list_row["name"] or "").strip() or None,
            preserve_existing_metadata=True,
            allow_empty_result=True,
        )
        sync_food_inventory_orders_for_shopping_list(conn, shopping_list_id, actor_user_id=user.id)
        detail = load_shopping_list_detail(conn, shopping_list_id)
        conn.commit()

    detail["refreshed"] = True
    return detail


@app.post("/api/shopping-lists/{shopping_list_id}/carry-forward")
def carry_forward_shopping_list(
    shopping_list_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
    payload: ShoppingListCarryForwardPayload | None = None,
) -> dict[str, Any]:
    name_override = payload.name if payload else None
    phase_override = payload.phase if payload else None
    with get_connection() as conn:
        detail = create_carry_forward_shopping_list(
            conn,
            source_list_id=shopping_list_id,
            name_override=name_override,
            phase_override=phase_override,
        )
        conn.commit()
    return detail


@app.post("/api/shopping-lists/{shopping_list_id}/apply-inventory")
def apply_shopping_list_inventory(
    shopping_list_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        list_row = conn.execute(
            """
            SELECT id, name, phase
            FROM shopping_lists
            WHERE id = ?
            """,
            (shopping_list_id,),
        ).fetchone()
        if not list_row:
            raise HTTPException(status_code=404, detail="Shopping list not found")

        phase = str(list_row["phase"] or "").strip().lower()
        if phase not in {"fresh", "daily"}:
            raise HTTPException(
                status_code=400,
                detail="Apply to inventory is enabled only for fresh and daily shopping lists.",
            )

        rows = conn.execute(
            """
            SELECT ingredient_id, in_stock_qty, in_stock_unit, required_unit
            FROM shopping_list_items
            WHERE shopping_list_id = ?
            """,
            (shopping_list_id,),
        ).fetchall()

        ingredient_ids = sorted({int(row["ingredient_id"]) for row in rows})
        if ingredient_ids:
            placeholders = ",".join("?" for _ in ingredient_ids)
            conn.execute(
                f"DELETE FROM inventory_items WHERE source = ? AND ingredient_id IN ({placeholders})",
                (MANUAL_INVENTORY_SOURCE, *ingredient_ids),
            )

        applied_count = 0
        for row in rows:
            qty = float(row["in_stock_qty"] or 0.0)
            unit = normalize_unit(str(row["in_stock_unit"] or row["required_unit"] or "").strip())
            if qty <= 0 or not unit:
                continue
            conn.execute(
                """
                INSERT INTO inventory_items(ingredient_id, quantity, unit, source, updated_at)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                """,
                (
                    int(row["ingredient_id"]),
                    round(qty, 4),
                    unit,
                    MANUAL_INVENTORY_SOURCE,
                ),
            )
            applied_count += 1

        conn.commit()

    return {
        "status": "ok",
        "shopping_list_id": shopping_list_id,
        "shopping_list_name": list_row["name"],
        "phase": phase,
        "inventory_source": MANUAL_INVENTORY_SOURCE,
        "applied_count": applied_count,
    }


KITCHEN_INVENTORY_NAME_HEADERS = {"ingredient", "ingredient name", "item", "item name", "name", "product"}
KITCHEN_INVENTORY_QTY_HEADERS = {"qty", "quantity", "count", "amount", "on hand", "on-hand", "stock", "storage", "inventory"}
KITCHEN_INVENTORY_UNIT_HEADERS = {"unit", "units", "uom", "measure"}


def load_kitchen_inventory_upload_grid(filename: str, content: bytes) -> list[list[Any]]:
    lowered = (filename or "").lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}")
        try:
            sheet = workbook.active
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("latin-1")
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not decode file as text: {exc}")
    return [list(row) for row in csv.reader(io.StringIO(text))]


def detect_kitchen_inventory_columns(rows: list[list[Any]]) -> tuple[int, int, int | None, int]:
    """Find (name_col, qty_col, unit_col, first_data_row) from a header row, else assume columns A/B/C."""
    for index, row in enumerate(rows[:10]):
        cells = [str(cell if cell is not None else "").strip().lower() for cell in row]
        name_col = qty_col = unit_col = None
        for col, cell in enumerate(cells):
            if name_col is None and cell in KITCHEN_INVENTORY_NAME_HEADERS:
                name_col = col
            elif qty_col is None and cell in KITCHEN_INVENTORY_QTY_HEADERS:
                qty_col = col
            elif unit_col is None and cell in KITCHEN_INVENTORY_UNIT_HEADERS:
                unit_col = col
        if name_col is not None and qty_col is not None:
            return name_col, qty_col, unit_col, index + 1
    return 0, 1, 2, 0


def load_kitchen_inventory_list_detail(conn: Any, list_id: int) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT id, name, inventory_date, source_filename, notes, created_at
        FROM kitchen_inventory_lists
        WHERE id = ?
        """,
        (list_id,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Kitchen inventory list not found")

    item_rows = conn.execute(
        """
        SELECT
            kii.id,
            kii.ingredient_id,
            i.name AS ingredient_name,
            kii.input_name,
            kii.input_qty,
            kii.input_unit,
            kii.canonical_qty,
            kii.canonical_unit,
            kii.conversion_note
        FROM kitchen_inventory_list_items kii
        LEFT JOIN ingredients i ON i.id = kii.ingredient_id
        WHERE kii.list_id = ?
        ORDER BY lower(kii.input_name), kii.id
        """,
        (list_id,),
    ).fetchall()

    items = [
        {
            "id": int(item["id"]),
            "ingredient_id": int(item["ingredient_id"]) if item["ingredient_id"] is not None else None,
            "ingredient_name": item["ingredient_name"],
            "input_name": item["input_name"],
            "input_qty": item["input_qty"],
            "input_unit": item["input_unit"],
            "canonical_qty": item["canonical_qty"],
            "canonical_unit": item["canonical_unit"],
            "conversion_note": item["conversion_note"],
        }
        for item in item_rows
    ]
    matched_count = sum(1 for item in items if item["ingredient_id"] is not None)
    converted_count = sum(1 for item in items if item["canonical_qty"] is not None)
    return {
        "id": int(row["id"]),
        "name": row["name"],
        "inventory_date": row["inventory_date"],
        "source_filename": row["source_filename"],
        "notes": row["notes"],
        "created_at": row["created_at"],
        "item_count": len(items),
        "matched_count": matched_count,
        "unmatched_count": len(items) - matched_count,
        "converted_count": converted_count,
        "items": items,
    }


@app.post("/api/kitchen-inventory/upload")
async def upload_kitchen_inventory_list(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
    file: Annotated[UploadFile, File()],
    name: Annotated[str, Form()] = "",
    inventoryDate: Annotated[str, Form()] = "",
    notes: Annotated[str, Form()] = "",
) -> dict[str, Any]:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    rows = load_kitchen_inventory_upload_grid(file.filename or "", content)
    name_col, qty_col, unit_col, data_start = detect_kitchen_inventory_columns(rows)

    inventory_date_text = (inventoryDate or "").strip()
    if inventory_date_text:
        try:
            inventory_date = date.fromisoformat(inventory_date_text).isoformat()
        except ValueError:
            raise HTTPException(status_code=400, detail="Inventory date must be formatted YYYY-MM-DD.")
    else:
        inventory_date = date.today().isoformat()

    list_name = (name or "").strip() or f"Kitchen Inventory {inventory_date}"

    with get_connection() as conn:
        ingredient_rows = conn.execute("SELECT id, name, canonical_unit FROM ingredients").fetchall()
        ingredient_by_name = {str(row["name"] or "").strip().lower(): row for row in ingredient_rows}
        alias_rows = conn.execute("SELECT ingredient_name, alias_name FROM ingredient_aliases").fetchall()
        alias_to_ingredient_name = {
            str(row["alias_name"] or "").strip().lower(): str(row["ingredient_name"] or "").strip().lower()
            for row in alias_rows
        }

        parsed_items: list[dict[str, Any]] = []
        skipped_rows: list[dict[str, Any]] = []
        for row_index in range(data_start, len(rows)):
            row = rows[row_index]

            def cell_text(col: int | None) -> str:
                if col is None or col >= len(row) or row[col] is None:
                    return ""
                return str(row[col]).strip()

            raw_name = cell_text(name_col)
            raw_qty = cell_text(qty_col)
            raw_unit = cell_text(unit_col)
            if not raw_name and not raw_qty:
                continue
            if not raw_name:
                skipped_rows.append({"row": row_index + 1, "reason": "Missing ingredient name."})
                continue
            try:
                qty = float(raw_qty.replace(",", ""))
            except ValueError:
                skipped_rows.append({"row": row_index + 1, "reason": f"Quantity '{raw_qty}' is not a number."})
                continue
            if qty < 0:
                skipped_rows.append({"row": row_index + 1, "reason": "Quantity is negative."})
                continue

            lookup_key = raw_name.lower()
            ingredient = ingredient_by_name.get(lookup_key)
            if ingredient is None:
                alias_target = alias_to_ingredient_name.get(lookup_key)
                if alias_target:
                    ingredient = ingredient_by_name.get(alias_target)

            unit = normalize_unit(raw_unit) if raw_unit else ""
            canonical_qty: float | None = None
            canonical_unit: str | None = None
            note: str | None = None
            if ingredient is None:
                note = "No matching ingredient in the catalog."
            else:
                if not unit:
                    unit = normalize_unit(str(ingredient["canonical_unit"] or "").strip())
                if not unit:
                    note = "Missing unit and the ingredient has no canonical unit."
                else:
                    canonical_qty, canonical_unit, note = to_canonical(str(ingredient["name"]), qty, unit)
                    if canonical_qty is not None:
                        canonical_qty = round(float(canonical_qty), 4)

            parsed_items.append(
                {
                    "ingredient_id": int(ingredient["id"]) if ingredient is not None else None,
                    "input_name": raw_name,
                    "input_qty": qty,
                    "input_unit": unit or (raw_unit or None),
                    "canonical_qty": canonical_qty,
                    "canonical_unit": canonical_unit,
                    "conversion_note": note,
                }
            )

        if not parsed_items:
            raise HTTPException(
                status_code=400,
                detail="No inventory rows found. Expected columns: ingredient name, quantity, unit.",
            )

        created = conn.execute(
            """
            INSERT INTO kitchen_inventory_lists(name, inventory_date, source_filename, notes)
            VALUES (?, ?, ?, ?)
            RETURNING id
            """,
            (list_name, inventory_date, file.filename or None, (notes or "").strip() or None),
        ).fetchone()
        list_id = int(created["id"])
        for item in parsed_items:
            conn.execute(
                """
                INSERT INTO kitchen_inventory_list_items(
                    list_id, ingredient_id, input_name, input_qty, input_unit,
                    canonical_qty, canonical_unit, conversion_note
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    list_id,
                    item["ingredient_id"],
                    item["input_name"],
                    item["input_qty"],
                    item["input_unit"],
                    item["canonical_qty"],
                    item["canonical_unit"],
                    item["conversion_note"],
                ),
            )
        conn.commit()
        detail = load_kitchen_inventory_list_detail(conn, list_id)

    detail["skipped_rows"] = skipped_rows
    return detail


class KitchenInventoryItemInput(BaseModel):
    ingredientId: int = Field(gt=0)
    qty: float = Field(ge=0)
    unit: str = ""


class KitchenInventoryCreatePayload(BaseModel):
    name: str = ""
    inventoryDate: str = ""
    notes: str = ""
    items: list[KitchenInventoryItemInput] = Field(min_length=1)


@app.post("/api/kitchen-inventory")
def create_kitchen_inventory_list(
    payload: KitchenInventoryCreatePayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    inventory_date_text = (payload.inventoryDate or "").strip()
    if inventory_date_text:
        try:
            inventory_date = date.fromisoformat(inventory_date_text).isoformat()
        except ValueError:
            raise HTTPException(status_code=400, detail="Inventory date must be formatted YYYY-MM-DD.")
    else:
        inventory_date = date.today().isoformat()

    list_name = (payload.name or "").strip() or f"Kitchen Inventory {inventory_date}"

    with get_connection() as conn:
        ingredient_ids = sorted({int(item.ingredientId) for item in payload.items})
        placeholders = ",".join("?" for _ in ingredient_ids)
        ingredient_rows = conn.execute(
            f"SELECT id, name, canonical_unit FROM ingredients WHERE id IN ({placeholders})",
            tuple(ingredient_ids),
        ).fetchall()
        ingredient_by_id = {int(row["id"]): row for row in ingredient_rows}
        missing_ids = [str(ingredient_id) for ingredient_id in ingredient_ids if ingredient_id not in ingredient_by_id]
        if missing_ids:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown ingredient id(s): {', '.join(missing_ids)}",
            )

        created = conn.execute(
            """
            INSERT INTO kitchen_inventory_lists(name, inventory_date, source_filename, notes)
            VALUES (?, ?, NULL, ?)
            RETURNING id
            """,
            (list_name, inventory_date, (payload.notes or "").strip() or None),
        ).fetchone()
        list_id = int(created["id"])

        for item in payload.items:
            ingredient = ingredient_by_id[int(item.ingredientId)]
            unit = normalize_unit((item.unit or "").strip()) or normalize_unit(
                str(ingredient["canonical_unit"] or "").strip()
            )
            canonical_qty: float | None = None
            canonical_unit: str | None = None
            note: str | None = None
            if not unit:
                note = "Missing unit and the ingredient has no canonical unit."
            else:
                canonical_qty, canonical_unit, note = to_canonical(str(ingredient["name"]), float(item.qty), unit)
                if canonical_qty is not None:
                    canonical_qty = round(float(canonical_qty), 4)
            conn.execute(
                """
                INSERT INTO kitchen_inventory_list_items(
                    list_id, ingredient_id, input_name, input_qty, input_unit,
                    canonical_qty, canonical_unit, conversion_note
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    list_id,
                    int(item.ingredientId),
                    str(ingredient["name"]),
                    float(item.qty),
                    unit or None,
                    canonical_qty,
                    canonical_unit,
                    note,
                ),
            )
        conn.commit()
        detail = load_kitchen_inventory_list_detail(conn, list_id)

    detail["skipped_rows"] = []
    return detail


@app.get("/api/kitchen-inventory")
def list_kitchen_inventory_lists() -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                kl.id,
                kl.name,
                kl.inventory_date,
                kl.source_filename,
                kl.created_at,
                COUNT(kii.id) AS item_count,
                COALESCE(SUM(CASE WHEN kii.ingredient_id IS NOT NULL THEN 1 ELSE 0 END), 0) AS matched_count,
                COALESCE(SUM(CASE WHEN kii.canonical_qty IS NOT NULL THEN 1 ELSE 0 END), 0) AS converted_count
            FROM kitchen_inventory_lists kl
            LEFT JOIN kitchen_inventory_list_items kii ON kii.list_id = kl.id
            GROUP BY kl.id
            ORDER BY kl.inventory_date DESC, kl.id DESC
            """
        ).fetchall()

    return [
        {
            "id": int(row["id"]),
            "name": row["name"],
            "inventory_date": row["inventory_date"],
            "source_filename": row["source_filename"],
            "created_at": row["created_at"],
            "item_count": int(row["item_count"] or 0),
            "matched_count": int(row["matched_count"] or 0),
            "converted_count": int(row["converted_count"] or 0),
        }
        for row in rows
    ]


@app.get("/api/kitchen-inventory/{list_id}")
def get_kitchen_inventory_list(list_id: int) -> dict[str, Any]:
    with get_connection() as conn:
        return load_kitchen_inventory_list_detail(conn, list_id)


@app.delete("/api/kitchen-inventory/{list_id}")
def delete_kitchen_inventory_list(
    list_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id, name FROM kitchen_inventory_lists WHERE id = ?",
            (list_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Kitchen inventory list not found")
        conn.execute("DELETE FROM kitchen_inventory_list_items WHERE list_id = ?", (list_id,))
        conn.execute("DELETE FROM kitchen_inventory_lists WHERE id = ?", (list_id,))
        conn.commit()
    return {"status": "deleted", "id": list_id, "name": row["name"]}


class ShoppingListApplyInventoryListPayload(BaseModel):
    inventoryListId: int = Field(gt=0)


@app.post("/api/shopping-lists/{shopping_list_id}/apply-inventory-list")
def apply_kitchen_inventory_list_to_shopping_list(
    shopping_list_id: int,
    payload: ShoppingListApplyInventoryListPayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        list_row = conn.execute(
            "SELECT id, name FROM shopping_lists WHERE id = ?",
            (shopping_list_id,),
        ).fetchone()
        if not list_row:
            raise HTTPException(status_code=404, detail="Shopping list not found")

        inventory_row = conn.execute(
            "SELECT id, name, inventory_date FROM kitchen_inventory_lists WHERE id = ?",
            (payload.inventoryListId,),
        ).fetchone()
        if not inventory_row:
            raise HTTPException(status_code=404, detail="Kitchen inventory list not found")

        inventory_by_key: dict[tuple[int, str], float] = {}
        for row in conn.execute(
            """
            SELECT ingredient_id, canonical_qty, canonical_unit
            FROM kitchen_inventory_list_items
            WHERE list_id = ?
              AND ingredient_id IS NOT NULL
              AND canonical_qty IS NOT NULL
              AND canonical_unit IS NOT NULL
            """,
            (payload.inventoryListId,),
        ).fetchall():
            key = (int(row["ingredient_id"]), str(row["canonical_unit"]))
            inventory_by_key[key] = inventory_by_key.get(key, 0.0) + float(row["canonical_qty"] or 0.0)

        item_rows = conn.execute(
            """
            SELECT id, ingredient_id, required_qty, required_unit
            FROM shopping_list_items
            WHERE shopping_list_id = ?
            """,
            (shopping_list_id,),
        ).fetchall()

        matched_count = 0
        zeroed_count = 0
        for item in item_rows:
            required_qty = float(item["required_qty"] or 0.0)
            required_unit = normalize_unit(str(item["required_unit"] or "").strip())
            _required_canonical_qty, item_canonical_unit = quantity_to_canonical(required_qty, required_unit)
            in_stock_canonical = float(
                inventory_by_key.get((int(item["ingredient_id"]), item_canonical_unit or ""), 0.0)
            )
            if in_stock_canonical > 0 and item_canonical_unit:
                in_stock_qty = canonical_qty_to_unit(in_stock_canonical, item_canonical_unit, required_unit)
                matched_count += 1
            else:
                in_stock_qty = 0.0
                zeroed_count += 1
            to_buy_qty = max(required_qty - in_stock_qty, 0.0)
            conn.execute(
                """
                UPDATE shopping_list_items
                SET in_stock_qty = ?, in_stock_unit = ?, to_buy_qty = ?, to_buy_unit = ?
                WHERE id = ?
                """,
                (
                    round(in_stock_qty, 4),
                    required_unit,
                    round(to_buy_qty, 4),
                    required_unit,
                    int(item["id"]),
                ),
            )
        conn.commit()

    return {
        "status": "ok",
        "shopping_list_id": shopping_list_id,
        "shopping_list_name": list_row["name"],
        "inventory_list_id": int(inventory_row["id"]),
        "inventory_list_name": inventory_row["name"],
        "inventory_date": inventory_row["inventory_date"],
        "matched_count": matched_count,
        "zeroed_count": zeroed_count,
    }


@app.patch("/api/shopping-lists/{shopping_list_id}/items/{item_id}")
def update_shopping_list_item(
    shopping_list_id: int,
    item_id: int,
    payload: ShoppingListItemUpdatePayload,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    fields = payload.model_fields_set
    if not fields:
        raise HTTPException(status_code=400, detail="No fields supplied")

    with get_connection() as conn:
        item_row = conn.execute(
            """
            SELECT
                sli.id,
                sli.shopping_list_id,
                sli.required_qty,
                sli.required_unit,
                sli.in_stock_qty,
                sli.in_stock_unit,
                sli.ordered_qty,
                sli.ordered_unit,
                sli.to_buy_unit,
                sli.vendor_id,
                sli.ordered,
                sli.ordered_at,
                sli.received,
                sli.received_at,
                sli.notes,
                sl.phase
            FROM shopping_list_items sli
            JOIN shopping_lists sl ON sl.id = sli.shopping_list_id
            WHERE sli.id = ? AND sli.shopping_list_id = ?
            """,
            (item_id, shopping_list_id),
        ).fetchone()
        if not item_row:
            raise HTTPException(status_code=404, detail="Shopping list item not found")

        required_qty = float(item_row["required_qty"] or 0.0)
        required_unit = normalize_unit(str(item_row["required_unit"] or "").strip())
        in_stock_unit = normalize_unit(str(item_row["in_stock_unit"] or required_unit).strip() or required_unit)
        ordered_unit = normalize_unit(str(item_row["ordered_unit"] or "").strip()) or None
        if (
            ordered_unit in {"g", "kg", "ml", "l"}
            and item_row["ordered_qty"] is None
            and not bool(item_row["ordered"])
            and not bool(item_row["received"])
        ):
            ordered_unit = None
        to_buy_unit = normalize_unit(str(item_row["to_buy_unit"] or required_unit).strip() or required_unit)
        in_stock_qty = float(item_row["in_stock_qty"] or 0.0)
        ordered_qty = float(item_row["ordered_qty"]) if item_row["ordered_qty"] is not None else None
        if "inStockQty" in fields:
            in_stock_qty = float(payload.inStockQty or 0.0)
        required_canonical_qty, required_canonical_unit = quantity_to_canonical(required_qty, required_unit)
        in_stock_canonical_qty, in_stock_canonical_unit = quantity_to_canonical(in_stock_qty, in_stock_unit)
        if required_canonical_qty is not None and required_canonical_unit == in_stock_canonical_unit:
            to_buy_canonical_qty = max(required_canonical_qty - (in_stock_canonical_qty or 0.0), 0.0)
            converted = canonical_qty_to_unit_or_none(to_buy_canonical_qty, required_canonical_unit, required_unit)
            to_buy_qty = converted if converted is not None else max(required_qty - in_stock_qty, 0.0)
        else:
            to_buy_qty = max(required_qty - in_stock_qty, 0.0)
        in_stock_unit = required_unit
        to_buy_unit = required_unit

        vendor_id = item_row["vendor_id"]
        now_iso = datetime.now(timezone.utc).isoformat()
        ordered = bool(item_row["ordered"])
        received = bool(item_row["received"])
        ordered_at = item_row["ordered_at"]
        received_at = item_row["received_at"]
        normalized_vendor_allocations: list[dict[str, Any]] | None = None

        if "vendorAllocations" in fields:
            existing_allocations = load_shopping_list_item_vendor_allocations_by_item_id(conn, [item_id]).get(item_id, [])
            existing_allocations_by_id = {
                int(entry["id"]): entry
                for entry in existing_allocations
                if entry.get("id") is not None
            }
            base_allocation_unit = (
                normalize_unit(str(to_buy_unit or required_unit or "").strip())
                or normalize_unit(str(item_row["ordered_unit"] or "").strip())
                or "each"
            )
            seen_vendor_keys: set[int] = set()
            normalized_vendor_allocations = []
            raw_allocations = payload.vendorAllocations or []
            for index, entry in enumerate(raw_allocations):
                allocation_id = int(entry.id) if entry.id is not None else None
                existing_allocation = (
                    existing_allocations_by_id.get(allocation_id)
                    if allocation_id is not None
                    else None
                )
                if allocation_id is not None and not existing_allocation:
                    raise HTTPException(status_code=400, detail="Vendor allocation not found for this item.")

                next_vendor_id = int(entry.vendorId) if entry.vendorId is not None else None
                if next_vendor_id is not None:
                    vendor_exists = conn.execute(
                        "SELECT id FROM vendors WHERE id = ?",
                        (next_vendor_id,),
                    ).fetchone()
                    if not vendor_exists:
                        raise HTTPException(status_code=400, detail="Vendor not found")

                if next_vendor_id is not None and next_vendor_id in seen_vendor_keys:
                    raise HTTPException(status_code=400, detail="Each shopping item source can appear only once.")

                next_allocated_qty = round(float(entry.allocatedQty or 0.0), 4)
                if next_allocated_qty < 0:
                    raise HTTPException(status_code=400, detail="Source amount must be non-negative.")
                next_allocated_unit = (
                    normalize_unit(str(entry.allocatedUnit or "").strip())
                    or base_allocation_unit
                )
                next_ordered = bool(entry.ordered)
                next_received = bool(entry.received)
                if next_received:
                    next_ordered = True
                if (next_ordered or next_received) and next_allocated_qty <= 0:
                    raise HTTPException(status_code=400, detail="Ordered or received source lines need an amount.")

                existing_ordered = bool(existing_allocation["ordered"]) if existing_allocation else False
                existing_received = bool(existing_allocation["received"]) if existing_allocation else False
                next_ordered_at = existing_allocation["ordered_at"] if existing_ordered and existing_allocation else None
                next_received_at = existing_allocation["received_at"] if existing_received and existing_allocation else None
                if next_ordered and not existing_ordered:
                    next_ordered_at = now_iso
                if not next_ordered:
                    next_ordered_at = None
                    next_received = False
                    next_received_at = None
                if next_received and not existing_received:
                    next_received_at = now_iso
                if not next_received:
                    next_received_at = None

                normalized_vendor_allocations.append(
                    {
                        "vendor_id": next_vendor_id,
                        "allocated_qty": next_allocated_qty,
                        "allocated_unit": next_allocated_unit,
                        "ordered": next_ordered,
                        "ordered_at": next_ordered_at,
                        "received": next_received,
                        "received_at": next_received_at,
                        "sort_order": index,
                    }
                )
                if next_vendor_id is not None:
                    seen_vendor_keys.add(next_vendor_id)

            allocation_summary = summarize_shopping_vendor_allocations(
                normalized_vendor_allocations,
                preferred_unit=normalize_unit(str(required_unit or to_buy_unit or "").strip()) or None,
            )
            vendor_id = allocation_summary["vendor_id"]
            ordered_qty = allocation_summary["ordered_qty"]
            ordered_unit = allocation_summary["ordered_unit"]
            ordered = bool(allocation_summary["ordered"])
            ordered_at = allocation_summary["ordered_at"]
            received = bool(allocation_summary["received"])
            received_at = allocation_summary["received_at"]
            status = allocation_summary["status"]
        else:
            if "vendorId" in fields:
                vendor_id = payload.vendorId
                if vendor_id is not None:
                    vendor_exists = conn.execute(
                        "SELECT id FROM vendors WHERE id = ?",
                        (vendor_id,),
                    ).fetchone()
                    if not vendor_exists:
                        raise HTTPException(status_code=400, detail="Vendor not found")

            if "orderedQty" in fields:
                if payload.orderedQty is None:
                    ordered_qty = None
                else:
                    ordered_qty = float(payload.orderedQty)
                    if ordered_qty < 0:
                        raise HTTPException(status_code=400, detail="Amount ordered must be non-negative.")
                    ordered_qty = round(ordered_qty, 4)
                if ordered_qty is not None and ordered_qty > 0 and not received:
                    ordered = True

            if "orderedUnit" in fields:
                ordered_unit = normalize_unit(str(payload.orderedUnit or "").strip()) or None

            if "ordered" in fields:
                ordered = bool(payload.ordered)
                if not ordered:
                    received = False

            if "received" in fields:
                received = bool(payload.received)
                if received:
                    ordered = True

            if ordered and not bool(item_row["ordered"]):
                ordered_at = now_iso
            if not ordered:
                ordered_at = None

            if received and not bool(item_row["received"]):
                received_at = now_iso
            if not received:
                received_at = None

            default_ordered_qty, default_ordered_unit = preferred_ordered_quantity_and_unit(
                to_buy_qty or required_qty or 0.0,
                to_buy_unit or required_unit,
                preferred_unit=ordered_unit,
            )

            if not ordered and not received:
                ordered_qty = None
            if ordered and (ordered_qty is None or ordered_qty <= 0):
                ordered_qty = default_ordered_qty
            if ordered_qty is not None and ordered_qty <= 0:
                ordered_qty = None
            if received and (ordered_qty is None or ordered_qty <= 0):
                ordered_qty = default_ordered_qty
            if ordered_qty is not None and ordered_qty > 0 and not ordered_unit:
                ordered_unit = default_ordered_unit or to_buy_unit or required_unit
            status = derive_shopping_item_status(ordered=ordered, received=received)

        notes = item_row["notes"]
        if "notes" in fields:
            notes = payload.notes.strip() if payload.notes and payload.notes.strip() else None

        conn.execute(
            """
            UPDATE shopping_list_items
            SET vendor_id = ?,
                in_stock_qty = ?,
                in_stock_unit = ?,
                to_buy_qty = ?,
                to_buy_unit = ?,
                ordered_qty = ?,
                ordered_unit = ?,
                ordered = ?,
                ordered_at = ?,
                received = ?,
                received_at = ?,
                status = ?,
                notes = ?
            WHERE id = ? AND shopping_list_id = ?
            """,
            (
                vendor_id,
                round(in_stock_qty, 4),
                in_stock_unit,
                round(to_buy_qty, 4),
                to_buy_unit,
                ordered_qty,
                ordered_unit,
                1 if ordered else 0,
                ordered_at,
                1 if received else 0,
                received_at,
                status,
                notes,
                item_id,
                shopping_list_id,
            ),
        )
        if normalized_vendor_allocations is not None:
            replace_shopping_item_vendor_allocations(conn, item_id, normalized_vendor_allocations)

        refresh_shopping_list_status(conn, shopping_list_id)
        sync_food_inventory_orders_for_shopping_list(conn, shopping_list_id, actor_user_id=user.id)
        detail = load_shopping_list_detail(conn, shopping_list_id)
        conn.commit()
    return detail


@app.put("/api/ingredients/{ingredient_id}")
def update_ingredient(
    ingredient_id: int,
    payload: IngredientUpdatePayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    ingredient_name = payload.name.strip()
    if not ingredient_name:
        raise HTTPException(status_code=400, detail="Ingredient name cannot be blank")

    canonical_unit = payload.canonical_unit.strip() if payload.canonical_unit and payload.canonical_unit.strip() else None
    if canonical_unit:
        canonical_unit = normalize_unit(canonical_unit)
        allowed_canonical_units = {"g", "kg", "ml", "l", "each"} | COUNT_UNITS
        if canonical_unit not in allowed_canonical_units:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"'{payload.canonical_unit}' is not a valid canonical unit. "
                    f"Allowed: {', '.join(sorted(allowed_canonical_units))}."
                ),
            )
    notes = payload.notes.strip() if payload.notes and payload.notes.strip() else None
    category = payload.category.strip() if payload.category and payload.category.strip() else None
    purchase_tier = payload.purchase_tier.strip() if payload.purchase_tier and payload.purchase_tier.strip() else None

    with get_connection() as conn:
        existing = conn.execute("SELECT id FROM ingredients WHERE id = ?", (ingredient_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Ingredient not found")

        duplicate = conn.execute(
            "SELECT id FROM ingredients WHERE lower(name) = lower(?) AND id != ?",
            (ingredient_name, ingredient_id),
        ).fetchone()
        if duplicate:
            raise HTTPException(status_code=400, detail=f"An ingredient named '{ingredient_name}' already exists")

        conn.execute(
            """
            UPDATE ingredients
            SET name = ?, canonical_unit = ?, grams_per_cup = ?, notes = ?, category = ?, purchase_tier = ?
            WHERE id = ?
            """,
            (ingredient_name, canonical_unit, payload.grams_per_cup, notes, category, purchase_tier, ingredient_id),
        )
        updated = conn.execute(
            "SELECT id, name, category, purchase_tier, canonical_unit, grams_per_cup, notes FROM ingredients WHERE id = ?",
            (ingredient_id,),
        ).fetchone()
        conn.commit()

    return dict(updated)


@app.get("/api/unit-conversions")
def list_unit_conversions() -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                id,
                item_name,
                quantity_from,
                unit_from,
                quantity_to,
                unit_to,
                context,
                source_sheet,
                source_row,
                notes
            FROM unit_conversions
            ORDER BY context, item_name, unit_from, id
            """
        ).fetchall()
    return [
        {
            "id": int(row["id"]),
            "item_name": row["item_name"],
            "quantity_from": float(row["quantity_from"]),
            "unit_from": row["unit_from"],
            "quantity_to": float(row["quantity_to"]),
            "unit_to": row["unit_to"],
            "context": row["context"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
            "notes": row["notes"],
        }
        for row in rows
    ]


@app.get("/api/recipe-categories")
def list_recipe_categories() -> list[str]:
    return RECIPE_CATEGORIES


@app.post("/api/recipes")
def create_recipe(
    payload: RecipeCreate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    recipe_name = payload.name.strip()
    if not recipe_name:
        raise HTTPException(status_code=400, detail="Recipe name cannot be blank")
    recipe_category = normalize_recipe_category(payload.category)
    recipe_notes = payload.notes.strip() if payload.notes and payload.notes.strip() else None

    with get_connection() as conn:
        try:
            recipe_row = conn.execute(
                """
                INSERT INTO recipes(name, category, base_servings, notes)
                VALUES (?, ?, ?, ?)
                RETURNING id, name, category, base_servings, notes
                """,
                (recipe_name, recipe_category, payload.base_servings, recipe_notes),
            ).fetchone()
            recipe_id = int(recipe_row["id"])
            replace_recipe_ingredients(conn, recipe_id, payload.ingredients)
            replace_recipe_steps(conn, recipe_id, payload.steps)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not create recipe: {exc}") from exc

        conn.commit()

    return {
        "id": recipe_row["id"],
        "name": recipe_row["name"],
        "category": recipe_row["category"],
        "base_servings": recipe_row["base_servings"],
        "notes": recipe_row["notes"],
    }


@app.put("/api/recipes/{recipe_id}")
def update_recipe(
    recipe_id: int,
    payload: RecipeCreate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    recipe_name = payload.name.strip()
    if not recipe_name:
        raise HTTPException(status_code=400, detail="Recipe name cannot be blank")
    recipe_category = normalize_recipe_category(payload.category)
    recipe_notes = payload.notes.strip() if payload.notes and payload.notes.strip() else None

    with get_connection() as conn:
        existing = conn.execute("SELECT id FROM recipes WHERE id = ?", (recipe_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Recipe not found")

        try:
            conn.execute(
                """
                UPDATE recipes
                SET name = ?, category = ?, base_servings = ?, notes = ?
                WHERE id = ?
                """,
                (recipe_name, recipe_category, payload.base_servings, recipe_notes, recipe_id),
            )
            replace_recipe_ingredients(conn, recipe_id, payload.ingredients)
            replace_recipe_steps(conn, recipe_id, payload.steps)
            updated = conn.execute(
                "SELECT id, name, category, base_servings, notes FROM recipes WHERE id = ?",
                (recipe_id,),
            ).fetchone()
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not update recipe: {exc}") from exc

        conn.commit()

    return {
        "id": int(updated["id"]),
        "name": updated["name"],
        "category": updated["category"],
        "base_servings": float(updated["base_servings"]),
        "notes": updated["notes"],
    }


@app.get("/api/recipes")
def list_recipes() -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, name, category, base_servings, notes, created_at
            FROM recipes
            ORDER BY created_at DESC
            """
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/recipes/full")
def list_recipes_full() -> list[dict[str, Any]]:
    with get_connection() as conn:
        recipes = conn.execute(
            """
            SELECT id, name, category, base_servings, notes, created_at
            FROM recipes
            ORDER BY name
            """
        ).fetchall()

        ingredient_rows = conn.execute(
            """
            SELECT
                ri.recipe_id,
                i.name AS ingredient_name,
                ri.quantity,
                ri.unit,
                ri.prep_notes,
                i.category AS ingredient_category,
                i.grams_per_cup,
                i.canonical_unit AS ingredient_canonical_unit
            FROM recipe_ingredients ri
            JOIN ingredients i ON i.id = ri.ingredient_id
            ORDER BY ri.recipe_id, ri.id
            """
        ).fetchall()

        step_rows = conn.execute(
            """
            SELECT recipe_id, step_order, instruction
            FROM recipe_steps
            ORDER BY recipe_id, step_order, id
            """
        ).fetchall()

    ingredients_by_recipe: dict[int, list[dict[str, Any]]] = {}
    for row in ingredient_rows:
        recipe_id = int(row["recipe_id"])
        entry: dict[str, Any] = {
            "name": row["ingredient_name"],
            "quantity": float(row["quantity"]),
            "unit": row["unit"],
            "prep_notes": row["prep_notes"],
        }
        if row["ingredient_category"]:
            entry["category"] = row["ingredient_category"]
        if row["grams_per_cup"]:
            entry["grams_per_cup"] = float(row["grams_per_cup"])
        if row["ingredient_canonical_unit"]:
            entry["canonical_unit"] = normalize_unit(str(row["ingredient_canonical_unit"]))
        ingredients_by_recipe.setdefault(recipe_id, []).append(entry)

    steps_by_recipe: dict[int, list[str]] = {}
    for row in step_rows:
        recipe_id = int(row["recipe_id"])
        steps_by_recipe.setdefault(recipe_id, []).append(row["instruction"])

    output: list[dict[str, Any]] = []
    for recipe in recipes:
        recipe_id = int(recipe["id"])
        output.append(
            {
                "id": recipe_id,
                "name": recipe["name"],
                "category": recipe["category"],
                "base_servings": float(recipe["base_servings"]),
                "notes": recipe["notes"],
                "created_at": recipe["created_at"],
                "ingredients": ingredients_by_recipe.get(recipe_id, []),
                "steps": steps_by_recipe.get(recipe_id, []),
            }
        )

    return output


@app.post("/api/scale-preview", response_model=ScalePreviewResponse)
def scale_preview(payload: ScalePreviewRequest) -> ScalePreviewResponse:
    factor = payload.target_servings / payload.base_servings
    scaled_items: list[ScaledIngredient] = []

    for item in payload.ingredients:
        unit = normalize_unit(item.unit)
        scaled_qty = item.quantity * factor
        canonical_qty, canonical_unit, note = require_canonical_conversion(
            item.name,
            scaled_qty,
            unit,
            context="Scale preview conversion failed",
        )
        shopping_qty, shopping_unit = to_shopping_unit(canonical_qty, canonical_unit)

        scaled_items.append(
            ScaledIngredient(
                name=item.name,
                input_quantity=round(item.quantity, 4),
                input_unit=unit,
                scaled_quantity=round(scaled_qty, 4),
                scaled_unit=unit,
                canonical_quantity=round(canonical_qty, 4) if canonical_qty is not None else None,
                canonical_unit=canonical_unit,
                shopping_quantity=round(shopping_qty, 4) if shopping_qty is not None else None,
                shopping_unit=shopping_unit,
                note=note,
            )
        )

    return ScalePreviewResponse(scale_factor=round(factor, 4), ingredients=scaled_items)


@app.post("/api/recipes/{recipe_id}/scale", response_model=ScalePreviewResponse)
def scale_recipe(recipe_id: int, target_servings: float = Query(..., gt=0)) -> ScalePreviewResponse:
    with get_connection() as conn:
        recipe = conn.execute(
            "SELECT id, name, base_servings FROM recipes WHERE id = ?", (recipe_id,)
        ).fetchone()
        if not recipe:
            raise HTTPException(status_code=404, detail="Recipe not found")

        items = conn.execute(
            """
            SELECT i.name AS ingredient_name, ri.quantity, ri.unit
            FROM recipe_ingredients ri
            JOIN ingredients i ON i.id = ri.ingredient_id
            WHERE ri.recipe_id = ?
            ORDER BY ri.id
            """,
            (recipe_id,),
        ).fetchall()

    payload = ScalePreviewRequest(
        base_servings=float(recipe["base_servings"]),
        target_servings=target_servings,
        ingredients=[
            IngredientInput(name=row["ingredient_name"], quantity=row["quantity"], unit=row["unit"])
            for row in items
        ],
    )
    return scale_preview(payload)


@app.get("/api/retreat-plans")
def list_retreat_plans() -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, name, start_date, day_count, default_people, created_at, updated_at
            FROM retreat_plans
            ORDER BY updated_at DESC, id DESC
            """
        ).fetchall()
    return [dict(r) for r in rows]


def hydrate_retreat_plan_payload(row: Any) -> dict[str, Any]:
    payload = json.loads(row["plan_json"]) if row["plan_json"] else {}
    if not isinstance(payload, dict):
        payload = {}

    retreat_meals = payload.get("retreatMeals")
    if not isinstance(retreat_meals, list):
        retreat_meals = payload.get("meals") if isinstance(payload.get("meals"), list) else []

    test_meals = payload.get("testMeals")
    if not isinstance(test_meals, list):
        test_meals = []

    default_people = float(payload.get("defaultPeople") or row["default_people"])
    retreat_default_people = float(payload.get("retreatDefaultPeople") or default_people)
    test_default_people = float(payload.get("testDefaultPeople") or DEFAULT_TEST_HEADCOUNT)

    active_profile = payload.get("activeProfile")
    if active_profile not in HEADCOUNT_PROFILES:
        active_profile = "retreat"

    shopping_profile = payload.get("shoppingProfile")
    if shopping_profile not in HEADCOUNT_PROFILES:
        shopping_profile = DEFAULT_SHOPPING_PROFILE

    payload["id"] = int(row["id"])
    payload["name"] = payload.get("name") or row["name"]
    payload["startDate"] = payload.get("startDate", row["start_date"])
    payload["dayCount"] = int(payload.get("dayCount") or row["day_count"])
    payload["defaultPeople"] = default_people
    payload["retreatDefaultPeople"] = retreat_default_people
    payload["testDefaultPeople"] = test_default_people
    payload["activeProfile"] = active_profile
    payload["shoppingProfile"] = shopping_profile
    payload["retreatMeals"] = retreat_meals
    payload["testMeals"] = test_meals
    payload["meals"] = retreat_meals
    payload["created_at"] = row["created_at"]
    payload["updated_at"] = row["updated_at"]
    return payload


@app.get("/api/retreat-plans/{plan_id}")
def get_retreat_plan(plan_id: int) -> dict[str, Any]:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT id, name, start_date, day_count, default_people, plan_json, created_at, updated_at
            FROM retreat_plans
            WHERE id = ?
            """,
            (plan_id,),
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Retreat plan not found")

    return hydrate_retreat_plan_payload(row)


@app.get("/api/retreat-plans/{plan_id}/ingredients-by-retreat")
def get_retreat_plan_ingredients_by_retreat(
    plan_id: int,
    tier: Literal["bulk", "fresh", "daily"] = Query(default="bulk"),
) -> dict[str, Any]:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT id, name, start_date, day_count, default_people, plan_json, created_at, updated_at
            FROM retreat_plans
            WHERE id = ?
            """,
            (plan_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Retreat plan not found")

        plan_payload = hydrate_retreat_plan_payload(row)
        aggregate, missing_recipes, _dish_breakdown = build_required_ingredients_from_plan(
            conn,
            plan_payload=plan_payload,
            profile="retreat",
            purchase_tiers=resolve_purchase_tiers_for_shopping(tier, None),
        )

        category_by_ingredient_id: dict[int, str | None] = {}
        ingredient_ids = sorted({int(entry["ingredient_id"]) for entry in aggregate.values()})
        if ingredient_ids:
            placeholders = ",".join("?" for _ in ingredient_ids)
            category_rows = conn.execute(
                f"SELECT id, category FROM ingredients WHERE id IN ({placeholders})",
                tuple(ingredient_ids),
            ).fetchall()
            category_by_ingredient_id = {
                int(category_row["id"]): (
                    str(category_row["category"]).strip() if category_row["category"] is not None else None
                )
                for category_row in category_rows
            }

    items_by_category: dict[str, list[dict[str, Any]]] = {}
    ingredient_name_counts: dict[str, int] = {}
    for entry in aggregate.values():
        ingredient_name = str(entry["ingredient_name"] or "").strip() or "Unknown ingredient"
        ingredient_name_counts[ingredient_name.lower()] = ingredient_name_counts.get(ingredient_name.lower(), 0) + 1

    for entry in aggregate.values():
        ingredient_id = int(entry["ingredient_id"])
        ingredient_name = str(entry["ingredient_name"] or "").strip() or "Unknown ingredient"
        required_unit = str(entry["canonical_unit"] or "").strip()
        if ingredient_name_counts.get(ingredient_name.lower(), 0) > 1 and required_unit:
            ingredient_name = f"{ingredient_name} ({required_unit})"

        category_name = (
            str(category_by_ingredient_id.get(ingredient_id) or "").strip()
            or "Uncategorized"
        )
        items_by_category.setdefault(category_name, []).append(
            {
                "ingredient_id": ingredient_id,
                "ingredient_name": ingredient_name,
                "ingredient_category": category_name,
                "required_qty": round(float(entry["required_qty"]), 4),
                "required_unit": required_unit,
            }
        )

    def category_sort_key(name: str) -> tuple[int, str]:
        return (1, "") if name == "Uncategorized" else (0, name.lower())

    categories: list[dict[str, Any]] = []
    total_items = 0
    for category_name in sorted(items_by_category.keys(), key=category_sort_key):
        category_items = sorted(
            items_by_category[category_name],
            key=lambda item: (
                str(item["ingredient_name"]).lower(),
                int(item["ingredient_id"]),
            ),
        )
        total_items += len(category_items)
        categories.append(
            {
                "category": category_name,
                "item_count": len(category_items),
                "items": category_items,
            }
        )

    return {
        "retreat_plan": {
            "id": int(plan_payload["id"]),
            "name": plan_payload["name"],
            "start_date": plan_payload.get("startDate"),
            "updated_at": plan_payload.get("updated_at"),
        },
        "tier": tier,
        "item_count": total_items,
        "category_count": len(categories),
        "missing_recipes": sorted(missing_recipes),
        "categories": categories,
    }


@app.post("/api/retreat-plans")
def upsert_retreat_plan(
    payload: RetreatPlanPayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    plan_name = payload.name.strip()
    if not plan_name:
        raise HTTPException(status_code=400, detail="Retreat name cannot be blank")
    requested_plan_id = int(payload.id) if payload.id else None

    retreat_meals = payload.retreatMeals if payload.retreatMeals is not None else payload.meals
    test_meals = payload.testMeals if payload.testMeals is not None else []
    meal_lists = [payload.meals, retreat_meals, test_meals]
    if any(meal.day > payload.dayCount for meal_list in meal_lists for meal in meal_list):
        raise HTTPException(status_code=400, detail="Meal day cannot exceed dayCount")

    payload_dict = payload.model_dump()
    payload_dict.pop("id", None)
    payload_dict["name"] = plan_name
    payload_dict["defaultPeople"] = float(payload.defaultPeople)
    payload_dict["retreatDefaultPeople"] = float(payload.retreatDefaultPeople or payload.defaultPeople)
    payload_dict["testDefaultPeople"] = float(payload.testDefaultPeople)
    payload_dict["activeProfile"] = payload.activeProfile
    payload_dict["shoppingProfile"] = DEFAULT_SHOPPING_PROFILE
    payload_dict["retreatMeals"] = [meal.model_dump() for meal in retreat_meals]
    payload_dict["testMeals"] = [meal.model_dump() for meal in test_meals]
    payload_dict["meals"] = [meal.model_dump() for meal in retreat_meals]
    payload_json = json.dumps(payload_dict)

    with get_connection() as conn:
        if requested_plan_id:
            existing = conn.execute(
                "SELECT id FROM retreat_plans WHERE id = ?",
                (requested_plan_id,),
            ).fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="Retreat plan not found")

            name_conflict = conn.execute(
                "SELECT id FROM retreat_plans WHERE lower(name) = lower(?) AND id <> ?",
                (plan_name, requested_plan_id),
            ).fetchone()
            if name_conflict:
                raise HTTPException(status_code=409, detail="Retreat name already exists")

            conn.execute(
                """
                UPDATE retreat_plans
                SET name = ?, start_date = ?, day_count = ?, default_people = ?, plan_json = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    plan_name,
                    payload.startDate,
                    payload.dayCount,
                    payload_dict["defaultPeople"],
                    payload_json,
                    requested_plan_id,
                ),
            )
            plan_id = requested_plan_id
            action = "updated"
        else:
            existing = conn.execute(
                "SELECT id FROM retreat_plans WHERE lower(name) = lower(?)",
                (plan_name,),
            ).fetchone()

            if existing:
                conn.execute(
                    """
                    UPDATE retreat_plans
                    SET name = ?, start_date = ?, day_count = ?, default_people = ?, plan_json = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (
                        plan_name,
                        payload.startDate,
                        payload.dayCount,
                        payload_dict["defaultPeople"],
                        payload_json,
                        existing["id"],
                    ),
                )
                plan_id = int(existing["id"])
                action = "updated"
            else:
                created = conn.execute(
                    """
                    INSERT INTO retreat_plans(name, start_date, day_count, default_people, plan_json)
                    VALUES (?, ?, ?, ?, ?)
                    RETURNING id
                    """,
                    (
                        plan_name,
                        payload.startDate,
                        payload.dayCount,
                        payload_dict["defaultPeople"],
                        payload_json,
                    ),
                ).fetchone()
                plan_id = int(created["id"])
                action = "created"

        row = conn.execute(
            "SELECT id, name, updated_at FROM retreat_plans WHERE id = ?",
            (plan_id,),
        ).fetchone()
        conn.commit()

    return {
        "id": int(row["id"]),
        "name": row["name"],
        "status": action,
        "updated_at": row["updated_at"],
    }


@app.post("/api/retreat-plans/{plan_id}/duplicate")
def duplicate_retreat_plan(
    plan_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
    payload: RetreatPlanDuplicatePayload | None = None,
) -> dict[str, Any]:
    with get_connection() as conn:
        source = conn.execute(
            """
            SELECT id, name, start_date, day_count, default_people, plan_json
            FROM retreat_plans
            WHERE id = ?
            """,
            (plan_id,),
        ).fetchone()
        if not source:
            raise HTTPException(status_code=404, detail="Retreat plan not found")

        source_payload = json.loads(source["plan_json"]) if source["plan_json"] else {}
        if not isinstance(source_payload, dict):
            source_payload = {}

        requested_name = payload.name.strip() if payload and payload.name and payload.name.strip() else None
        base_name = requested_name or f"{source['name']} (Copy)"
        copy_name = unique_retreat_plan_name(conn, base_name)

        source_payload["name"] = copy_name
        source_payload["startDate"] = source_payload.get("startDate", source["start_date"])
        source_payload["dayCount"] = int(source_payload.get("dayCount") or source["day_count"])
        source_payload["defaultPeople"] = float(
            source_payload.get("defaultPeople") or source["default_people"]
        )
        source_payload["retreatDefaultPeople"] = float(
            source_payload.get("retreatDefaultPeople") or source_payload["defaultPeople"]
        )
        source_payload["testDefaultPeople"] = float(
            source_payload.get("testDefaultPeople") or DEFAULT_TEST_HEADCOUNT
        )
        active_profile = source_payload.get("activeProfile")
        source_payload["activeProfile"] = active_profile if active_profile in HEADCOUNT_PROFILES else "retreat"
        shopping_profile = source_payload.get("shoppingProfile")
        source_payload["shoppingProfile"] = (
            shopping_profile if shopping_profile in HEADCOUNT_PROFILES else DEFAULT_SHOPPING_PROFILE
        )

        retreat_meals = source_payload.get("retreatMeals")
        if not isinstance(retreat_meals, list):
            retreat_meals = source_payload.get("meals") if isinstance(source_payload.get("meals"), list) else []
        test_meals = source_payload.get("testMeals")
        if not isinstance(test_meals, list):
            test_meals = []
        source_payload["retreatMeals"] = retreat_meals
        source_payload["testMeals"] = test_meals
        source_payload["meals"] = retreat_meals

        created = conn.execute(
            """
            INSERT INTO retreat_plans(name, start_date, day_count, default_people, plan_json)
            VALUES (?, ?, ?, ?, ?)
            RETURNING id, name, created_at, updated_at
            """,
            (
                copy_name,
                source_payload["startDate"],
                source_payload["dayCount"],
                source_payload["defaultPeople"],
                json.dumps(source_payload),
            ),
        ).fetchone()
        conn.commit()

    return {
        "id": int(created["id"]),
        "name": created["name"],
        "status": "duplicated",
        "source_plan_id": int(source["id"]),
        "created_at": created["created_at"],
        "updated_at": created["updated_at"],
    }


@app.delete("/api/retreat-plans/{plan_id}")
def delete_retreat_plan(
    plan_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id, name FROM retreat_plans WHERE id = ?",
            (plan_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Retreat plan not found")

        conn.execute("DELETE FROM retreat_plans WHERE id = ?", (plan_id,))
        conn.commit()

    return {
        "id": int(existing["id"]),
        "name": existing["name"],
        "status": "deleted",
    }


@app.post("/api/service-snapshots")
def create_service_snapshot(
    payload: ServiceSnapshotPayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        if payload.retreatPlanId is not None:
            plan = conn.execute(
                "SELECT id FROM retreat_plans WHERE id = ?", (payload.retreatPlanId,)
            ).fetchone()
            if not plan:
                raise HTTPException(status_code=400, detail="Referenced retreat plan not found")

        row = conn.execute(
            """
            INSERT INTO service_snapshots(retreat_name, payload_json, retreat_plan_id)
            VALUES (?, ?, ?)
            RETURNING id, retreat_name, created_at
            """,
            (payload.retreatName.strip(), payload.model_dump_json(), payload.retreatPlanId),
        ).fetchone()
        conn.commit()

    return {
        "id": int(row["id"]),
        "retreat_name": row["retreat_name"],
        "created_at": row["created_at"],
        "meal_count": len(payload.meals),
    }


@app.get("/api/service-snapshots/latest")
def get_latest_service_snapshot() -> dict[str, Any]:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT id, retreat_name, payload_json, created_at
            FROM service_snapshots
            ORDER BY id DESC
            LIMIT 1
            """
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="No service snapshot found")

    payload = json.loads(row["payload_json"])
    return {
        "snapshot_id": int(row["id"]),
        "retreat_name": row["retreat_name"],
        "created_at": row["created_at"],
        "payload": payload,
    }


@app.get("/api/service-snapshots/by-plan/{plan_id}")
def get_service_snapshot_by_plan(plan_id: int) -> dict[str, Any]:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT id, retreat_name, payload_json, created_at
            FROM service_snapshots
            WHERE retreat_plan_id = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (plan_id,),
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="No service snapshot found for this retreat plan")

    payload = json.loads(row["payload_json"])
    return {
        "snapshot_id": int(row["id"]),
        "retreat_name": row["retreat_name"],
        "created_at": row["created_at"],
        "payload": payload,
    }


def resolve_purchase_tiers_for_shopping(phase: str, purchase_tiers: list[str] | None) -> set[str]:
    if purchase_tiers:
        cleaned = {
            str(tier).strip().lower()
            for tier in purchase_tiers
            if str(tier).strip().lower() in PURCHASE_TIERS
        }
        if not cleaned:
            raise HTTPException(
                status_code=400,
                detail=f"purchaseTiers must contain one of: {', '.join(PURCHASE_TIERS)}",
            )
        return cleaned

    normalized_phase = phase.strip().lower()
    if normalized_phase in PURCHASE_TIERS:
        return {normalized_phase}
    return set()


def resolve_requested_shopping_plan_ids(payload: ShoppingListGeneratePayload) -> list[int]:
    requested_plan_ids: list[int] = []
    if payload.retreatPlanIds:
        for raw_id in payload.retreatPlanIds:
            plan_id = int(raw_id)
            if plan_id <= 0:
                raise HTTPException(status_code=400, detail="retreatPlanIds must contain positive integers")
            requested_plan_ids.append(plan_id)

    if payload.retreatPlanId is not None:
        plan_id = int(payload.retreatPlanId)
        if plan_id <= 0:
            raise HTTPException(status_code=400, detail="retreatPlanId must be a positive integer")
        requested_plan_ids.append(plan_id)

    return sorted(set(requested_plan_ids))


def normalize_shopping_list_generation_config(payload: ShoppingListGeneratePayload) -> dict[str, Any]:
    requested_plan_ids = resolve_requested_shopping_plan_ids(payload)
    purchase_tiers = (
        sorted(
            {
                str(tier).strip().lower()
                for tier in (payload.purchaseTiers or [])
                if str(tier).strip().lower() in PURCHASE_TIERS
            }
        )
        or None
    )
    return {
        "retreatPlanIds": requested_plan_ids,
        "allRetreats": bool(payload.allRetreats),
        "phase": str(payload.phase or "bulk").strip().lower() or "bulk",
        "purchaseTiers": purchase_tiers,
        "profile": str(payload.profile or DEFAULT_SHOPPING_PROFILE).strip().lower() or DEFAULT_SHOPPING_PROFILE,
        "subtractInventory": bool(payload.subtractInventory),
        "includeZeroToBuy": bool(payload.includeZeroToBuy),
    }


def shopping_list_name_implies_all_retreats(name: Any) -> bool:
    return str(name or "").strip().lower().startswith("all retreats")


def load_shopping_list_source_plan_ids(conn: Any, shopping_list_id: int) -> list[int]:
    rows = conn.execute(
        """
        SELECT DISTINCT slis.retreat_plan_id
        FROM shopping_list_item_sources slis
        JOIN shopping_list_items sli ON sli.id = slis.shopping_list_item_id
        WHERE sli.shopping_list_id = ?
          AND slis.retreat_plan_id IS NOT NULL
        ORDER BY slis.retreat_plan_id
        """,
        (shopping_list_id,),
    ).fetchall()
    return [int(row["retreat_plan_id"]) for row in rows if row["retreat_plan_id"] is not None]


def infer_shopping_list_generation_payload(conn: Any, shopping_list_id: int) -> ShoppingListGeneratePayload:
    list_row = conn.execute(
        """
        SELECT
            id,
            name,
            phase,
            retreat_plan_id,
            generation_config_json
        FROM shopping_lists
        WHERE id = ?
        """,
        (shopping_list_id,),
    ).fetchone()
    if not list_row:
        raise HTTPException(status_code=404, detail="Shopping list not found")

    def build_payload_from_config(config: dict[str, Any]) -> ShoppingListGeneratePayload:
        retreat_plan_ids = []
        for raw_id in config.get("retreatPlanIds") or []:
            try:
                plan_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if plan_id > 0 and plan_id not in retreat_plan_ids:
                retreat_plan_ids.append(plan_id)
        retreat_plan_ids.sort()
        if not retreat_plan_ids and list_row["retreat_plan_id"] is not None:
            retreat_plan_ids = [int(list_row["retreat_plan_id"])]

        raw_purchase_tiers = config.get("purchaseTiers")
        purchase_tiers = None
        if isinstance(raw_purchase_tiers, list):
            purchase_tiers = [
                str(tier).strip().lower()
                for tier in raw_purchase_tiers
                if str(tier).strip().lower() in PURCHASE_TIERS
            ] or None

        phase = str(config.get("phase") or list_row["phase"] or "bulk").strip().lower() or "bulk"
        profile = str(config.get("profile") or DEFAULT_SHOPPING_PROFILE).strip().lower() or DEFAULT_SHOPPING_PROFILE
        if profile not in HEADCOUNT_PROFILES:
            profile = DEFAULT_SHOPPING_PROFILE

        all_retreats = bool(config.get("allRetreats"))
        return ShoppingListGeneratePayload(
            retreatPlanId=retreat_plan_ids[0] if len(retreat_plan_ids) == 1 else None,
            retreatPlanIds=retreat_plan_ids,
            allRetreats=all_retreats,
            phase=phase,
            purchaseTiers=purchase_tiers,
            profile=profile,
            subtractInventory=bool(config.get("subtractInventory", True)),
            includeZeroToBuy=bool(config.get("includeZeroToBuy", False)),
        )

    raw_config = str(list_row["generation_config_json"] or "").strip()
    if raw_config:
        try:
            parsed_config = json.loads(raw_config)
        except json.JSONDecodeError:
            parsed_config = None
        if isinstance(parsed_config, dict):
            return build_payload_from_config(parsed_config)

    source_plan_ids = load_shopping_list_source_plan_ids(conn, shopping_list_id)
    if not source_plan_ids and list_row["retreat_plan_id"] is not None:
        source_plan_ids = [int(list_row["retreat_plan_id"])]

    current_plan_ids = [
        int(row["id"])
        for row in conn.execute("SELECT id FROM retreat_plans ORDER BY id").fetchall()
    ]
    all_retreats = shopping_list_name_implies_all_retreats(list_row["name"])
    if not all_retreats and source_plan_ids and current_plan_ids and source_plan_ids == current_plan_ids:
        all_retreats = True

    if not all_retreats and not source_plan_ids:
        raise HTTPException(
            status_code=400,
            detail="This shopping list does not have enough source metadata to refresh.",
        )

    phase = str(list_row["phase"] or "bulk").strip().lower() or "bulk"
    return ShoppingListGeneratePayload(
        retreatPlanId=source_plan_ids[0] if len(source_plan_ids) == 1 and not all_retreats else None,
        retreatPlanIds=[] if all_retreats else source_plan_ids,
        allRetreats=all_retreats,
        phase=phase,
        profile=DEFAULT_SHOPPING_PROFILE,
        subtractInventory=True,
        includeZeroToBuy=False,
    )


def meals_for_profile(plan_payload: dict[str, Any], profile: str) -> list[dict[str, Any]]:
    if profile == "test":
        test_meals = plan_payload.get("testMeals")
        return test_meals if isinstance(test_meals, list) else []

    retreat_meals = plan_payload.get("retreatMeals")
    if isinstance(retreat_meals, list):
        return retreat_meals
    meals = plan_payload.get("meals")
    return meals if isinstance(meals, list) else []


def merge_required_ingredient_aggregate(
    target: dict[tuple[int, str], dict[str, Any]],
    source: dict[tuple[int, str], dict[str, Any]],
) -> None:
    for key, source_entry in source.items():
        existing = target.get(key)
        if not existing:
            target[key] = {
                "ingredient_id": int(source_entry["ingredient_id"]),
                "ingredient_name": source_entry["ingredient_name"],
                "canonical_unit": source_entry["canonical_unit"],
                "required_qty": float(source_entry["required_qty"]),
            }
            continue
        existing["required_qty"] += float(source_entry["required_qty"])


def build_required_ingredients_from_plan(
    conn: Any,
    plan_payload: dict[str, Any],
    profile: str,
    purchase_tiers: set[str],
) -> tuple[dict[tuple[int, str], dict[str, Any]], set[str], dict[tuple[int, str], dict[str, float]]]:
    meals = meals_for_profile(plan_payload, profile)
    dish_names: set[str] = set()
    for meal in meals:
        if not isinstance(meal, dict):
            continue
        dishes = meal.get("dishes")
        if not isinstance(dishes, list):
            continue
        for dish in dishes:
            dish_name = str(dish or "").strip()
            if dish_name:
                dish_names.add(dish_name)

    if not dish_names:
        return {}, set(), {}

    placeholders = ",".join("?" for _ in dish_names)
    recipe_rows = conn.execute(
        f"""
        SELECT id, name, base_servings
        FROM recipes
        WHERE lower(name) IN ({placeholders})
        """,
        tuple(dish.lower() for dish in dish_names),
    ).fetchall()

    recipes_by_name: dict[str, dict[str, Any]] = {
        str(row["name"]).strip().lower(): {
            "id": int(row["id"]),
            "name": row["name"],
            "base_servings": float(row["base_servings"]),
        }
        for row in recipe_rows
    }

    recipe_ids = [recipe["id"] for recipe in recipes_by_name.values()]
    ingredients_by_recipe: dict[int, list[dict[str, Any]]] = {}
    if recipe_ids:
        placeholders = ",".join("?" for _ in recipe_ids)
        ingredient_rows = conn.execute(
            f"""
            SELECT
                ri.recipe_id,
                i.id AS ingredient_id,
                i.name AS ingredient_name,
                i.purchase_tier,
                ri.quantity,
                ri.unit
            FROM recipe_ingredients ri
            JOIN ingredients i ON i.id = ri.ingredient_id
            WHERE ri.recipe_id IN ({placeholders})
              AND lower(i.name) NOT IN ('water', 'hot water')
            """,
            tuple(recipe_ids),
        ).fetchall()

        for row in ingredient_rows:
            recipe_id = int(row["recipe_id"])
            ingredients_by_recipe.setdefault(recipe_id, []).append(
                {
                    "ingredient_id": int(row["ingredient_id"]),
                    "ingredient_name": row["ingredient_name"],
                    "purchase_tier": str(row["purchase_tier"] or "").strip().lower() or None,
                    "quantity": float(row["quantity"]),
                    "unit": str(row["unit"] or "").strip(),
                }
            )

    aggregate: dict[tuple[int, str], dict[str, Any]] = {}
    dish_aggregate: dict[tuple[int, str], dict[str, float]] = {}
    missing_recipes: set[str] = set()
    for meal in meals:
        if not isinstance(meal, dict):
            continue
        people = float(meal.get("people") or 0)
        if people <= 0:
            continue
        dishes = meal.get("dishes")
        if not isinstance(dishes, list):
            continue

        for dish in dishes:
            dish_name = str(dish or "").strip()
            if not dish_name:
                continue

            recipe = recipes_by_name.get(dish_name.lower())
            if not recipe:
                missing_recipes.add(dish_name)
                continue

            base_servings = recipe["base_servings"]
            if base_servings <= 0:
                missing_recipes.add(dish_name)
                continue

            factor = people / base_servings
            for ingredient in ingredients_by_recipe.get(recipe["id"], []):
                tier = ingredient["purchase_tier"]
                if purchase_tiers and tier and tier not in purchase_tiers:
                    continue

                normalized_unit = normalize_unit(ingredient["unit"])
                scaled_qty = ingredient["quantity"] * factor
                canonical_qty, canonical_unit, _note = require_canonical_conversion(
                    ingredient["ingredient_name"],
                    scaled_qty,
                    normalized_unit,
                    context=f"Shopping generation conversion failed for recipe '{recipe['name']}'",
                )

                key = (ingredient["ingredient_id"], canonical_unit)
                entry = aggregate.get(key)
                if not entry:
                    entry = {
                        "ingredient_id": ingredient["ingredient_id"],
                        "ingredient_name": ingredient["ingredient_name"],
                        "canonical_unit": canonical_unit,
                        "required_qty": 0.0,
                    }
                    aggregate[key] = entry
                entry["required_qty"] += canonical_qty

                dish_bucket = dish_aggregate.setdefault(key, {})
                dish_bucket[dish_name] = dish_bucket.get(dish_name, 0.0) + canonical_qty

    return aggregate, missing_recipes, dish_aggregate


def load_inventory_canonical_by_key(conn: Any) -> dict[tuple[int, str], float]:
    rows = conn.execute(
        """
        SELECT
            ii.ingredient_id,
            i.name AS ingredient_name,
            ii.quantity,
            ii.unit,
            ii.source
        FROM inventory_items ii
        JOIN ingredients i ON i.id = ii.ingredient_id
        """
    ).fetchall()

    base_aggregate: dict[tuple[int, str], float] = {}
    override_aggregate: dict[tuple[int, str], float] = {}
    for row in rows:
        quantity = float(row["quantity"] or 0)
        if quantity <= 0:
            continue

        unit = normalize_unit(str(row["unit"] or ""))
        canonical_qty, canonical_unit, _note = require_canonical_conversion(
            row["ingredient_name"],
            quantity,
            unit,
            context="Inventory conversion failed",
        )

        key = (int(row["ingredient_id"]), canonical_unit)
        source = str(row["source"] or "").strip()
        if source == MANUAL_INVENTORY_SOURCE:
            override_aggregate[key] = override_aggregate.get(key, 0.0) + canonical_qty
        else:
            base_aggregate[key] = base_aggregate.get(key, 0.0) + canonical_qty

    for key, value in override_aggregate.items():
        base_aggregate[key] = value

    return base_aggregate


def to_metric_display_unit(canonical_qty: float, canonical_unit: str) -> tuple[float, str]:
    if canonical_unit == "g":
        if canonical_qty >= 1000:
            return canonical_qty / 1000.0, "kg"
        return canonical_qty, "g"

    if canonical_unit == "ml":
        if canonical_qty >= 1000:
            return canonical_qty / 1000.0, "l"
        return canonical_qty, "ml"

    return canonical_qty, canonical_unit


def preferred_metric_unit(canonical_qty: float, canonical_unit: str) -> str:
    if canonical_unit == "g":
        return "kg" if canonical_qty >= 1000 else "g"
    if canonical_unit == "ml":
        return "l" if canonical_qty >= 1000 else "ml"
    return canonical_unit


def canonical_qty_to_unit_or_none(canonical_qty: float, canonical_unit: str, target_unit: str) -> float | None:
    target = normalize_unit(target_unit)
    if canonical_unit == "g" and target in MASS_TO_G:
        return canonical_qty / MASS_TO_G[target]
    if canonical_unit == "ml" and target in VOLUME_TO_ML:
        return canonical_qty / VOLUME_TO_ML[target]
    if canonical_unit == target:
        return canonical_qty
    return None


def canonical_qty_to_unit(canonical_qty: float, canonical_unit: str, target_unit: str) -> float:
    converted = canonical_qty_to_unit_or_none(canonical_qty, canonical_unit, target_unit)
    if converted is None:
        return canonical_qty
    return converted


def quantity_to_canonical(quantity: float, unit: str) -> tuple[float | None, str | None]:
    normalized = normalize_unit(unit)
    if normalized in MASS_TO_G:
        return quantity * MASS_TO_G[normalized], "g"
    if normalized in VOLUME_TO_ML:
        return quantity * VOLUME_TO_ML[normalized], "ml"
    if normalized:
        return quantity, normalized
    return None, None


def normalize_count_style_purchase_unit(unit: str | None) -> str | None:
    normalized = normalize_unit(str(unit or "").strip())
    if not normalized:
        return None
    if normalized in {"piece", "sprig", "leaf", "pinch"}:
        return "each"
    if normalized in {
        "each",
        "bag",
        "box",
        "case",
        "can",
        "packet",
        "pack",
        "bottle",
        "jug",
        "jar",
        "carton",
        "tub",
        "package",
        "bunch",
        "loaf",
    }:
        return normalized
    return None


def convert_quantity_between_units(
    quantity: float | None,
    from_unit: str | None,
    to_unit: str | None,
) -> float | None:
    try:
        numeric_quantity = float(quantity)
    except (TypeError, ValueError):
        return None

    normalized_from = normalize_unit(str(from_unit or "").strip())
    normalized_to = normalize_unit(str(to_unit or "").strip())
    if not normalized_from or not normalized_to:
        return None
    if normalized_from == normalized_to:
        return numeric_quantity

    canonical_qty, canonical_unit = quantity_to_canonical(numeric_quantity, normalized_from)
    if canonical_qty is None or not canonical_unit:
        return None
    return canonical_qty_to_unit_or_none(canonical_qty, canonical_unit, normalized_to)


def preferred_ordered_quantity_and_unit(
    quantity: float | None,
    unit: str | None,
    preferred_unit: str | None = None,
) -> tuple[float | None, str | None]:
    try:
        numeric_quantity = float(quantity)
    except (TypeError, ValueError):
        numeric_quantity = 0.0
    normalized_unit = normalize_unit(str(unit or "").strip())
    normalized_preferred_unit = normalize_unit(str(preferred_unit or "").strip()) or None

    if normalized_preferred_unit:
        if numeric_quantity > 0:
            converted = convert_quantity_between_units(
                numeric_quantity,
                normalized_unit,
                normalized_preferred_unit,
            )
            if converted is not None:
                return round(converted, 4), normalized_preferred_unit
            if normalize_count_style_purchase_unit(normalized_preferred_unit):
                fallback_qty = (
                    1.0
                    if normalized_unit in MASS_TO_G or normalized_unit in VOLUME_TO_ML
                    else numeric_quantity
                )
                return round(fallback_qty, 4), normalized_preferred_unit
            return round(numeric_quantity, 4), normalized_preferred_unit
        return None, normalized_preferred_unit

    canonical_qty, canonical_unit = quantity_to_canonical(numeric_quantity, normalized_unit)
    if numeric_quantity > 0 and canonical_qty is not None and canonical_unit == "g":
        target_unit = "lb" if canonical_qty >= MASS_TO_G["lb"] else "oz"
        converted = canonical_qty_to_unit_or_none(canonical_qty, canonical_unit, target_unit)
        if converted is not None:
            return round(converted, 4), target_unit
    if numeric_quantity > 0 and canonical_qty is not None and canonical_unit == "ml":
        if canonical_qty >= VOLUME_TO_ML["gal"]:
            target_unit = "gal"
        elif canonical_qty >= VOLUME_TO_ML["qt"]:
            target_unit = "qt"
        else:
            target_unit = "fl oz"
        converted = canonical_qty_to_unit_or_none(canonical_qty, canonical_unit, target_unit)
        if converted is not None:
            return round(converted, 4), target_unit

    return (
        round(numeric_quantity, 4) if numeric_quantity > 0 else None,
        normalize_count_style_purchase_unit(normalized_unit) or "each",
    )


def derive_shopping_item_status(ordered: bool, received: bool) -> str:
    if received:
        return "received"
    if ordered:
        return "ordered"
    return "open"


def derive_progress_status(total_count: int, ordered_count: int, received_count: int) -> str:
    if total_count <= 0:
        return "draft"
    if received_count >= total_count:
        return "received"
    if ordered_count > 0:
        return "in_progress"
    return "draft"


def shopping_item_source_key(
    ingredient_id: Any,
    quantity: Any,
    unit: Any,
) -> tuple[int, str] | None:
    if ingredient_id is None:
        return None
    canonical_qty, canonical_unit = quantity_to_canonical(
        float(quantity or 0.0),
        str(unit or "").strip(),
    )
    if canonical_qty is None or not canonical_unit:
        return None
    return int(ingredient_id), canonical_unit


def ensure_vendor_exists(conn: Any, vendor_id: int) -> Any:
    row = conn.execute(
        "SELECT id, name FROM vendors WHERE id = ?",
        (vendor_id,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="Vendor not found")
    return row


def ensure_shopping_list_exists(conn: Any, shopping_list_id: int) -> Any:
    row = conn.execute(
        """
        SELECT id, name, phase, retreat_plan_id
        FROM shopping_lists
        WHERE id = ?
        """,
        (shopping_list_id,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Shopping list not found")
    return row


def ensure_shopping_pickup_list_exists(conn: Any, pickup_list_id: int) -> Any:
    row = conn.execute(
        """
        SELECT id, shopping_list_id, name, vendor_id, assignee, pickup_date, notes, created_at, updated_at
        FROM shopping_pickup_lists
        WHERE id = ?
        """,
        (pickup_list_id,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Pickup list not found")
    return row


def shopping_vendor_allocations_supported(conn: Any) -> bool:
    return table_exists(conn, "shopping_list_item_vendor_allocations")


def load_shopping_list_item_vendor_allocations_by_item_id(
    conn: Any,
    shopping_list_item_ids: list[int],
) -> dict[int, list[dict[str, Any]]]:
    normalized_ids = [int(item_id) for item_id in shopping_list_item_ids if int(item_id) > 0]
    if not normalized_ids or not shopping_vendor_allocations_supported(conn):
        return {}

    placeholders = ",".join("?" for _ in normalized_ids)
    rows = conn.execute(
        f"""
        SELECT
            sliva.id,
            sliva.shopping_list_item_id,
            sliva.vendor_id,
            v.name AS vendor_name,
            sliva.allocated_qty,
            sliva.allocated_unit,
            sliva.ordered,
            sliva.ordered_at,
            sliva.received,
            sliva.received_at,
            sliva.sort_order
        FROM shopping_list_item_vendor_allocations sliva{" NOT INDEXED" if conn.backend == "sqlite" else ""}
        LEFT JOIN vendors v ON v.id = sliva.vendor_id
        WHERE sliva.shopping_list_item_id IN ({placeholders})
        ORDER BY sliva.id
        """,
        tuple(normalized_ids),
    ).fetchall()

    allocations_by_item_id: dict[int, list[dict[str, Any]]] = {}
    for row in sorted(
        rows,
        key=lambda entry: (
            int(entry["shopping_list_item_id"]),
            int(entry["sort_order"] or 0),
            int(entry["id"]),
        ),
    ):
        item_id = int(row["shopping_list_item_id"])
        allocations_by_item_id.setdefault(item_id, []).append(
            {
                "id": int(row["id"]),
                "vendor_id": int(row["vendor_id"]) if row["vendor_id"] is not None else None,
                "vendor_name": normalize_optional_text(row["vendor_name"]),
                "allocated_qty": float(row["allocated_qty"] or 0.0),
                "allocated_unit": normalize_unit(str(row["allocated_unit"] or "").strip()) or None,
                "ordered": bool(row["ordered"]),
                "ordered_at": row["ordered_at"],
                "received": bool(row["received"]),
                "received_at": row["received_at"],
                "sort_order": int(row["sort_order"] or 0),
            }
        )
    return allocations_by_item_id


def fallback_vendor_allocation_from_item_row(row: Any) -> list[dict[str, Any]]:
    base_unit = normalize_unit(
        str(row["ordered_unit"] or row["to_buy_unit"] or row["required_unit"] or "").strip()
    ) or None
    if not base_unit:
        return []

    allocated_qty = rounded_quantity(
        row["ordered_qty"]
        if row["ordered_qty"] is not None
        else row["to_buy_qty"]
        if row["to_buy_qty"] is not None
        else row["required_qty"]
    )
    if allocated_qty <= 0 and row["vendor_id"] is None and not bool(row["ordered"]) and not bool(row["received"]):
        return []

    return [
        {
            "id": None,
            "vendor_id": int(row["vendor_id"]) if row["vendor_id"] is not None else None,
            "vendor_name": normalize_optional_text(row["vendor_name"]),
            "allocated_qty": allocated_qty,
            "allocated_unit": base_unit,
            "ordered": bool(row["ordered"]),
            "ordered_at": row["ordered_at"],
            "received": bool(row["received"]),
            "received_at": row["received_at"],
            "sort_order": 0,
        }
    ]


def summarize_shopping_vendor_allocations(
    allocations: list[dict[str, Any]],
    *,
    preferred_unit: str | None,
) -> dict[str, Any]:
    rows = [dict(entry) for entry in allocations if isinstance(entry, dict)]
    if not rows:
        return {
            "vendor_id": None,
            "ordered_qty": None,
            "ordered_unit": None,
            "ordered": False,
            "ordered_at": None,
            "received": False,
            "received_at": None,
            "status": "open",
        }

    normalized_preferred_unit = normalize_unit(str(preferred_unit or "").strip()) or None
    vendor_ids = [row.get("vendor_id") for row in rows if row.get("vendor_id") is not None]
    positive_rows = [
        row for row in rows
        if rounded_quantity(row.get("allocated_qty")) > 0 or bool(row.get("ordered")) or bool(row.get("received"))
    ]
    ordered_rows = [row for row in rows if bool(row.get("ordered")) or bool(row.get("received"))]
    received_rows = [row for row in rows if bool(row.get("received"))]

    ordered = bool(ordered_rows)
    received = bool(positive_rows) and len(received_rows) == len(positive_rows)

    total_qty: float | None = 0.0 if normalized_preferred_unit else None
    for row in rows:
        allocated_qty = rounded_quantity(row.get("allocated_qty"))
        allocated_unit = normalize_unit(str(row.get("allocated_unit") or "").strip()) or None
        if allocated_qty <= 0:
            continue
        if normalized_preferred_unit is None:
            total_qty = None
            break
        converted = convert_quantity_between_units(
            allocated_qty,
            allocated_unit,
            normalized_preferred_unit,
        )
        if converted is None:
            total_qty = None
            break
        total_qty += converted

    ordered_at_values = [row.get("ordered_at") for row in ordered_rows if row.get("ordered_at")]
    received_at_values = [row.get("received_at") for row in received_rows if row.get("received_at")]

    ordered_unit = normalized_preferred_unit
    ordered_qty: float | None = round(total_qty, 4) if total_qty is not None and total_qty > 0 else None
    if ordered_qty is None and len(rows) == 1:
        single_unit = normalize_unit(str(rows[0].get("allocated_unit") or "").strip()) or None
        single_qty = rounded_quantity(rows[0].get("allocated_qty"))
        if single_unit and single_qty > 0:
            ordered_unit = single_unit
            ordered_qty = single_qty

    return {
        "vendor_id": vendor_ids[0] if len(vendor_ids) == 1 and len(rows) == 1 else None,
        "ordered_qty": ordered_qty,
        "ordered_unit": ordered_unit if ordered_qty is not None else None,
        "ordered": ordered,
        "ordered_at": min(ordered_at_values) if ordered_at_values else None,
        "received": received,
        "received_at": max(received_at_values) if received and received_at_values else None,
        "status": derive_shopping_item_status(ordered=ordered, received=received),
    }


def replace_shopping_item_vendor_allocations(
    conn: Any,
    shopping_list_item_id: int,
    allocations: list[dict[str, Any]],
) -> None:
    if not shopping_vendor_allocations_supported(conn):
        return
    if conn.backend == "sqlite":
        existing_rows = conn.execute(
            """
            SELECT id
            FROM shopping_list_item_vendor_allocations NOT INDEXED
            WHERE shopping_list_item_id = ?
            ORDER BY id
            """,
            (shopping_list_item_id,),
        ).fetchall()
        existing_ids = [int(row["id"]) for row in existing_rows if row["id"] is not None]
        if existing_ids:
            placeholders = ",".join("?" for _ in existing_ids)
            conn.execute(
                f"DELETE FROM shopping_list_item_vendor_allocations WHERE id IN ({placeholders})",
                tuple(existing_ids),
            )
    else:
        conn.execute(
            "DELETE FROM shopping_list_item_vendor_allocations WHERE shopping_list_item_id = ?",
            (shopping_list_item_id,),
        )
    for index, allocation in enumerate(allocations):
        conn.execute(
            """
            INSERT INTO shopping_list_item_vendor_allocations(
                shopping_list_item_id,
                vendor_id,
                allocated_qty,
                allocated_unit,
                ordered,
                ordered_at,
                received,
                received_at,
                sort_order,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (
                shopping_list_item_id,
                allocation.get("vendor_id"),
                round(float(allocation.get("allocated_qty") or 0.0), 4),
                normalize_unit(str(allocation.get("allocated_unit") or "").strip()) or "each",
                1 if bool(allocation.get("ordered")) else 0,
                allocation.get("ordered_at"),
                1 if bool(allocation.get("received")) else 0,
                allocation.get("received_at"),
                int(allocation.get("sort_order") if allocation.get("sort_order") is not None else index),
            ),
        )


def load_existing_shopping_list_item_metadata(
    conn: Any,
    shopping_list_id: int,
) -> dict[tuple[int, str], dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT
            id,
            ingredient_id,
            required_qty,
            required_unit,
            vendor_id,
            ordered_qty,
            ordered_unit,
            owner,
            pickup_date,
            ordered,
            ordered_at,
            received,
            received_at,
            notes
        FROM shopping_list_items
        WHERE shopping_list_id = ?
        ORDER BY id
        """,
        (shopping_list_id,),
    ).fetchall()
    allocation_rows_by_item_id = load_shopping_list_item_vendor_allocations_by_item_id(
        conn,
        [int(row["id"]) for row in rows if row["id"] is not None],
    )

    preserved: dict[tuple[int, str], dict[str, Any]] = {}
    duplicate_keys: set[tuple[int, str]] = set()
    for row in rows:
        required_qty = float(row["required_qty"] or 0.0)
        required_unit = str(row["required_unit"] or "").strip()
        _canonical_qty, canonical_unit = quantity_to_canonical(required_qty, required_unit)
        if not canonical_unit:
            continue
        key = (int(row["ingredient_id"]), canonical_unit)
        if key in preserved:
            duplicate_keys.add(key)
            continue
        ordered = bool(row["ordered"])
        received = bool(row["received"])
        ordered_unit = normalize_unit(str(row["ordered_unit"] or "").strip()) or None
        if ordered_unit in {"g", "kg", "ml", "l"} and row["ordered_qty"] is None and not ordered and not received:
            ordered_unit = None
        item_id = int(row["id"])
        vendor_allocations = allocation_rows_by_item_id.get(item_id, [])
        preserved[key] = {
            "vendor_id": int(row["vendor_id"]) if row["vendor_id"] is not None else None,
            "ordered_qty": float(row["ordered_qty"]) if row["ordered_qty"] is not None else None,
            "ordered_unit": ordered_unit,
            "owner": str(row["owner"] or "").strip() or None,
            "pickup_date": str(row["pickup_date"] or "").strip() or None,
            "ordered": ordered,
            "ordered_at": row["ordered_at"] if ordered else None,
            "received": received,
            "received_at": row["received_at"] if received else None,
            "notes": str(row["notes"] or "").strip() or None,
            "vendor_allocations": [
                {
                    "vendor_id": entry.get("vendor_id"),
                    "allocated_qty": round(float(entry.get("allocated_qty") or 0.0), 4),
                    "allocated_unit": normalize_unit(str(entry.get("allocated_unit") or "").strip()) or ordered_unit or required_unit,
                    "ordered": bool(entry.get("ordered")),
                    "ordered_at": entry.get("ordered_at") if bool(entry.get("ordered")) else None,
                    "received": bool(entry.get("received")),
                    "received_at": entry.get("received_at") if bool(entry.get("received")) else None,
                    "sort_order": int(entry.get("sort_order") or 0),
                }
                for entry in vendor_allocations
            ],
        }

    for key in duplicate_keys:
        preserved.pop(key, None)
    return preserved


def unique_shopping_pickup_list_name(
    conn: Any,
    shopping_list_id: int,
    base_name: str,
    exclude_pickup_list_id: int | None = None,
) -> str:
    seed = " ".join(str(base_name or "").strip().split()) or "Pickup List"
    candidate = seed
    suffix = 2
    while True:
        if exclude_pickup_list_id is None:
            exists = conn.execute(
                """
                SELECT 1
                FROM shopping_pickup_lists
                WHERE shopping_list_id = ?
                  AND lower(name) = lower(?)
                """,
                (shopping_list_id, candidate),
            ).fetchone()
        else:
            exists = conn.execute(
                """
                SELECT 1
                FROM shopping_pickup_lists
                WHERE shopping_list_id = ?
                  AND lower(name) = lower(?)
                  AND id != ?
                """,
                (shopping_list_id, candidate, exclude_pickup_list_id),
            ).fetchone()
        if not exists:
            return candidate
        candidate = f"{seed} ({suffix})"
        suffix += 1


def load_shopping_list_item_source_key_map(
    conn: Any,
    shopping_list_id: int,
) -> dict[tuple[int, str], int]:
    rows = conn.execute(
        """
        SELECT id, ingredient_id, required_qty, required_unit
        FROM shopping_list_items
        WHERE shopping_list_id = ?
        ORDER BY id
        """,
        (shopping_list_id,),
    ).fetchall()

    mapping: dict[tuple[int, str], int] = {}
    duplicates: set[tuple[int, str]] = set()
    for row in rows:
        key = shopping_item_source_key(
            row["ingredient_id"],
            row["required_qty"],
            row["required_unit"],
        )
        if key is None:
            continue
        if key in mapping:
            duplicates.add(key)
            continue
        mapping[key] = int(row["id"])

    for key in duplicates:
        mapping.pop(key, None)
    return mapping


def relink_shopping_pickup_list_items(conn: Any, shopping_list_id: int) -> None:
    if not table_exists(conn, "shopping_pickup_list_items"):
        return

    item_id_by_source_key = load_shopping_list_item_source_key_map(conn, shopping_list_id)
    rows = conn.execute(
        """
        SELECT
            ppli.id,
            ppli.source_ingredient_id,
            ppli.source_canonical_unit
        FROM shopping_pickup_list_items ppli
        JOIN shopping_pickup_lists ppl
          ON ppl.id = ppli.shopping_pickup_list_id
        WHERE ppl.shopping_list_id = ?
        """,
        (shopping_list_id,),
    ).fetchall()

    for row in rows:
        source_key = (
            int(row["source_ingredient_id"]),
            str(row["source_canonical_unit"] or "").strip(),
        )
        next_item_id = item_id_by_source_key.get(source_key)
        conn.execute(
            """
            UPDATE shopping_pickup_list_items
            SET shopping_list_item_id = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (next_item_id, int(row["id"])),
        )


def derive_pickup_list_status(
    total_count: int,
    ordered_count: int,
    received_count: int,
    missing_item_count: int,
) -> str:
    if missing_item_count > 0:
        return "needs_review"
    return derive_progress_status(total_count, ordered_count, received_count)


def load_shopping_pickup_list_detail(conn: Any, pickup_list_id: int) -> dict[str, Any]:
    list_row = conn.execute(
        """
        SELECT
            ppl.id,
            ppl.shopping_list_id,
            ppl.name,
            ppl.vendor_id,
            v.name AS vendor_name,
            ppl.assignee,
            ppl.pickup_date,
            ppl.notes,
            ppl.created_at,
            ppl.updated_at
        FROM shopping_pickup_lists ppl
        LEFT JOIN vendors v ON v.id = ppl.vendor_id
        WHERE ppl.id = ?
        """,
        (pickup_list_id,),
    ).fetchone()
    if not list_row:
        raise HTTPException(status_code=404, detail="Pickup list not found")

    item_rows = conn.execute(
        """
        SELECT
            ppli.id,
            ppli.shopping_list_item_id,
            ppli.source_ingredient_id,
            ppli.source_canonical_unit,
            ppli.sort_order,
            sli.ordered,
            sli.received,
            COALESCE(i.name, ('Unknown ingredient #' || ppli.source_ingredient_id)) AS ingredient_name
        FROM shopping_pickup_list_items ppli
        LEFT JOIN shopping_list_items sli ON sli.id = ppli.shopping_list_item_id
        LEFT JOIN ingredients i ON i.id = ppli.source_ingredient_id
        WHERE ppli.shopping_pickup_list_id = ?
        ORDER BY ppli.sort_order, ppli.id
        """,
        (pickup_list_id,),
    ).fetchall()

    item_ids: list[int] = []
    missing_items: list[dict[str, Any]] = []
    ordered_count = 0
    received_count = 0
    for row in item_rows:
        linked_item_id = row["shopping_list_item_id"]
        if linked_item_id is None:
            missing_items.append(
                {
                    "ingredient_id": int(row["source_ingredient_id"]),
                    "ingredient_name": str(row["ingredient_name"] or "").strip() or "Unknown ingredient",
                    "canonical_unit": str(row["source_canonical_unit"] or "").strip() or None,
                }
            )
            continue
        item_ids.append(int(linked_item_id))
        if bool(row["ordered"]):
            ordered_count += 1
        if bool(row["received"]):
            received_count += 1

    total_count = len(item_rows)
    missing_item_count = len(missing_items)
    return {
        "id": int(list_row["id"]),
        "shopping_list_id": int(list_row["shopping_list_id"]),
        "name": list_row["name"],
        "vendor_id": int(list_row["vendor_id"]) if list_row["vendor_id"] is not None else None,
        "vendor_name": normalize_optional_text(list_row["vendor_name"]),
        "assignee": normalize_optional_text(list_row["assignee"]),
        "pickup_date": normalize_optional_text(list_row["pickup_date"]),
        "notes": normalize_optional_text(list_row["notes"]),
        "created_at": list_row["created_at"],
        "updated_at": list_row["updated_at"],
        "item_count": total_count,
        "ordered_count": ordered_count,
        "received_count": received_count,
        "missing_item_count": missing_item_count,
        "status": derive_pickup_list_status(
            total_count,
            ordered_count,
            received_count,
            missing_item_count,
        ),
        "item_ids": item_ids,
        "missing_items": missing_items,
    }


def load_shopping_pickup_lists_for_shopping_list(
    conn: Any,
    shopping_list_id: int,
) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT id
        FROM shopping_pickup_lists
        WHERE shopping_list_id = ?
        ORDER BY lower(name), id
        """,
        (shopping_list_id,),
    ).fetchall()
    return [load_shopping_pickup_list_detail(conn, int(row["id"])) for row in rows]


def create_saved_shopping_pickup_list(
    conn: Any,
    *,
    shopping_list_id: int,
    item_ids: list[int],
    requested_name: str | None,
    vendor_id: int | None,
    assignee: str | None,
    pickup_date: str | None,
    notes: str | None,
) -> dict[str, Any]:
    shopping_list = ensure_shopping_list_exists(conn, shopping_list_id)
    if vendor_id is not None:
        vendor_row = ensure_vendor_exists(conn, vendor_id)
        vendor_name = str(vendor_row["name"] or "").strip() or None
    else:
        vendor_name = None

    normalized_item_ids: list[int] = []
    seen_ids: set[int] = set()
    for item_id in item_ids:
        numeric = int(item_id)
        if numeric <= 0 or numeric in seen_ids:
            continue
        normalized_item_ids.append(numeric)
        seen_ids.add(numeric)
    if not normalized_item_ids:
        raise HTTPException(status_code=400, detail="Select at least one shopping item.")

    placeholders = ",".join("?" for _ in normalized_item_ids)
    item_rows = conn.execute(
        f"""
        SELECT id, ingredient_id, required_qty, required_unit
        FROM shopping_list_items
        WHERE shopping_list_id = ?
          AND id IN ({placeholders})
        ORDER BY id
        """,
        (shopping_list_id, *normalized_item_ids),
    ).fetchall()
    if len(item_rows) != len(normalized_item_ids):
        raise HTTPException(status_code=400, detail="One or more selected items are no longer in this shopping list.")

    item_row_by_id = {int(row["id"]): row for row in item_rows}
    default_name = (
        f"{vendor_name} Run"
        if vendor_name
        else f"{shopping_list['name']} Pickup"
    )
    final_name = unique_shopping_pickup_list_name(
        conn,
        shopping_list_id,
        requested_name or default_name,
    )

    created = conn.execute(
        """
        INSERT INTO shopping_pickup_lists(
            shopping_list_id,
            name,
            vendor_id,
            assignee,
            pickup_date,
            notes
        )
        VALUES (?, ?, ?, ?, ?, ?)
        RETURNING id
        """,
        (
            shopping_list_id,
            final_name,
            vendor_id,
            assignee,
            pickup_date,
            notes,
        ),
    ).fetchone()
    pickup_list_id = int(created["id"])

    seen_source_keys: set[tuple[int, str]] = set()
    for sort_order, item_id in enumerate(normalized_item_ids):
        row = item_row_by_id[item_id]
        source_key = shopping_item_source_key(
            row["ingredient_id"],
            row["required_qty"],
            row["required_unit"],
        )
        if source_key is None:
            raise HTTPException(status_code=400, detail="A selected shopping item could not be saved.")
        if source_key in seen_source_keys:
            raise HTTPException(
                status_code=400,
                detail="Selected items include duplicate ingredient/unit lines that cannot be saved in one pickup list.",
            )
        source_ingredient_id, source_canonical_unit = source_key
        seen_source_keys.add(source_key)
        conn.execute(
            """
            INSERT INTO shopping_pickup_list_items(
                shopping_pickup_list_id,
                shopping_list_item_id,
                source_ingredient_id,
                source_canonical_unit,
                sort_order
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                pickup_list_id,
                item_id,
                source_ingredient_id,
                source_canonical_unit,
                sort_order,
            ),
        )

    return load_shopping_pickup_list_detail(conn, pickup_list_id)


def materialize_shopping_list(
    conn: Any,
    payload: ShoppingListGeneratePayload,
    *,
    shopping_list_id: int | None = None,
    fixed_name: str | None = None,
    preserve_existing_metadata: bool = False,
    allow_empty_result: bool = False,
) -> dict[str, Any]:
    purchase_tiers = resolve_purchase_tiers_for_shopping(payload.phase, payload.purchaseTiers)
    aggregate: dict[tuple[int, str], dict[str, Any]] = {}
    missing_recipes: set[str] = set()
    included_plan_ids: list[int] = []
    plan_dish_breakdown_by_plan_id: dict[int, dict[tuple[int, str], dict[str, float]]] = {}
    plan_name_by_id: dict[int, str] = {}
    retreat_plan_id_for_list: int | None = None

    if payload.allRetreats:
        plan_rows = conn.execute(
            """
            SELECT id, name, plan_json
            FROM retreat_plans
            ORDER BY id
            """
        ).fetchall()
        if not plan_rows:
            raise HTTPException(status_code=404, detail="No retreat plans found")

        for plan_row in plan_rows:
            try:
                plan_payload = json.loads(plan_row["plan_json"]) if plan_row["plan_json"] else {}
            except json.JSONDecodeError as exc:
                raise HTTPException(
                    status_code=400,
                    detail=f"Retreat plan #{int(plan_row['id'])} payload is not valid JSON",
                ) from exc
            if not isinstance(plan_payload, dict):
                raise HTTPException(
                    status_code=400,
                    detail=f"Retreat plan #{int(plan_row['id'])} payload is malformed",
                )

            plan_aggregate, plan_missing, plan_dish_breakdown = build_required_ingredients_from_plan(
                conn,
                plan_payload=plan_payload,
                profile=payload.profile,
                purchase_tiers=purchase_tiers,
            )
            if plan_aggregate:
                plan_id = int(plan_row["id"])
                merge_required_ingredient_aggregate(aggregate, plan_aggregate)
                included_plan_ids.append(plan_id)
                plan_dish_breakdown_by_plan_id[plan_id] = plan_dish_breakdown
                plan_name_by_id[plan_id] = (
                    str(plan_row["name"] or f"Retreat #{plan_id}").strip() or f"Retreat #{plan_id}"
                )
            missing_recipes.update(plan_missing)
    else:
        requested_plan_ids = resolve_requested_shopping_plan_ids(payload)
        if not requested_plan_ids:
            raise HTTPException(
                status_code=400,
                detail="Select at least one retreat plan unless allRetreats=true",
            )

        placeholders = ",".join("?" for _ in requested_plan_ids)
        plan_rows = conn.execute(
            f"""
            SELECT id, name, plan_json
            FROM retreat_plans
            WHERE id IN ({placeholders})
            ORDER BY id
            """,
            tuple(requested_plan_ids),
        ).fetchall()
        found_ids = sorted(int(row["id"]) for row in plan_rows)
        if found_ids != requested_plan_ids:
            missing_ids = sorted(set(requested_plan_ids) - set(found_ids))
            raise HTTPException(
                status_code=404,
                detail=f"Retreat plan(s) not found: {', '.join(str(x) for x in missing_ids)}",
            )

        for plan_row in plan_rows:
            try:
                plan_payload = json.loads(plan_row["plan_json"]) if plan_row["plan_json"] else {}
            except json.JSONDecodeError as exc:
                raise HTTPException(
                    status_code=400,
                    detail=f"Retreat plan #{int(plan_row['id'])} payload is not valid JSON",
                ) from exc
            if not isinstance(plan_payload, dict):
                raise HTTPException(
                    status_code=400,
                    detail=f"Retreat plan #{int(plan_row['id'])} payload is malformed",
                )

            plan_aggregate, plan_missing, plan_dish_breakdown = build_required_ingredients_from_plan(
                conn,
                plan_payload=plan_payload,
                profile=payload.profile,
                purchase_tiers=purchase_tiers,
            )
            if plan_aggregate:
                plan_id = int(plan_row["id"])
                merge_required_ingredient_aggregate(aggregate, plan_aggregate)
                included_plan_ids.append(plan_id)
                plan_dish_breakdown_by_plan_id[plan_id] = plan_dish_breakdown
                plan_name_by_id[plan_id] = (
                    str(plan_row["name"] or f"Retreat #{plan_id}").strip() or f"Retreat #{plan_id}"
                )
            missing_recipes.update(plan_missing)

        if len(requested_plan_ids) == 1:
            retreat_plan_id_for_list = requested_plan_ids[0]

    if not aggregate and shopping_list_id is None:
        raise HTTPException(
            status_code=400,
            detail="No ingredients found for this profile and tier filter.",
        )

    inventory_by_key = load_inventory_canonical_by_key(conn) if payload.subtractInventory else {}
    config_json = json.dumps(
        normalize_shopping_list_generation_config(payload),
        ensure_ascii=True,
        separators=(",", ":"),
    )

    existing_metadata_by_key: dict[tuple[int, str], dict[str, Any]] = {}
    if shopping_list_id is None:
        label = fixed_name.strip() if fixed_name and fixed_name.strip() else None
        if not label:
            if payload.allRetreats:
                label = f"All Retreats - {payload.phase.title()} Order"
            else:
                if len(included_plan_ids) == 1:
                    plan_id_for_label = included_plan_ids[0]
                    plan_name_row = conn.execute(
                        "SELECT name FROM retreat_plans WHERE id = ?",
                        (plan_id_for_label,),
                    ).fetchone()
                    plan_name = plan_name_row["name"] if plan_name_row else "Retreat"
                    label = f"{plan_name} - {payload.phase.title()} Order"
                else:
                    label = f"Selected Retreats - {payload.phase.title()} Order"
        label = unique_shopping_list_name(conn, label)

        created = conn.execute(
            """
            INSERT INTO shopping_lists(retreat_plan_id, name, phase, generation_config_json, status)
            VALUES (?, ?, ?, ?, 'draft')
            RETURNING id
            """,
            (retreat_plan_id_for_list, label, payload.phase, config_json),
        ).fetchone()
        shopping_list_id = int(created["id"])
    else:
        existing_list = conn.execute(
            """
            SELECT id, name
            FROM shopping_lists
            WHERE id = ?
            """,
            (shopping_list_id,),
        ).fetchone()
        if not existing_list:
            raise HTTPException(status_code=404, detail="Shopping list not found")

        if preserve_existing_metadata:
            existing_metadata_by_key = load_existing_shopping_list_item_metadata(conn, shopping_list_id)

        label = fixed_name.strip() if fixed_name and fixed_name.strip() else str(existing_list["name"] or "").strip()
        if not label:
            label = f"Shopping List #{shopping_list_id}"

        conn.execute("DELETE FROM shopping_list_items WHERE shopping_list_id = ?", (shopping_list_id,))
        conn.execute(
            """
            UPDATE shopping_lists
            SET retreat_plan_id = ?,
                name = ?,
                phase = ?,
                generation_config_json = ?,
                status = 'draft'
            WHERE id = ?
            """,
            (retreat_plan_id_for_list, label, payload.phase, config_json, shopping_list_id),
        )

    inserted_items = 0
    ingredient_category_by_id: dict[int, str | None] = {}
    ingredient_ids = sorted({int(entry["ingredient_id"]) for entry in aggregate.values()})
    if ingredient_ids:
        placeholders = ",".join("?" for _ in ingredient_ids)
        category_rows = conn.execute(
            f"SELECT id, category FROM ingredients WHERE id IN ({placeholders})",
            tuple(ingredient_ids),
        ).fetchall()
        ingredient_category_by_id = {
            int(row["id"]): (str(row["category"] or "").strip() or None)
            for row in category_rows
        }

    for key, entry in sorted(
        aggregate.items(),
        key=lambda item: item[1]["ingredient_name"].lower(),
    ):
        ingredient_id, canonical_unit = key
        required_canonical = float(entry["required_qty"])
        in_stock_canonical = float(inventory_by_key.get(key, 0.0))
        to_buy_canonical = max(required_canonical - in_stock_canonical, 0.0)
        if not payload.includeZeroToBuy and to_buy_canonical <= 0:
            continue

        ingredient_category = ingredient_category_by_id.get(int(ingredient_id))
        if canonical_unit == "g" and is_spice_or_seasoning_category(ingredient_category):
            row_unit = "g"
        elif canonical_unit == "g" and is_produce_category(ingredient_category):
            row_unit = "kg"
        else:
            row_unit = preferred_metric_unit(required_canonical, canonical_unit)
        required_qty = canonical_qty_to_unit(required_canonical, canonical_unit, row_unit)
        in_stock_qty = canonical_qty_to_unit(in_stock_canonical, canonical_unit, row_unit)
        to_buy_qty = canonical_qty_to_unit(to_buy_canonical, canonical_unit, row_unit)
        preserved = existing_metadata_by_key.get((int(ingredient_id), canonical_unit))
        preserved_vendor_allocations: list[dict[str, Any]] = []
        if preserved:
            for allocation in preserved.get("vendor_allocations") or []:
                allocated_qty = round(float(allocation.get("allocated_qty") or 0.0), 4)
                allocated_unit = normalize_unit(str(allocation.get("allocated_unit") or "").strip()) or row_unit
                if allocated_qty > 0 and allocated_unit != row_unit:
                    converted_qty = convert_quantity_between_units(allocated_qty, allocated_unit, row_unit)
                    if converted_qty is not None:
                        allocated_qty = round(converted_qty, 4)
                        allocated_unit = row_unit
                preserved_vendor_allocations.append(
                    {
                        "vendor_id": allocation.get("vendor_id"),
                        "allocated_qty": allocated_qty,
                        "allocated_unit": allocated_unit or row_unit,
                        "ordered": bool(allocation.get("ordered")),
                        "ordered_at": allocation.get("ordered_at") if bool(allocation.get("ordered")) else None,
                        "received": bool(allocation.get("received")),
                        "received_at": allocation.get("received_at") if bool(allocation.get("received")) else None,
                        "sort_order": int(allocation.get("sort_order") or len(preserved_vendor_allocations)),
                    }
                )
        allocation_summary = summarize_shopping_vendor_allocations(
            preserved_vendor_allocations,
            preferred_unit=row_unit,
        ) if preserved_vendor_allocations else None
        ordered = bool(preserved["ordered"]) if preserved else False
        received = bool(preserved["received"]) if preserved else False
        if allocation_summary:
            ordered = bool(allocation_summary["ordered"])
            received = bool(allocation_summary["received"])

        created_item = conn.execute(
            """
            INSERT INTO shopping_list_items(
                shopping_list_id,
                ingredient_id,
                required_qty,
                required_unit,
                in_stock_qty,
                in_stock_unit,
                to_buy_qty,
                to_buy_unit,
                ordered_qty,
                ordered_unit,
                vendor_id,
                owner,
                pickup_date,
                ordered,
                ordered_at,
                received,
                received_at,
                status,
                notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            RETURNING id
            """,
            (
                shopping_list_id,
                ingredient_id,
                round(required_qty, 4),
                row_unit,
                round(in_stock_qty, 4),
                row_unit,
                round(to_buy_qty, 4),
                row_unit,
                (
                    allocation_summary["ordered_qty"]
                    if allocation_summary and allocation_summary["ordered_qty"] is not None
                    else round(float(preserved["ordered_qty"]), 4)
                    if preserved and preserved["ordered_qty"] is not None
                    else None
                ),
                (
                    allocation_summary["ordered_unit"]
                    if allocation_summary and allocation_summary["ordered_unit"]
                    else preserved["ordered_unit"]
                    if preserved and preserved["ordered_unit"]
                    else None
                ),
                (
                    allocation_summary["vendor_id"]
                    if allocation_summary
                    else preserved["vendor_id"] if preserved else None
                ),
                preserved["owner"] if preserved else None,
                preserved["pickup_date"] if preserved else None,
                1 if ordered else 0,
                (
                    allocation_summary["ordered_at"]
                    if allocation_summary and ordered
                    else preserved["ordered_at"] if preserved and ordered else None
                ),
                1 if received else 0,
                (
                    allocation_summary["received_at"]
                    if allocation_summary and received
                    else preserved["received_at"] if preserved and received else None
                ),
                (
                    allocation_summary["status"]
                    if allocation_summary
                    else derive_shopping_item_status(ordered=ordered, received=received)
                ),
                preserved["notes"] if preserved else None,
            ),
        ).fetchone()
        shopping_list_item_id = int(created_item["id"])
        if preserved_vendor_allocations:
            replace_shopping_item_vendor_allocations(conn, shopping_list_item_id, preserved_vendor_allocations)

        for plan_id, plan_dish_breakdown in plan_dish_breakdown_by_plan_id.items():
            dish_required_by_name = plan_dish_breakdown.get(key) or {}
            if not dish_required_by_name:
                continue

            for dish_name, dish_required_canonical in sorted(
                dish_required_by_name.items(),
                key=lambda item: item[0].lower(),
            ):
                if dish_required_canonical <= 0:
                    continue
                dish_required_qty = canonical_qty_to_unit(
                    float(dish_required_canonical),
                    canonical_unit,
                    row_unit,
                )
                conn.execute(
                    """
                    INSERT INTO shopping_list_item_sources(
                        shopping_list_item_id,
                        retreat_plan_id,
                        retreat_plan_name,
                        dish_name,
                        required_qty,
                        required_unit
                    )
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        shopping_list_item_id,
                        plan_id,
                        plan_name_by_id.get(plan_id) or f"Retreat #{plan_id}",
                        dish_name,
                        round(dish_required_qty, 4),
                        row_unit,
                    ),
                )
        inserted_items += 1

    if inserted_items <= 0 and shopping_list_id is not None and not allow_empty_result:
        raise HTTPException(
            status_code=400,
            detail="All filtered ingredients are already fully covered by inventory.",
        )
    if inserted_items <= 0 and shopping_list_id is None:
        conn.execute("DELETE FROM shopping_lists WHERE id = ?", (shopping_list_id,))
        raise HTTPException(
            status_code=400,
            detail="All filtered ingredients are already fully covered by inventory.",
        )

    relink_shopping_pickup_list_items(conn, shopping_list_id)
    refresh_shopping_list_status(conn, shopping_list_id)
    detail = load_shopping_list_detail(conn, shopping_list_id)
    detail["missing_recipes"] = sorted(missing_recipes)
    detail["source_retreat_plan_ids"] = sorted(included_plan_ids)
    detail["source_retreat_plan_count"] = len(included_plan_ids)
    return detail


def refresh_shopping_list_status(conn: Any, shopping_list_id: int) -> None:
    counts = conn.execute(
        """
        SELECT
            COUNT(*) AS total_count,
            COALESCE(SUM(CASE WHEN COALESCE(ordered, 0) = 1 THEN 1 ELSE 0 END), 0) AS ordered_count,
            COALESCE(SUM(CASE WHEN COALESCE(received, 0) = 1 THEN 1 ELSE 0 END), 0) AS received_count
        FROM shopping_list_items
        WHERE shopping_list_id = ?
        """,
        (shopping_list_id,),
    ).fetchone()
    total_count = int(counts["total_count"] or 0)
    ordered_count = int(counts["ordered_count"] or 0)
    received_count = int(counts["received_count"] or 0)
    status = derive_progress_status(total_count, ordered_count, received_count)

    conn.execute(
        "UPDATE shopping_lists SET status = ? WHERE id = ?",
        (status, shopping_list_id),
    )


def load_shopping_list_detail(conn: Any, shopping_list_id: int) -> dict[str, Any]:
    list_row = conn.execute(
        """
        SELECT
            sl.id,
            sl.name,
            sl.phase,
            sl.status,
            sl.created_at,
            sl.retreat_plan_id,
            rp.name AS retreat_plan_name
        FROM shopping_lists sl
        LEFT JOIN retreat_plans rp ON rp.id = sl.retreat_plan_id
        WHERE sl.id = ?
        """,
        (shopping_list_id,),
    ).fetchone()
    if not list_row:
        raise HTTPException(status_code=404, detail="Shopping list not found")

    item_rows = conn.execute(
        """
        SELECT
            sli.id,
            sli.ingredient_id,
            COALESCE(i.name, ('Unknown ingredient #' || sli.ingredient_id)) AS ingredient_name,
            i.category AS ingredient_category,
            sli.required_qty,
            sli.required_unit,
            sli.in_stock_qty,
            sli.in_stock_unit,
            sli.to_buy_qty,
            sli.to_buy_unit,
            sli.ordered_qty,
            sli.ordered_unit,
            sli.vendor_id,
            v.name AS vendor_name,
            sli.ordered,
            sli.ordered_at,
            sli.received,
            sli.received_at,
            sli.status,
            sli.owner,
            sli.pickup_date,
            sli.notes
        FROM shopping_list_items sli
        LEFT JOIN ingredients i ON i.id = sli.ingredient_id
        LEFT JOIN vendors v ON v.id = sli.vendor_id
        WHERE sli.shopping_list_id = ?
        ORDER BY
            CASE WHEN i.name IS NULL THEN 1 ELSE 0 END,
            lower(COALESCE(i.name, '')),
            sli.id
        """,
        (shopping_list_id,),
    ).fetchall()

    source_breakdown_by_item: dict[int, list[dict[str, Any]]] = {}
    top_source_by_item: dict[int, dict[str, Any]] = {}
    vendor_allocations_by_item_id = load_shopping_list_item_vendor_allocations_by_item_id(
        conn,
        [int(row["id"]) for row in item_rows if row["id"] is not None],
    )
    if item_rows:
        source_table_exists = table_exists(conn, "shopping_list_item_sources")
        if source_table_exists:
            source_columns = table_columns(conn, "shopping_list_item_sources")
            dish_name_select = "slis.dish_name" if "dish_name" in source_columns else "NULL"
            item_ids = [int(row["id"]) for row in item_rows]
            placeholders = ",".join("?" for _ in item_ids)
            source_rows = conn.execute(
                f"""
                SELECT
                    slis.shopping_list_item_id,
                    slis.retreat_plan_id,
                    COALESCE(rp.name, slis.retreat_plan_name) AS retreat_plan_name,
                    {dish_name_select} AS dish_name,
                    slis.required_qty,
                    slis.required_unit
                FROM shopping_list_item_sources slis
                LEFT JOIN retreat_plans rp ON rp.id = slis.retreat_plan_id
                WHERE slis.shopping_list_item_id IN ({placeholders})
                ORDER BY
                    slis.shopping_list_item_id,
                    lower(COALESCE(rp.name, slis.retreat_plan_name)),
                    slis.id
                """,
                tuple(item_ids),
            ).fetchall()

            retreat_totals_by_item: dict[int, dict[tuple[int | None, str, str], float]] = {}
            for source_row in source_rows:
                item_id = int(source_row["shopping_list_item_id"])
                retreat_id = int(source_row["retreat_plan_id"]) if source_row["retreat_plan_id"] is not None else None
                retreat_name = str(source_row["retreat_plan_name"] or "").strip()
                if not retreat_name:
                    retreat_name = f"Retreat #{retreat_id}" if retreat_id is not None else "Unknown retreat"

                required_qty = float(source_row["required_qty"] or 0.0)
                required_unit = str(source_row["required_unit"] or "").strip()
                if required_qty <= 0 or not required_unit:
                    continue

                retreat_key = (retreat_id, retreat_name, required_unit)
                totals = retreat_totals_by_item.setdefault(item_id, {})
                totals[retreat_key] = totals.get(retreat_key, 0.0) + required_qty

                dish_name = str(source_row["dish_name"] or "").strip()
                if dish_name:
                    current_top = top_source_by_item.get(item_id)
                    if not current_top or required_qty > float(current_top["required_qty"]):
                        top_source_by_item[item_id] = {
                            "retreat_plan_id": retreat_id,
                            "retreat_plan_name": retreat_name,
                            "dish_name": dish_name,
                            "required_qty": required_qty,
                            "required_unit": required_unit,
                        }

            for item_id, totals in retreat_totals_by_item.items():
                breakdown = [
                    {
                        "retreat_plan_id": retreat_id,
                        "retreat_plan_name": retreat_name,
                        "required_qty": round(total_qty, 4),
                        "required_unit": required_unit,
                    }
                    for (retreat_id, retreat_name, required_unit), total_qty in totals.items()
                ]
                breakdown.sort(
                    key=lambda entry: (
                        -float(entry["required_qty"] or 0.0),
                        str(entry["retreat_plan_name"] or "").lower(),
                    )
                )
                source_breakdown_by_item[item_id] = breakdown
                if item_id not in top_source_by_item and breakdown:
                    lead = breakdown[0]
                    top_source_by_item[item_id] = {
                        "retreat_plan_id": lead["retreat_plan_id"],
                        "retreat_plan_name": lead["retreat_plan_name"],
                        "dish_name": None,
                        "required_qty": lead["required_qty"],
                        "required_unit": lead["required_unit"],
                    }

    ingredient_row_counts: dict[int, int] = {}
    for row in item_rows:
        ingredient_id = int(row["ingredient_id"])
        ingredient_row_counts[ingredient_id] = ingredient_row_counts.get(ingredient_id, 0) + 1

    items: list[dict[str, Any]] = []
    ordered_count = 0
    received_count = 0
    for row in item_rows:
        item_id = int(row["id"])
        vendor_allocations = vendor_allocations_by_item_id.get(item_id) or fallback_vendor_allocation_from_item_row(row)
        allocation_summary = summarize_shopping_vendor_allocations(
            vendor_allocations,
            preferred_unit=normalize_unit(str(row["required_unit"] or row["to_buy_unit"] or "").strip()) or None,
        ) if vendor_allocations else {
            "vendor_id": int(row["vendor_id"]) if row["vendor_id"] is not None else None,
            "ordered_qty": float(row["ordered_qty"]) if row["ordered_qty"] is not None else None,
            "ordered_unit": normalize_unit(str(row["ordered_unit"] or "").strip()) or None,
            "ordered": bool(row["ordered"]),
            "ordered_at": row["ordered_at"] if bool(row["ordered"]) else None,
            "received": bool(row["received"]),
            "received_at": row["received_at"] if bool(row["received"]) else None,
            "status": row["status"] or derive_shopping_item_status(bool(row["ordered"]), bool(row["received"])),
        }

        ordered = bool(allocation_summary["ordered"])
        received = bool(allocation_summary["received"])
        if ordered:
            ordered_count += 1
        if received:
            received_count += 1

        ingredient_id = int(row["ingredient_id"])
        ingredient_name = row["ingredient_name"]
        if ingredient_row_counts.get(ingredient_id, 0) > 1:
            qualifier = str(row["required_unit"] or row["to_buy_unit"] or "").strip()
            # Keep names clean for count-style units (e.g., "Cardamom" instead of
            # "Cardamom (piece)") while still qualifying other duplicate unit splits.
            if qualifier and qualifier.lower() not in {"piece", "pieces"}:
                ingredient_name = f"{ingredient_name} ({qualifier})"

        source_breakdown = source_breakdown_by_item.get(item_id, [])
        if not source_breakdown and list_row["retreat_plan_id"] is not None:
            fallback_retreat_id = int(list_row["retreat_plan_id"])
            fallback_retreat_name = (
                str(list_row["retreat_plan_name"] or "").strip() or f"Retreat #{fallback_retreat_id}"
            )
            fallback_required_qty = float(row["required_qty"]) if row["required_qty"] is not None else None
            fallback_required_unit = str(row["required_unit"] or "").strip()
            if fallback_required_qty is not None and fallback_required_unit:
                source_breakdown = [
                    {
                        "retreat_plan_id": fallback_retreat_id,
                        "retreat_plan_name": fallback_retreat_name,
                        "required_qty": round(fallback_required_qty, 4),
                        "required_unit": fallback_required_unit,
                    }
                ]

        top_source = top_source_by_item.get(item_id)
        if not top_source and source_breakdown:
            lead = source_breakdown[0]
            top_source = {
                "retreat_plan_id": lead["retreat_plan_id"],
                "retreat_plan_name": lead["retreat_plan_name"],
                "dish_name": None,
                "required_qty": lead["required_qty"],
                "required_unit": lead["required_unit"],
            }

        items.append(
            {
                "id": item_id,
                "ingredient_id": ingredient_id,
                "ingredient_name": ingredient_name,
                "ingredient_category": row["ingredient_category"],
                "required_qty": float(row["required_qty"]) if row["required_qty"] is not None else None,
                "required_unit": row["required_unit"],
                "in_stock_qty": float(row["in_stock_qty"]) if row["in_stock_qty"] is not None else None,
                "in_stock_unit": row["in_stock_unit"],
                "to_buy_qty": float(row["to_buy_qty"]) if row["to_buy_qty"] is not None else None,
                "to_buy_unit": row["to_buy_unit"],
                "ordered_qty": allocation_summary["ordered_qty"],
                "ordered_unit": allocation_summary["ordered_unit"],
                "vendor_id": allocation_summary["vendor_id"],
                "vendor_name": (
                    next(
                        (
                            entry.get("vendor_name")
                            for entry in vendor_allocations
                            if entry.get("vendor_id") is not None
                            and entry.get("vendor_id") == allocation_summary["vendor_id"]
                            and entry.get("vendor_name")
                        ),
                        None,
                    )
                    or row["vendor_name"]
                ),
                "ordered": ordered,
                "ordered_at": allocation_summary["ordered_at"],
                "received": received,
                "received_at": allocation_summary["received_at"],
                "status": allocation_summary["status"],
                "owner": row["owner"],
                "pickup_date": row["pickup_date"],
                "notes": row["notes"],
                "source_breakdown": source_breakdown,
                "top_source": top_source,
                "vendor_allocations": vendor_allocations,
            }
        )

    return {
        "id": int(list_row["id"]),
        "name": list_row["name"],
        "phase": list_row["phase"] or "bulk",
        "status": list_row["status"] or "draft",
        "created_at": list_row["created_at"],
        "retreat_plan_id": int(list_row["retreat_plan_id"]) if list_row["retreat_plan_id"] is not None else None,
        "retreat_plan_name": list_row["retreat_plan_name"],
        "item_count": len(items),
        "ordered_count": ordered_count,
        "received_count": received_count,
        "items": items,
    }


def create_carry_forward_shopping_list(
    conn: Any,
    source_list_id: int,
    name_override: str | None = None,
    phase_override: str | None = None,
) -> dict[str, Any]:
    source_list = conn.execute(
        """
        SELECT id, name, phase, retreat_plan_id
        FROM shopping_lists
        WHERE id = ?
        """,
        (source_list_id,),
    ).fetchone()
    if not source_list:
        raise HTTPException(status_code=404, detail="Source shopping list not found")

    phase = str(source_list["phase"] or "custom").strip().lower() or "custom"
    if phase_override is not None:
        candidate_phase = str(phase_override or "").strip().lower()
        if candidate_phase not in SHOPPING_PHASES:
            raise HTTPException(
                status_code=400,
                detail=f"phase must be one of: {', '.join(SHOPPING_PHASES)}",
            )
        phase = candidate_phase

    source_items = conn.execute(
        """
        SELECT
            ingredient_id,
            required_qty,
            required_unit,
            to_buy_qty,
            to_buy_unit,
            vendor_id,
            notes
        FROM shopping_list_items
        WHERE shopping_list_id = ?
          AND COALESCE(received, 0) = 0
        ORDER BY id
        """,
        (source_list_id,),
    ).fetchall()

    pending_items: list[dict[str, Any]] = []
    for row in source_items:
        pending_qty = float(row["to_buy_qty"]) if row["to_buy_qty"] is not None else float(row["required_qty"] or 0)
        pending_unit = str(row["to_buy_unit"] or row["required_unit"] or "").strip()
        if pending_qty <= 0 or not pending_unit:
            continue
        pending_items.append(
            {
                "ingredient_id": int(row["ingredient_id"]),
                "pending_qty": pending_qty,
                "pending_unit": pending_unit,
                "vendor_id": int(row["vendor_id"]) if row["vendor_id"] is not None else None,
                "notes": row["notes"],
            }
        )

    if not pending_items:
        raise HTTPException(
            status_code=400,
            detail="No unreceived items with remaining quantity to carry forward.",
        )

    name = name_override.strip() if name_override and name_override.strip() else None
    if not name:
        name = f"{source_list['name']} - Step 2"
    name = unique_shopping_list_name(conn, name)

    created = conn.execute(
        """
        INSERT INTO shopping_lists(retreat_plan_id, name, phase, status)
        VALUES (?, ?, ?, 'draft')
        RETURNING id
        """,
        (source_list["retreat_plan_id"], name, phase),
    ).fetchone()
    new_list_id = int(created["id"])

    for item in pending_items:
        notes = (
            f"Carry-forward from list #{source_list_id}: {item['notes']}"
            if item["notes"]
            else f"Carry-forward from list #{source_list_id}"
        )
        conn.execute(
            """
            INSERT INTO shopping_list_items(
                shopping_list_id,
                ingredient_id,
                required_qty,
                required_unit,
                in_stock_qty,
                in_stock_unit,
                to_buy_qty,
                to_buy_unit,
                ordered_qty,
                ordered_unit,
                vendor_id,
                status,
                ordered,
                received,
                notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, ?, 'open', 0, 0, ?)
            """,
            (
                new_list_id,
                item["ingredient_id"],
                round(item["pending_qty"], 4),
                item["pending_unit"],
                0,
                item["pending_unit"],
                round(item["pending_qty"], 4),
                item["pending_unit"],
                item["vendor_id"],
                notes,
            ),
        )

    refresh_shopping_list_status(conn, new_list_id)
    detail = load_shopping_list_detail(conn, new_list_id)
    detail["carried_from_list_id"] = int(source_list["id"])
    detail["carried_item_count"] = len(pending_items)
    return detail


def unique_shopping_list_name(
    conn: Any,
    base_name: str,
    exclude_shopping_list_id: int | None = None,
) -> str:
    seed = " ".join(str(base_name or "").strip().split()) or "Shopping List"
    candidate = seed
    suffix = 2
    while True:
        if exclude_shopping_list_id is None:
            exists = conn.execute(
                "SELECT 1 FROM shopping_lists WHERE lower(name) = lower(?)",
                (candidate,),
            ).fetchone()
        else:
            exists = conn.execute(
                """
                SELECT 1
                FROM shopping_lists
                WHERE lower(name) = lower(?)
                  AND id != ?
                """,
                (candidate, exclude_shopping_list_id),
            ).fetchone()
        if not exists:
            break
        candidate = f"{seed} ({suffix})"
        suffix += 1
    return candidate


def unique_retreat_plan_name(conn: Any, base_name: str) -> str:
    seed = " ".join(base_name.strip().split()) or "Untitled Retreat (Copy)"
    candidate = seed
    suffix = 2
    while conn.execute(
        "SELECT 1 FROM retreat_plans WHERE lower(name) = lower(?)",
        (candidate,),
    ).fetchone():
        candidate = f"{seed} ({suffix})"
        suffix += 1
    return candidate


def normalize_unit(unit: str) -> str:
    value = unit.strip().lower()
    aliases = {
        "cups": "cup",
        "gms": "g",
        "kilogram": "kg",
        "kilograms": "kg",
        "kilo": "kg",
        "kilos": "kg",
        "gram": "g",
        "grams": "g",
        "liter": "l",
        "liters": "l",
        "litre": "l",
        "litres": "l",
        "milliliter": "ml",
        "milliliters": "ml",
        "millilitre": "ml",
        "millilitres": "ml",
        "tablespoon": "tbsp",
        "tablespoons": "tbsp",
        "tbs": "tbsp",
        "teaspoon": "tsp",
        "teaspoons": "tsp",
        "tsb": "tsp",
        "pound": "lb",
        "pounds": "lb",
        "lbs": "lb",
        "ounce": "oz",
        "ounces": "oz",
        "fluid ounce": "fl oz",
        "fluid ounces": "fl oz",
        "floz": "fl oz",
        "fl. oz.": "fl oz",
        "quart": "qt",
        "quarts": "qt",
        "gallon": "gal",
        "gallons": "gal",
        "eaches": "each",
        "ea": "each",
        "pieces": "piece",
        "packets": "packet",
        "packs": "pack",
        "pk": "pack",
        "pks": "pack",
        "cans": "can",
        "bunches": "bunch",
        "loaves": "loaf",
        "sprigs": "sprig",
        "springs": "sprig",
        "leaves": "leaf",
        "bags": "bag",
        "boxes": "box",
        "cases": "case",
        "bottles": "bottle",
        "jugs": "jug",
        "jars": "jar",
        "cartons": "carton",
        "tubs": "tub",
        "packages": "package",
        "pinches": "pinch",
        "pod": "piece",
        "pods": "piece",
        "clove": "piece",
        "cloves": "piece",
    }
    return aliases.get(value, value)


def normalize_ingredient_name(ingredient_name: str) -> str:
    candidate = " ".join(ingredient_name.strip().split())
    aliases = {
        "kidney bean": "Rajma",
        "kidney beans": "Rajma",
        "red kidney bean": "Rajma",
        "red kidney beans": "Rajma",
        "lemon": "Lemon juice concentrate",
        "lime": "Lemon juice concentrate",
        "lemon or lime": "Lemon juice concentrate",
        "english cucumber": "Cucumber",
        "english cucumbers": "Cucumber",
        "purple cabbage": "Red Cabbage",
        "red cabbage": "Red Cabbage",
        "red onion": "Onion",
        "red onions": "Onion",
        "extra firm tofu": "extra-firm tofu",
        "extra firm tofy": "extra-firm tofu",
        "curry leaf": "Curry leaves",
        "cardamom pod": "Cardamom",
        "cardamom pods": "Cardamom",
        "green cardamom": "Cardamom",
        "green cardamom pod": "Cardamom",
        "green cardamom pods": "Cardamom",
        "mung dal": "Yellow Mung Dal",
        "yellow mung dal": "Yellow Mung Dal",
        "yellow moong dal": "Yellow Mung Dal",
    }
    return aliases.get(candidate.lower(), candidate)


def normalize_recipe_category(category: str | None) -> str | None:
    if category is None:
        return None
    candidate = " ".join(str(category).strip().split())
    if not candidate:
        return None

    matches = [known for known in RECIPE_CATEGORIES if known.lower() == candidate.lower()]
    if matches:
        return matches[0]

    raise HTTPException(
        status_code=400,
        detail=(
            f"Unknown recipe category '{candidate}'. "
            f"Allowed categories: {', '.join(RECIPE_CATEGORIES)}"
        ),
    )


def get_or_create_ingredient_id(conn: Any, ingredient_name: str) -> int:
    ing_name = normalize_ingredient_name(ingredient_name)
    if not ing_name:
        raise HTTPException(status_code=400, detail="Ingredient name cannot be blank")

    ing_row = conn.execute(
        "SELECT id FROM ingredients WHERE lower(name) = lower(?)",
        (ing_name,),
    ).fetchone()
    if ing_row:
        return int(ing_row["id"])

    created = conn.execute(
        "INSERT INTO ingredients(name) VALUES (?) RETURNING id",
        (ing_name,),
    ).fetchone()
    ingredient_id = int(created["id"])

    # Automatically fetch USDA density data for the new ingredient.
    try:
        populate_ingredient_conversions(conn, ingredient_id, ing_name)
    except Exception:
        import logging
        logging.getLogger(__name__).warning(
            "USDA lookup failed for %r — ingredient created without density data",
            ing_name,
            exc_info=True,
        )

    return ingredient_id


def replace_recipe_ingredients(
    conn: Any, recipe_id: int, ingredients: list[RecipeIngredientCreate]
) -> None:
    conn.execute("DELETE FROM recipe_ingredients WHERE recipe_id = ?", (recipe_id,))
    for item in ingredients:
        ingredient_id = get_or_create_ingredient_id(conn, item.ingredient_name)
        prep_notes = item.prep_notes.strip() if item.prep_notes and item.prep_notes.strip() else None
        normalized_unit = normalize_unit(item.unit)
        normalized_quantity = float(item.quantity)

        grams_per_cup, canonical_unit, category = ingredient_profile(item.ingredient_name)
        if is_dal_or_rice_ingredient(item.ingredient_name, category):
            if normalized_unit in VOLUME_TO_ML and normalized_unit != "cup":
                normalized_quantity = (
                    normalized_quantity * VOLUME_TO_ML[normalized_unit] / VOLUME_TO_ML["cup"]
                )
                normalized_unit = "cup"
            elif normalized_unit in MASS_TO_G:
                grams_per_cup_value = grams_per_cup
                if not grams_per_cup_value:
                    cup_to_g = ingredient_specific_unit_conversion_factor(
                        item.ingredient_name,
                        "cup",
                        "g",
                    )
                    cup_to_kg = ingredient_specific_unit_conversion_factor(
                        item.ingredient_name,
                        "cup",
                        "kg",
                    )
                    if cup_to_g is not None:
                        grams_per_cup_value = cup_to_g
                    elif cup_to_kg is not None:
                        grams_per_cup_value = cup_to_kg * 1000.0
                if grams_per_cup_value and grams_per_cup_value > 0:
                    grams = normalized_quantity * MASS_TO_G[normalized_unit]
                    normalized_quantity = grams / grams_per_cup_value
                    normalized_unit = "cup"

        if canonical_unit in COUNT_UNITS and normalized_unit in COUNT_UNITS and normalized_unit != canonical_unit:
            factor = ingredient_specific_unit_conversion_factor(item.ingredient_name, normalized_unit, canonical_unit)
            if factor is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Missing conversion for ingredient '{item.ingredient_name}': "
                        f"'{normalized_unit}' -> '{canonical_unit}'."
                    ),
                )
            normalized_quantity *= factor
            normalized_unit = canonical_unit
        elif not is_dal_or_rice_ingredient(item.ingredient_name, category) and canonical_unit in MASS_TO_G and normalized_unit != canonical_unit:
            canonical_qty, canonical_base_unit, _note = to_canonical(
                item.ingredient_name,
                normalized_quantity,
                normalized_unit,
            )
            if canonical_qty is None:
                raise HTTPException(
                    status_code=400,
                    detail=_note
                    or (
                        f"Missing conversion for ingredient '{item.ingredient_name}': "
                        f"'{normalized_unit}' -> '{canonical_unit}'."
                    ),
                )
            if canonical_base_unit == "g":
                normalized_quantity = canonical_qty / MASS_TO_G[canonical_unit]
                normalized_unit = canonical_unit
            elif canonical_base_unit in COUNT_UNITS:
                # Preserve count-unit recipes for packaged items when no
                # ingredient-specific weight mapping exists.
                normalized_quantity = canonical_qty
                normalized_unit = canonical_base_unit
            else:
                raise HTTPException(
                    status_code=400,
                    detail=_note
                    or (
                        f"Missing conversion for ingredient '{item.ingredient_name}': "
                        f"'{normalized_unit}' -> '{canonical_unit}'."
                    ),
                )

        conn.execute(
            """
            INSERT INTO recipe_ingredients(recipe_id, ingredient_id, quantity, unit, prep_notes)
            VALUES (?, ?, ?, ?, ?)
            """,
            (recipe_id, ingredient_id, normalized_quantity, normalized_unit, prep_notes),
        )


def replace_recipe_steps(conn: Any, recipe_id: int, steps: list[str] | None) -> None:
    conn.execute("DELETE FROM recipe_steps WHERE recipe_id = ?", (recipe_id,))
    clean_steps = [step.strip() for step in (steps or []) if step and step.strip()]
    for index, instruction in enumerate(clean_steps, start=1):
        conn.execute(
            """
            INSERT INTO recipe_steps(recipe_id, step_order, instruction)
            VALUES (?, ?, ?)
            """,
            (recipe_id, index, instruction),
        )


def ingredient_profile(ingredient_name: str) -> tuple[float | None, str | None, str | None]:
    normalized_name = normalize_ingredient_name(ingredient_name)
    with get_connection() as conn:
        row = conn.execute(
            "SELECT grams_per_cup, canonical_unit, category FROM ingredients WHERE lower(name) = lower(?)",
            (normalized_name,),
        ).fetchone()
        if not row:
            return None, None, None
        grams_per_cup = float(row["grams_per_cup"]) if row["grams_per_cup"] else None
        canonical_unit = normalize_unit(str(row["canonical_unit"] or "").strip()) if row["canonical_unit"] else None
        category = str(row["category"] or "").strip() if row["category"] else None
        return grams_per_cup, canonical_unit, category


def is_spice_or_seasoning_category(category: str | None) -> bool:
    if not category:
        return False
    normalized = category.strip().lower()
    return "spice" in normalized or "seasoning" in normalized


def is_prepared_packaged_category(category: str | None) -> bool:
    if not category:
        return False
    normalized = category.strip().lower()
    return "prepared" in normalized or "packaged" in normalized


def is_produce_category(category: str | None) -> bool:
    if not category:
        return False
    return category.strip().lower() in {"produce", "fruits"}


def is_pulse_or_legume_category(category: str | None) -> bool:
    if not category:
        return False
    normalized = category.strip().lower()
    return "pulse" in normalized or "legume" in normalized


def is_dal_or_rice_ingredient(ingredient_name: str, category: str | None = None) -> bool:
    normalized_name = normalize_ingredient_name(ingredient_name).lower()
    if not normalized_name:
        return False
    if "vinegar" in normalized_name:
        return False
    if DAL_RICE_TOKEN_RE.search(normalized_name):
        return True

    if is_pulse_or_legume_category(category):
        return True
    return False


def ingredient_specific_unit_conversion_factor(
    ingredient_name: str,
    unit_from: str,
    unit_to: str,
) -> float | None:
    normalized_name = normalize_ingredient_name(ingredient_name)
    from_unit = normalize_unit(unit_from)
    to_unit = normalize_unit(unit_to)
    if not normalized_name or not from_unit or not to_unit:
        return None
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT quantity_from, quantity_to
            FROM unit_conversions
            WHERE lower(COALESCE(item_name, '')) = lower(?)
              AND lower(unit_from) = lower(?)
              AND lower(unit_to) = lower(?)
            ORDER BY CASE WHEN context = 'ingredient_specific' THEN 0 ELSE 1 END, id
            LIMIT 1
            """,
            (normalized_name, from_unit, to_unit),
        ).fetchone()
    if not row:
        return None
    quantity_from = float(row["quantity_from"] or 0)
    quantity_to = float(row["quantity_to"] or 0)
    if quantity_from <= 0 or quantity_to <= 0:
        return None
    return quantity_to / quantity_from


def ingredient_specific_g_per_unit(ingredient_name: str, unit: str) -> float | None:
    return ingredient_specific_unit_conversion_factor(ingredient_name, unit, "g")


def to_canonical(ingredient_name: str, qty: float, unit: str) -> tuple[float | None, str | None, str | None]:
    unit = normalize_unit(unit)
    if unit in MASS_TO_G:
        grams = qty * MASS_TO_G[unit]
        density_grams_per_cup, mass_canonical_unit, _mass_category = ingredient_profile(ingredient_name)
        if (
            mass_canonical_unit in {"ml", "l"}
            and density_grams_per_cup is not None
            and float(density_grams_per_cup) > 0
        ):
            grams_per_ml = float(density_grams_per_cup) / VOLUME_TO_ML["cup"]
            return grams / grams_per_ml, "ml", "Converted mass to volume using ingredient density."
        return grams, "g", None

    specific_g_per_unit = ingredient_specific_g_per_unit(ingredient_name, unit)
    if specific_g_per_unit is not None:
        return qty * specific_g_per_unit, "g", "Converted using ingredient-specific unit conversion."

    grams_per_cup, canonical_unit, _ingredient_category = ingredient_profile(ingredient_name)

    if unit in VOLUME_TO_ML:
        if canonical_unit in {"ml", "l"}:
            return qty * VOLUME_TO_ML[unit], "ml", "Canonical volume unit preference applied."

        if grams_per_cup is not None:
            cups = (qty * VOLUME_TO_ML[unit]) / VOLUME_TO_ML["cup"]
            grams = cups * grams_per_cup
            return grams, "g", "Converted using ingredient-specific grams_per_cup."

        if canonical_unit in MASS_TO_G:
            return None, None, (
                f"Missing conversion for ingredient '{ingredient_name}': "
                f"'{unit}' -> '{canonical_unit}' requires ingredient-specific density mapping."
            )
        if canonical_unit in COUNT_UNITS:
            return None, None, (
                f"Missing conversion for ingredient '{ingredient_name}': "
                f"'{unit}' -> '{canonical_unit}' requires ingredient-specific unit mapping."
            )

        return qty * VOLUME_TO_ML[unit], "ml", "No ingredient density found; kept as volume."

    if unit in COUNT_UNITS:
        if canonical_unit in COUNT_UNITS:
            target_unit = normalize_unit(canonical_unit)
            if unit == target_unit:
                return qty, target_unit, "Canonical count unit preference applied."

            factor = ingredient_specific_unit_conversion_factor(ingredient_name, unit, target_unit)
            if factor is not None:
                return qty * factor, target_unit, "Converted using ingredient-specific count-unit conversion."
            return None, None, (
                f"Missing conversion for ingredient '{ingredient_name}': "
                f"'{unit}' -> '{target_unit}'."
            )
        if canonical_unit in MASS_TO_G or canonical_unit in {"ml", "l"}:
            if is_prepared_packaged_category(_ingredient_category):
                return qty, unit, (
                    "Missing mass/volume conversion; kept as count unit for prepared/packaged ingredient."
                )
            return None, None, (
                f"Missing conversion for ingredient '{ingredient_name}': "
                f"'{unit}' -> '{canonical_unit}'."
            )
        return qty, unit, None

    return None, None, f"Unknown unit '{unit}' for ingredient '{ingredient_name}'."


def require_canonical_conversion(
    ingredient_name: str,
    quantity: float,
    unit: str,
    *,
    context: str | None = None,
) -> tuple[float, str, str | None]:
    canonical_qty, canonical_unit, note = to_canonical(ingredient_name, quantity, unit)
    if canonical_qty is None or canonical_unit is None:
        detail = note or f"Could not convert ingredient '{ingredient_name}' with unit '{unit}'."
        if context:
            detail = f"{context}: {detail}"
        raise HTTPException(status_code=400, detail=detail)
    return canonical_qty, canonical_unit, note


def to_shopping_unit(canonical_qty: float | None, canonical_unit: str | None) -> tuple[float | None, str | None]:
    if canonical_qty is None or canonical_unit is None:
        return None, None

    if canonical_unit == "g":
        if canonical_qty >= 1000:
            return canonical_qty / 1000.0, "kg"
        if canonical_qty >= 453.59237:
            return canonical_qty / 453.59237, "lb"
        return canonical_qty, "g"

    if canonical_unit == "ml":
        if canonical_qty >= 1000:
            return canonical_qty / 1000.0, "l"
        return canonical_qty, "ml"

    return canonical_qty, canonical_unit


def normalize_optional_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text if text else None


def normalize_required_text(value: Any, *, field_name: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail=f"{field_name} is required.")
    return text


def load_app_setting(conn: Any, setting_key: str) -> Any | None:
    if not table_exists(conn, "app_settings"):
        return None
    return conn.execute(
        """
        SELECT
            app_settings.setting_key,
            app_settings.setting_value,
            app_settings.updated_by_user_id,
            app_settings.created_at,
            app_settings.updated_at,
            users.username AS updated_by_username
        FROM app_settings
        LEFT JOIN users
          ON users.id = app_settings.updated_by_user_id
        WHERE app_settings.setting_key = ?
        """,
        (setting_key,),
    ).fetchone()


def save_app_setting(
    conn: Any,
    *,
    setting_key: str,
    setting_value: str | None,
    updated_by_user_id: int | None = None,
) -> None:
    existing = load_app_setting(conn, setting_key)
    normalized_value = setting_value if setting_value is not None else None
    if existing:
        conn.execute(
            """
            UPDATE app_settings
            SET setting_value = ?,
                updated_by_user_id = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE setting_key = ?
            """,
            (normalized_value, updated_by_user_id, setting_key),
        )
        return

    conn.execute(
        """
        INSERT INTO app_settings(
            setting_key,
            setting_value,
            updated_by_user_id,
            created_at,
            updated_at
        )
        VALUES (?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """,
        (setting_key, normalized_value, updated_by_user_id),
    )


def resolve_shared_access_state(
    conn: Any,
    *,
    setting_key: str,
    environment_value: str | None = None,
) -> dict[str, Any]:
    row = load_app_setting(conn, setting_key)
    if row is not None:
        access_code = str(row["setting_value"] or "").strip() or None
        return {
            "access_code": access_code,
            "guest_access_enabled": bool(access_code),
            "source": "admin" if access_code else "admin_disabled",
            "updated_at": row["updated_at"],
            "updated_by_user_id": int(row["updated_by_user_id"]) if row["updated_by_user_id"] is not None else None,
            "updated_by_username": normalize_optional_text(row["updated_by_username"]),
        }

    configured = str(environment_value or "").strip() or None
    return {
        "access_code": configured,
        "guest_access_enabled": bool(configured),
        "source": "environment" if configured else "none",
        "updated_at": None,
        "updated_by_user_id": None,
        "updated_by_username": None,
    }


def resolve_inventory_withdraw_access_state(conn: Any) -> dict[str, Any]:
    return resolve_shared_access_state(
        conn,
        setting_key=APP_SETTING_INVENTORY_WITHDRAW_ACCESS_CODE,
        environment_value=os.getenv(INVENTORY_WITHDRAW_ACCESS_CODE_ENV),
    )


def get_inventory_withdraw_access_code(conn: Any) -> str | None:
    return resolve_inventory_withdraw_access_state(conn).get("access_code")


def resolve_kitchen_guest_access_state(conn: Any, scope_slug: str) -> dict[str, Any]:
    details = resolve_kitchen_access_scope(scope_slug)
    return resolve_shared_access_state(conn, setting_key=details["setting_key"])


def get_kitchen_guest_access_code(conn: Any, scope_slug: str) -> str | None:
    return resolve_kitchen_guest_access_state(conn, scope_slug).get("access_code")


def load_optional_session_user(conn: Any, request: Request) -> AuthUser | None:
    existing = get_request_user(request)
    if existing:
        return existing
    raw_token = request.cookies.get(SESSION_COOKIE_NAME)
    if not raw_token:
        return None
    return authenticate_session_token(conn, raw_token)


def load_inventory_withdraw_actor(
    conn: Any,
    request: Request,
    *,
    access_code: str | None = None,
    withdrawn_by: str | None = None,
) -> tuple[AuthUser | None, str]:
    user = load_optional_session_user(conn, request)
    if user:
        if user.role not in {ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN}:
            raise HTTPException(status_code=403, detail="This account cannot withdraw inventory.")
        return user, user.username

    configured_code = get_inventory_withdraw_access_code(conn)
    if not configured_code:
        raise HTTPException(
            status_code=403,
            detail="Guest withdraw access is not configured. Sign in or ask an admin to set the access code.",
        )

    provided_code = normalize_required_text(access_code, field_name="Access code")
    if not secrets.compare_digest(provided_code, configured_code):
        raise HTTPException(status_code=403, detail="Invalid withdraw access code.")

    return None, normalize_required_text(withdrawn_by, field_name="Withdrawn by")


def authorize_inventory_withdraw_access(
    conn: Any,
    request: Request,
    *,
    access_code: str | None = None,
) -> AuthUser | None:
    user = load_optional_session_user(conn, request)
    if user:
        if user.role not in {ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN}:
            raise HTTPException(status_code=403, detail="This account cannot withdraw inventory.")
        return user

    configured_code = get_inventory_withdraw_access_code(conn)
    if not configured_code:
        raise HTTPException(
            status_code=403,
            detail="Guest withdraw access is not configured. Sign in or ask an admin to set the access code.",
        )

    provided_code = normalize_required_text(access_code, field_name="Access code")
    if not secrets.compare_digest(provided_code, configured_code):
        raise HTTPException(status_code=403, detail="Invalid withdraw access code.")
    return None


def normalize_storage_grid_location(value: Any) -> str:
    text = normalize_required_text(value, field_name="Storage location")
    match = STORAGE_GRID_LOCATION_RE.fullmatch(text)
    if not match:
        raise HTTPException(
            status_code=400,
            detail="Storage location must use shelf grid format like A1 through Z99 (A1-A20 preferred).",
        )
    return f"{match.group(1).upper()}{match.group(2)}"


def normalize_optional_storage_grid_location(value: Any) -> str | None:
    text = normalize_optional_text(value)
    if not text:
        return None
    return normalize_storage_grid_location(text)


def normalize_inventory_barcode(value: Any) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    if not INVENTORY_BARCODE_RE.fullmatch(digits):
        raise HTTPException(status_code=400, detail="Barcode must be 8-14 digits.")
    return digits


def infer_inventory_unit_from_text(value: Any) -> str | None:
    text = normalize_optional_text(value)
    if not text:
        return None
    lowered = text.lower()
    match = re.search(r"\b(kg|g|lb|lbs|oz|l|ml|ct|count|pack|packs|ea|each)\b", lowered)
    if match:
        token = match.group(1)
        normalized = {
            "lbs": "lb",
            "ct": "each",
            "count": "each",
            "ea": "each",
            "packs": "pack",
        }.get(token, token)
        return normalized
    if re.search(r"\b\d+\b", lowered):
        return "each"
    return None


def is_infra_category_hint(value: Any) -> bool:
    text = normalize_optional_text(value)
    if not text:
        return False
    lowered = text.lower()
    compact = re.sub(r"\s+", " ", lowered).strip()
    if compact in INFRA_CATEGORY_EXACT:
        return True
    return False


def infer_inventory_category_from_text(*values: Any) -> str | None:
    for value in values:
        text = normalize_optional_text(value)
        if not text:
            continue
        lowered = text.lower()
        if any(hint in lowered for hint in INFRA_ITEM_HINTS):
            return INFRA_CATEGORY_NAME
    return None


def normalize_inventory_category(value: Any) -> str | None:
    text = normalize_optional_text(value)
    if not text:
        return None
    normalized = re.sub(r"\s+", " ", text).strip()
    if is_infra_category_hint(normalized):
        return INFRA_CATEGORY_NAME
    return normalized


def normalize_lookup_category(value: Any) -> str | None:
    text = normalize_optional_text(value)
    if not text:
        return None
    normalized = text.split(",")[0].split(":")[-1].replace("-", " ").replace("_", " ")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    if not normalized:
        return None
    return normalize_inventory_category(normalized.title())


def normalize_lookup_query(value: Any) -> str | None:
    text = normalize_optional_text(value)
    if not text:
        return None
    normalized = re.sub(r"\s+", " ", text).strip()
    if len(normalized) < 2:
        return None
    return normalized[:120]


def normalize_lookup_barcode_candidate(value: Any) -> str | None:
    digits = re.sub(r"\D+", "", str(value or ""))
    if not INVENTORY_BARCODE_RE.fullmatch(digits):
        return None
    return digits


def parse_inventory_product_catalog_row(row: Any) -> dict[str, Any] | None:
    source = normalize_optional_text(row["source"]) or "catalog"
    barcode = normalize_lookup_barcode_candidate(row["barcode"])
    name = normalize_optional_text(row["product_name"])
    brand = normalize_optional_text(row["brand"])
    category = normalize_inventory_category(row["category"])
    if not category:
        category = infer_inventory_category_from_text(
            name,
            brand,
            row["source_sku"],
            row["notes"],
            row["product_url"],
        )
    unit = normalize_optional_text(row["unit"]) or infer_inventory_unit_from_text(name)
    image_url = normalize_optional_text(row["image_url"])
    product_url = normalize_optional_text(row["product_url"])
    source_sku = normalize_optional_text(row["source_sku"])

    if not any([barcode, name, category, unit, image_url, product_url]):
        return None

    return {
        "source": source,
        "barcode": barcode,
        "name": name,
        "brand": brand,
        "category": category,
        "unit": unit,
        "image_url": image_url,
        "product_url": product_url,
        "source_sku": source_sku,
    }


def lookup_product_inventory_catalog(barcode: str) -> dict[str, Any] | None:
    normalized_barcode = normalize_lookup_barcode_candidate(barcode)
    if not normalized_barcode:
        return None

    used_ikea_prefix_fallback = False
    with get_connection() as conn:
        if not table_exists(conn, "inventory_product_catalog"):
            return None
        rows = conn.execute(
            """
            SELECT
                source,
                barcode,
                product_name,
                brand,
                category,
                unit,
                image_url,
                product_url,
                source_sku,
                notes
            FROM inventory_product_catalog
            WHERE barcode = ?
            ORDER BY
                CASE WHEN lower(source) = 'webstaurantstore' THEN 0 ELSE 1 END,
                updated_at DESC,
                id DESC
            LIMIT 10
            """,
            (normalized_barcode,),
        ).fetchall()
        if not rows and len(normalized_barcode) == 14:
            ikea_prefix = normalized_barcode[:8]
            rows = conn.execute(
                """
                SELECT
                    source,
                    barcode,
                    product_name,
                    brand,
                    category,
                    unit,
                    image_url,
                    product_url,
                    source_sku,
                    notes
                FROM inventory_product_catalog
                WHERE lower(source) = 'ikea'
                  AND (
                    barcode = ?
                    OR replace(replace(replace(COALESCE(source_sku, ''), '.', ''), '-', ''), ' ', '') = ?
                    OR lower(COALESCE(product_url, '')) LIKE ?
                  )
                ORDER BY
                    CASE WHEN barcode = ? THEN 0 ELSE 1 END,
                    updated_at DESC,
                    id DESC
                LIMIT 10
                """,
                (ikea_prefix, ikea_prefix, f"%-{ikea_prefix}/%", ikea_prefix),
            ).fetchall()
            used_ikea_prefix_fallback = bool(rows)

    hits: list[dict[str, Any]] = []
    for row in rows:
        parsed = parse_inventory_product_catalog_row(row)
        if not parsed:
            continue
        hits.append(parsed)
    if not hits:
        return None

    merged: dict[str, Any] = {"source": None}
    for hit in hits:
        merged["source"] = merge_lookup_source_names(merged.get("source"), hit.get("source"))
        for key in ("name", "category", "unit", "image_url"):
            if not merged.get(key) and hit.get(key):
                merged[key] = hit[key]
    if used_ikea_prefix_fallback:
        merged["source"] = merge_lookup_source_names(merged.get("source"), "ikea-sku-prefix")
    if not any(merged.get(key) for key in ("name", "category", "unit", "image_url")):
        return None
    return merged


def search_inventory_product_catalog(query: str, *, limit: int = 12) -> list[dict[str, Any]]:
    search_query = normalize_lookup_query(query)
    if not search_query:
        return []
    bounded_limit = max(1, min(int(limit), 60))
    like_pattern = f"%{search_query.lower()}%"
    barcode_digits = re.sub(r"\D+", "", search_query)
    barcode_like = f"%{barcode_digits}%"

    with get_connection() as conn:
        if not table_exists(conn, "inventory_product_catalog"):
            return []
        rows = conn.execute(
            """
            SELECT
                source,
                barcode,
                product_name,
                brand,
                category,
                unit,
                image_url,
                product_url,
                source_sku,
                notes
            FROM inventory_product_catalog
            WHERE
                lower(COALESCE(product_name, '')) LIKE ?
                OR lower(COALESCE(brand, '')) LIKE ?
                OR lower(COALESCE(category, '')) LIKE ?
                OR lower(COALESCE(source_sku, '')) LIKE ?
                OR lower(COALESCE(product_url, '')) LIKE ?
                OR (? != '' AND barcode LIKE ?)
            ORDER BY
                CASE WHEN lower(source) = 'webstaurantstore' THEN 0 ELSE 1 END,
                updated_at DESC,
                id DESC
            LIMIT ?
            """,
            (
                like_pattern,
                like_pattern,
                like_pattern,
                like_pattern,
                like_pattern,
                barcode_digits,
                barcode_like,
                bounded_limit,
            ),
        ).fetchall()

    hits: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        parsed = parse_inventory_product_catalog_row(row)
        if not parsed:
            continue
        barcode_key = str(parsed.get("barcode") or "")
        name_key = str(parsed.get("name") or "").strip().lower()
        dedupe_key = (barcode_key, name_key)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        hits.append(parsed)
        if len(hits) >= bounded_limit:
            break
    return hits


def parse_open_facts_product_row(
    product: dict[str, Any],
    *,
    source_name: str,
) -> dict[str, Any] | None:
    name = (
        normalize_optional_text(product.get("product_name"))
        or normalize_optional_text(product.get("product_name_en"))
        or normalize_optional_text(product.get("generic_name"))
    )
    category_tags = product.get("categories_tags")
    primary_tag = category_tags[0] if isinstance(category_tags, list) and category_tags else None
    category = normalize_lookup_category(primary_tag) or normalize_lookup_category(product.get("categories"))
    if not category:
        category = infer_inventory_category_from_text(
            name,
            product.get("categories"),
            product.get("generic_name"),
            product.get("quantity"),
            product.get("packaging"),
            product.get("packaging_text"),
        )
    image_url = (
        normalize_optional_text(product.get("image_front_url"))
        or normalize_optional_text(product.get("image_url"))
        or normalize_optional_text(product.get("image_front_small_url"))
    )
    unit = infer_inventory_unit_from_text(
        product.get("quantity")
        or product.get("packaging_text")
        or product.get("packaging")
    )
    barcode = normalize_lookup_barcode_candidate(product.get("code"))

    if not any([name, category, unit, image_url, barcode]):
        return None

    return {
        "source": source_name,
        "barcode": barcode,
        "name": name,
        "category": category,
        "unit": unit,
        "image_url": image_url,
    }


def lookup_product_open_facts_catalog(
    barcode: str,
    *,
    endpoint_template: str,
    source_name: str,
) -> dict[str, Any] | None:
    endpoint = endpoint_template.format(barcode=barcode)
    request = urllib_request.Request(endpoint, method="GET")
    try:
        with urllib_request.urlopen(request, timeout=8) as response:
            raw = response.read().decode("utf-8")
    except (urllib_error.HTTPError, urllib_error.URLError):
        return None

    try:
        payload = json.loads(raw)
    except Exception:
        return None

    if int(payload.get("status") or 0) != 1:
        return None

    product = payload.get("product")
    if not isinstance(product, dict):
        return None

    parsed = parse_open_facts_product_row(product, source_name=source_name)
    if not parsed:
        return None
    parsed.pop("barcode", None)
    return parsed


def search_products_open_facts_catalog(
    query: str,
    *,
    endpoint: str,
    source_name: str,
    limit: int = 6,
) -> list[dict[str, Any]]:
    search_query = normalize_lookup_query(query)
    if not search_query:
        return []
    bounded_limit = max(1, min(int(limit), 20))
    params = {
        "search_terms": search_query,
        "search_simple": 1,
        "action": "process",
        "json": 1,
        "page_size": bounded_limit,
        "fields": (
            "code,product_name,product_name_en,generic_name,categories,categories_tags,"
            "image_front_url,image_url,image_front_small_url,quantity,packaging,packaging_text"
        ),
    }
    endpoint_with_query = f"{endpoint}?{urllib_parse.urlencode(params)}"
    request = urllib_request.Request(endpoint_with_query, method="GET")
    try:
        with urllib_request.urlopen(request, timeout=8) as response:
            raw = response.read().decode("utf-8")
    except (urllib_error.HTTPError, urllib_error.URLError):
        return []

    try:
        payload = json.loads(raw)
    except Exception:
        return []

    products = payload.get("products")
    if not isinstance(products, list):
        return []

    hits: list[dict[str, Any]] = []
    for product in products:
        if not isinstance(product, dict):
            continue
        parsed = parse_open_facts_product_row(product, source_name=source_name)
        if not parsed:
            continue
        hits.append(parsed)
        if len(hits) >= bounded_limit:
            break
    return hits


def lookup_product_open_products_facts(barcode: str) -> dict[str, Any] | None:
    return lookup_product_open_facts_catalog(
        barcode,
        endpoint_template=OPEN_PRODUCTS_FACTS_PRODUCT_ENDPOINT,
        source_name="openproductsfacts",
    )


def lookup_product_open_beauty_facts(barcode: str) -> dict[str, Any] | None:
    return lookup_product_open_facts_catalog(
        barcode,
        endpoint_template=OPEN_BEAUTY_FACTS_PRODUCT_ENDPOINT,
        source_name="openbeautyfacts",
    )


def lookup_product_open_food_facts(barcode: str) -> dict[str, Any] | None:
    return lookup_product_open_facts_catalog(
        barcode,
        endpoint_template=OPEN_FOOD_FACTS_PRODUCT_ENDPOINT,
        source_name="openfoodfacts",
    )


def lookup_product_upcitemdb(barcode: str) -> dict[str, Any] | None:
    endpoint = UPCITEMDB_LOOKUP_ENDPOINT.format(barcode=barcode)
    headers: dict[str, str] = {}
    api_key = str(os.getenv(UPCITEMDB_API_KEY_ENV, "") or "").strip()
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    request = urllib_request.Request(endpoint, headers=headers, method="GET")
    try:
        with urllib_request.urlopen(request, timeout=8) as response:
            raw = response.read().decode("utf-8")
    except (urllib_error.HTTPError, urllib_error.URLError):
        return None

    try:
        payload = json.loads(raw)
    except Exception:
        return None

    rows = payload.get("items")
    if not isinstance(rows, list) or not rows:
        return None
    row = rows[0]
    if not isinstance(row, dict):
        return None

    images = row.get("images")
    image_url = normalize_optional_text(images[0]) if isinstance(images, list) and images else None
    name = normalize_optional_text(row.get("title"))
    category = normalize_lookup_category(row.get("category"))
    if not category:
        category = infer_inventory_category_from_text(
            name,
            row.get("description"),
            row.get("title"),
            row.get("category"),
        )
    unit = infer_inventory_unit_from_text(name)

    if not any([name, category, unit, image_url]):
        return None

    return {
        "source": "upcitemdb",
        "name": name,
        "category": category,
        "unit": unit,
        "image_url": image_url,
    }


def lookup_inventory_product_metadata(barcode: str) -> dict[str, Any] | None:
    providers = [
        lookup_product_inventory_catalog,
        lookup_product_open_products_facts,
        lookup_product_open_beauty_facts,
        lookup_product_open_food_facts,
        lookup_product_upcitemdb,
    ]
    hits: list[dict[str, Any]] = []
    for provider in providers:
        hit = provider(barcode)
        if hit:
            hits.append(hit)
        merged_preview = {
            key: next((source.get(key) for source in hits if source.get(key)), None)
            for key in ("name", "category", "unit", "image_url")
        }
        if all(merged_preview.values()):
            break
    if not hits:
        return None

    merged: dict[str, Any] = {"source": ", ".join(source["source"] for source in hits)}
    for key in ("name", "category", "unit", "image_url"):
        merged[key] = next((source.get(key) for source in hits if source.get(key)), None)
    return merged


def merge_lookup_source_names(existing_source: Any, incoming_source: Any) -> str | None:
    ordered_parts: list[str] = []
    for raw in (existing_source, incoming_source):
        for part in str(raw or "").split(","):
            token = part.strip()
            if token and token not in ordered_parts:
                ordered_parts.append(token)
    if not ordered_parts:
        return None
    return ", ".join(ordered_parts)


def search_inventory_product_metadata(query: str, *, limit: int = 12) -> list[dict[str, Any]]:
    search_query = normalize_lookup_query(query)
    if not search_query:
        return []
    bounded_limit = max(1, min(int(limit), 60))
    provider_limit = max(2, min(8, bounded_limit))
    providers = [
        lambda q: search_inventory_product_catalog(q, limit=provider_limit),
        lambda q: search_products_open_facts_catalog(
            q,
            endpoint=OPEN_PRODUCTS_FACTS_SEARCH_ENDPOINT,
            source_name="openproductsfacts",
            limit=provider_limit,
        ),
        lambda q: search_products_open_facts_catalog(
            q,
            endpoint=OPEN_BEAUTY_FACTS_SEARCH_ENDPOINT,
            source_name="openbeautyfacts",
            limit=provider_limit,
        ),
        lambda q: search_products_open_facts_catalog(
            q,
            endpoint=OPEN_FOOD_FACTS_SEARCH_ENDPOINT,
            source_name="openfoodfacts",
            limit=provider_limit,
        ),
    ]
    merged_hits: list[dict[str, Any]] = []
    index_by_key: dict[tuple[str, str], int] = {}
    for provider in providers:
        try:
            provider_hits = provider(search_query)
        except Exception:
            continue
        for hit in provider_hits:
            barcode_key = str(hit.get("barcode") or "")
            name_key = str(hit.get("name") or "").strip().lower()
            if not barcode_key and not name_key:
                continue
            key = (barcode_key, name_key)
            existing_index = index_by_key.get(key)
            if existing_index is None:
                index_by_key[key] = len(merged_hits)
                merged_hits.append(dict(hit))
                if len(merged_hits) >= bounded_limit:
                    break
                continue
            existing = merged_hits[existing_index]
            existing["source"] = merge_lookup_source_names(existing.get("source"), hit.get("source"))
            for field in ("name", "category", "unit", "image_url", "barcode"):
                if not existing.get(field) and hit.get(field):
                    existing[field] = hit[field]
        if len(merged_hits) >= bounded_limit:
            break
    return merged_hits[:bounded_limit]


def normalize_item_unit(value: Any) -> str:
    text = str(value or "").strip().lower()
    return text if text else "each"


def as_active_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return int(value) != 0
    text = str(value).strip().lower()
    return text not in {"", "0", "false", "no", "off"}


def format_retreat_inventory_location_row(row: Any) -> dict[str, Any]:
    return {
        "id": int(row["id"]),
        "name": row["name"],
        "description": row["description"],
        "active": as_active_flag(row["active"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def ensure_retreat_inventory_location(conn: Any, location_name: str) -> int:
    clean_name = validate_shelf_location_name(str(location_name or ""))

    existing = conn.execute(
        """
        SELECT id
        FROM retreat_inventory_locations
        WHERE deleted_at IS NULL
          AND lower(name) = lower(?)
        """,
        (clean_name,),
    ).fetchone()
    if existing:
        return int(existing["id"])

    row = conn.execute(
        """
        INSERT INTO retreat_inventory_locations(
            name,
            active,
            created_at,
            updated_at
        )
        VALUES (?, 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        RETURNING id
        """,
        (clean_name,),
    ).fetchone()
    return int(row["id"])


def load_retreat_inventory_item_locations(conn: Any, item_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not item_ids:
        return {}
    placeholders = ", ".join("?" for _ in item_ids)
    rows = conn.execute(
        f"""
        SELECT
            il.item_id,
            il.location_id,
            il.quantity,
            il.updated_at,
            l.name AS location_name,
            l.description AS location_description
        FROM retreat_inventory_item_locations il
        JOIN retreat_inventory_locations l ON l.id = il.location_id
        WHERE il.item_id IN ({placeholders})
          AND l.deleted_at IS NULL
        ORDER BY il.item_id, lower(l.name), il.id
        """,
        tuple(item_ids),
    ).fetchall()

    by_item: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        item_id = int(row["item_id"])
        by_item.setdefault(item_id, []).append(
            {
                "location_id": int(row["location_id"]),
                "location_name": row["location_name"],
                "location_description": row["location_description"],
                "quantity": int(row["quantity"] or 0),
                "updated_at": row["updated_at"],
            }
        )
    return by_item


def validate_retreat_item_location_inputs(
    conn: Any,
    payload_locations: list[RetreatInventoryItemLocationInput],
) -> list[dict[str, Any]]:
    if not payload_locations:
        return []

    normalized: list[dict[str, Any]] = []
    seen_ids: set[int] = set()
    for entry in payload_locations:
        location_id = int(entry.locationId)
        if location_id in seen_ids:
            raise HTTPException(status_code=400, detail=f"Duplicate location id in payload: {location_id}.")
        seen_ids.add(location_id)
        normalized.append(
            {
                "location_id": location_id,
                "quantity": int(entry.quantity),
            }
        )

    placeholders = ", ".join("?" for _ in normalized)
    location_rows = conn.execute(
        f"""
        SELECT id, name, active
        FROM retreat_inventory_locations
        WHERE deleted_at IS NULL
          AND id IN ({placeholders})
        """,
        tuple(entry["location_id"] for entry in normalized),
    ).fetchall()
    location_by_id = {int(row["id"]): row for row in location_rows}

    for entry in normalized:
        location_row = location_by_id.get(entry["location_id"])
        if not location_row:
            raise HTTPException(status_code=404, detail=f"Location {entry['location_id']} not found.")
        if not as_active_flag(location_row["active"]):
            raise HTTPException(status_code=400, detail=f"Location {location_row['name']} is inactive.")
        entry["location_name"] = location_row["name"]
    return normalized


def replace_retreat_item_locations(conn: Any, item_id: int, locations: list[dict[str, Any]]) -> None:
    conn.execute("DELETE FROM retreat_inventory_item_locations WHERE item_id = ?", (item_id,))
    for entry in locations:
        conn.execute(
            """
            INSERT INTO retreat_inventory_item_locations(
                item_id,
                location_id,
                quantity,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (
                int(item_id),
                int(entry["location_id"]),
                int(entry["quantity"]),
            ),
        )


def sync_retreat_item_level_from_locations(conn: Any, item_id: int) -> None:
    tracking_row = conn.execute(
        """
        SELECT c.tracking_mode
        FROM retreat_inventory_items i
        JOIN retreat_inventory_categories c ON c.id = i.category_id
        WHERE i.id = ?
          AND i.deleted_at IS NULL
          AND c.deleted_at IS NULL
        """,
        (item_id,),
    ).fetchone()
    if not tracking_row:
        return
    if str(tracking_row["tracking_mode"] or "ITEM").upper() != "ITEM":
        return

    total_row = conn.execute(
        """
        SELECT COALESCE(SUM(il.quantity), 0) AS total_quantity
        FROM retreat_inventory_item_locations il
        JOIN retreat_inventory_locations l ON l.id = il.location_id
        WHERE il.item_id = ?
          AND l.deleted_at IS NULL
        """,
        (item_id,),
    ).fetchone()
    total_quantity = int(total_row["total_quantity"] or 0)

    level_row = conn.execute(
        """
        SELECT id
        FROM retreat_inventory_levels
        WHERE item_id = ?
          AND category_id IS NULL
        """,
        (item_id,),
    ).fetchone()
    if level_row:
        conn.execute(
            """
            UPDATE retreat_inventory_levels
            SET quantity = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (total_quantity, int(level_row["id"])),
        )
        return

    conn.execute(
        """
        INSERT INTO retreat_inventory_levels(
            item_id,
            category_id,
            quantity,
            min_threshold,
            updated_at
        )
        VALUES (?, NULL, ?, 0, CURRENT_TIMESTAMP)
        """,
        (item_id, total_quantity),
    )


def format_retreat_inventory_category_row(row: Any) -> dict[str, Any]:
    return {
        "id": int(row["id"]),
        "name": row["name"],
        "tracking_mode": row["tracking_mode"],
        "image_url": row["image_url"],
        "active": as_active_flag(row["active"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def format_retreat_inventory_item_row(row: Any) -> dict[str, Any]:
    return {
        "id": int(row["id"]),
        "name": row["name"],
        "barcode": row["barcode"],
        "category_id": int(row["category_id"]),
        "category_name": row["category_name"],
        "tracking_mode": row["tracking_mode"],
        "shelf_location": row["shelf_location"],
        "unit": row["unit"],
        "purchase_url": row["purchase_url"],
        "brand": row["brand"],
        "description": row["description"],
        "image_url": row["image_url"],
        "active": as_active_flag(row["active"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def attach_retreat_inventory_item_locations(conn: Any, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    item_ids = [int(item["id"]) for item in items if item.get("id") is not None]
    location_map = load_retreat_inventory_item_locations(conn, item_ids)
    for item in items:
        item_id = int(item["id"])
        locations = location_map.get(item_id, [])
        item["locations"] = locations
        if locations:
            item["location_summary"] = ", ".join(
                f"{loc['location_name']} ({int(loc['quantity'])})" for loc in locations
            )
            # Backward compatibility field.
            item["shelf_location"] = locations[0]["location_name"]
        else:
            item["location_summary"] = None
            item["shelf_location"] = item.get("shelf_location")
    return items


@app.get("/api/retreat-inventory/categories")
def list_retreat_inventory_categories(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
    include_inactive: bool = Query(default=False),
) -> list[dict[str, Any]]:
    filters = ["deleted_at IS NULL"]
    if not include_inactive:
        filters.append("active = 1")
    where_sql = f"WHERE {' AND '.join(filters)}"
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT id, name, tracking_mode, image_url, active, created_at, updated_at
            FROM retreat_inventory_categories
            {where_sql}
            ORDER BY lower(name), id
            """
        ).fetchall()
    return [format_retreat_inventory_category_row(row) for row in rows]


@app.post("/api/retreat-inventory/categories", status_code=201)
def create_retreat_inventory_category(
    payload: RetreatInventoryCategoryCreate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    name = payload.name.strip()
    image_url = normalize_optional_text(payload.imageUrl)
    with get_connection() as conn:
        try:
            row = conn.execute(
                """
                INSERT INTO retreat_inventory_categories(
                    name, tracking_mode, image_url, active, created_at, updated_at
                )
                VALUES (?, ?, ?, 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id, name, tracking_mode, image_url, active, created_at, updated_at
                """,
                (name, payload.trackingMode, image_url),
            ).fetchone()
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not create category: {exc}") from exc
        conn.commit()
    return format_retreat_inventory_category_row(row)


@app.patch("/api/retreat-inventory/categories/{category_id}")
def update_retreat_inventory_category(
    category_id: int,
    payload: RetreatInventoryCategoryUpdate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    updates: list[str] = []
    params: list[Any] = []

    if payload.name is not None:
        clean_name = payload.name.strip()
        if not clean_name:
            raise HTTPException(status_code=400, detail="Category name cannot be blank.")
        updates.append("name = ?")
        params.append(clean_name)
    if payload.trackingMode is not None:
        updates.append("tracking_mode = ?")
        params.append(payload.trackingMode)
    if payload.imageUrl is not None:
        updates.append("image_url = ?")
        params.append(normalize_optional_text(payload.imageUrl))
    if payload.active is not None:
        updates.append("active = ?")
        params.append(1 if payload.active else 0)

    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM retreat_inventory_categories WHERE id = ? AND deleted_at IS NULL",
            (category_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Category not found")

        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(category_id)
            try:
                conn.execute(
                    f"""
                    UPDATE retreat_inventory_categories
                    SET {', '.join(updates)}
                    WHERE id = ?
                    """,
                    tuple(params),
                )
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"Could not update category: {exc}") from exc

        row = conn.execute(
            """
            SELECT id, name, tracking_mode, image_url, active, created_at, updated_at
            FROM retreat_inventory_categories
            WHERE id = ?
            """,
            (category_id,),
        ).fetchone()
        conn.commit()

    return format_retreat_inventory_category_row(row)


@app.get("/api/retreat-inventory/locations")
def list_retreat_inventory_locations(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
    include_inactive: bool = Query(default=False),
) -> list[dict[str, Any]]:
    filters = ["deleted_at IS NULL"]
    if not include_inactive:
        filters.append("active = 1")
    where_sql = f"WHERE {' AND '.join(filters)}"
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT id, name, description, active, created_at, updated_at
            FROM retreat_inventory_locations
            {where_sql}
            ORDER BY lower(name), id
            """
        ).fetchall()
    return [format_retreat_inventory_location_row(row) for row in rows]


@app.post("/api/retreat-inventory/locations", status_code=201)
def create_retreat_inventory_location(
    payload: RetreatInventoryLocationCreate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    name = validate_shelf_location_name(payload.name)

    with get_connection() as conn:
        try:
            row = conn.execute(
                """
                INSERT INTO retreat_inventory_locations(
                    name,
                    description,
                    active,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id, name, description, active, created_at, updated_at
                """,
                (
                    name,
                    normalize_optional_text(payload.description),
                    1 if payload.active else 0,
                ),
            ).fetchone()
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not create location: {exc}") from exc
        conn.commit()
    return format_retreat_inventory_location_row(row)


@app.patch("/api/retreat-inventory/locations/{location_id}")
def update_retreat_inventory_location(
    location_id: int,
    payload: RetreatInventoryLocationUpdate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    updates: list[str] = []
    params: list[Any] = []

    if payload.name is not None:
        name = validate_shelf_location_name(payload.name)
        updates.append("name = ?")
        params.append(name)
    if payload.description is not None:
        updates.append("description = ?")
        params.append(normalize_optional_text(payload.description))
    if payload.active is not None:
        updates.append("active = ?")
        params.append(1 if payload.active else 0)

    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM retreat_inventory_locations WHERE id = ? AND deleted_at IS NULL",
            (location_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Location not found")

        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(location_id)
            try:
                conn.execute(
                    f"""
                    UPDATE retreat_inventory_locations
                    SET {', '.join(updates)}
                    WHERE id = ?
                    """,
                    tuple(params),
                )
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"Could not update location: {exc}") from exc

        row = conn.execute(
            """
            SELECT id, name, description, active, created_at, updated_at
            FROM retreat_inventory_locations
            WHERE id = ?
            """,
            (location_id,),
        ).fetchone()
        conn.commit()
    return format_retreat_inventory_location_row(row)


@app.get("/api/retreat-inventory/items")
def list_retreat_inventory_items(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
    category_id: int | None = Query(default=None, ge=1),
    search: str | None = Query(default=None),
    barcode: str | None = Query(default=None),
    include_inactive: bool = Query(default=False),
) -> list[dict[str, Any]]:
    filters = ["i.deleted_at IS NULL", "c.deleted_at IS NULL"]
    params: list[Any] = []
    if category_id is not None:
        filters.append("i.category_id = ?")
        params.append(category_id)
    if search and search.strip():
        term = f"%{search.strip().lower()}%"
        filters.append(
            """(
                lower(i.name) LIKE ?
                OR lower(COALESCE(i.brand, '')) LIKE ?
                OR lower(COALESCE(i.purchase_url, '')) LIKE ?
                OR lower(COALESCE(i.unit, '')) LIKE ?
                OR EXISTS (
                    SELECT 1
                    FROM retreat_inventory_item_locations il
                    JOIN retreat_inventory_locations l ON l.id = il.location_id
                    WHERE il.item_id = i.id
                      AND l.deleted_at IS NULL
                      AND lower(l.name) LIKE ?
                )
            )"""
        )
        params.extend([term, term, term, term, term])
    if barcode and barcode.strip():
        filters.append("i.barcode = ?")
        params.append(barcode.strip())
    if not include_inactive:
        filters.append("i.active = 1")
        filters.append("c.active = 1")

    where_sql = f"WHERE {' AND '.join(filters)}"
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT
                i.id,
                i.name,
                i.barcode,
                i.category_id,
                i.shelf_location,
                i.unit,
                i.purchase_url,
                i.brand,
                i.description,
                i.image_url,
                i.active,
                i.created_at,
                i.updated_at,
                c.name AS category_name,
                c.tracking_mode
            FROM retreat_inventory_items i
            JOIN retreat_inventory_categories c ON c.id = i.category_id
            {where_sql}
            ORDER BY lower(i.name), i.id
            """,
            tuple(params),
        ).fetchall()
        items = [format_retreat_inventory_item_row(row) for row in rows]
        return attach_retreat_inventory_item_locations(conn, items)


@app.post("/api/retreat-inventory/items", status_code=201)
def create_retreat_inventory_item(
    payload: RetreatInventoryItemCreate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    name = payload.name.strip()
    barcode = payload.barcode.strip()
    if not barcode:
        raise HTTPException(status_code=400, detail="Barcode is required.")

    with get_connection() as conn:
        category = conn.execute(
            """
            SELECT id
            FROM retreat_inventory_categories
            WHERE id = ? AND deleted_at IS NULL
            """,
            (payload.categoryId,),
        ).fetchone()
        if not category:
            raise HTTPException(status_code=404, detail="Category not found")

        normalized_locations = validate_retreat_item_location_inputs(conn, payload.locations)
        legacy_shelf_location = normalize_optional_text(payload.shelfLocation)
        if not normalized_locations and legacy_shelf_location:
            location_id = ensure_retreat_inventory_location(conn, legacy_shelf_location)
            normalized_locations = [
                {
                    "location_id": location_id,
                    "location_name": legacy_shelf_location,
                    "quantity": 0,
                }
            ]
        shelf_location_value = (
            normalized_locations[0]["location_name"] if normalized_locations else legacy_shelf_location
        )

        try:
            row = conn.execute(
                """
                INSERT INTO retreat_inventory_items(
                    name,
                    barcode,
                    category_id,
                    shelf_location,
                    unit,
                    purchase_url,
                    brand,
                    description,
                    image_url,
                    active,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id
                """,
                (
                    name,
                    barcode,
                    payload.categoryId,
                    shelf_location_value,
                    normalize_item_unit(payload.unit),
                    normalize_optional_text(payload.purchaseUrl),
                    normalize_optional_text(payload.brand),
                    normalize_optional_text(payload.description),
                    normalize_optional_text(payload.imageUrl),
                    1 if payload.active else 0,
                ),
            ).fetchone()
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not create item: {exc}") from exc

        item_id = int(row["id"])
        replace_retreat_item_locations(conn, item_id, normalized_locations)
        sync_retreat_item_level_from_locations(conn, item_id)
        item = conn.execute(
            """
            SELECT
                i.id,
                i.name,
                i.barcode,
                i.category_id,
                i.shelf_location,
                i.unit,
                i.purchase_url,
                i.brand,
                i.description,
                i.image_url,
                i.active,
                i.created_at,
                i.updated_at,
                c.name AS category_name,
                c.tracking_mode
            FROM retreat_inventory_items i
            JOIN retreat_inventory_categories c ON c.id = i.category_id
            WHERE i.id = ?
            """,
            (item_id,),
        ).fetchone()
        item_payload = format_retreat_inventory_item_row(item)
        attach_retreat_inventory_item_locations(conn, [item_payload])
        conn.commit()

    return item_payload


@app.patch("/api/retreat-inventory/items/{item_id}")
def update_retreat_inventory_item(
    item_id: int,
    payload: RetreatInventoryItemUpdate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_ADMIN))],
) -> dict[str, Any]:
    updates: list[str] = []
    params: list[Any] = []
    replace_locations = False
    normalized_locations: list[dict[str, Any]] | None = None

    if payload.name is not None:
        clean_name = payload.name.strip()
        if not clean_name:
            raise HTTPException(status_code=400, detail="Item name cannot be blank.")
        updates.append("name = ?")
        params.append(clean_name)
    if payload.barcode is not None:
        clean_barcode = payload.barcode.strip()
        if not clean_barcode:
            raise HTTPException(status_code=400, detail="Barcode cannot be blank.")
        updates.append("barcode = ?")
        params.append(clean_barcode)
    if payload.categoryId is not None:
        updates.append("category_id = ?")
        params.append(payload.categoryId)
    if payload.unit is not None:
        updates.append("unit = ?")
        params.append(normalize_item_unit(payload.unit))
    if payload.purchaseUrl is not None:
        updates.append("purchase_url = ?")
        params.append(normalize_optional_text(payload.purchaseUrl))
    if payload.brand is not None:
        updates.append("brand = ?")
        params.append(normalize_optional_text(payload.brand))
    if payload.description is not None:
        updates.append("description = ?")
        params.append(normalize_optional_text(payload.description))
    if payload.imageUrl is not None:
        updates.append("image_url = ?")
        params.append(normalize_optional_text(payload.imageUrl))
    if payload.active is not None:
        updates.append("active = ?")
        params.append(1 if payload.active else 0)

    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM retreat_inventory_items WHERE id = ? AND deleted_at IS NULL",
            (item_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Item not found")

        if payload.categoryId is not None:
            category = conn.execute(
                """
                SELECT id
                FROM retreat_inventory_categories
                WHERE id = ? AND deleted_at IS NULL
                """,
                (payload.categoryId,),
            ).fetchone()
            if not category:
                raise HTTPException(status_code=404, detail="Category not found")

        if payload.locations is not None:
            normalized_locations = validate_retreat_item_location_inputs(conn, payload.locations)
            replace_locations = True
        elif payload.shelfLocation is not None:
            legacy_shelf_location = normalize_optional_text(payload.shelfLocation)
            if legacy_shelf_location:
                location_id = ensure_retreat_inventory_location(conn, legacy_shelf_location)
                normalized_locations = [
                    {
                        "location_id": location_id,
                        "location_name": legacy_shelf_location,
                        "quantity": 0,
                    }
                ]
            else:
                normalized_locations = []
            replace_locations = True

        if replace_locations:
            shelf_location_value = (
                normalized_locations[0]["location_name"] if normalized_locations else None
            )
            updates.append("shelf_location = ?")
            params.append(shelf_location_value)

        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(item_id)
            try:
                conn.execute(
                    f"""
                    UPDATE retreat_inventory_items
                    SET {', '.join(updates)}
                    WHERE id = ?
                    """,
                    tuple(params),
                )
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"Could not update item: {exc}") from exc

        if replace_locations and normalized_locations is not None:
            replace_retreat_item_locations(conn, item_id, normalized_locations)
            sync_retreat_item_level_from_locations(conn, item_id)

        row = conn.execute(
            """
            SELECT
                i.id,
                i.name,
                i.barcode,
                i.category_id,
                i.shelf_location,
                i.unit,
                i.purchase_url,
                i.brand,
                i.description,
                i.image_url,
                i.active,
                i.created_at,
                i.updated_at,
                c.name AS category_name,
                c.tracking_mode
            FROM retreat_inventory_items i
            JOIN retreat_inventory_categories c ON c.id = i.category_id
            WHERE i.id = ?
            """,
            (item_id,),
        ).fetchone()
        item_payload = format_retreat_inventory_item_row(row)
        attach_retreat_inventory_item_locations(conn, [item_payload])
        conn.commit()

    return item_payload


@app.get("/api/retreat-inventory/levels")
def list_retreat_inventory_levels(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                ril.id,
                ril.item_id,
                ril.category_id,
                ril.quantity,
                ril.min_threshold,
                ril.updated_at,
                i.name AS item_name,
                i.barcode AS item_barcode,
                i.unit AS item_unit,
                i.image_url AS item_image_url,
                c.id AS category_row_id,
                c.name AS category_name,
                c.tracking_mode
            FROM retreat_inventory_levels ril
            LEFT JOIN retreat_inventory_items i ON i.id = ril.item_id
            LEFT JOIN retreat_inventory_categories c
              ON c.id = COALESCE(ril.category_id, i.category_id)
            WHERE (i.deleted_at IS NULL OR i.id IS NULL)
              AND (c.deleted_at IS NULL OR c.id IS NULL)
            ORDER BY lower(COALESCE(i.name, c.name, '')), ril.id
            """
        ).fetchall()
        item_ids = [int(row["item_id"]) for row in rows if row["item_id"] is not None]
        location_map = load_retreat_inventory_item_locations(conn, item_ids)

    result: list[dict[str, Any]] = []
    for row in rows:
        item_id_value = row["item_id"]
        category_id_value = row["category_id"] if row["category_id"] is not None else row["category_row_id"]
        entity_type = "ITEM" if item_id_value is not None else "CATEGORY"
        item_locations = location_map.get(int(item_id_value), []) if item_id_value is not None else []
        item_shelf_location = item_locations[0]["location_name"] if item_locations else None
        location_summary = (
            ", ".join(f"{loc['location_name']} ({int(loc['quantity'])})" for loc in item_locations)
            if item_locations
            else None
        )
        result.append(
            {
                "id": int(row["id"]),
                "entity_type": entity_type,
                "item_id": int(item_id_value) if item_id_value is not None else None,
                "category_id": int(category_id_value) if category_id_value is not None else None,
                "item_name": row["item_name"],
                "item_barcode": row["item_barcode"],
                "item_shelf_location": item_shelf_location,
                "item_locations": item_locations,
                "location_summary": location_summary,
                "item_unit": row["item_unit"],
                "item_image_url": row["item_image_url"],
                "category_name": row["category_name"],
                "tracking_mode": row["tracking_mode"],
                "quantity": int(row["quantity"] or 0),
                "min_threshold": int(row["min_threshold"] or 0),
                "is_low_stock": int(row["quantity"] or 0) <= int(row["min_threshold"] or 0),
                "updated_at": row["updated_at"],
            }
        )
    return result


@app.get("/api/retreat-inventory/transactions")
def list_retreat_inventory_transactions(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
    limit: int = Query(default=100, ge=1, le=500),
) -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                rit.id,
                rit.entity_type,
                rit.entity_id,
                rit.transaction_type,
                rit.quantity,
                rit.reason,
                rit.barcode,
                rit.created_at,
                u.username AS user_name,
                i.name AS item_name,
                c.name AS category_name
            FROM retreat_inventory_transactions rit
            LEFT JOIN users u ON u.id = rit.user_id
            LEFT JOIN retreat_inventory_items i
              ON rit.entity_type = 'ITEM'
             AND i.id = rit.entity_id
            LEFT JOIN retreat_inventory_categories c
              ON rit.entity_type = 'CATEGORY'
             AND c.id = rit.entity_id
            ORDER BY rit.id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [
        {
            "id": int(row["id"]),
            "entity_type": row["entity_type"],
            "entity_id": int(row["entity_id"]),
            "entity_name": row["item_name"] if row["entity_type"] == "ITEM" else row["category_name"],
            "transaction_type": row["transaction_type"],
            "quantity": int(row["quantity"]),
            "reason": row["reason"],
            "barcode": row["barcode"],
            "user_name": row["user_name"],
            "created_at": row["created_at"],
        }
        for row in rows
    ]


@app.post("/api/retreat-inventory/scan")
def scan_retreat_inventory(
    payload: RetreatInventoryScanPayload,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    barcode = payload.barcode.strip()
    if not barcode:
        raise HTTPException(status_code=400, detail="Barcode is required.")

    tx_type = payload.transactionType
    quantity = int(payload.quantity)
    if tx_type in {"IN", "OUT"} and quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than zero for IN/OUT scans.")
    if tx_type == "ADJUSTMENT" and quantity == 0:
        raise HTTPException(status_code=400, detail="Quantity cannot be zero for adjustments.")

    with get_connection() as conn:
        item = conn.execute(
            """
            SELECT
                i.id,
                i.name,
                i.barcode,
                i.image_url,
                i.shelf_location,
                i.unit,
                i.purchase_url,
                i.active AS item_active,
                c.id AS category_id,
                c.name AS category_name,
                c.tracking_mode,
                c.active AS category_active
            FROM retreat_inventory_items i
            JOIN retreat_inventory_categories c ON c.id = i.category_id
            WHERE i.barcode = ?
              AND i.deleted_at IS NULL
              AND c.deleted_at IS NULL
            """,
            (barcode,),
        ).fetchone()
        if not item:
            raise HTTPException(status_code=404, detail=f"Barcode {barcode} not found.")
        if not as_active_flag(item["item_active"]):
            raise HTTPException(status_code=400, detail="Item is inactive.")
        if not as_active_flag(item["category_active"]):
            raise HTTPException(status_code=400, detail="Item category is inactive.")

        tracking_mode = str(item["tracking_mode"] or "ITEM").upper()
        entity_type = "CATEGORY" if tracking_mode == "CATEGORY" else "ITEM"
        entity_id = int(item["category_id"]) if entity_type == "CATEGORY" else int(item["id"])
        item_id = int(item["id"])

        item_locations = load_retreat_inventory_item_locations(conn, [item_id]).get(item_id, [])
        selected_location_id: int | None = None
        selected_location_name: str | None = None
        if entity_type == "ITEM":
            if payload.locationId is None and len(item_locations) > 1:
                raise HTTPException(
                    status_code=400,
                    detail="Select a location when scanning an item stored in multiple locations.",
                )
            if payload.locationId is not None:
                location = conn.execute(
                    """
                    SELECT id, name, active
                    FROM retreat_inventory_locations
                    WHERE id = ?
                      AND deleted_at IS NULL
                    """,
                    (payload.locationId,),
                ).fetchone()
                if not location:
                    raise HTTPException(status_code=404, detail=f"Location {payload.locationId} not found.")
                if not as_active_flag(location["active"]):
                    raise HTTPException(status_code=400, detail=f"Location {location['name']} is inactive.")
                selected_location_id = int(location["id"])
                selected_location_name = location["name"]
            elif len(item_locations) == 1:
                selected_location_id = int(item_locations[0]["location_id"])
                selected_location_name = item_locations[0]["location_name"]
        elif payload.locationId is not None:
            raise HTTPException(
                status_code=400,
                detail="Location cannot be applied when this item uses CATEGORY tracking mode.",
            )

        lock_clause = " FOR UPDATE" if getattr(conn, "backend", "sqlite") == "postgres" else ""
        if entity_type == "ITEM":
            level_where = "item_id = ? AND category_id IS NULL"
            level_params: tuple[Any, ...] = (entity_id,)
            insert_params: tuple[Any, ...] = (entity_id, None)
        else:
            level_where = "category_id = ? AND item_id IS NULL"
            level_params = (entity_id,)
            insert_params = (None, entity_id)

        level = conn.execute(
            f"""
            SELECT id, quantity, min_threshold
            FROM retreat_inventory_levels
            WHERE {level_where}{lock_clause}
            """,
            level_params,
        ).fetchone()
        if not level:
            try:
                conn.execute(
                    """
                    INSERT INTO retreat_inventory_levels(item_id, category_id, quantity, min_threshold, updated_at)
                    VALUES (?, ?, 0, 0, CURRENT_TIMESTAMP)
                    """,
                    insert_params,
                )
            except Exception:
                # Another request can create the row concurrently; read it again below.
                pass
            level = conn.execute(
                f"""
                SELECT id, quantity, min_threshold
                FROM retreat_inventory_levels
                WHERE {level_where}{lock_clause}
                """,
                level_params,
            ).fetchone()
        if not level:
            raise HTTPException(status_code=500, detail="Could not initialize inventory level row.")

        previous_quantity = int(level["quantity"] or 0)
        if tx_type == "IN":
            delta = quantity
        elif tx_type == "OUT":
            delta = -quantity
        else:
            delta = quantity
        next_quantity = previous_quantity + delta
        if next_quantity < 0:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient inventory: current quantity is {previous_quantity}.",
            )

        conn.execute(
            """
            UPDATE retreat_inventory_levels
            SET quantity = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (next_quantity, int(level["id"])),
        )

        location_quantity_after: int | None = None
        if entity_type == "ITEM" and selected_location_id is not None:
            level_location = conn.execute(
                f"""
                SELECT id, quantity
                FROM retreat_inventory_item_locations
                WHERE item_id = ?
                  AND location_id = ?{lock_clause}
                """,
                (item_id, selected_location_id),
            ).fetchone()
            if not level_location:
                conn.execute(
                    """
                    INSERT INTO retreat_inventory_item_locations(
                        item_id,
                        location_id,
                        quantity,
                        created_at,
                        updated_at
                    )
                    VALUES (?, ?, 0, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    """,
                    (item_id, selected_location_id),
                )
                level_location = conn.execute(
                    f"""
                    SELECT id, quantity
                    FROM retreat_inventory_item_locations
                    WHERE item_id = ?
                      AND location_id = ?{lock_clause}
                    """,
                    (item_id, selected_location_id),
                ).fetchone()

            previous_location_quantity = int(level_location["quantity"] or 0)
            location_quantity_after = previous_location_quantity + delta
            if location_quantity_after < 0:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Insufficient inventory at location {selected_location_name}: "
                        f"current quantity is {previous_location_quantity}."
                    ),
                )
            conn.execute(
                """
                UPDATE retreat_inventory_item_locations
                SET quantity = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (location_quantity_after, int(level_location["id"])),
            )

        item_locations = load_retreat_inventory_item_locations(conn, [item_id]).get(item_id, [])
        item_shelf_location = item_locations[0]["location_name"] if item_locations else item["shelf_location"]

        tx_row = conn.execute(
            """
            INSERT INTO retreat_inventory_transactions(
                entity_type,
                entity_id,
                transaction_type,
                quantity,
                reason,
                user_id,
                barcode,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            RETURNING id, created_at
            """,
            (
                entity_type,
                entity_id,
                tx_type,
                quantity,
                normalize_optional_text(payload.reason),
                user.id,
                barcode,
            ),
        ).fetchone()
        conn.commit()

    entity_name = item["category_name"] if entity_type == "CATEGORY" else item["name"]
    return {
        "transaction_id": int(tx_row["id"]),
        "created_at": tx_row["created_at"],
        "barcode": barcode,
        "item": {
            "id": int(item["id"]),
            "name": item["name"],
            "image_url": item["image_url"],
            "shelf_location": item_shelf_location,
            "locations": item_locations,
            "unit": item["unit"],
            "purchase_url": item["purchase_url"],
            "category_id": int(item["category_id"]),
            "category_name": item["category_name"],
            "tracking_mode": tracking_mode,
        },
        "location": (
            {
                "id": selected_location_id,
                "name": selected_location_name,
                "quantity_after": location_quantity_after,
            }
            if selected_location_id is not None
            else None
        ),
        "entity": {
            "type": entity_type,
            "id": entity_id,
            "name": entity_name,
        },
        "transaction_type": tx_type,
        "quantity": quantity,
        "delta": delta,
        "quantity_before": previous_quantity,
        "quantity_after": next_quantity,
    }


@app.get("/api/retreat-inventory/locations/by-shelf")
def list_retreat_inventory_locations_by_shelf(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT l.id, l.name, l.description, l.active,
                   COUNT(il.id) AS item_count
            FROM retreat_inventory_locations l
            LEFT JOIN retreat_inventory_item_locations il ON il.location_id = l.id
            WHERE l.deleted_at IS NULL AND l.active = 1
            GROUP BY l.id, l.name, l.description, l.active
            ORDER BY lower(l.name), l.id
            """
        ).fetchall()

    shelves: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        name = row["name"] or ""
        shelf_letter = name[0].upper() if name else "?"
        shelves.setdefault(shelf_letter, []).append(
            {
                "id": int(row["id"]),
                "name": row["name"],
                "description": row["description"],
                "item_count": int(row["item_count"]),
            }
        )

    return [
        {"shelf": letter, "locations": locs}
        for letter, locs in sorted(shelves.items())
    ]


def format_purchase_order_row(row: Any) -> dict[str, Any]:
    return {
        "id": int(row["id"]),
        "supplier_name": row["supplier_name"],
        "status": row["status"],
        "expected_date": row["expected_date"],
        "notes": row["notes"],
        "created_by_user_id": int(row["created_by_user_id"]) if row["created_by_user_id"] else None,
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def load_purchase_order_items(conn: Any, order_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not order_ids:
        return {}
    placeholders = ", ".join("?" for _ in order_ids)
    rows = conn.execute(
        f"""
        SELECT
            poi.id,
            poi.purchase_order_id,
            poi.entity_type,
            poi.entity_id,
            poi.ordered_quantity,
            poi.received_quantity,
            poi.created_at,
            poi.updated_at,
            CASE WHEN poi.entity_type = 'ITEM' THEN i.name ELSE c.name END AS entity_name,
            i.barcode AS item_barcode,
            i.unit AS item_unit,
            i.purchase_url AS item_purchase_url
        FROM retreat_inventory_purchase_order_items poi
        LEFT JOIN retreat_inventory_items i
          ON poi.entity_type = 'ITEM' AND i.id = poi.entity_id
        LEFT JOIN retreat_inventory_categories c
          ON poi.entity_type = 'CATEGORY' AND c.id = poi.entity_id
        WHERE poi.purchase_order_id IN ({placeholders})
        ORDER BY poi.purchase_order_id, poi.id
        """,
        tuple(order_ids),
    ).fetchall()

    by_order: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        oid = int(row["purchase_order_id"])
        by_order.setdefault(oid, []).append(
            {
                "id": int(row["id"]),
                "entity_type": row["entity_type"],
                "entity_id": int(row["entity_id"]),
                "entity_name": row["entity_name"],
                "item_barcode": row["item_barcode"],
                "item_unit": row["item_unit"],
                "item_purchase_url": row["item_purchase_url"],
                "ordered_quantity": int(row["ordered_quantity"]),
                "received_quantity": int(row["received_quantity"]),
            }
        )
    return by_order


def normalize_inventory_barcode_list(values: list[Any] | None) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for value in values or []:
        barcode = normalize_inventory_barcode(value)
        if barcode in seen:
            continue
        seen.add(barcode)
        normalized.append(barcode)
    return normalized


def build_inventory_barcode_payload(
    primary_barcode: Any | None,
    *,
    requested_barcodes: list[Any] | None = None,
    existing_barcodes: list[str] | None = None,
) -> tuple[str | None, list[str]]:
    primary = normalize_inventory_barcode(primary_barcode) if primary_barcode is not None else None
    merged_values: list[str] = []
    if primary:
        merged_values.append(primary)
    if requested_barcodes is None:
        merged_values.extend(existing_barcodes or [])
    else:
        merged_values.extend(normalize_inventory_barcode_list(requested_barcodes))

    ordered: list[str] = []
    seen: set[str] = set()
    for barcode in merged_values:
        if barcode in seen:
            continue
        seen.add(barcode)
        ordered.append(barcode)
    return primary or (ordered[0] if ordered else None), ordered


def load_standalone_inventory_barcodes_map(
    conn: Any,
    item_ids: list[int],
    *,
    fallback_primary_by_item: dict[int, Any] | None = None,
) -> dict[int, list[str]]:
    normalized_ids = [int(item_id) for item_id in item_ids if int(item_id) > 0]
    barcode_map: dict[int, list[str]] = {item_id: [] for item_id in normalized_ids}
    primary_by_item = {
        int(item_id): normalize_lookup_barcode_candidate(barcode)
        for item_id, barcode in (fallback_primary_by_item or {}).items()
        if int(item_id) > 0
    }

    if normalized_ids and table_exists(conn, "standalone_inventory_barcodes"):
        placeholders = ", ".join("?" for _ in normalized_ids)
        rows = conn.execute(
            f"""
            SELECT inventory_item_id, barcode
            FROM standalone_inventory_barcodes
            WHERE inventory_item_id IN ({placeholders})
            ORDER BY inventory_item_id, id
            """,
            tuple(normalized_ids),
        ).fetchall()
        for row in rows:
            item_id = int(row["inventory_item_id"])
            barcode = normalize_lookup_barcode_candidate(row["barcode"])
            if not barcode:
                continue
            barcode_map.setdefault(item_id, [])
            if barcode not in barcode_map[item_id]:
                barcode_map[item_id].append(barcode)

    for item_id in normalized_ids:
        ordered: list[str] = []
        primary = primary_by_item.get(item_id)
        if primary:
            ordered.append(primary)
        for barcode in barcode_map.get(item_id, []):
            if barcode not in ordered:
                ordered.append(barcode)
        barcode_map[item_id] = ordered
    return barcode_map


def load_standalone_inventory_item_barcodes(
    conn: Any,
    item_id: int,
    *,
    fallback_primary_barcode: Any | None = None,
) -> list[str]:
    barcode_map = load_standalone_inventory_barcodes_map(
        conn,
        [item_id],
        fallback_primary_by_item={int(item_id): fallback_primary_barcode},
    )
    return barcode_map.get(int(item_id), [])


def format_standalone_inventory_row(
    row: Any,
    *,
    barcode_map: dict[int, list[str]] | None = None,
) -> dict[str, Any]:
    entry = dict(row)
    item_id = int(entry.get("id") or 0)
    primary_barcode = normalize_lookup_barcode_candidate(entry.get("barcode"))
    barcodes = list((barcode_map or {}).get(item_id, []))
    if primary_barcode and primary_barcode not in barcodes:
        barcodes.insert(0, primary_barcode)
    entry["barcode"] = primary_barcode or (barcodes[0] if barcodes else None)
    entry["primary_barcode"] = entry["barcode"]
    entry["barcodes"] = barcodes
    entry["barcode_count"] = len(barcodes)
    entry["category"] = normalize_inventory_category(entry.get("category"))
    return entry


def format_standalone_inventory_rows(conn: Any, rows: list[Any]) -> list[dict[str, Any]]:
    row_list = list(rows)
    if not row_list:
        return []
    fallback_primary_by_item = {
        int(row["id"]): row["barcode"]
        for row in row_list
        if row["id"] is not None
    }
    barcode_map = load_standalone_inventory_barcodes_map(
        conn,
        [int(row["id"]) for row in row_list if row["id"] is not None],
        fallback_primary_by_item=fallback_primary_by_item,
    )
    return [format_standalone_inventory_row(row, barcode_map=barcode_map) for row in row_list]


def find_standalone_inventory_item_by_barcode(conn: Any, barcode: str) -> Any | None:
    normalized_barcode = normalize_inventory_barcode(barcode)
    if table_exists(conn, "standalone_inventory_barcodes"):
        row = conn.execute(
            """
            SELECT si.*
            FROM standalone_inventory_barcodes sib
            JOIN standalone_inventory si
              ON si.id = sib.inventory_item_id
            WHERE sib.barcode = ?
            ORDER BY
                CASE WHEN COALESCE(si.barcode, '') = ? THEN 0 ELSE 1 END,
                si.updated_at DESC,
                si.id DESC
            LIMIT 1
            """,
            (normalized_barcode, normalized_barcode),
        ).fetchone()
        if row:
            return row
    return conn.execute(
        """
        SELECT *
        FROM standalone_inventory
        WHERE barcode = ?
        ORDER BY updated_at DESC, id DESC
        LIMIT 1
        """,
        (normalized_barcode,),
    ).fetchone()


def find_standalone_inventory_barcode_owner(
    conn: Any,
    barcode: str,
    *,
    exclude_item_id: int | None = None,
) -> Any | None:
    normalized_barcode = normalize_inventory_barcode(barcode)
    if table_exists(conn, "standalone_inventory_barcodes"):
        filters = ["sib.barcode = ?"]
        params: list[Any] = [normalized_barcode]
        if exclude_item_id is not None:
            filters.append("si.id != ?")
            params.append(int(exclude_item_id))
        row = conn.execute(
            f"""
            SELECT si.id, si.item_name
            FROM standalone_inventory_barcodes sib
            JOIN standalone_inventory si
              ON si.id = sib.inventory_item_id
            WHERE {' AND '.join(filters)}
            LIMIT 1
            """,
            tuple(params),
        ).fetchone()
        if row:
            return row

    filters = ["barcode = ?"]
    params = [normalized_barcode]
    if exclude_item_id is not None:
        filters.append("id != ?")
        params.append(int(exclude_item_id))
    return conn.execute(
        f"""
        SELECT id, item_name
        FROM standalone_inventory
        WHERE {' AND '.join(filters)}
        LIMIT 1
        """,
        tuple(params),
    ).fetchone()


def sync_standalone_inventory_item_barcodes(
    conn: Any,
    *,
    item_id: int,
    primary_barcode: str | None,
    barcodes: list[str],
) -> None:
    normalized_item_id = int(item_id)
    normalized_primary = normalize_inventory_barcode(primary_barcode) if primary_barcode else None
    ordered_barcodes: list[str] = []
    seen: set[str] = set()
    for raw_barcode in ([normalized_primary] if normalized_primary else []) + list(barcodes or []):
        barcode = normalize_inventory_barcode(raw_barcode)
        if barcode in seen:
            continue
        seen.add(barcode)
        ordered_barcodes.append(barcode)
    effective_primary = normalized_primary or (ordered_barcodes[0] if ordered_barcodes else None)

    for barcode in ordered_barcodes:
        duplicate = find_standalone_inventory_barcode_owner(conn, barcode, exclude_item_id=normalized_item_id)
        if duplicate:
            raise HTTPException(
                status_code=409,
                detail=f'Barcode {barcode} already exists for "{duplicate["item_name"]}".',
            )

    conn.execute(
        """
        UPDATE standalone_inventory
        SET barcode = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (effective_primary, normalized_item_id),
    )

    if not table_exists(conn, "standalone_inventory_barcodes"):
        return

    if ordered_barcodes:
        placeholders = ", ".join("?" for _ in ordered_barcodes)
        conn.execute(
            f"""
            DELETE FROM standalone_inventory_barcodes
            WHERE inventory_item_id = ?
              AND barcode NOT IN ({placeholders})
            """,
            (normalized_item_id, *ordered_barcodes),
        )
    else:
        conn.execute(
            "DELETE FROM standalone_inventory_barcodes WHERE inventory_item_id = ?",
            (normalized_item_id,),
        )

    for barcode in ordered_barcodes:
        existing = conn.execute(
            """
            SELECT id
            FROM standalone_inventory_barcodes
            WHERE barcode = ?
            """,
            (barcode,),
        ).fetchone()
        if existing:
            conn.execute(
                """
                UPDATE standalone_inventory_barcodes
                SET inventory_item_id = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (normalized_item_id, int(existing["id"])),
            )
            continue
        conn.execute(
            """
            INSERT INTO standalone_inventory_barcodes(
                inventory_item_id,
                barcode,
                created_at,
                updated_at
            )
            VALUES (?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (normalized_item_id, barcode),
        )


def merge_distinct_pipe_text(*values: Any) -> str | None:
    parts: list[str] = []
    seen: set[str] = set()
    for raw_value in values:
        for part in str(raw_value or "").split("|"):
            clean = normalize_optional_text(part)
            if not clean or clean in seen:
                continue
            seen.add(clean)
            parts.append(clean)
    return " | ".join(parts) if parts else None


def normalize_inventory_unit_for_merge(value: Any) -> str | None:
    text = normalize_optional_text(value)
    if not text:
        return None
    lowered = text.lower()
    aliases = {
        "ea": "each",
        "eaches": "each",
        "each": "each",
        "unit": "each",
        "units": "each",
        "count": "each",
        "counts": "each",
        "ct": "each",
        "pk": "pack",
        "pks": "pack",
        "packs": "pack",
        "cases": "case",
        "boxes": "box",
        "bags": "bag",
        "bottles": "bottle",
        "rolls": "roll",
        "pairs": "pair",
        "sets": "set",
        "jars": "jar",
        "cans": "can",
        "cartons": "carton",
        "tubs": "tub",
        "jugs": "jug",
        "pieces": "piece",
        "packets": "packet",
        "bundles": "bundle",
    }
    return aliases.get(lowered, lowered)


def normalize_inventory_location_for_merge(value: Any) -> str | None:
    text = normalize_optional_text(value)
    if not text:
        return None
    try:
        return normalize_storage_grid_location(text)
    except HTTPException:
        return text.upper()


def ensure_standalone_inventory_merge_safe(source: Any, target: Any) -> None:
    source_name = normalize_optional_text(source["item_name"]) or f'item {source["id"]}'
    target_name = normalize_optional_text(target["item_name"]) or f'item {target["id"]}'
    source_unit_raw = normalize_optional_text(source["unit"])
    target_unit_raw = normalize_optional_text(target["unit"])
    source_unit = normalize_inventory_unit_for_merge(source_unit_raw)
    target_unit = normalize_inventory_unit_for_merge(target_unit_raw)
    if source_unit and target_unit and source_unit != target_unit:
        raise HTTPException(
            status_code=400,
            detail=(
                f'Cannot merge "{source_name}" into "{target_name}" because units differ '
                f'({source_unit_raw} vs {target_unit_raw}). Align the units first.'
            ),
        )

    source_location_raw = normalize_optional_text(source["location"])
    target_location_raw = normalize_optional_text(target["location"])
    source_location = normalize_inventory_location_for_merge(source_location_raw)
    target_location = normalize_inventory_location_for_merge(target_location_raw)
    if source_location and target_location and source_location != target_location:
        raise HTTPException(
            status_code=400,
            detail=(
                f'Cannot merge "{source_name}" into "{target_name}" because locations differ '
                f'({source_location_raw} vs {target_location_raw}). Move them to one location first '
                "or keep them as separate rows."
            ),
        )


def ensure_standalone_inventory_order_merge_safe(source_order_row: Any, target_order_row: Any) -> None:
    order_id = int(source_order_row["order_id"])
    source_unit_raw = normalize_optional_text(source_order_row["unit_snapshot"])
    target_unit_raw = normalize_optional_text(target_order_row["unit_snapshot"])
    source_unit = normalize_inventory_unit_for_merge(source_unit_raw)
    target_unit = normalize_inventory_unit_for_merge(target_unit_raw)
    if source_unit and target_unit and source_unit != target_unit:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Cannot merge inventory items because order {order_id} has conflicting unit snapshots "
                f"({source_unit_raw} vs {target_unit_raw}). Resolve the order rows first."
            ),
        )

    source_purchase_unit = normalize_purchase_unit(source_order_row["purchase_unit"])
    target_purchase_unit = normalize_purchase_unit(target_order_row["purchase_unit"])
    source_units_per_purchase = round(normalize_units_per_purchase(source_order_row["units_per_purchase"]), 6)
    target_units_per_purchase = round(normalize_units_per_purchase(target_order_row["units_per_purchase"]), 6)
    if source_purchase_unit != target_purchase_unit or source_units_per_purchase != target_units_per_purchase:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Cannot merge inventory items because order {order_id} uses different purchase packaging "
                f"({source_purchase_unit} x {source_units_per_purchase:g} vs "
                f"{target_purchase_unit} x {target_units_per_purchase:g}). Resolve the order rows first."
            ),
        )


def build_standalone_inventory_filters(
    *,
    category: str | None = None,
    search: str | None = None,
) -> tuple[list[str], list[Any]]:
    filters: list[str] = []
    params: list[Any] = []

    category_filter = normalize_inventory_category(category)
    if category_filter:
        filters.append("lower(category) = lower(?)")
        params.append(category_filter)

    if search and search.strip():
        needle = search.strip().lower()
        filters.append(
            (
                "(lower(item_name) LIKE ? OR lower(COALESCE(category, '')) LIKE ? "
                "OR lower(COALESCE(location, '')) LIKE ? OR COALESCE(barcode, '') LIKE ? "
                "OR EXISTS ("
                "  SELECT 1 FROM standalone_inventory_barcodes sib "
                "  WHERE sib.inventory_item_id = standalone_inventory.id AND sib.barcode LIKE ?"
                ") "
                "OR lower(COALESCE(notes, '')) LIKE ? OR lower(COALESCE(order_url, '')) LIKE ?)"
            )
        )
        params.extend(
            [
                f"%{needle}%",
                f"%{needle}%",
                f"%{needle}%",
                f"%{needle}%",
                f"%{needle}%",
                f"%{needle}%",
                f"%{needle}%",
            ]
        )

    return filters, params


def query_standalone_inventory_rows(
    conn: Any,
    *,
    category: str | None = None,
    search: str | None = None,
    limit: int | None = None,
) -> list[dict[str, Any]]:
    filters, params = build_standalone_inventory_filters(category=category, search=search)
    where_sql = f"WHERE {' AND '.join(filters)}" if filters else ""
    sql = f"SELECT * FROM standalone_inventory {where_sql} ORDER BY lower(item_name)"
    if limit is not None:
        bounded_limit = max(1, min(int(limit), 200))
        sql = f"{sql} LIMIT ?"
        params.append(bounded_limit)
    rows = conn.execute(sql, tuple(params)).fetchall()
    return format_standalone_inventory_rows(conn, rows)


def query_standalone_inventory_withdraw_matches(
    conn: Any,
    *,
    query: str,
    limit: int = 12,
) -> list[dict[str, Any]]:
    normalized_query = normalize_required_text(query, field_name="Query")
    bounded_limit = max(1, min(int(limit), 50))
    matches: list[dict[str, Any]] = []
    seen_item_ids: set[int] = set()

    barcode_candidate = re.sub(r"\D+", "", normalized_query)
    if INVENTORY_BARCODE_RE.fullmatch(barcode_candidate):
        exact_row = find_standalone_inventory_item_by_barcode(conn, barcode_candidate)
        if exact_row is not None:
            exact_payload = format_standalone_inventory_rows(conn, [exact_row])[0]
            matches.append(exact_payload)
            seen_item_ids.add(int(exact_payload["id"]))

    for payload in query_standalone_inventory_rows(conn, search=normalized_query, limit=bounded_limit):
        item_id = int(payload["id"])
        if item_id in seen_item_ids:
            continue
        matches.append(payload)
        seen_item_ids.add(item_id)
        if len(matches) >= bounded_limit:
            break
    return matches


def as_non_negative_quantity(value: Any) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return 0.0
    if numeric <= 0:
        return 0.0
    return float(numeric)


def rounded_quantity(value: Any) -> float:
    return round(as_non_negative_quantity(value), 3)


def normalize_purchase_unit(value: Any) -> str:
    normalized = normalize_optional_text(value)
    if not normalized:
        return "unit"
    key = normalized.lower()
    aliases = {
        "unit": "unit",
        "units": "unit",
        "each": "unit",
        "ea": "unit",
        "case": "case",
        "cases": "case",
        "cs": "case",
        "pack": "pack",
        "packs": "pack",
    }
    return aliases.get(key, key)


def normalize_units_per_purchase(value: Any) -> float:
    units = rounded_quantity(value)
    if units <= 0:
        return 1.0
    return units


def normalize_inventory_order_domain(value: Any) -> str | None:
    normalized = normalize_optional_text(value)
    if not normalized:
        return None
    key = normalized.upper().replace("-", "_").replace(" ", "_")
    allowed = {"FOOD", "NON_FOOD"}
    if key not in allowed:
        raise HTTPException(status_code=400, detail="Invalid inventory order domain.")
    return key


def normalize_inventory_order_source_type(value: Any) -> str | None:
    normalized = normalize_optional_text(value)
    if not normalized:
        return None
    key = normalized.upper().replace("-", "_").replace(" ", "_")
    allowed = {"SHOPPING_LIST", "NON_FOOD_PLAN", "MANUAL", "LEGACY"}
    if key not in allowed:
        raise HTTPException(status_code=400, detail="Invalid inventory order source type.")
    return key


def normalize_inventory_order_workflow_stage(value: Any) -> str | None:
    normalized = normalize_optional_text(value)
    if not normalized:
        return None
    key = normalized.upper().replace("-", "_").replace(" ", "_")
    allowed = {"PLANNING", "PURCHASING", "RECEIVING", "COMPLETE"}
    if key not in allowed:
        raise HTTPException(status_code=400, detail="Invalid inventory order workflow stage.")
    return key


def default_inventory_order_workflow_stage(domain: Any) -> str:
    normalized_domain = normalize_inventory_order_domain(domain)
    return "RECEIVING" if normalized_domain == "FOOD" else "PLANNING"


def inventory_order_workspace_label(workflow_stage: str) -> str:
    labels = {
        "PLANNING": "Planning",
        "PURCHASING": "Purchasing",
        "RECEIVING": "Receiving",
        "COMPLETE": "Completed Orders",
    }
    return labels.get(str(workflow_stage or "").strip().upper(), "Order")


def resolve_inventory_order_update_workflow_stage(
    payload: InventoryOrderUpdate,
    *,
    current_stage: str,
    domain: Any,
) -> str:
    expected_stage = normalize_inventory_order_workflow_stage(payload.expectedWorkflowStage)
    if expected_stage is None:
        raise HTTPException(status_code=400, detail="expectedWorkflowStage is required when updating an order.")

    if current_stage == "COMPLETE":
        raise HTTPException(status_code=409, detail="Completed orders are read-only.")

    if expected_stage != current_stage:
        workspace = inventory_order_workspace_label(current_stage)
        raise HTTPException(
            status_code=409,
            detail=f"Order is currently in {current_stage}. Reload the {workspace} page to continue.",
        )

    requested_stage = (
        normalize_inventory_order_workflow_stage(payload.workflowStage)
        if payload.workflowStage is not None
        else current_stage
    ) or default_inventory_order_workflow_stage(domain)
    if requested_stage == current_stage:
        return requested_stage

    if current_stage == "PLANNING" and requested_stage == "PURCHASING":
        return requested_stage
    if current_stage == "PURCHASING" and requested_stage == "RECEIVING":
        return requested_stage
    if current_stage == "RECEIVING" and requested_stage == "COMPLETE":
        raise HTTPException(
            status_code=400,
            detail="Use finalize-receiving to move a receiving order to COMPLETE.",
        )

    if current_stage == "RECEIVING":
        raise HTTPException(
            status_code=400,
            detail="Receiving orders can only stay in RECEIVING until finalized.",
        )

    next_stage = "PURCHASING" if current_stage == "PLANNING" else "RECEIVING"
    raise HTTPException(
        status_code=400,
        detail=f"Orders in {current_stage} can only stay in {current_stage} or move to {next_stage}.",
    )


def format_inventory_order_row(row: Any) -> dict[str, Any]:
    workflow_stage = normalize_inventory_order_workflow_stage(row["workflow_stage"])
    if workflow_stage is None:
        workflow_stage = default_inventory_order_workflow_stage(row["domain"])
    return {
        "id": int(row["id"]),
        "domain": row["domain"],
        "source_type": row["source_type"],
        "source_id": int(row["source_id"]) if row["source_id"] is not None else None,
        "name": row["name"],
        "status": row["status"],
        "workflow_stage": workflow_stage,
        "supplier_name": normalize_optional_text(row["supplier_name"]),
        "notes": normalize_optional_text(row["notes"]),
        "created_by_user_id": int(row["created_by_user_id"]) if row["created_by_user_id"] is not None else None,
        "created_by_username": normalize_optional_text(row["created_by_username"]),
        "ordered_by_user_id": int(row["ordered_by_user_id"]) if row["ordered_by_user_id"] is not None else None,
        "ordered_by_username": normalize_optional_text(row["ordered_by_username"]),
        "ordered_at": row["ordered_at"],
        "received_by_user_id": int(row["received_by_user_id"]) if row["received_by_user_id"] is not None else None,
        "received_by_username": normalize_optional_text(row["received_by_username"]),
        "received_at": row["received_at"],
        "put_away_by_user_id": int(row["put_away_by_user_id"]) if row["put_away_by_user_id"] is not None else None,
        "put_away_by_username": normalize_optional_text(row["put_away_by_username"]),
        "put_away_at": row["put_away_at"],
        "completed_by_user_id": int(row["completed_by_user_id"]) if row["completed_by_user_id"] is not None else None,
        "completed_by_username": normalize_optional_text(row["completed_by_username"]),
        "completed_at": row["completed_at"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def derive_inventory_order_item_status(
    *,
    required_quantity: float,
    ordered_quantity: float,
    received_quantity: float,
) -> str:
    required = max(0.0, float(required_quantity))
    ordered = max(0.0, float(ordered_quantity))
    received = max(0.0, float(received_quantity))
    target = max(required, ordered)

    if target <= 0 and received <= 0:
        return "OPEN"
    if target > 0 and received >= target:
        return "RECEIVED"
    if received > 0:
        return "PARTIAL"
    if ordered > 0:
        return "ORDERED"
    return "OPEN"


def derive_inventory_order_status(items: list[dict[str, Any]]) -> str:
    relevant = [
        item
        for item in items
        if max(
            float(item.get("required_quantity") or 0.0),
            float(item.get("ordered_quantity") or 0.0),
            float(item.get("received_quantity") or 0.0),
        ) > 0
    ]
    if not relevant:
        return "DRAFT"

    all_received = all(item.get("status") == "RECEIVED" for item in relevant)
    if all_received:
        return "RECEIVED"

    any_received = any(item.get("status") in {"RECEIVED", "PARTIAL"} for item in relevant)
    if any_received:
        return "PARTIAL"

    any_ordered = any(item.get("status") == "ORDERED" for item in relevant)
    if any_ordered:
        return "ORDERED"

    return "DRAFT"


def inventory_order_item_is_relevant(domain: str | None, item: dict[str, Any]) -> bool:
    normalized_domain = str(domain or "").strip().upper()
    if normalized_domain == "NON_FOOD":
        return max(
            float(item.get("to_order_quantity") or 0.0),
            float(item.get("ordered_quantity") or 0.0),
            float(item.get("received_quantity") or 0.0),
        ) > 0
    return max(
        float(item.get("required_quantity") or 0.0),
        float(item.get("ordered_quantity") or 0.0),
        float(item.get("received_quantity") or 0.0),
    ) > 0


def resolve_inventory_order_item_snapshot(
    conn: Any,
    *,
    domain: str,
    item_type: str,
    item_id: int,
    requested_unit: str | None,
    ingredient_inventory_by_key: dict[tuple[int, str], float] | None = None,
    preserve_requested_unit: bool = False,
) -> dict[str, Any]:
    normalized_domain = str(domain or "").strip().upper()
    normalized_item_type = str(item_type or "").strip().upper()
    normalized_requested_unit = normalize_unit(str(requested_unit or "").strip()) or None

    if normalized_domain == "FOOD":
        if normalized_item_type != "INGREDIENT":
            raise HTTPException(status_code=400, detail="Food orders currently support ingredient lines only.")
        row = conn.execute(
            """
            SELECT id, name, category, canonical_unit
            FROM ingredients
            WHERE id = ?
            """,
            (item_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=400, detail=f"Ingredient {item_id} does not exist.")

        canonical_unit = normalize_unit(str(row["canonical_unit"] or "").strip()) or "each"
        unit_snapshot = normalized_requested_unit or canonical_unit
        _sample_qty, sample_canonical_unit = quantity_to_canonical(1.0, unit_snapshot)
        if sample_canonical_unit != canonical_unit and not preserve_requested_unit:
            unit_snapshot = canonical_unit
            _sample_qty, sample_canonical_unit = quantity_to_canonical(1.0, unit_snapshot)

        current_quantity_snapshot = 0.0
        if sample_canonical_unit == canonical_unit and ingredient_inventory_by_key is not None:
            current_canonical_qty = float(ingredient_inventory_by_key.get((item_id, sample_canonical_unit), 0.0))
            current_quantity_snapshot = round(
                canonical_qty_to_unit(current_canonical_qty, sample_canonical_unit, unit_snapshot),
                3,
            )

        return {
            "item_name_snapshot": str(row["name"] or "").strip() or f"Ingredient #{item_id}",
            "category_snapshot": normalize_inventory_category(row["category"]),
            "unit_snapshot": unit_snapshot,
            "current_quantity_snapshot": current_quantity_snapshot,
            "live_quantity": current_quantity_snapshot,
            "live_unit": unit_snapshot,
            "live_location": None,
        }

    if normalized_domain == "NON_FOOD":
        if normalized_item_type != "STANDALONE_INVENTORY":
            raise HTTPException(status_code=400, detail="Non-food orders currently support standalone inventory lines only.")
        row = conn.execute(
            """
            SELECT id, item_name, category, quantity, unit, location, barcode, order_url
            FROM standalone_inventory
            WHERE id = ?
            """,
            (item_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=400, detail=f"Inventory item {item_id} does not exist.")

        base_unit = normalize_unit(str(row["unit"] or "").strip()) or "each"
        unit_snapshot = normalized_requested_unit or base_unit
        current_quantity_snapshot = rounded_quantity(row["quantity"])
        if unit_snapshot != base_unit:
            converted = convert_quantity_between_units(current_quantity_snapshot, base_unit, unit_snapshot)
            if converted is not None:
                current_quantity_snapshot = round(converted, 3)
            else:
                unit_snapshot = base_unit

        return {
            "item_name_snapshot": str(row["item_name"] or "").strip() or f"Inventory Item #{item_id}",
            "category_snapshot": normalize_inventory_category(row["category"]),
            "unit_snapshot": unit_snapshot,
            "current_quantity_snapshot": current_quantity_snapshot,
            "live_quantity": rounded_quantity(row["quantity"]),
            "live_unit": base_unit,
            "live_location": normalize_optional_text(row["location"]),
            "live_barcode": normalize_optional_text(row["barcode"]),
            "live_order_url": normalize_optional_text(row["order_url"]),
        }

    raise HTTPException(status_code=400, detail="Invalid inventory order domain.")


def load_inventory_order_items(conn: Any, order_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not order_ids:
        return {}

    placeholders = ", ".join("?" for _ in order_ids)
    rows = conn.execute(
        f"""
        SELECT
            ioi.id,
            ioi.order_id,
            ioi.item_type,
            ioi.item_id,
            ioi.item_name_snapshot,
            ioi.category_snapshot,
            ioi.unit_snapshot,
            ioi.current_quantity_snapshot,
            ioi.required_quantity,
            ioi.ordered_quantity,
            ioi.received_quantity,
            ioi.applied_quantity,
            ioi.purchase_unit,
            ioi.units_per_purchase,
            ioi.draft_purchase_unit,
            ioi.draft_units_per_purchase,
            ioi.draft_ordered_purchase_quantity,
            ioi.ordered_purchase_quantity,
            ioi.received_purchase_quantity,
            ioi.source_shopping_list_item_id,
            ioi.order_url_snapshot,
            ioi.order_url_override,
            ioi.notes,
            ioi.ordered_by_user_id,
            ioi.ordered_at,
            ioi.received_by_user_id,
            ioi.received_at,
            ioi.created_at,
            ioi.updated_at,
            ing.name AS ingredient_name,
            ing.category AS ingredient_category,
            ing.canonical_unit AS ingredient_canonical_unit,
            si.item_name AS standalone_item_name,
            si.category AS standalone_category,
            si.quantity AS standalone_quantity,
            si.unit AS standalone_unit,
            si.location AS standalone_location,
            si.barcode AS standalone_barcode,
            si.order_url AS standalone_order_url,
            sli.shopping_list_id AS source_shopping_list_id,
            sl.name AS source_shopping_list_name
        FROM inventory_order_items ioi
        LEFT JOIN ingredients ing
          ON ioi.item_type = 'INGREDIENT'
         AND ing.id = ioi.item_id
        LEFT JOIN standalone_inventory si
          ON ioi.item_type = 'STANDALONE_INVENTORY'
         AND si.id = ioi.item_id
        LEFT JOIN shopping_list_items sli
          ON sli.id = ioi.source_shopping_list_item_id
        LEFT JOIN shopping_lists sl
          ON sl.id = sli.shopping_list_id
        WHERE ioi.order_id IN ({placeholders})
        ORDER BY lower(ioi.item_name_snapshot), ioi.id
        """,
        tuple(order_ids),
    ).fetchall()

    ingredient_inventory_by_key = load_inventory_canonical_by_key(conn)
    grouped: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        item_type = str(row["item_type"] or "").strip().upper()
        item_id = int(row["item_id"])
        unit_snapshot = normalize_unit(str(row["unit_snapshot"] or "").strip()) or "each"
        current_quantity_snapshot = rounded_quantity(row["current_quantity_snapshot"])
        required_quantity = rounded_quantity(row["required_quantity"])
        ordered_quantity = rounded_quantity(row["ordered_quantity"])
        received_quantity = rounded_quantity(row["received_quantity"])
        applied_quantity = rounded_quantity(row["applied_quantity"])
        purchase_unit = normalize_purchase_unit(row["purchase_unit"])
        units_per_purchase = normalize_units_per_purchase(row["units_per_purchase"])
        draft_purchase_unit = normalize_purchase_unit(row["draft_purchase_unit"])
        draft_units_per_purchase = normalize_units_per_purchase(row["draft_units_per_purchase"])
        draft_ordered_purchase_quantity = rounded_quantity(row["draft_ordered_purchase_quantity"])
        ordered_purchase_quantity = rounded_quantity(row["ordered_purchase_quantity"])
        received_purchase_quantity = rounded_quantity(row["received_purchase_quantity"])
        if draft_purchase_unit == "unit":
            draft_units_per_purchase = 1.0
        if draft_ordered_purchase_quantity <= 0 and ordered_purchase_quantity > 0:
            draft_purchase_unit = purchase_unit
            draft_units_per_purchase = units_per_purchase
            draft_ordered_purchase_quantity = ordered_purchase_quantity
        if ordered_purchase_quantity <= 0 and ordered_quantity > 0:
            ordered_purchase_quantity = round(ordered_quantity / units_per_purchase, 3)
        if received_purchase_quantity <= 0 and received_quantity > 0:
            received_purchase_quantity = round(received_quantity / units_per_purchase, 3)
        draft_ordered_quantity = (
            draft_ordered_purchase_quantity
            if draft_purchase_unit == "unit"
            else round(draft_ordered_purchase_quantity * draft_units_per_purchase, 3)
        )

        live_quantity: float | None = None
        live_unit: str | None = None
        live_location: str | None = None
        inventory_barcode: str | None = None
        inventory_order_url: str | None = None
        if item_type == "INGREDIENT":
            canonical_unit = normalize_unit(str(row["ingredient_canonical_unit"] or "").strip()) or None
            if canonical_unit:
                current_canonical_qty = float(ingredient_inventory_by_key.get((item_id, canonical_unit), 0.0))
                converted_live_quantity = canonical_qty_to_unit_or_none(current_canonical_qty, canonical_unit, unit_snapshot)
                if converted_live_quantity is not None:
                    live_quantity = round(converted_live_quantity, 3)
                    live_unit = unit_snapshot
                else:
                    live_quantity = round(current_canonical_qty, 3)
                    live_unit = canonical_unit
        else:
            live_quantity = rounded_quantity(row["standalone_quantity"])
            live_unit = normalize_unit(str(row["standalone_unit"] or "").strip()) or unit_snapshot
            live_location = normalize_optional_text(row["standalone_location"])
            inventory_barcode = normalize_optional_text(row["standalone_barcode"])
            inventory_order_url = normalize_optional_text(row["standalone_order_url"])

        putaway_pending_quantity = round(max(0.0, received_quantity - applied_quantity), 3)
        to_order_quantity = max(0.0, round(required_quantity - current_quantity_snapshot, 3))
        if item_type == "STANDALONE_INVENTORY":
            status = derive_standalone_inventory_item_status(
                to_order_quantity=to_order_quantity,
                ordered_quantity=ordered_quantity,
                received_quantity=received_quantity,
            )
        else:
            status = derive_inventory_order_item_status(
                required_quantity=required_quantity,
                ordered_quantity=ordered_quantity,
                received_quantity=received_quantity,
            )
        order_url_snapshot = normalize_optional_text(row["order_url_snapshot"])
        order_url_override = normalize_optional_text(row["order_url_override"])
        effective_order_url = order_url_override or order_url_snapshot or inventory_order_url
        payload = {
            "id": int(row["id"]),
            "order_id": int(row["order_id"]),
            "item_type": item_type,
            "item_id": item_id,
            "item_name_snapshot": row["item_name_snapshot"],
            "category_snapshot": normalize_inventory_category(row["category_snapshot"]),
            "unit_snapshot": unit_snapshot,
            "current_quantity_snapshot": current_quantity_snapshot,
            "required_quantity": required_quantity,
            "to_order_quantity": to_order_quantity,
            "ordered_quantity": ordered_quantity,
            "received_quantity": received_quantity,
            "applied_quantity": applied_quantity,
            "applied_received_quantity": applied_quantity,
            "purchase_unit": purchase_unit,
            "units_per_purchase": units_per_purchase,
            "draft_purchase_unit": draft_purchase_unit,
            "draft_units_per_purchase": draft_units_per_purchase,
            "draft_ordered_purchase_quantity": draft_ordered_purchase_quantity,
            "draft_ordered_quantity": draft_ordered_quantity,
            "ordered_purchase_quantity": ordered_purchase_quantity,
            "received_purchase_quantity": received_purchase_quantity,
            "source_shopping_list_item_id": (
                int(row["source_shopping_list_item_id"])
                if row["source_shopping_list_item_id"] is not None
                else None
            ),
            "source_shopping_list_id": (
                int(row["source_shopping_list_id"])
                if row["source_shopping_list_id"] is not None
                else None
            ),
            "source_shopping_list_name": normalize_optional_text(row["source_shopping_list_name"]),
            "order_url_snapshot": order_url_snapshot,
            "order_url_override": order_url_override,
            "effective_order_url": effective_order_url,
            "notes": normalize_optional_text(row["notes"]),
            "ordered_by_user_id": int(row["ordered_by_user_id"]) if row["ordered_by_user_id"] is not None else None,
            "ordered_at": row["ordered_at"],
            "received_by_user_id": int(row["received_by_user_id"]) if row["received_by_user_id"] is not None else None,
            "received_at": row["received_at"],
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
            "live_quantity": live_quantity,
            "live_unit": live_unit,
            "live_location": live_location,
            "inventory_item_id": item_id if item_type == "STANDALONE_INVENTORY" else None,
            "inventory_item_name": row["standalone_item_name"] if item_type == "STANDALONE_INVENTORY" else None,
            "inventory_category": (
                normalize_inventory_category(row["standalone_category"])
                if item_type == "STANDALONE_INVENTORY"
                else None
            ),
            "inventory_quantity": live_quantity if item_type == "STANDALONE_INVENTORY" else None,
            "inventory_unit": live_unit if item_type == "STANDALONE_INVENTORY" else None,
            "inventory_location": live_location if item_type == "STANDALONE_INVENTORY" else None,
            "inventory_barcode": inventory_barcode if item_type == "STANDALONE_INVENTORY" else None,
            "inventory_order_url": inventory_order_url if item_type == "STANDALONE_INVENTORY" else None,
            "putaway_pending_quantity": putaway_pending_quantity,
            "status": status,
        }
        grouped.setdefault(int(row["order_id"]), []).append(payload)
    return grouped


def refresh_inventory_order_status(conn: Any, order_id: int, *, actor_user_id: int | None = None) -> str:
    items_map = load_inventory_order_items(conn, [order_id])
    items = items_map.get(order_id, [])
    status = derive_inventory_order_status(items)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    existing = conn.execute(
        """
        SELECT
            ordered_at,
            ordered_by_user_id,
            received_at,
            received_by_user_id,
            put_away_at,
            put_away_by_user_id
        FROM inventory_orders
        WHERE id = ? AND deleted_at IS NULL
        """,
        (order_id,),
    ).fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Inventory order not found")

    any_ordered = any(as_non_negative_quantity(item.get("ordered_quantity")) > 0 for item in items)
    any_received = any(as_non_negative_quantity(item.get("received_quantity")) > 0 for item in items)
    any_put_away = any(as_non_negative_quantity(item.get("applied_quantity")) > 0 for item in items)

    ordered_at = existing["ordered_at"]
    ordered_by_user_id = existing["ordered_by_user_id"]
    received_at = existing["received_at"]
    received_by_user_id = existing["received_by_user_id"]
    put_away_at = existing["put_away_at"]
    put_away_by_user_id = existing["put_away_by_user_id"]

    if not any_ordered:
        ordered_at = None
        ordered_by_user_id = None

    if any_ordered and not ordered_at:
        ordered_at = now
        ordered_by_user_id = actor_user_id

    if not any_received:
        received_at = None
        received_by_user_id = None

    if any_received and not received_at:
        received_at = now
        received_by_user_id = actor_user_id

    if not any_put_away:
        put_away_at = None
        put_away_by_user_id = None

    if any_put_away and not put_away_at:
        put_away_at = now
        put_away_by_user_id = actor_user_id

    conn.execute(
        """
        UPDATE inventory_orders
        SET status = ?,
            ordered_at = ?,
            ordered_by_user_id = ?,
            received_at = ?,
            received_by_user_id = ?,
            put_away_at = ?,
            put_away_by_user_id = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (
            status,
            ordered_at,
            ordered_by_user_id,
            received_at,
            received_by_user_id,
            put_away_at,
            put_away_by_user_id,
            order_id,
        ),
    )
    return status


def finalize_inventory_order_receiving(conn: Any, order_id: int, *, actor_user_id: int) -> dict[str, Any]:
    detail = load_inventory_order_detail(conn, order_id)
    current_stage = normalize_inventory_order_workflow_stage(detail.get("workflow_stage")) or default_inventory_order_workflow_stage(
        detail.get("domain")
    )
    if current_stage != "RECEIVING":
        raise HTTPException(status_code=400, detail="Only receiving-stage orders can be finalized here.")
    if detail.get("status") != "RECEIVED" or not detail.get("fully_put_away"):
        raise HTTPException(
            status_code=400,
            detail="Finalize is available only after the order is fully received and fully put away.",
        )
    if detail.get("received_by_user_id") is None:
        raise HTTPException(status_code=400, detail="Capture receipt ownership before finalizing this order.")
    if detail.get("put_away_by_user_id") is None:
        raise HTTPException(status_code=400, detail="Capture putaway ownership before finalizing this order.")
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    conn.execute(
        """
        UPDATE inventory_orders
        SET workflow_stage = 'COMPLETE',
            completed_by_user_id = ?,
            completed_at = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (actor_user_id, now, order_id),
    )
    return load_inventory_order_detail(conn, order_id)


def upsert_inventory_order_items(
    conn: Any,
    order_id: int,
    *,
    domain: str,
    items: list[InventoryOrderItemInput],
    actor_user_id: int | None,
    preserve_requested_unit: bool = False,
    preserve_empty_items: bool = False,
) -> None:
    existing_rows = conn.execute(
        """
        SELECT *
        FROM inventory_order_items
        WHERE order_id = ?
        """,
        (order_id,),
    ).fetchall()
    existing_by_key = {
        (
            str(row["item_type"] or "").strip().upper(),
            int(row["item_id"]),
            int(row["source_shopping_list_item_id"]) if row["source_shopping_list_item_id"] is not None else None,
        ): row
        for row in existing_rows
    }
    keep_keys: set[tuple[str, int, int | None]] = set()
    ingredient_inventory_by_key = load_inventory_canonical_by_key(conn) if str(domain or "").strip().upper() == "FOOD" else {}
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    for item in items:
        item_type = str(item.itemType or "").strip().upper()
        item_id = int(item.itemId)
        source_shopping_list_item_id = int(item.sourceShoppingListItemId) if item.sourceShoppingListItemId is not None else None
        lookup_key = (item_type, item_id, source_shopping_list_item_id)
        existing = existing_by_key.get(lookup_key)
        try:
            snapshot = resolve_inventory_order_item_snapshot(
                conn,
                domain=domain,
                item_type=item_type,
                item_id=item_id,
                requested_unit=item.unit,
                ingredient_inventory_by_key=ingredient_inventory_by_key,
                preserve_requested_unit=preserve_requested_unit,
            )
        except HTTPException:
            # Item was deleted upstream but this line already exists on the order.
            # Preserve the stored snapshot so the user can still save/advance the order.
            if existing is None:
                raise
            snapshot = {
                "item_name_snapshot": existing["item_name_snapshot"],
                "category_snapshot": existing["category_snapshot"],
                "unit_snapshot": existing["unit_snapshot"],
                "current_quantity_snapshot": rounded_quantity(existing["current_quantity_snapshot"]),
                "live_quantity": rounded_quantity(existing["current_quantity_snapshot"]),
                "live_unit": existing["unit_snapshot"],
                "live_location": None,
                "live_order_url": existing["order_url_snapshot"] if "order_url_snapshot" in existing.keys() else None,
            }

        required_quantity = rounded_quantity(item.requiredQuantity)
        ordered_quantity = rounded_quantity(item.orderedQuantity)
        received_quantity = rounded_quantity(item.receivedQuantity)
        applied_quantity = rounded_quantity(item.appliedQuantity)
        purchase_unit = normalize_purchase_unit(item.purchaseUnit)
        units_per_purchase = normalize_units_per_purchase(item.unitsPerPurchase)
        ordered_purchase_quantity = rounded_quantity(item.orderedPurchaseQuantity)
        received_purchase_quantity = rounded_quantity(item.receivedPurchaseQuantity)

        if ordered_purchase_quantity > 0:
            ordered_quantity = round(ordered_purchase_quantity * units_per_purchase, 3)
        elif ordered_quantity > 0:
            ordered_purchase_quantity = round(ordered_quantity / units_per_purchase, 3)

        if received_purchase_quantity > 0:
            received_quantity = round(received_purchase_quantity * units_per_purchase, 3)
        elif received_quantity > 0:
            received_purchase_quantity = round(received_quantity / units_per_purchase, 3)

        if received_quantity > ordered_quantity:
            ordered_quantity = received_quantity
        if received_purchase_quantity > ordered_purchase_quantity:
            ordered_purchase_quantity = received_purchase_quantity
        if ordered_quantity <= 0:
            received_quantity = 0.0
            applied_quantity = 0.0
            ordered_purchase_quantity = 0.0
            received_purchase_quantity = 0.0
        if applied_quantity > received_quantity:
            applied_quantity = received_quantity

        notes = normalize_optional_text(item.notes)
        order_url_override = normalize_optional_text(item.orderUrlOverride)
        if source_shopping_list_item_id is not None:
            if str(domain or "").strip().upper() != "FOOD" or item_type != "INGREDIENT":
                raise HTTPException(
                    status_code=400,
                    detail="sourceShoppingListItemId is only valid for food ingredient lines.",
                )
            source_item = conn.execute(
                """
                SELECT ingredient_id
                FROM shopping_list_items
                WHERE id = ?
                """,
                (source_shopping_list_item_id,),
            ).fetchone()
            if not source_item:
                raise HTTPException(status_code=400, detail="Source shopping list item does not exist.")
            if int(source_item["ingredient_id"]) != item_id:
                raise HTTPException(
                    status_code=400,
                    detail="Source shopping list item does not match the ingredient on this order line.",
                )
        if order_url_override and (str(domain or "").strip().upper() != "NON_FOOD" or item_type != "STANDALONE_INVENTORY"):
            raise HTTPException(
                status_code=400,
                detail="orderUrlOverride is only valid for non-food inventory lines.",
            )

        key = lookup_key
        item_fields = set(getattr(item, "model_fields_set", set()))
        draft_purchase_unit = (
            normalize_purchase_unit(item.draftPurchaseUnit)
            if "draftPurchaseUnit" in item_fields and item.draftPurchaseUnit is not None
            else normalize_purchase_unit(existing["draft_purchase_unit"])
            if existing and existing["draft_purchase_unit"] is not None
            else purchase_unit
        )
        draft_units_per_purchase = (
            normalize_units_per_purchase(item.draftUnitsPerPurchase)
            if "draftUnitsPerPurchase" in item_fields and item.draftUnitsPerPurchase is not None
            else normalize_units_per_purchase(existing["draft_units_per_purchase"])
            if existing and existing["draft_units_per_purchase"] is not None
            else units_per_purchase
        )
        if draft_purchase_unit == "unit":
            draft_units_per_purchase = 1.0
        draft_ordered_purchase_quantity = (
            rounded_quantity(item.draftOrderedPurchaseQuantity)
            if "draftOrderedPurchaseQuantity" in item_fields and item.draftOrderedPurchaseQuantity is not None
            else rounded_quantity(existing["draft_ordered_purchase_quantity"])
            if existing and existing["draft_ordered_purchase_quantity"] is not None
            else ordered_purchase_quantity
        )
        if ordered_quantity > 0:
            draft_purchase_unit = purchase_unit
            draft_units_per_purchase = units_per_purchase
            draft_ordered_purchase_quantity = ordered_purchase_quantity

        if (
            not preserve_empty_items
            and required_quantity <= 0
            and ordered_quantity <= 0
            and received_quantity <= 0
            and applied_quantity <= 0
            and draft_ordered_purchase_quantity <= 0
            and not order_url_override
            and not notes
            and source_shopping_list_item_id is None
        ):
            continue

        if key in keep_keys:
            raise HTTPException(status_code=400, detail="Duplicate order lines are not allowed.")
        old_ordered = as_non_negative_quantity(existing["ordered_quantity"]) if existing else 0.0
        old_received = as_non_negative_quantity(existing["received_quantity"]) if existing else 0.0
        ordered_at = existing["ordered_at"] if existing else None
        ordered_by_user_id = existing["ordered_by_user_id"] if existing else None
        received_at = existing["received_at"] if existing else None
        received_by_user_id = existing["received_by_user_id"] if existing else None

        if ordered_quantity > 0 and old_ordered <= 0:
            ordered_at = now
            ordered_by_user_id = actor_user_id
        if ordered_quantity <= 0:
            ordered_at = None
            ordered_by_user_id = None
            received_at = None
            received_by_user_id = None

        if received_quantity > 0 and old_received <= 0:
            received_at = now
            received_by_user_id = actor_user_id
        if received_quantity <= 0:
            received_at = None
            received_by_user_id = None

        if existing:
            conn.execute(
                """
                UPDATE inventory_order_items
                SET item_name_snapshot = ?,
                    category_snapshot = ?,
                    unit_snapshot = ?,
                    current_quantity_snapshot = ?,
                    required_quantity = ?,
                    ordered_quantity = ?,
                    received_quantity = ?,
                    applied_quantity = ?,
                    purchase_unit = ?,
                    units_per_purchase = ?,
                    draft_purchase_unit = ?,
                    draft_units_per_purchase = ?,
                    draft_ordered_purchase_quantity = ?,
                    ordered_purchase_quantity = ?,
                    received_purchase_quantity = ?,
                    source_shopping_list_item_id = ?,
                    order_url_snapshot = ?,
                    order_url_override = ?,
                    notes = ?,
                    ordered_by_user_id = ?,
                    ordered_at = ?,
                    received_by_user_id = ?,
                    received_at = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    snapshot["item_name_snapshot"],
                    snapshot["category_snapshot"],
                    snapshot["unit_snapshot"],
                    snapshot["current_quantity_snapshot"],
                    required_quantity,
                    ordered_quantity,
                    received_quantity,
                    applied_quantity,
                    purchase_unit,
                    units_per_purchase,
                    draft_purchase_unit,
                    draft_units_per_purchase,
                    draft_ordered_purchase_quantity,
                    ordered_purchase_quantity,
                    received_purchase_quantity,
                    source_shopping_list_item_id,
                    snapshot.get("live_order_url"),
                    order_url_override,
                    notes,
                    ordered_by_user_id,
                    ordered_at,
                    received_by_user_id,
                    received_at,
                    int(existing["id"]),
                ),
            )
        else:
            conn.execute(
                """
                INSERT INTO inventory_order_items(
                    order_id,
                    item_type,
                    item_id,
                    item_name_snapshot,
                    category_snapshot,
                    unit_snapshot,
                    current_quantity_snapshot,
                    required_quantity,
                    ordered_quantity,
                    received_quantity,
                    applied_quantity,
                    purchase_unit,
                    units_per_purchase,
                    draft_purchase_unit,
                    draft_units_per_purchase,
                    draft_ordered_purchase_quantity,
                    ordered_purchase_quantity,
                    received_purchase_quantity,
                    source_shopping_list_item_id,
                    order_url_snapshot,
                    order_url_override,
                    notes,
                    ordered_by_user_id,
                    ordered_at,
                    received_by_user_id,
                    received_at,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """,
                (
                    order_id,
                    item_type,
                    item_id,
                    snapshot["item_name_snapshot"],
                    snapshot["category_snapshot"],
                    snapshot["unit_snapshot"],
                    snapshot["current_quantity_snapshot"],
                    required_quantity,
                    ordered_quantity,
                    received_quantity,
                    applied_quantity,
                    purchase_unit,
                    units_per_purchase,
                    draft_purchase_unit,
                    draft_units_per_purchase,
                    draft_ordered_purchase_quantity,
                    ordered_purchase_quantity,
                    received_purchase_quantity,
                    source_shopping_list_item_id,
                    snapshot.get("live_order_url"),
                    order_url_override,
                    notes,
                    ordered_by_user_id,
                    ordered_at,
                    received_by_user_id,
                    received_at,
                ),
            )
        keep_keys.add(key)

    for row in existing_rows:
        key = (
            str(row["item_type"] or "").strip().upper(),
            int(row["item_id"]),
            int(row["source_shopping_list_item_id"]) if row["source_shopping_list_item_id"] is not None else None,
        )
        if key in keep_keys:
            continue
        conn.execute("DELETE FROM inventory_order_items WHERE id = ?", (int(row["id"]),))


def load_inventory_order_detail(conn: Any, order_id: int) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT
            io.id,
            io.domain,
            io.source_type,
            io.source_id,
            io.name,
            io.status,
            io.workflow_stage,
            io.supplier_name,
            io.notes,
            io.created_by_user_id,
            creator.username AS created_by_username,
            io.ordered_by_user_id,
            ordered_user.username AS ordered_by_username,
            io.ordered_at,
            io.received_by_user_id,
            received_user.username AS received_by_username,
            io.received_at,
            io.put_away_by_user_id,
            putaway_user.username AS put_away_by_username,
            io.put_away_at,
            io.completed_by_user_id,
            completed_user.username AS completed_by_username,
            io.completed_at,
            io.created_at,
            io.updated_at
        FROM inventory_orders io
        LEFT JOIN users creator
          ON creator.id = io.created_by_user_id
        LEFT JOIN users ordered_user
          ON ordered_user.id = io.ordered_by_user_id
        LEFT JOIN users received_user
          ON received_user.id = io.received_by_user_id
        LEFT JOIN users putaway_user
          ON putaway_user.id = io.put_away_by_user_id
        LEFT JOIN users completed_user
          ON completed_user.id = io.completed_by_user_id
        WHERE io.id = ?
          AND io.deleted_at IS NULL
        """,
        (order_id,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Inventory order not found")

    payload = format_inventory_order_row(row)
    items_map = load_inventory_order_items(conn, [order_id])
    items = items_map.get(order_id, [])
    payload["items"] = items

    domain = str(payload.get("domain") or "").strip().upper()
    relevant = [item for item in items if inventory_order_item_is_relevant(domain, item)]
    payload["item_count"] = len(items)
    payload["relevant_item_count"] = len(relevant)
    payload["to_order_count"] = sum(1 for item in relevant if item.get("status") == "NOT_ORDERED")
    payload["ordered_count"] = sum(1 for item in relevant if item.get("status") in {"ORDERED", "PARTIAL"})
    payload["received_count"] = sum(1 for item in relevant if item.get("status") == "RECEIVED")
    payload["putaway_pending_count"] = sum(
        1 for item in items if float(item.get("putaway_pending_quantity") or 0.0) > 0
    )
    payload["fully_put_away"] = all(
        float(item.get("putaway_pending_quantity") or 0.0) <= 0
        for item in items
    ) if items else False
    payload["total_required_quantity"] = round(sum(float(item.get("required_quantity") or 0.0) for item in items), 3)
    payload["total_to_order_quantity"] = round(sum(float(item.get("to_order_quantity") or 0.0) for item in items), 3)
    payload["total_ordered_quantity"] = round(sum(float(item.get("ordered_quantity") or 0.0) for item in items), 3)
    payload["total_received_quantity"] = round(sum(float(item.get("received_quantity") or 0.0) for item in items), 3)
    payload["total_applied_quantity"] = round(sum(float(item.get("applied_quantity") or 0.0) for item in items), 3)
    return payload


def record_inventory_order_item_event(
    conn: Any,
    *,
    order_id: int,
    order_item_id: int | None,
    event_type: str,
    quantity: float | None = None,
    purchase_quantity: float | None = None,
    location: str | None = None,
    notes: str | None = None,
    user_id: int | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO inventory_order_item_events(
            order_id,
            order_item_id,
            event_type,
            quantity,
            purchase_quantity,
            location,
            notes,
            user_id,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """,
        (
            order_id,
            order_item_id,
            event_type,
            rounded_quantity(quantity) if quantity is not None else None,
            rounded_quantity(purchase_quantity) if purchase_quantity is not None else None,
            normalize_optional_storage_grid_location(location),
            normalize_optional_text(notes),
            user_id,
        ),
    )


def record_inventory_movement(
    conn: Any,
    *,
    domain: str,
    item_type: str,
    item_id: int,
    order_id: int | None,
    order_item_id: int | None,
    movement_type: str,
    quantity_delta: float,
    unit: str | None,
    location: str | None = None,
    reason: str | None = None,
    user_id: int | None = None,
    actor_name: str | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO inventory_movements(
            domain,
            item_type,
            item_id,
            order_id,
            order_item_id,
            movement_type,
            quantity_delta,
            unit,
            location,
            reason,
            user_id,
            actor_name,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """,
        (
            normalize_inventory_order_domain(domain),
            str(item_type or "").strip().upper(),
            int(item_id),
            order_id,
            order_item_id,
            str(movement_type or "").strip().upper(),
            round(float(quantity_delta or 0.0), 3),
            normalize_unit(str(unit or "").strip()) or None,
            normalize_optional_storage_grid_location(location),
            normalize_optional_text(reason),
            user_id,
            normalize_optional_text(actor_name),
        ),
    )


def list_inventory_order_putaway_queue_rows(
    conn: Any,
    *,
    order_id: int | None = None,
    domain: str | None = None,
) -> list[dict[str, Any]]:
    filters = [
        "io.deleted_at IS NULL",
        "ioi.received_quantity > ioi.applied_quantity",
    ]
    params: list[Any] = []
    normalized_domain = normalize_inventory_order_domain(domain)
    if normalized_domain:
        filters.append("io.domain = ?")
        params.append(normalized_domain)
    if order_id is not None:
        filters.append("io.id = ?")
        params.append(int(order_id))

    where_sql = f"WHERE {' AND '.join(filters)}"
    rows = conn.execute(
        f"""
        SELECT
            io.id AS order_id,
            io.domain AS order_domain,
            io.name AS order_name,
            io.status AS order_status,
            io.updated_at AS order_updated_at,
            ioi.id AS order_item_id,
            ioi.item_type,
            ioi.item_id,
            ioi.source_shopping_list_item_id,
            ioi.item_name_snapshot,
            ioi.category_snapshot,
            ioi.unit_snapshot,
            ioi.purchase_unit,
            ioi.units_per_purchase,
            ioi.received_quantity,
            ioi.applied_quantity,
            ioi.notes,
            si.quantity AS standalone_inventory_quantity,
            si.unit AS standalone_inventory_unit,
            si.location AS standalone_inventory_location,
            si.barcode AS standalone_inventory_barcode
        FROM inventory_order_items ioi
        JOIN inventory_orders io
          ON io.id = ioi.order_id
        LEFT JOIN standalone_inventory si
          ON ioi.item_type = 'STANDALONE_INVENTORY'
         AND si.id = ioi.item_id
        {where_sql}
        ORDER BY io.updated_at DESC, io.id DESC, lower(ioi.item_name_snapshot), ioi.id
        """,
        tuple(params),
    ).fetchall()

    queue: list[dict[str, Any]] = []
    for row in rows:
        received_quantity = rounded_quantity(row["received_quantity"])
        applied_quantity = rounded_quantity(row["applied_quantity"])
        remaining_quantity = round(max(0.0, received_quantity - applied_quantity), 3)
        if remaining_quantity <= 0:
            continue
        unit = normalize_unit(str(row["unit_snapshot"] or "").strip()) or "each"
        queue.append(
            {
                "order_id": int(row["order_id"]),
                "order_domain": row["order_domain"],
                "order_name": row["order_name"],
                "order_status": row["order_status"],
                "order_updated_at": row["order_updated_at"],
                "order_item_id": int(row["order_item_id"]),
                "item_type": str(row["item_type"] or "").strip().upper(),
                "item_id": int(row["item_id"]),
                "source_shopping_list_item_id": (
                    int(row["source_shopping_list_item_id"])
                    if row["source_shopping_list_item_id"] is not None
                    else None
                ),
                "item_name": row["item_name_snapshot"],
                "category": normalize_inventory_category(row["category_snapshot"]),
                "unit": unit,
                "purchase_unit": normalize_purchase_unit(row["purchase_unit"]),
                "units_per_purchase": normalize_units_per_purchase(row["units_per_purchase"]),
                "received_quantity": received_quantity,
                "applied_quantity": applied_quantity,
                "remaining_putaway_quantity": remaining_quantity,
                "inventory_quantity": rounded_quantity(row["standalone_inventory_quantity"]),
                "inventory_unit": normalize_unit(str(row["standalone_inventory_unit"] or "").strip()) or None,
                "inventory_location": normalize_optional_text(row["standalone_inventory_location"]),
                "inventory_barcode": normalize_optional_text(row["standalone_inventory_barcode"]),
                "notes": normalize_optional_text(row["notes"]),
            }
        )
    return queue


def build_food_inventory_order_name(shopping_list_name: Any, vendor_name: Any) -> str:
    base_name = normalize_optional_text(shopping_list_name) or "Shopping List"
    supplier_label = normalize_optional_text(vendor_name) or "Pending Vendor"
    return f"{base_name} - {supplier_label}"


def build_food_inventory_order_item_input(
    row: Any,
    *,
    source_item_id: int | None = None,
) -> InventoryOrderItemInput | None:
    required_qty = rounded_quantity(row["required_qty"])
    required_unit = normalize_unit(str(row["required_unit"] or "").strip()) or None
    to_buy_qty = rounded_quantity(row["to_buy_qty"])
    to_buy_unit = normalize_unit(str(row["to_buy_unit"] or row["required_unit"] or "").strip()) or required_unit
    ordered_qty = rounded_quantity(row["ordered_qty"])
    ordered_unit = normalize_unit(str(row["ordered_unit"] or "").strip()) or None
    ordered = bool(row["ordered"])
    received = bool(row["received"])

    target_qty = to_buy_qty if to_buy_qty > 0 else required_qty
    target_unit = to_buy_unit or required_unit
    order_unit = ordered_unit or target_unit or required_unit or "each"

    required_quantity = 0.0
    if target_qty > 0 and target_unit and order_unit:
        converted_required = convert_quantity_between_units(target_qty, target_unit, order_unit)
        if converted_required is not None:
            required_quantity = round(converted_required, 3)

    if ordered_qty <= 0 and (ordered or received):
        fallback_ordered_qty, fallback_ordered_unit = preferred_ordered_quantity_and_unit(
            target_qty,
            target_unit,
            preferred_unit=order_unit,
        )
        if fallback_ordered_unit:
            order_unit = fallback_ordered_unit
        ordered_qty = rounded_quantity(fallback_ordered_qty)

    if ordered_qty <= 0 and required_quantity > 0 and (ordered or received):
        ordered_qty = required_quantity

    if ordered_qty <= 0 and not ordered and not received:
        return None

    received_qty = ordered_qty if received and ordered_qty > 0 else 0.0
    notes = normalize_optional_text(row["notes"])
    resolved_source_item_id = int(source_item_id) if source_item_id is not None else int(row["id"])

    return InventoryOrderItemInput(
        itemType="INGREDIENT",
        itemId=int(row["ingredient_id"]),
        requiredQuantity=required_quantity,
        orderedQuantity=ordered_qty,
        receivedQuantity=received_qty,
        appliedQuantity=0,
        unit=order_unit,
        purchaseUnit="unit",
        unitsPerPurchase=1,
        orderedPurchaseQuantity=ordered_qty,
        receivedPurchaseQuantity=received_qty,
        sourceShoppingListItemId=resolved_source_item_id,
        notes=notes,
    )


def archive_food_inventory_orders_for_shopping_list(conn: Any, shopping_list_id: int) -> int:
    rows = conn.execute(
        """
        SELECT id
        FROM inventory_orders
        WHERE domain = 'FOOD'
          AND source_type = 'SHOPPING_LIST'
          AND source_id = ?
          AND deleted_at IS NULL
        """,
        (shopping_list_id,),
    ).fetchall()
    order_ids = [int(row["id"]) for row in rows]
    for order_id in order_ids:
        conn.execute(
            """
            UPDATE inventory_orders
            SET deleted_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (order_id,),
        )
    return len(order_ids)


def sync_food_inventory_orders_for_shopping_list(
    conn: Any,
    shopping_list_id: int,
    *,
    actor_user_id: int | None = None,
) -> list[int]:
    shopping_list_row = conn.execute(
        """
        SELECT id, name
        FROM shopping_lists
        WHERE id = ?
        """,
        (shopping_list_id,),
    ).fetchone()
    if not shopping_list_row:
        return []

    shopping_rows = conn.execute(
        """
        SELECT
            sli.id,
            sli.ingredient_id,
            sli.required_qty,
            sli.required_unit,
            sli.to_buy_qty,
            sli.to_buy_unit,
            sli.ordered_qty,
            sli.ordered_unit,
            sli.vendor_id,
            v.name AS vendor_name,
            sli.ordered,
            sli.received,
            sli.notes
        FROM shopping_list_items sli
        LEFT JOIN vendors v ON v.id = sli.vendor_id
        WHERE sli.shopping_list_id = ?
        ORDER BY sli.id
        """,
        (shopping_list_id,),
    ).fetchall()
    vendor_allocations_by_item_id = load_shopping_list_item_vendor_allocations_by_item_id(
        conn,
        [int(row["id"]) for row in shopping_rows if row["id"] is not None],
    )

    grouped_rows: dict[int | None, dict[str, Any]] = {}
    for row in shopping_rows:
        item_id = int(row["id"])
        allocation_rows = vendor_allocations_by_item_id.get(item_id, [])
        effective_rows: list[dict[str, Any]] = []
        if allocation_rows:
            for allocation in allocation_rows:
                effective_rows.append(
                    {
                        "id": item_id,
                        "ingredient_id": int(row["ingredient_id"]),
                        "required_qty": round(float(allocation.get("allocated_qty") or 0.0), 4),
                        "required_unit": allocation.get("allocated_unit") or row["required_unit"],
                        "to_buy_qty": round(float(allocation.get("allocated_qty") or 0.0), 4),
                        "to_buy_unit": allocation.get("allocated_unit") or row["to_buy_unit"] or row["required_unit"],
                        "ordered_qty": (
                            round(float(allocation.get("allocated_qty") or 0.0), 4)
                            if bool(allocation.get("ordered")) or bool(allocation.get("received"))
                            else 0.0
                        ),
                        "ordered_unit": allocation.get("allocated_unit") or row["ordered_unit"] or row["to_buy_unit"] or row["required_unit"],
                        "vendor_id": allocation.get("vendor_id"),
                        "vendor_name": allocation.get("vendor_name"),
                        "ordered": bool(allocation.get("ordered")),
                        "received": bool(allocation.get("received")),
                        "notes": row["notes"],
                    }
                )
        else:
            effective_rows.append(dict(row))

        for effective_row in effective_rows:
            ordered_qty = rounded_quantity(effective_row["ordered_qty"])
            if not bool(effective_row["ordered"]) and not bool(effective_row["received"]) and ordered_qty <= 0:
                continue
            vendor_id = int(effective_row["vendor_id"]) if effective_row["vendor_id"] is not None else None
            group = grouped_rows.setdefault(
                vendor_id,
                {
                    "vendor_id": vendor_id,
                    "vendor_name": normalize_optional_text(effective_row["vendor_name"]),
                    "rows": [],
                },
            )
            if not group["vendor_name"]:
                group["vendor_name"] = normalize_optional_text(effective_row["vendor_name"])
            group["rows"].append(effective_row)

    existing_orders = conn.execute(
        """
        SELECT id, supplier_name
        FROM inventory_orders
        WHERE domain = 'FOOD'
          AND source_type = 'SHOPPING_LIST'
          AND source_id = ?
          AND deleted_at IS NULL
        ORDER BY id
        """,
        (shopping_list_id,),
    ).fetchall()
    existing_order_ids = [int(row["id"]) for row in existing_orders]
    existing_source_ids_by_order: dict[int, set[int]] = {order_id: set() for order_id in existing_order_ids}
    if existing_order_ids:
        placeholders = ", ".join("?" for _ in existing_order_ids)
        existing_line_rows = conn.execute(
            f"""
            SELECT order_id, source_shopping_list_item_id
            FROM inventory_order_items
            WHERE order_id IN ({placeholders})
            """,
            tuple(existing_order_ids),
        ).fetchall()
        for row in existing_line_rows:
            if row["source_shopping_list_item_id"] is None:
                continue
            existing_source_ids_by_order.setdefault(int(row["order_id"]), set()).add(int(row["source_shopping_list_item_id"]))

    synced_order_ids: list[int] = []
    used_order_ids: set[int] = set()
    for group in grouped_rows.values():
        payload_items = [
            payload
            for payload in (build_food_inventory_order_item_input(row, source_item_id=int(row["id"])) for row in group["rows"])
            if payload is not None
        ]
        if not payload_items:
            continue

        group_source_item_ids = {int(item.sourceShoppingListItemId) for item in payload_items if item.sourceShoppingListItemId}
        chosen_order: Any | None = None
        expected_supplier_name = normalize_optional_text(group["vendor_name"])
        if expected_supplier_name:
            for order_row in existing_orders:
                order_id = int(order_row["id"])
                if order_id in used_order_ids:
                    continue
                if normalize_optional_text(order_row["supplier_name"]) == expected_supplier_name:
                    chosen_order = order_row
                    break

        if chosen_order is None:
            best_overlap = 0
            for order_row in existing_orders:
                order_id = int(order_row["id"])
                if order_id in used_order_ids:
                    continue
                overlap = len(group_source_item_ids & existing_source_ids_by_order.get(order_id, set()))
                if overlap > best_overlap:
                    best_overlap = overlap
                    chosen_order = order_row

        order_name = build_food_inventory_order_name(shopping_list_row["name"], group["vendor_name"])
        supplier_name = normalize_optional_text(group["vendor_name"])
        if chosen_order is None:
            created = conn.execute(
                """
                INSERT INTO inventory_orders(
                    domain,
                    source_type,
                    source_id,
                    name,
                    status,
                    workflow_stage,
                    supplier_name,
                    created_by_user_id,
                    created_at,
                    updated_at
                )
                VALUES ('FOOD', 'SHOPPING_LIST', ?, ?, 'DRAFT', 'RECEIVING', ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id
                """,
                (shopping_list_id, order_name, supplier_name, actor_user_id),
            ).fetchone()
            order_id = int(created["id"])
        else:
            order_id = int(chosen_order["id"])
            conn.execute(
                """
                UPDATE inventory_orders
                SET name = ?,
                    workflow_stage = 'RECEIVING',
                    supplier_name = ?,
                    deleted_at = NULL,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (order_name, supplier_name, order_id),
            )

        upsert_inventory_order_items(
            conn,
            order_id,
            domain="FOOD",
            items=payload_items,
            actor_user_id=actor_user_id,
            preserve_requested_unit=True,
        )
        refresh_inventory_order_status(conn, order_id, actor_user_id=actor_user_id)
        synced_order_ids.append(order_id)
        used_order_ids.add(order_id)

    for order_row in existing_orders:
        order_id = int(order_row["id"])
        if order_id in used_order_ids:
            continue
        conn.execute(
            """
            UPDATE inventory_orders
            SET deleted_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (order_id,),
        )

    return synced_order_ids


def derive_standalone_inventory_item_status(
    *,
    to_order_quantity: float,
    ordered_quantity: float,
    received_quantity: float,
) -> str:
    to_order = max(0.0, float(to_order_quantity))
    ordered = max(0.0, float(ordered_quantity))
    received = max(0.0, float(received_quantity))

    if to_order <= 0:
        return "SUFFICIENT"
    if received >= to_order:
        return "RECEIVED"
    if received > 0:
        return "PARTIAL"
    if ordered > 0:
        return "ORDERED"
    return "NOT_ORDERED"


def tokenize_similarity_terms(value: Any) -> list[str]:
    text = normalize_optional_text(value)
    if not text:
        return []
    cleaned = re.sub(r"[^a-z0-9]+", " ", text.lower())
    stop_tokens = {"the", "and", "for", "with", "from", "item"}
    seen: set[str] = set()
    tokens: list[str] = []
    for token in cleaned.split():
        if len(token) < 2 or token in stop_tokens or token in seen:
            continue
        seen.add(token)
        tokens.append(token)
        if len(tokens) >= 12:
            break
    return tokens


def score_inventory_similarity(
    row: dict[str, Any],
    *,
    name_hint: str | None = None,
    category_hint: str | None = None,
    unit_hint: str | None = None,
) -> float:
    row_name = str(row.get("item_name") or "").strip().lower()
    row_category = str(row.get("category") or "").strip().lower()
    row_notes = str(row.get("notes") or "").strip().lower()
    row_location = str(row.get("location") or "").strip().lower()
    searchable_blob = f"{row_name} {row_category} {row_notes} {row_location}"
    score = 0.0

    normalized_name_hint = normalize_lookup_query(name_hint)
    if normalized_name_hint:
        name_phrase = normalized_name_hint.lower()
        name_tokens = tokenize_similarity_terms(name_phrase)
        if row_name and row_name == name_phrase:
            score += 12.0
        elif row_name and name_phrase in row_name:
            score += 8.0
        elif row_name and row_name in name_phrase and len(row_name) >= 5:
            score += 4.0
        for token in name_tokens:
            if token in row_name:
                score += 2.0
            elif token in searchable_blob:
                score += 0.8

    normalized_category_hint = normalize_optional_text(category_hint)
    if normalized_category_hint and row_category:
        category_phrase = normalized_category_hint.lower()
        if row_category == category_phrase:
            score += 2.5
        elif category_phrase in row_category or row_category in category_phrase:
            score += 1.4

    normalized_unit_hint = normalize_optional_text(unit_hint)
    if normalized_unit_hint:
        unit_phrase = normalized_unit_hint.lower()
        row_unit = str(row.get("unit") or "").strip().lower()
        if row_unit and row_unit == unit_phrase:
            score += 0.75

    return score


def search_similar_inventory_items(
    conn: Any,
    *,
    name_hint: str | None = None,
    category_hint: str | None = None,
    unit_hint: str | None = None,
    exclude_item_ids: set[int] | None = None,
    limit: int = 12,
) -> list[dict[str, Any]]:
    if not any([normalize_optional_text(name_hint), normalize_optional_text(category_hint), normalize_optional_text(unit_hint)]):
        return []
    excluded = exclude_item_ids or set()
    bounded_limit = max(1, min(int(limit), 80))
    rows = conn.execute("SELECT * FROM standalone_inventory ORDER BY lower(item_name)").fetchall()
    payload_rows = format_standalone_inventory_rows(conn, rows)

    scored_rows: list[dict[str, Any]] = []
    for payload in payload_rows:
        row_id = int(payload.get("id") or 0)
        if row_id and row_id in excluded:
            continue
        score = score_inventory_similarity(
            payload,
            name_hint=name_hint,
            category_hint=category_hint,
            unit_hint=unit_hint,
        )
        if score < 1.1:
            continue
        payload["similarity_score"] = round(score, 3)
        scored_rows.append(payload)

    scored_rows.sort(
        key=lambda item: (
            -float(item.get("similarity_score") or 0.0),
            str(item.get("item_name") or "").lower(),
            int(item.get("id") or 0),
        )
    )
    return scored_rows[:bounded_limit]


@app.get("/api/retreat-inventory/purchase-orders")
def list_purchase_orders(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
    status: str | None = Query(default=None),
) -> list[dict[str, Any]]:
    filters = ["deleted_at IS NULL"]
    params: list[Any] = []
    if status and status.strip():
        filters.append("status = ?")
        params.append(status.strip().upper())
    where_sql = f"WHERE {' AND '.join(filters)}"
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT id, supplier_name, status, expected_date, notes,
                   created_by_user_id, created_at, updated_at
            FROM retreat_inventory_purchase_orders
            {where_sql}
            ORDER BY id DESC
            """,
            tuple(params),
        ).fetchall()
        order_ids = [int(r["id"]) for r in rows]
        items_map = load_purchase_order_items(conn, order_ids)
    result = []
    for row in rows:
        order = format_purchase_order_row(row)
        order["items"] = items_map.get(int(row["id"]), [])
        result.append(order)
    return result


@app.post("/api/retreat-inventory/purchase-orders", status_code=201)
def create_purchase_order(
    payload: PurchaseOrderCreate,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        row = conn.execute(
            """
            INSERT INTO retreat_inventory_purchase_orders(
                supplier_name, status, expected_date, notes,
                created_by_user_id, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id, supplier_name, status, expected_date, notes,
                      created_by_user_id, created_at, updated_at
            """,
            (
                normalize_optional_text(payload.supplierName),
                payload.status,
                normalize_optional_text(payload.expectedDate),
                normalize_optional_text(payload.notes),
                user.id,
            ),
        ).fetchone()
        order_id = int(row["id"])

        for item in payload.items:
            conn.execute(
                """
                INSERT INTO retreat_inventory_purchase_order_items(
                    purchase_order_id, entity_type, entity_id,
                    ordered_quantity, received_quantity,
                    created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """,
                (order_id, item.entityType, item.entityId, item.orderedQuantity, item.receivedQuantity),
            )

        conn.commit()
        items_map = load_purchase_order_items(conn, [order_id])

    order = format_purchase_order_row(row)
    order["items"] = items_map.get(order_id, [])
    return order


@app.patch("/api/retreat-inventory/purchase-orders/{order_id}")
def update_purchase_order(
    order_id: int,
    payload: PurchaseOrderUpdate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    updates: list[str] = []
    params: list[Any] = []
    if payload.supplierName is not None:
        updates.append("supplier_name = ?")
        params.append(normalize_optional_text(payload.supplierName))
    if payload.status is not None:
        updates.append("status = ?")
        params.append(payload.status)
    if payload.expectedDate is not None:
        updates.append("expected_date = ?")
        params.append(normalize_optional_text(payload.expectedDate))
    if payload.notes is not None:
        updates.append("notes = ?")
        params.append(normalize_optional_text(payload.notes))

    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM retreat_inventory_purchase_orders WHERE id = ? AND deleted_at IS NULL",
            (order_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Purchase order not found")

        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(order_id)
            conn.execute(
                f"""
                UPDATE retreat_inventory_purchase_orders
                SET {', '.join(updates)}
                WHERE id = ?
                """,
                tuple(params),
            )

        if payload.items is not None:
            conn.execute(
                "DELETE FROM retreat_inventory_purchase_order_items WHERE purchase_order_id = ?",
                (order_id,),
            )
            for item in payload.items:
                conn.execute(
                    """
                    INSERT INTO retreat_inventory_purchase_order_items(
                        purchase_order_id, entity_type, entity_id,
                        ordered_quantity, received_quantity,
                        created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    """,
                    (order_id, item.entityType, item.entityId, item.orderedQuantity, item.receivedQuantity),
                )

        row = conn.execute(
            """
            SELECT id, supplier_name, status, expected_date, notes,
                   created_by_user_id, created_at, updated_at
            FROM retreat_inventory_purchase_orders
            WHERE id = ?
            """,
            (order_id,),
        ).fetchone()
        conn.commit()
        items_map = load_purchase_order_items(conn, [order_id])

    order = format_purchase_order_row(row)
    order["items"] = items_map.get(order_id, [])
    return order


@app.get("/api/inventory")
def list_inventory(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
    category: str | None = Query(default=None),
    search: str | None = Query(default=None),
) -> list[dict[str, Any]]:
    with get_connection() as conn:
        return query_standalone_inventory_rows(conn, category=category, search=search)


@app.get("/api/inventory/withdraw-config")
def get_inventory_withdraw_config() -> dict[str, Any]:
    with get_connection() as conn:
        state = resolve_inventory_withdraw_access_state(conn)
        conn.commit()
    return {
        "guest_access_enabled": bool(state.get("guest_access_enabled")),
        "access_code_required_for_guests": True,
    }


@app.post("/api/inventory/withdraw-search")
def search_inventory_withdraw_candidates(
    payload: InventoryWithdrawSearchPayload,
    request: Request,
) -> dict[str, Any]:
    with get_connection() as conn:
        actor_user = authorize_inventory_withdraw_access(
            conn,
            request,
            access_code=payload.accessCode,
        )
        items = query_standalone_inventory_withdraw_matches(conn, query=payload.query, limit=payload.limit)
        conn.commit()

    return {
        "query": normalize_required_text(payload.query, field_name="Query"),
        "items": items,
        "actor": {
            "mode": "user" if actor_user else "guest",
            "name": actor_user.username if actor_user else None,
            "role": actor_user.role if actor_user else None,
        },
    }


@app.post("/api/inventory/withdraw-complete")
def complete_inventory_withdrawal(
    payload: InventoryWithdrawCompletePayload,
    request: Request,
) -> dict[str, Any]:
    if not payload.items:
        raise HTTPException(status_code=400, detail="Add at least one item before completing the withdrawal.")

    requested_by_item_id: dict[int, float] = {}
    for item in payload.items:
        item_id = int(item.itemId)
        if item_id in requested_by_item_id:
            raise HTTPException(status_code=400, detail=f"Item {item_id} is duplicated in this withdrawal.")
        requested_by_item_id[item_id] = rounded_quantity(item.quantity)

    reason = normalize_optional_text(payload.reason)
    requested_item_ids = sorted(requested_by_item_id.keys())
    placeholders = ", ".join("?" for _ in requested_item_ids)

    with get_connection() as conn:
        actor_user, actor_name = load_inventory_withdraw_actor(
            conn,
            request,
            access_code=payload.accessCode,
            withdrawn_by=payload.withdrawnBy,
        )

        rows = conn.execute(
            f"""
            SELECT *
            FROM standalone_inventory
            WHERE id IN ({placeholders})
            """,
            tuple(requested_item_ids),
        ).fetchall()
        rows_by_id = {int(row["id"]): row for row in rows}
        missing_ids = [item_id for item_id in requested_item_ids if item_id not in rows_by_id]
        if missing_ids:
            raise HTTPException(status_code=404, detail=f"Inventory item(s) not found: {', '.join(str(item_id) for item_id in missing_ids)}")

        results: list[dict[str, Any]] = []
        total_quantity = 0.0
        for item_id in requested_item_ids:
            row = rows_by_id[item_id]
            quantity_before = rounded_quantity(row["quantity"])
            quantity_requested = requested_by_item_id[item_id]
            if quantity_requested <= 0:
                raise HTTPException(status_code=400, detail=f"Withdrawal quantity must be greater than zero for item {item_id}.")
            if quantity_requested > quantity_before:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f'Insufficient inventory for "{row["item_name"]}": '
                        f"{quantity_before:g} available, {quantity_requested:g} requested."
                    ),
                )

            quantity_after = round(quantity_before - quantity_requested, 3)
            conn.execute(
                """
                UPDATE standalone_inventory
                SET quantity = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (quantity_after, item_id),
            )
            record_inventory_movement(
                conn,
                domain="NON_FOOD",
                item_type="STANDALONE_INVENTORY",
                item_id=item_id,
                order_id=None,
                order_item_id=None,
                movement_type="OUT",
                quantity_delta=-quantity_requested,
                unit=normalize_optional_text(row["unit"]) or "each",
                location=normalize_optional_storage_grid_location(row["location"]) if normalize_optional_text(row["location"]) else None,
                reason=reason,
                user_id=actor_user.id if actor_user else None,
                actor_name=actor_name,
            )
            total_quantity += quantity_requested
            results.append(
                {
                    "item_id": item_id,
                    "item_name": row["item_name"],
                    "quantity_withdrawn": quantity_requested,
                    "quantity_before": quantity_before,
                    "quantity_after": quantity_after,
                    "unit": normalize_optional_text(row["unit"]) or "each",
                    "category": normalize_inventory_category(row["category"]),
                    "location": normalize_optional_text(row["location"]),
                }
            )

        conn.commit()

    return {
        "status": "ok",
        "withdrawn_by": actor_name,
        "line_count": len(results),
        "total_quantity": round(total_quantity, 3),
        "reason": reason,
        "items": results,
    }


@app.post("/api/inventory/order-draft-item", status_code=201)
def create_inventory_order_draft_item(
    payload: StandaloneInventoryOrderDraftItemCreate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    item_name = normalize_required_text(payload.itemName, field_name="Item name")
    category = normalize_inventory_category(payload.category)
    unit = normalize_optional_text(payload.unit) or "each"
    unit = normalize_required_text(unit, field_name="Unit")
    primary_barcode, barcodes = build_inventory_barcode_payload(payload.barcode)
    location = normalize_optional_storage_grid_location(payload.location)
    image_url = normalize_optional_text(payload.imageUrl)
    order_url = normalize_optional_text(payload.orderUrl)
    notes = normalize_optional_text(payload.notes)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    with get_connection() as conn:
        row = conn.execute(
            """
            INSERT INTO standalone_inventory(
                item_name,
                barcode,
                quantity,
                unit,
                category,
                location,
                image_url,
                notes,
                order_url,
                import_source,
                created_at,
                updated_at
            )
            VALUES (?, ?, 0, ?, ?, ?, ?, ?, ?, 'order-draft', ?, ?)
            RETURNING *
            """,
            (
                item_name,
                primary_barcode,
                unit,
                category,
                location,
                image_url,
                notes,
                order_url,
                now,
                now,
            ),
        ).fetchone()
        sync_standalone_inventory_item_barcodes(
            conn,
            item_id=int(row["id"]),
            primary_barcode=primary_barcode,
            barcodes=barcodes,
        )
        row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (int(row["id"]),)).fetchone()
        payload_row = format_standalone_inventory_rows(conn, [row])[0]
        conn.commit()

    return payload_row


@app.get("/api/orders")
def list_inventory_orders(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
    domain: str | None = Query(default=None),
    status: str | None = Query(default=None),
    source_type: str | None = Query(default=None, alias="sourceType"),
    workflow_stage: str | None = Query(default=None, alias="workflowStage"),
) -> list[dict[str, Any]]:
    filters = ["deleted_at IS NULL"]
    params: list[Any] = []

    normalized_domain = normalize_inventory_order_domain(domain)
    if normalized_domain:
        filters.append("domain = ?")
        params.append(normalized_domain)

    normalized_status = normalize_optional_text(status)
    if normalized_status:
        upper_status = normalized_status.upper()
        if upper_status not in {"DRAFT", "ORDERED", "PARTIAL", "RECEIVED"}:
            raise HTTPException(status_code=400, detail="Invalid order status filter.")
        filters.append("status = ?")
        params.append(upper_status)

    normalized_source_type = normalize_inventory_order_source_type(source_type)
    if normalized_source_type:
        filters.append("source_type = ?")
        params.append(normalized_source_type)

    normalized_workflow_stage = normalize_inventory_order_workflow_stage(workflow_stage)
    if normalized_workflow_stage:
        filters.append("workflow_stage = ?")
        params.append(normalized_workflow_stage)

    where_sql = f"WHERE {' AND '.join(filters)}"
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT
                io.id,
                io.domain,
                io.source_type,
                io.source_id,
                io.name,
                io.status,
                io.workflow_stage,
                io.supplier_name,
                io.notes,
                io.created_by_user_id,
                creator.username AS created_by_username,
                io.ordered_by_user_id,
                ordered_user.username AS ordered_by_username,
                io.ordered_at,
                io.received_by_user_id,
                received_user.username AS received_by_username,
                io.received_at,
                io.put_away_by_user_id,
                putaway_user.username AS put_away_by_username,
                io.put_away_at,
                io.completed_by_user_id,
                completed_user.username AS completed_by_username,
                io.completed_at,
                io.created_at,
                io.updated_at
            FROM inventory_orders io
            LEFT JOIN users creator
              ON creator.id = io.created_by_user_id
            LEFT JOIN users ordered_user
              ON ordered_user.id = io.ordered_by_user_id
            LEFT JOIN users received_user
              ON received_user.id = io.received_by_user_id
            LEFT JOIN users putaway_user
              ON putaway_user.id = io.put_away_by_user_id
            LEFT JOIN users completed_user
              ON completed_user.id = io.completed_by_user_id
            {where_sql}
            ORDER BY io.updated_at DESC, io.id DESC
            """,
            tuple(params),
        ).fetchall()
        order_ids = [int(row["id"]) for row in rows]
        items_map = load_inventory_order_items(conn, order_ids)

    payloads: list[dict[str, Any]] = []
    for row in rows:
        order_id = int(row["id"])
        payload = format_inventory_order_row(row)
        items = items_map.get(order_id, [])
        domain_name = str(payload.get("domain") or "").strip().upper()
        relevant = [item for item in items if inventory_order_item_is_relevant(domain_name, item)]
        payload["item_count"] = len(items)
        payload["relevant_item_count"] = len(relevant)
        payload["to_order_count"] = sum(1 for item in relevant if item.get("status") == "NOT_ORDERED")
        payload["ordered_count"] = sum(
            1 for item in relevant if item.get("status") in {"ORDERED", "PARTIAL"}
        )
        payload["received_count"] = sum(1 for item in relevant if item.get("status") == "RECEIVED")
        payload["putaway_pending_count"] = sum(
            1 for item in items if float(item.get("putaway_pending_quantity") or 0.0) > 0
        )
        payload["fully_put_away"] = all(
            float(item.get("putaway_pending_quantity") or 0.0) <= 0
            for item in items
        ) if items else False
        payload["total_required_quantity"] = round(sum(float(item.get("required_quantity") or 0.0) for item in items), 3)
        payload["total_to_order_quantity"] = round(sum(float(item.get("to_order_quantity") or 0.0) for item in items), 3)
        payload["total_ordered_quantity"] = round(sum(float(item.get("ordered_quantity") or 0.0) for item in items), 3)
        payload["total_received_quantity"] = round(sum(float(item.get("received_quantity") or 0.0) for item in items), 3)
        payload["total_applied_quantity"] = round(sum(float(item.get("applied_quantity") or 0.0) for item in items), 3)
        payloads.append(payload)
    return payloads


@app.post("/api/orders", status_code=201)
def create_inventory_order(
    payload: InventoryOrderCreate,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    domain = normalize_inventory_order_domain(payload.domain)
    if not domain:
        raise HTTPException(status_code=400, detail="Order domain is required.")

    source_type = normalize_inventory_order_source_type(payload.sourceType)
    source_id = int(payload.sourceId) if payload.sourceId is not None else None
    workflow_stage = normalize_inventory_order_workflow_stage(payload.workflowStage) or default_inventory_order_workflow_stage(domain)
    if source_id is not None and not source_type:
        raise HTTPException(status_code=400, detail="sourceId requires sourceType.")

    name = normalize_optional_text(payload.name) or ("Food Order" if domain == "FOOD" else "Non-Food Order")
    supplier_name = normalize_optional_text(payload.supplierName)
    notes = normalize_optional_text(payload.notes)

    with get_connection() as conn:
        row = conn.execute(
            """
            INSERT INTO inventory_orders(
                domain,
                source_type,
                source_id,
                name,
                status,
                workflow_stage,
                supplier_name,
                notes,
                created_by_user_id,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, 'DRAFT', ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id
            """,
            (domain, source_type, source_id, name, workflow_stage, supplier_name, notes, user.id),
        ).fetchone()
        order_id = int(row["id"])

        if payload.items:
            upsert_inventory_order_items(
                conn,
                order_id,
                domain=domain,
                items=payload.items,
                actor_user_id=user.id,
            )
        refresh_inventory_order_status(conn, order_id, actor_user_id=user.id)
        detail = load_inventory_order_detail(conn, order_id)
        conn.commit()
    return detail


@app.get("/api/orders/{order_id}")
def get_inventory_order(
    order_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        return load_inventory_order_detail(conn, order_id)


@app.patch("/api/orders/{order_id}")
def update_inventory_order(
    order_id: int,
    payload: InventoryOrderUpdate,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    fields = set(payload.model_fields_set)
    if not fields or fields <= {"expectedWorkflowStage"}:
        raise HTTPException(status_code=400, detail="Provide at least one field to update.")

    updates: list[str] = []
    params: list[Any] = []

    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id, domain, source_type, source_id, workflow_stage
            FROM inventory_orders
            WHERE id = ? AND deleted_at IS NULL
            """,
            (order_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Inventory order not found")

        current_stage = normalize_inventory_order_workflow_stage(existing["workflow_stage"]) or default_inventory_order_workflow_stage(
            existing["domain"]
        )
        next_workflow_stage = resolve_inventory_order_update_workflow_stage(
            payload,
            current_stage=current_stage,
            domain=existing["domain"],
        )

        if "name" in fields:
            clean_name = normalize_optional_text(payload.name)
            if not clean_name:
                raise HTTPException(status_code=400, detail="Order name cannot be empty.")
            updates.append("name = ?")
            params.append(clean_name)
        if "supplierName" in fields:
            updates.append("supplier_name = ?")
            params.append(normalize_optional_text(payload.supplierName))
        if "notes" in fields:
            updates.append("notes = ?")
            params.append(normalize_optional_text(payload.notes))
        if "workflowStage" in fields:
            updates.append("workflow_stage = ?")
            params.append(next_workflow_stage)

        if "sourceType" in fields or "sourceId" in fields:
            next_source_type = (
                normalize_inventory_order_source_type(payload.sourceType)
                if "sourceType" in fields
                else normalize_inventory_order_source_type(existing["source_type"])
            )
            next_source_id = (
                int(payload.sourceId)
                if "sourceId" in fields and payload.sourceId is not None
                else int(existing["source_id"])
                if existing["source_id"] is not None
                else None
            )
            if next_source_type is None:
                if "sourceId" in fields and next_source_id is not None:
                    raise HTTPException(status_code=400, detail="sourceId requires sourceType.")
                next_source_id = None

            updates.append("source_type = ?")
            params.append(next_source_type)
            updates.append("source_id = ?")
            params.append(next_source_id)

        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(order_id)
            conn.execute(
                f"""
                UPDATE inventory_orders
                SET {', '.join(updates)}
                WHERE id = ?
                """,
                tuple(params),
            )

        if payload.items is not None:
            upsert_inventory_order_items(
                conn,
                order_id,
                domain=str(existing["domain"] or "").strip().upper(),
                items=payload.items,
                actor_user_id=user.id,
                preserve_empty_items=bool(payload.preserveEmptyItems),
            )

        refresh_inventory_order_status(conn, order_id, actor_user_id=user.id)
        detail = load_inventory_order_detail(conn, order_id)
        conn.commit()
    return detail


@app.post("/api/orders/{order_id}/finalize-receiving")
def finalize_inventory_order_receiving_route(
    order_id: int,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        refresh_inventory_order_status(conn, order_id, actor_user_id=user.id)
        detail = finalize_inventory_order_receiving(conn, order_id, actor_user_id=user.id)
        conn.commit()
    return detail


@app.delete("/api/orders/{order_id}")
def delete_inventory_order(
    order_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id, name
            FROM inventory_orders
            WHERE id = ? AND deleted_at IS NULL
            """,
            (order_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Inventory order not found")

        conn.execute(
            """
            UPDATE inventory_orders
            SET deleted_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (order_id,),
        )
        conn.commit()

    return {"id": int(existing["id"]), "name": existing["name"], "status": "deleted"}


@app.get("/api/orders/putaway-queue")
def list_inventory_order_putaway_queue(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
    order_id: int | None = Query(default=None, ge=1),
    domain: str | None = Query(default=None),
) -> list[dict[str, Any]]:
    with get_connection() as conn:
        return list_inventory_order_putaway_queue_rows(conn, order_id=order_id, domain=domain)


@app.post("/api/orders/{order_id}/putaway")
def putaway_inventory_order_items(
    order_id: int,
    payload: InventoryOrderPutawayPayload,
    user: Annotated[AuthUser, Depends(require_roles(ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    if not payload.items:
        raise HTTPException(status_code=400, detail="Provide at least one order item to put away.")

    requested_by_item_id: dict[int, InventoryOrderPutawayItemInput] = {}
    for item in payload.items:
        order_item_id = int(item.orderItemId)
        if order_item_id in requested_by_item_id:
            raise HTTPException(status_code=400, detail=f"Order item {order_item_id} is duplicated in request.")
        requested_by_item_id[order_item_id] = item

    requested_ids = sorted(requested_by_item_id.keys())
    placeholders = ", ".join("?" for _ in requested_ids)
    applied_items: list[dict[str, Any]] = []
    applied_quantity_total = 0.0

    with get_connection() as conn:
        order_row = conn.execute(
            """
            SELECT id, domain
            FROM inventory_orders
            WHERE id = ? AND deleted_at IS NULL
            """,
            (order_id,),
        ).fetchone()
        if not order_row:
            raise HTTPException(status_code=404, detail="Inventory order not found")

        domain = str(order_row["domain"] or "").strip().upper()
        rows = conn.execute(
            f"""
            SELECT
                ioi.id,
                ioi.item_type,
                ioi.item_id,
                ioi.item_name_snapshot,
                ioi.unit_snapshot,
                ioi.purchase_unit,
                ioi.units_per_purchase,
                ioi.received_quantity,
                ioi.applied_quantity,
                ing.name AS ingredient_name,
                si.unit AS standalone_unit,
                si.location AS standalone_location
            FROM inventory_order_items ioi
            LEFT JOIN ingredients ing
              ON ioi.item_type = 'INGREDIENT'
             AND ing.id = ioi.item_id
            LEFT JOIN standalone_inventory si
              ON ioi.item_type = 'STANDALONE_INVENTORY'
             AND si.id = ioi.item_id
            WHERE ioi.order_id = ?
              AND ioi.id IN ({placeholders})
            ORDER BY ioi.id
            """,
            (order_id, *requested_ids),
        ).fetchall()

        rows_by_id = {int(row["id"]): row for row in rows}
        missing_ids = [order_item_id for order_item_id in requested_ids if order_item_id not in rows_by_id]
        if missing_ids:
            joined_missing = ", ".join(str(item_id) for item_id in missing_ids)
            raise HTTPException(status_code=400, detail=f"Order item(s) not found for order {order_id}: {joined_missing}.")

        for order_item_id in requested_ids:
            row = rows_by_id[order_item_id]
            request_item = requested_by_item_id[order_item_id]

            item_type = str(row["item_type"] or "").strip().upper()
            item_id = int(row["item_id"])
            item_name = str(row["item_name_snapshot"] or "").strip() or f"Item #{item_id}"
            unit_snapshot = normalize_unit(str(row["unit_snapshot"] or "").strip()) or "each"
            units_per_purchase = normalize_units_per_purchase(row["units_per_purchase"])

            received_quantity = rounded_quantity(row["received_quantity"])
            applied_quantity = rounded_quantity(row["applied_quantity"])
            remaining_quantity = round(max(0.0, received_quantity - applied_quantity), 3)
            if remaining_quantity <= 0:
                continue

            requested_quantity = (
                rounded_quantity(request_item.quantity) if request_item.quantity is not None else remaining_quantity
            )
            if requested_quantity <= 0:
                continue
            delta = round(min(remaining_quantity, requested_quantity), 3)
            if delta <= 0:
                continue

            location_override = normalize_optional_storage_grid_location(request_item.location)
            reason = normalize_optional_text(request_item.reason) or "Order putaway"
            purchase_quantity = round(delta / units_per_purchase, 3) if units_per_purchase > 0 else delta
            movement_quantity = delta
            movement_unit = unit_snapshot

            if domain == "FOOD":
                if item_type != "INGREDIENT":
                    raise HTTPException(status_code=400, detail="Food putaway supports ingredient lines only.")
                ingredient_name = str(row["ingredient_name"] or item_name).strip() or item_name
                require_canonical_conversion(
                    ingredient_name,
                    delta,
                    unit_snapshot,
                    context=f"Food putaway conversion failed for '{ingredient_name}'",
                )
                conn.execute(
                    """
                    INSERT INTO inventory_items(ingredient_id, quantity, unit, source, updated_at)
                    VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """,
                    (
                        item_id,
                        delta,
                        unit_snapshot,
                        ORDER_PUTAWAY_INVENTORY_SOURCE,
                    ),
                )
            elif domain == "NON_FOOD":
                if item_type != "STANDALONE_INVENTORY":
                    raise HTTPException(status_code=400, detail="Non-food putaway supports standalone inventory lines only.")
                base_unit = normalize_unit(str(row["standalone_unit"] or "").strip()) or unit_snapshot
                if base_unit != unit_snapshot:
                    converted_delta = convert_quantity_between_units(delta, unit_snapshot, base_unit)
                    if converted_delta is None:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Cannot put away '{item_name}' because {unit_snapshot} cannot be converted to {base_unit}.",
                        )
                    movement_quantity = round(converted_delta, 3)
                    movement_unit = base_unit
                if location_override is not None:
                    conn.execute(
                        """
                        UPDATE standalone_inventory
                        SET quantity = COALESCE(quantity, 0) + ?, location = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                        """,
                        (movement_quantity, location_override, item_id),
                    )
                else:
                    conn.execute(
                        """
                        UPDATE standalone_inventory
                        SET quantity = COALESCE(quantity, 0) + ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                        """,
                        (movement_quantity, item_id),
                    )
            else:
                raise HTTPException(status_code=400, detail="Invalid inventory order domain.")

            next_applied_quantity = round(min(received_quantity, applied_quantity + delta), 3)
            conn.execute(
                """
                UPDATE inventory_order_items
                SET applied_quantity = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (next_applied_quantity, order_item_id),
            )

            record_inventory_order_item_event(
                conn,
                order_id=order_id,
                order_item_id=order_item_id,
                event_type="PUT_AWAY",
                quantity=delta,
                purchase_quantity=purchase_quantity,
                location=location_override,
                notes=reason,
                user_id=user.id,
            )
            record_inventory_movement(
                conn,
                domain=domain,
                item_type=item_type,
                item_id=item_id,
                order_id=order_id,
                order_item_id=order_item_id,
                movement_type="IN",
                quantity_delta=movement_quantity,
                unit=movement_unit,
                location=location_override,
                reason=reason,
                user_id=user.id,
                actor_name=user.username,
            )

            applied_quantity_total += delta
            applied_items.append(
                {
                    "order_item_id": order_item_id,
                    "item_id": item_id,
                    "item_type": item_type,
                    "item_name": item_name,
                    "delta_applied": delta,
                    "movement_quantity": movement_quantity,
                    "movement_unit": movement_unit,
                }
            )

        refresh_inventory_order_status(conn, order_id, actor_user_id=user.id)
        detail = load_inventory_order_detail(conn, order_id)
        conn.commit()

    return {
        "order_id": order_id,
        "applied_item_count": len(applied_items),
        "applied_quantity_total": round(applied_quantity_total, 3),
        "items": applied_items,
        "order": detail,
    }


@app.get("/api/inventory/equivalent-search")
def search_inventory_equivalents(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
    query: str | None = Query(default=None),
    barcode: str | None = Query(default=None),
    limit: int = Query(default=20, ge=1, le=80),
) -> dict[str, Any]:
    normalized_query = normalize_lookup_query(query)
    normalized_barcode = normalize_lookup_barcode_candidate(barcode)
    if not normalized_query and not normalized_barcode:
        raise HTTPException(status_code=400, detail="Provide a search query or barcode.")

    bounded_limit = max(1, min(int(limit), 80))
    lookup_hit = lookup_inventory_product_metadata(normalized_barcode) if normalized_barcode else None

    inventory_matches: list[dict[str, Any]] = []
    seen_inventory_ids: set[int] = set()
    with get_connection() as conn:
        if normalized_barcode:
            exact = find_standalone_inventory_item_by_barcode(conn, normalized_barcode)
            if exact:
                payload = format_standalone_inventory_rows(conn, [exact])[0]
                payload["match_type"] = "exact_barcode"
                row_id = int(payload.get("id") or 0)
                if row_id:
                    seen_inventory_ids.add(row_id)
                inventory_matches.append(payload)

        if normalized_query:
            for payload in query_standalone_inventory_rows(conn, search=normalized_query, limit=bounded_limit):
                row_id = int(payload.get("id") or 0)
                if row_id and row_id in seen_inventory_ids:
                    continue
                payload["match_type"] = "text_search"
                if row_id:
                    seen_inventory_ids.add(row_id)
                inventory_matches.append(payload)
                if len(inventory_matches) >= bounded_limit:
                    break

        leading_inventory = inventory_matches[0] if inventory_matches else None
        similar_name_hint = (
            normalized_query
            or normalize_optional_text((lookup_hit or {}).get("name"))
            or normalize_optional_text((leading_inventory or {}).get("item_name"))
        )
        similar_category_hint = (
            normalize_optional_text((lookup_hit or {}).get("category"))
            or normalize_optional_text((leading_inventory or {}).get("category"))
        )
        similar_unit_hint = (
            normalize_optional_text((lookup_hit or {}).get("unit"))
            or normalize_optional_text((leading_inventory or {}).get("unit"))
        )
        similar_matches = search_similar_inventory_items(
            conn,
            name_hint=similar_name_hint,
            category_hint=similar_category_hint,
            unit_hint=similar_unit_hint,
            exclude_item_ids=seen_inventory_ids,
            limit=bounded_limit,
        )
        for payload in similar_matches:
            row_id = int(payload.get("id") or 0)
            if row_id and row_id in seen_inventory_ids:
                continue
            payload["match_type"] = "similar_lookup" if normalized_barcode else "similar_text"
            if row_id:
                seen_inventory_ids.add(row_id)
            inventory_matches.append(payload)
            if len(inventory_matches) >= bounded_limit:
                break

    raw_industry_hits: list[dict[str, Any]] = []
    if normalized_barcode and lookup_hit:
        hit = dict(lookup_hit)
        hit["barcode"] = normalized_barcode
        hit["match_type"] = "barcode_lookup"
        raw_industry_hits.append(hit)
    if normalized_query:
        try:
            industry_hits = search_inventory_product_metadata(normalized_query, limit=bounded_limit)
        except Exception:
            industry_hits = []
        for hit in industry_hits:
            payload = dict(hit)
            payload["match_type"] = "text_search"
            raw_industry_hits.append(payload)

    deduped_industry: list[dict[str, Any]] = []
    industry_index: dict[tuple[str, str], int] = {}
    for hit in raw_industry_hits:
        barcode_key = normalize_lookup_barcode_candidate(hit.get("barcode")) or ""
        name_value = normalize_optional_text(hit.get("name")) or ""
        key = (barcode_key, name_value.lower())
        if not key[0] and not key[1]:
            continue
        existing_idx = industry_index.get(key)
        if existing_idx is None:
            industry_index[key] = len(deduped_industry)
            deduped_industry.append(
                {
                    "source": normalize_optional_text(hit.get("source")),
                    "barcode": barcode_key or None,
                    "name": name_value or None,
                    "category": normalize_inventory_category(hit.get("category")),
                    "unit": normalize_optional_text(hit.get("unit")),
                    "image_url": normalize_optional_text(hit.get("image_url")),
                    "match_type": normalize_optional_text(hit.get("match_type")) or "text_search",
                }
            )
            continue
        existing = deduped_industry[existing_idx]
        existing["source"] = merge_lookup_source_names(existing.get("source"), hit.get("source"))
        for field in ("barcode", "name", "category", "unit", "image_url"):
            if existing.get(field):
                continue
            if field == "barcode":
                existing[field] = normalize_lookup_barcode_candidate(hit.get(field))
            elif field == "category":
                existing[field] = normalize_inventory_category(hit.get(field))
            else:
                existing[field] = normalize_optional_text(hit.get(field))

    return {
        "query": normalized_query,
        "barcode": normalized_barcode,
        "inventory_matches": inventory_matches[:bounded_limit],
        "industry_matches": deduped_industry[:bounded_limit],
    }


@app.get("/api/inventory/categories")
def list_inventory_categories(
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> list[str]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT DISTINCT category FROM standalone_inventory WHERE category IS NOT NULL AND category != ''"
        ).fetchall()
    seen: set[str] = set()
    categories: list[str] = []
    for row in rows:
        category = normalize_inventory_category(row["category"])
        if not category:
            continue
        key = category.lower()
        if key in seen:
            continue
        seen.add(key)
        categories.append(category)
    categories.sort(key=lambda value: value.lower())
    return categories


@app.get("/api/inventory/barcode-lookup/{barcode}")
def lookup_inventory_barcode(
    barcode: str,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    normalized_barcode = normalize_inventory_barcode(barcode)
    with get_connection() as conn:
        existing = find_standalone_inventory_item_by_barcode(conn, normalized_barcode)
        existing_item = format_standalone_inventory_rows(conn, [existing])[0] if existing else None
    lookup_hit = lookup_inventory_product_metadata(normalized_barcode)

    similar_matches: list[dict[str, Any]] = []
    with get_connection() as conn:
        excluded_ids = {int(existing_item["id"])} if existing_item and existing_item.get("id") else set()
        similar_matches = search_similar_inventory_items(
            conn,
            name_hint=(existing_item or {}).get("item_name") or (lookup_hit or {}).get("name"),
            category_hint=(existing_item or {}).get("category") or (lookup_hit or {}).get("category"),
            unit_hint=(existing_item or {}).get("unit") or (lookup_hit or {}).get("unit"),
            exclude_item_ids=excluded_ids,
            limit=12,
        )
        for payload in similar_matches:
            payload["match_type"] = "similar_lookup"
    return {
        "barcode": normalized_barcode,
        "existing_item": existing_item,
        "similar_matches": similar_matches,
        "lookup": lookup_hit,
    }


@app.post("/api/inventory/barcode-bind")
def bind_inventory_barcode(
    payload: StandaloneInventoryBarcodeBindPayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    item_id = int(payload.item_id)
    barcode = normalize_inventory_barcode(payload.barcode)

    with get_connection() as conn:
        item = conn.execute(
            "SELECT * FROM standalone_inventory WHERE id = ?",
            (item_id,),
        ).fetchone()
        if not item:
            raise HTTPException(status_code=404, detail="Inventory item not found")

        existing_barcodes = load_standalone_inventory_item_barcodes(
            conn,
            item_id,
            fallback_primary_barcode=item["barcode"],
        )
        primary_barcode = (
            normalize_lookup_barcode_candidate(item["barcode"])
            or (existing_barcodes[0] if existing_barcodes else None)
            or barcode
        )
        next_barcodes = list(existing_barcodes)
        if barcode not in next_barcodes:
            next_barcodes.append(barcode)
        sync_standalone_inventory_item_barcodes(
            conn,
            item_id=item_id,
            primary_barcode=primary_barcode,
            barcodes=next_barcodes,
        )
        row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        item_payload = format_standalone_inventory_rows(conn, [row])[0]
        conn.commit()

    return {
        "status": "already_bound" if barcode in existing_barcodes else "bound",
        "barcode": barcode,
        "item": item_payload,
    }


@app.post("/api/inventory/merge")
def merge_inventory_items(
    payload: StandaloneInventoryMergePayload,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    source_item_id = int(payload.source_item_id)
    target_item_id = int(payload.target_item_id)
    if source_item_id == target_item_id:
        raise HTTPException(status_code=400, detail="Source and target inventory items must be different.")

    impacted_order_ids: set[int] = set()

    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM standalone_inventory
            WHERE id IN (?, ?)
            """,
            (source_item_id, target_item_id),
        ).fetchall()
        rows_by_id = {int(row["id"]): row for row in rows}
        source = rows_by_id.get(source_item_id)
        target = rows_by_id.get(target_item_id)
        if not source or not target:
            raise HTTPException(status_code=404, detail="One or both inventory items were not found.")
        ensure_standalone_inventory_merge_safe(source, target)

        source_barcodes = load_standalone_inventory_item_barcodes(
            conn,
            source_item_id,
            fallback_primary_barcode=source["barcode"],
        )
        target_barcodes = load_standalone_inventory_item_barcodes(
            conn,
            target_item_id,
            fallback_primary_barcode=target["barcode"],
        )
        merged_primary = (
            normalize_lookup_barcode_candidate(target["barcode"])
            or normalize_lookup_barcode_candidate(source["barcode"])
            or (target_barcodes[0] if target_barcodes else None)
            or (source_barcodes[0] if source_barcodes else None)
        )
        merged_barcodes: list[str] = []
        for barcode in ([merged_primary] if merged_primary else []) + target_barcodes + source_barcodes:
            normalized = normalize_lookup_barcode_candidate(barcode)
            if normalized and normalized not in merged_barcodes:
                merged_barcodes.append(normalized)

        source_order_rows = conn.execute(
            """
            SELECT ioi.*
            FROM inventory_order_items ioi
            JOIN inventory_orders io
              ON io.id = ioi.order_id
            WHERE io.domain = 'NON_FOOD'
              AND io.deleted_at IS NULL
              AND ioi.item_type = 'STANDALONE_INVENTORY'
              AND ioi.item_id = ?
            ORDER BY ioi.order_id, ioi.id
            """,
            (source_item_id,),
        ).fetchall()

        for source_order_row in source_order_rows:
            source_order_item_id = int(source_order_row["id"])
            order_id = int(source_order_row["order_id"])
            impacted_order_ids.add(order_id)

            target_order_row = conn.execute(
                """
                SELECT *
                FROM inventory_order_items
                WHERE order_id = ?
                  AND item_type = 'STANDALONE_INVENTORY'
                  AND item_id = ?
                LIMIT 1
                """,
                (order_id, target_item_id),
            ).fetchone()

            if not target_order_row:
                conn.execute(
                    """
                    UPDATE inventory_order_items
                    SET item_id = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (target_item_id, source_order_item_id),
                )
                continue

            ensure_standalone_inventory_order_merge_safe(source_order_row, target_order_row)
            target_order_item_id = int(target_order_row["id"])
            merged_purchase_unit = normalize_purchase_unit(
                normalize_optional_text(target_order_row["purchase_unit"])
                or normalize_optional_text(source_order_row["purchase_unit"])
            )
            merged_units_per_purchase = normalize_units_per_purchase(
                target_order_row["units_per_purchase"]
                if float(target_order_row["units_per_purchase"] or 0) > 0
                else source_order_row["units_per_purchase"]
            )
            merged_draft_purchase_unit = normalize_purchase_unit(
                normalize_optional_text(target_order_row["draft_purchase_unit"])
                or normalize_optional_text(source_order_row["draft_purchase_unit"])
                or merged_purchase_unit
            )
            merged_draft_units_per_purchase = normalize_units_per_purchase(
                target_order_row["draft_units_per_purchase"]
                if float(target_order_row["draft_units_per_purchase"] or 0) > 0
                else source_order_row["draft_units_per_purchase"]
            )
            merged_required_quantity = round(
                rounded_quantity(target_order_row["required_quantity"])
                + rounded_quantity(source_order_row["required_quantity"]),
                3,
            )
            merged_ordered_quantity = round(
                rounded_quantity(target_order_row["ordered_quantity"])
                + rounded_quantity(source_order_row["ordered_quantity"]),
                3,
            )
            merged_received_quantity = round(
                rounded_quantity(target_order_row["received_quantity"])
                + rounded_quantity(source_order_row["received_quantity"]),
                3,
            )
            merged_ordered_purchase_quantity = round(
                rounded_quantity(target_order_row["ordered_purchase_quantity"])
                + rounded_quantity(source_order_row["ordered_purchase_quantity"]),
                3,
            )
            merged_received_purchase_quantity = round(
                rounded_quantity(target_order_row["received_purchase_quantity"])
                + rounded_quantity(source_order_row["received_purchase_quantity"]),
                3,
            )
            merged_draft_ordered_purchase_quantity = round(
                rounded_quantity(target_order_row["draft_ordered_purchase_quantity"])
                + rounded_quantity(source_order_row["draft_ordered_purchase_quantity"]),
                3,
            )
            merged_applied_quantity = round(
                rounded_quantity(target_order_row["applied_quantity"])
                + rounded_quantity(source_order_row["applied_quantity"]),
                3,
            )
            if merged_draft_purchase_unit == "unit":
                merged_draft_units_per_purchase = 1.0
            if merged_ordered_purchase_quantity > 0:
                merged_draft_purchase_unit = merged_purchase_unit
                merged_draft_units_per_purchase = merged_units_per_purchase
                merged_draft_ordered_purchase_quantity = merged_ordered_purchase_quantity
            conn.execute(
                """
                UPDATE inventory_order_items
                SET item_name_snapshot = ?,
                    category_snapshot = ?,
                    unit_snapshot = ?,
                    current_quantity_snapshot = ?,
                    required_quantity = ?,
                    ordered_quantity = ?,
                    received_quantity = ?,
                    purchase_unit = ?,
                    units_per_purchase = ?,
                    draft_purchase_unit = ?,
                    draft_units_per_purchase = ?,
                    draft_ordered_purchase_quantity = ?,
                    ordered_purchase_quantity = ?,
                    received_purchase_quantity = ?,
                    applied_quantity = ?,
                    order_url_snapshot = ?,
                    order_url_override = ?,
                    notes = ?,
                    ordered_by_user_id = ?,
                    ordered_at = ?,
                    received_by_user_id = ?,
                    received_at = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    normalize_optional_text(target_order_row["item_name_snapshot"])
                    or normalize_optional_text(target["item_name"])
                    or source_order_row["item_name_snapshot"],
                    normalize_inventory_category(target_order_row["category_snapshot"])
                    or normalize_inventory_category(target["category"])
                    or normalize_inventory_category(source_order_row["category_snapshot"]),
                    normalize_optional_text(target_order_row["unit_snapshot"])
                    or normalize_optional_text(target["unit"])
                    or normalize_optional_text(source_order_row["unit_snapshot"]),
                    round(
                        rounded_quantity(target_order_row["current_quantity_snapshot"])
                        + rounded_quantity(source_order_row["current_quantity_snapshot"]),
                        3,
                    ),
                    merged_required_quantity,
                    merged_ordered_quantity,
                    merged_received_quantity,
                    merged_purchase_unit,
                    merged_units_per_purchase,
                    merged_draft_purchase_unit,
                    merged_draft_units_per_purchase,
                    merged_draft_ordered_purchase_quantity,
                    merged_ordered_purchase_quantity,
                    merged_received_purchase_quantity,
                    merged_applied_quantity,
                    normalize_optional_text(target_order_row["order_url_snapshot"])
                    or normalize_optional_text(source_order_row["order_url_snapshot"]),
                    normalize_optional_text(target_order_row["order_url_override"])
                    or normalize_optional_text(source_order_row["order_url_override"]),
                    merge_distinct_pipe_text(target_order_row["notes"], source_order_row["notes"]),
                    target_order_row["ordered_by_user_id"] or source_order_row["ordered_by_user_id"],
                    target_order_row["ordered_at"] or source_order_row["ordered_at"],
                    target_order_row["received_by_user_id"] or source_order_row["received_by_user_id"],
                    target_order_row["received_at"] or source_order_row["received_at"],
                    target_order_item_id,
                ),
            )
            conn.execute(
                """
                UPDATE inventory_order_item_events
                SET order_item_id = ?
                WHERE order_item_id = ?
                """,
                (target_order_item_id, source_order_item_id),
            )
            conn.execute(
                """
                UPDATE inventory_movements
                SET order_item_id = ?,
                    order_id = ?
                WHERE order_item_id = ?
                """,
                (target_order_item_id, order_id, source_order_item_id),
            )
            conn.execute(
                "DELETE FROM inventory_order_items WHERE id = ?",
                (source_order_item_id,),
            )

        conn.execute(
            """
            UPDATE inventory_movements
            SET item_id = ?
            WHERE domain = 'NON_FOOD'
              AND item_type = 'STANDALONE_INVENTORY'
              AND item_id = ?
            """,
            (target_item_id, source_item_id),
        )
        if table_exists(conn, "standalone_inventory_barcodes"):
            conn.execute(
                """
                UPDATE standalone_inventory_barcodes
                SET inventory_item_id = ?, updated_at = CURRENT_TIMESTAMP
                WHERE inventory_item_id = ?
                """,
                (target_item_id, source_item_id),
            )
        conn.execute(
            """
            UPDATE standalone_inventory
            SET barcode = NULL, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (source_item_id,),
        )

        merged_quantity = round(
            as_non_negative_quantity(target["quantity"]) + as_non_negative_quantity(source["quantity"]),
            3,
        )
        conn.execute(
            """
            UPDATE standalone_inventory
            SET item_name = ?,
                barcode = ?,
                quantity = ?,
                unit = ?,
                category = ?,
                location = ?,
                image_url = ?,
                notes = ?,
                order_url = ?,
                import_source = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                normalize_optional_text(target["item_name"]) or normalize_optional_text(source["item_name"]),
                merged_primary,
                merged_quantity,
                normalize_optional_text(target["unit"]) or normalize_optional_text(source["unit"]),
                normalize_inventory_category(target["category"]) or normalize_inventory_category(source["category"]),
                normalize_optional_text(target["location"]) or normalize_optional_text(source["location"]),
                normalize_optional_text(target["image_url"]) or normalize_optional_text(source["image_url"]),
                merge_distinct_pipe_text(target["notes"], source["notes"]),
                normalize_optional_text(target["order_url"]) or normalize_optional_text(source["order_url"]),
                normalize_optional_text(target["import_source"]) or normalize_optional_text(source["import_source"]),
                target_item_id,
            ),
        )
        sync_standalone_inventory_item_barcodes(
            conn,
            item_id=target_item_id,
            primary_barcode=merged_primary,
            barcodes=merged_barcodes,
        )
        conn.execute("DELETE FROM standalone_inventory WHERE id = ?", (source_item_id,))

        for order_id in impacted_order_ids:
            refresh_inventory_order_status(conn, order_id)

        merged_row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (target_item_id,)).fetchone()
        merged_item = format_standalone_inventory_rows(conn, [merged_row])[0]
        conn.commit()

    return {
        "status": "merged",
        "source_item_id": source_item_id,
        "target_item_id": target_item_id,
        "impacted_order_count": len(impacted_order_ids),
        "item": merged_item,
    }


@app.post("/api/inventory", status_code=201)
def create_inventory_item(
    payload: StandaloneInventoryCreate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    item_name = normalize_required_text(payload.item_name, field_name="Item name")
    primary_barcode, barcodes = build_inventory_barcode_payload(
        payload.barcode,
        requested_barcodes=payload.barcodes,
    )
    unit = normalize_required_text(payload.unit, field_name="Unit")
    location = normalize_storage_grid_location(payload.location)
    image_url = normalize_required_text(payload.image_url, field_name="Product image")
    category = normalize_inventory_category(payload.category)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    with get_connection() as conn:
        row = conn.execute(
            """
            INSERT INTO standalone_inventory(
                item_name,
                barcode,
                quantity,
                unit,
                category,
                location,
                image_url,
                notes,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            RETURNING *
            """,
            (
                item_name,
                primary_barcode,
                payload.quantity,
                unit,
                category,
                location,
                image_url,
                normalize_optional_text(payload.notes),
                now,
                now,
            ),
        ).fetchone()
        sync_standalone_inventory_item_barcodes(
            conn,
            item_id=int(row["id"]),
            primary_barcode=primary_barcode,
            barcodes=barcodes,
        )
        row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (int(row["id"]),)).fetchone()
        payload_row = format_standalone_inventory_rows(conn, [row])[0]
        conn.commit()
    return payload_row


@app.put("/api/inventory/{item_id}")
def update_inventory_item(
    item_id: int,
    payload: StandaloneInventoryUpdate,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    item_name = normalize_required_text(payload.item_name, field_name="Item name")
    unit = normalize_required_text(payload.unit, field_name="Unit")
    location = normalize_storage_grid_location(payload.location)
    image_url = normalize_required_text(payload.image_url, field_name="Product image")
    category = normalize_inventory_category(payload.category)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    with get_connection() as conn:
        existing = conn.execute("SELECT id, barcode FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        existing_barcodes = load_standalone_inventory_item_barcodes(
            conn,
            item_id,
            fallback_primary_barcode=existing["barcode"],
        )
        primary_barcode, barcodes = build_inventory_barcode_payload(
            payload.barcode,
            requested_barcodes=payload.barcodes,
            existing_barcodes=existing_barcodes,
        )
        conn.execute(
            """
            UPDATE standalone_inventory
            SET item_name = ?, barcode = ?, quantity = ?, unit = ?, category = ?, location = ?, image_url = ?, notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                item_name,
                primary_barcode,
                payload.quantity,
                unit,
                category,
                location,
                image_url,
                normalize_optional_text(payload.notes),
                now,
                item_id,
            ),
        )
        sync_standalone_inventory_item_barcodes(
            conn,
            item_id=item_id,
            primary_barcode=primary_barcode,
            barcodes=barcodes,
        )
        row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        payload_row = format_standalone_inventory_rows(conn, [row])[0]
        conn.commit()
    return payload_row


@app.patch("/api/inventory/{item_id}")
def patch_inventory_item(
    item_id: int,
    payload: StandaloneInventoryPatch,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    fields = payload.model_fields_set
    if not fields:
        raise HTTPException(status_code=400, detail="Provide at least one field to update.")

    updates: list[str] = []
    params: list[Any] = []
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    with get_connection() as conn:
        existing = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Inventory item not found")

        current_primary_barcode = normalize_optional_text(existing["barcode"])
        current_barcodes = load_standalone_inventory_item_barcodes(
            conn,
            item_id,
            fallback_primary_barcode=current_primary_barcode,
        )
        barcode_fields_touched = "barcode" in fields or "barcodes" in fields
        next_primary_barcode: str | None = current_primary_barcode
        next_barcodes: list[str] = list(current_barcodes)

        if barcode_fields_touched:
            primary_input = current_primary_barcode
            if "barcode" in fields:
                primary_input = normalize_optional_text(payload.barcode)
            requested_barcodes = current_barcodes
            if "barcodes" in fields:
                requested_barcodes = payload.barcodes or []
            next_primary_barcode, next_barcodes = build_inventory_barcode_payload(
                primary_input,
                requested_barcodes=requested_barcodes,
            )
            updates.append("barcode = ?")
            params.append(next_primary_barcode)

        if "item_name" in fields:
            updates.append("item_name = ?")
            params.append(normalize_required_text(payload.item_name, field_name="Item name"))

        if "quantity" in fields:
            updates.append("quantity = ?")
            params.append(float(payload.quantity))

        if "unit" in fields:
            unit_value = normalize_optional_text(payload.unit)
            updates.append("unit = ?")
            params.append(normalize_required_text(unit_value, field_name="Unit") if unit_value is not None else None)

        if "category" in fields:
            updates.append("category = ?")
            params.append(normalize_inventory_category(payload.category))

        if "location" in fields:
            location_value = normalize_optional_text(payload.location)
            updates.append("location = ?")
            params.append(normalize_storage_grid_location(location_value) if location_value is not None else None)

        if "image_url" in fields:
            updates.append("image_url = ?")
            params.append(normalize_optional_text(payload.image_url))

        if "notes" in fields:
            updates.append("notes = ?")
            params.append(normalize_optional_text(payload.notes))

        updates.append("updated_at = ?")
        params.append(now)
        params.append(item_id)
        conn.execute(
            f"""
            UPDATE standalone_inventory
            SET {', '.join(updates)}
            WHERE id = ?
            """,
            tuple(params),
        )

        if barcode_fields_touched:
            sync_standalone_inventory_item_barcodes(
                conn,
                item_id=item_id,
                primary_barcode=next_primary_barcode,
                barcodes=next_barcodes,
            )

        row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        payload_row = format_standalone_inventory_rows(conn, [row])[0]
        conn.commit()
    return payload_row


@app.patch("/api/inventory/{item_id}/category")
def patch_inventory_item_category(
    item_id: int,
    payload: StandaloneInventoryCategoryPatch,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    category = normalize_inventory_category(payload.category)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM standalone_inventory WHERE id = ?",
            (item_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        conn.execute(
            """
            UPDATE standalone_inventory
            SET category = ?, updated_at = ?
            WHERE id = ?
            """,
            (category, now, item_id),
        )
        row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        payload_row = format_standalone_inventory_rows(conn, [row])[0]
        conn.commit()
    return payload_row


@app.patch("/api/inventory/{item_id}/notes")
def patch_inventory_item_notes(
    item_id: int,
    payload: StandaloneInventoryNotesPatch,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    notes = normalize_optional_text(payload.notes)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM standalone_inventory WHERE id = ?",
            (item_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        conn.execute(
            """
            UPDATE standalone_inventory
            SET notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (notes, now, item_id),
        )
        row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        payload_row = format_standalone_inventory_rows(conn, [row])[0]
        conn.commit()
    return payload_row


@app.patch("/api/inventory/{item_id}/item-name")
def patch_inventory_item_name(
    item_id: int,
    payload: StandaloneInventoryNamePatch,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    item_name = normalize_required_text(payload.item_name, field_name="Item name")
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM standalone_inventory WHERE id = ?",
            (item_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        conn.execute(
            """
            UPDATE standalone_inventory
            SET item_name = ?, updated_at = ?
            WHERE id = ?
            """,
            (item_name, now, item_id),
        )
        row = conn.execute("SELECT * FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        payload_row = format_standalone_inventory_rows(conn, [row])[0]
        conn.commit()
    return payload_row


@app.delete("/api/inventory/{item_id}")
def delete_inventory_item(
    item_id: int,
    _user: Annotated[AuthUser, Depends(require_roles(ROLE_VIEWER, ROLE_PLANNER, ROLE_ADMIN))],
) -> dict[str, Any]:
    with get_connection() as conn:
        existing = conn.execute("SELECT id, item_name FROM standalone_inventory WHERE id = ?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        conn.execute("DELETE FROM standalone_inventory WHERE id = ?", (item_id,))
        conn.commit()
    return {"id": item_id, "item_name": existing["item_name"], "status": "deleted"}


if FRONTEND_DIR.exists():
    # Static inventory pages are served from the repo-level frontend directory.
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")
