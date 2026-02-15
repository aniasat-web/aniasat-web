from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

# Allow `python scripts/import_all_recipes.py` from backend/.
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.db import get_connection, init_db


PEOPLE_UNITS = {"people", "person"}

METADATA_TITLES = {
    "common recipes used across multiple retreats",
    "spice mixes used in the retreats",
    "number of people",
    "total yield",
    "needed for retreats",
    "multiplier needed",
    "(per retreat)",
    "(all retreats)",
    "menu",
    "breakfast",
    "lunch",
    "dinner",
    "registration time",
    "day 1",
    "day 2",
    "day 3",
}

UNIT_ALIASES = {
    "kilogram": "kg",
    "kilograms": "kg",
    "gram": "g",
    "grams": "g",
    "liter": "l",
    "liters": "l",
    "litre": "l",
    "litres": "l",
    "milliliter": "ml",
    "milliliters": "ml",
    "millilitre": "ml",
    "millilitres": "ml",
    "tablespoon": "tbsp",
    "tablespoons": "tbsp",
    "tbs": "tbsp",
    "tbsp": "tbsp",
    "teaspoon": "tsp",
    "teaspoons": "tsp",
    "pound": "lb",
    "pounds": "lb",
}


@dataclass
class ParsedIngredient:
    name: str
    quantity: float
    unit: str


@dataclass
class ParsedRecipe:
    name: str
    base_servings: float = 6.0
    ingredients: list[ParsedIngredient] = field(default_factory=list)
    steps: list[str] = field(default_factory=list)
    source_sheet: str = ""


@dataclass
class RegionConfig:
    title_col: int
    ingredient_name_col: int
    quantity_col: int
    unit_col: int
    step_text_col: int
    lookahead_rows: int = 15


@dataclass
class SheetConfig:
    name: str
    regions: list[RegionConfig]


SHEET_CONFIGS: dict[str, SheetConfig] = {
    "Common Recipes": SheetConfig(
        name="Common Recipes",
        regions=[
            RegionConfig(
                title_col=2,
                ingredient_name_col=3,
                quantity_col=4,
                unit_col=5,
                step_text_col=3,
                lookahead_rows=12,
            ),
            RegionConfig(
                title_col=12,
                ingredient_name_col=12,
                quantity_col=13,
                unit_col=14,
                step_text_col=13,
                lookahead_rows=12,
            ),
        ],
    ),
    "KY1 + Upa": SheetConfig(
        name="KY1 + Upa",
        regions=[
            RegionConfig(2, 3, 4, 5, 3, 15),
            RegionConfig(11, 12, 13, 14, 12, 15),
            RegionConfig(20, 21, 22, 23, 21, 15),
        ],
    ),
    "KY2 + KY3": SheetConfig(
        name="KY2 + KY3",
        regions=[
            RegionConfig(2, 3, 4, 5, 3, 15),
            RegionConfig(11, 12, 13, 14, 12, 15),
            RegionConfig(20, 21, 22, 23, 21, 15),
        ],
    ),
    "Pranams": SheetConfig(
        name="Pranams",
        regions=[
            RegionConfig(2, 3, 4, 5, 3, 15),
        ],
    ),
}


def as_text(value: object) -> str | None:
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return None


def as_float(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def normalize_unit(unit: str) -> str:
    value = unit.strip().lower()
    return UNIT_ALIASES.get(value, value)


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def recipe_key(name: str) -> str:
    return normalize_spaces(name).lower()


def is_number_like(text: str) -> bool:
    try:
        float(text)
        return True
    except ValueError:
        return False


def has_upcoming_ingredients(ws, row_idx: int, title_col: int, lookahead_rows: int) -> bool:
    max_row = ws.max_row
    for rr in range(row_idx + 1, min(max_row, row_idx + lookahead_rows) + 1):
        maybe = as_text(ws.cell(rr, title_col).value)
        if maybe and maybe.lower() == "ingredients":
            return True
    return False


def is_recipe_title_candidate(title: str) -> bool:
    low = title.lower().strip()
    if not low:
        return False
    if low in METADATA_TITLES:
        return False
    if low.startswith("day "):
        return False
    if low == "ingredients":
        return False
    if "steps" in low:
        return False
    if is_number_like(low):
        return False
    return True


def parse_region(ws, sheet_name: str, cfg: RegionConfig) -> list[ParsedRecipe]:
    recipes: list[ParsedRecipe] = []
    current: ParsedRecipe | None = None
    mode: str | None = None

    def flush_current() -> None:
        nonlocal current, mode
        if current and current.ingredients:
            recipes.append(current)
        current = None
        mode = None

    for r in range(1, ws.max_row + 1):
        title = as_text(ws.cell(r, cfg.title_col).value)
        title_low = title.lower() if title else ""

        if title_low == "ingredients":
            if current:
                mode = "ingredients"
            continue

        if "steps" in title_low:
            if current:
                mode = "steps"
            continue

        ingredient_name = as_text(ws.cell(r, cfg.ingredient_name_col).value)
        ingredient_qty = as_float(ws.cell(r, cfg.quantity_col).value)
        ingredient_unit = as_text(ws.cell(r, cfg.unit_col).value)
        ingredient_row = bool(ingredient_name and ingredient_qty is not None and ingredient_unit)

        if current and mode == "ingredients" and ingredient_row:
            if ingredient_qty > 0:
                current.ingredients.append(
                    ParsedIngredient(
                        name=normalize_spaces(ingredient_name),
                        quantity=ingredient_qty,
                        unit=normalize_unit(ingredient_unit),
                    )
                )
            continue

        if current and mode == "steps":
            maybe_step = as_text(ws.cell(r, cfg.step_text_col).value)
            if maybe_step:
                step_low = maybe_step.lower()
                if "steps" not in step_low and step_low != "ingredients":
                    current.steps.append(normalize_spaces(maybe_step))

        if (
            title
            and is_recipe_title_candidate(title)
            and has_upcoming_ingredients(ws, r, cfg.title_col, cfg.lookahead_rows)
        ):
            flush_current()

            base_servings = 6.0
            qty = as_float(ws.cell(r, cfg.quantity_col).value)
            unit = as_text(ws.cell(r, cfg.unit_col).value)
            if qty is not None and unit and unit.lower() in PEOPLE_UNITS:
                base_servings = qty

            current = ParsedRecipe(
                name=normalize_spaces(title),
                base_servings=base_servings,
                source_sheet=sheet_name,
            )
            mode = None

    flush_current()
    return recipes


def dedupe_recipes(recipes: list[ParsedRecipe]) -> list[ParsedRecipe]:
    best: dict[str, ParsedRecipe] = {}

    def score(recipe: ParsedRecipe) -> tuple[int, int]:
        # Prefer richer ingredient coverage, then richer steps.
        return (len(recipe.ingredients), len(recipe.steps))

    for recipe in recipes:
        key = recipe_key(recipe.name)
        existing = best.get(key)
        if not existing or score(recipe) > score(existing):
            best[key] = recipe

    return list(best.values())


def upsert_ingredient(conn, ingredient_name: str) -> int:
    row = conn.execute(
        "SELECT id FROM ingredients WHERE lower(name) = lower(?)",
        (ingredient_name,),
    ).fetchone()
    if row:
        return int(row["id"])

    row = conn.execute(
        "INSERT INTO ingredients(name) VALUES (?) RETURNING id",
        (ingredient_name,),
    ).fetchone()
    return int(row["id"])


def import_recipes(recipes: list[ParsedRecipe], dry_run: bool, replace_existing: bool) -> None:
    init_db()

    created = 0
    replaced = 0
    skipped = 0

    with get_connection() as conn:
        for recipe in recipes:
            existing = conn.execute(
                "SELECT id FROM recipes WHERE lower(name) = lower(?)",
                (recipe.name,),
            ).fetchone()

            if existing and not replace_existing:
                skipped += 1
                continue

            if dry_run:
                if existing:
                    replaced += 1
                else:
                    created += 1
                continue

            if existing and replace_existing:
                conn.execute("DELETE FROM recipes WHERE id = ?", (existing["id"],))
                replaced += 1

            recipe_row = conn.execute(
                "INSERT INTO recipes(name, base_servings, notes) VALUES (?, ?, ?) RETURNING id",
                (
                    recipe.name,
                    recipe.base_servings,
                    f"Imported from workbook ({recipe.source_sheet})",
                ),
            ).fetchone()
            recipe_id = int(recipe_row["id"])
            created += 1

            for item in recipe.ingredients:
                ingredient_id = upsert_ingredient(conn, item.name)
                conn.execute(
                    """
                    INSERT INTO recipe_ingredients(recipe_id, ingredient_id, quantity, unit, prep_notes)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (recipe_id, ingredient_id, item.quantity, item.unit, None),
                )

            for idx, step_text in enumerate(recipe.steps, start=1):
                conn.execute(
                    "INSERT INTO recipe_steps(recipe_id, step_order, instruction) VALUES (?, ?, ?)",
                    (recipe_id, idx, step_text),
                )

        if not dry_run:
            conn.commit()

    print("Import summary")
    print(f"- unique recipes processed: {len(recipes)}")
    print(f"- recipes created: {created}")
    print(f"- recipes replaced: {replaced}")
    print(f"- recipes skipped (exists): {skipped}")


def parse_workbook(xlsx_path: Path, sheet_names: list[str]) -> list[ParsedRecipe]:
    wb = load_workbook(xlsx_path, data_only=True)

    parsed: list[ParsedRecipe] = []
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet not found in workbook: {sheet_name}")
            continue

        ws = wb[sheet_name]
        cfg = SHEET_CONFIGS[sheet_name]

        for region in cfg.regions:
            parsed.extend(parse_region(ws, sheet_name, region))

    return dedupe_recipes(parsed)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Import all recipes from the workbook into the Retreat Ops DB. "
            "Currently supports: Common Recipes, KY1 + Upa, KY2 + KY3, Pranams."
        )
    )
    parser.add_argument("--xlsx", type=Path, required=True, help="Path to source .xlsx workbook")
    parser.add_argument(
        "--sheets",
        nargs="*",
        default=None,
        help="Optional subset of sheets to parse (default: all supported sheets)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse/report only; no DB writes")
    parser.add_argument(
        "--replace-existing",
        action="store_true",
        help="Replace existing same-name recipes in DB",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Workbook not found: {args.xlsx}")

    if args.sheets:
        unknown = [name for name in args.sheets if name not in SHEET_CONFIGS]
        if unknown:
            raise SystemExit(
                "Unsupported sheet names: "
                + ", ".join(unknown)
                + ". Supported: "
                + ", ".join(SHEET_CONFIGS)
            )
        sheet_names = args.sheets
    else:
        sheet_names = list(SHEET_CONFIGS.keys())

    recipes = parse_workbook(args.xlsx, sheet_names)

    print("Parsed recipes")
    for recipe in sorted(recipes, key=lambda r: r.name.lower()):
        print(
            f"- {recipe.name} | sheet={recipe.source_sheet} | "
            f"base_servings={recipe.base_servings:g} | ingredients={len(recipe.ingredients)} | steps={len(recipe.steps)}"
        )

    import_recipes(recipes, dry_run=args.dry_run, replace_existing=args.replace_existing)


if __name__ == "__main__":
    main()
