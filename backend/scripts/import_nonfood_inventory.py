#!/usr/bin/env python3
"""Import non-food storage inventory from workbook into standalone_inventory.

Default source mapping targets the "Non-Food Inventory" sheet in:
/mnt/nas_home/Spring 2026 Inventory File.xlsx

Usage:
  cd backend
  . .venv/bin/activate
  python scripts/import_nonfood_inventory.py --xlsx /mnt/nas_home/Spring\\ 2026\\ Inventory\\ File.xlsx
  python scripts/import_nonfood_inventory.py --xlsx /mnt/nas_home/Spring\\ 2026\\ Inventory\\ File.xlsx --apply
"""

from __future__ import annotations

import argparse
import html
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DB_PATH = SCRIPT_DIR.parent / "data" / "retreat_ops.db"
DEFAULT_SHEET = "Non-Food Inventory"
DEFAULT_CATEGORY_SHEET = "anita temp"
IMPORT_SOURCE = "nonfood-inventory"
IMPORT_TAG = f"[import:{IMPORT_SOURCE}]"

URL_RE = re.compile(r"^https?://", re.IGNORECASE)
IMAGE_URL_RE = re.compile(r"^https?://.+\.(png|jpg|jpeg|webp|gif|bmp)(\?.*)?$", re.IGNORECASE)
META_TAG_RE = re.compile(r"<meta\s+[^>]*>", re.IGNORECASE)
META_ATTR_RE = re.compile(r'([a-zA-Z_:.-]+)\s*=\s*["\']([^"\']*)["\']')
LOCATION_CODE_RE = re.compile(r"\b([A-Za-z])\s*[- ]?\s*([1-9]|[1-9][0-9])\b")
NUMBER_RE = re.compile(r"-?\d+(?:\.\d+)?")
PACK_MULT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[a-zA-Z ]*[x*]\s*(\d+(?:\.\d+)?)")
CASE_PAREN_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[a-zA-Z ]*\(\s*(\d+(?:\.\d+)?)\s*\)")
INFRA_CATEGORY_NAME = "Infra"
INFRA_CATEGORY_EXACT = {
    "infra",
    "infrastructure",
    "maintenance",
    "facility maintenance",
    "facilities maintenance",
    "janitorial",
    "housekeeping",
}
INFRA_CATEGORY_HINTS = (
    "cleaning",
    "maintenance",
    "janitorial",
    "housekeeping",
    "facility",
    "facilities",
)
INFRA_ITEM_HINTS = (
    "all purpose cleaner",
    "cleaner",
    "detergent",
    "dish soap",
    "dishwashing",
    "disinfect",
    "sanitizer",
    "sanitiser",
    "toilet cleaner",
    "trash bag",
    "garbage bag",
    "paper towel",
    "broom",
    "mop",
    "vacuum",
)


@dataclass
class ParsedItem:
    item_name: str
    category: str | None
    quantity: float
    unit: str
    location: str | None
    image_url: str | None
    order_url: str | None
    notes: str | None
    import_source: str | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import non-food inventory tab to standalone_inventory.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Path to source XLSX file")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Worksheet name (default: {DEFAULT_SHEET})")
    parser.add_argument(
        "--category-sheet",
        default=DEFAULT_CATEGORY_SHEET,
        help=f"Optional category source sheet for overrides (default: {DEFAULT_CATEGORY_SHEET})",
    )
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH, help=f"SQLite DB path (default: {DEFAULT_DB_PATH})")
    parser.add_argument("--apply", action="store_true", help="Write changes (default is dry-run)")
    parser.add_argument(
        "--replace-existing-import",
        action="store_true",
        help=(
            "Delete rows previously imported by this script "
            f"(tracked by import_source={IMPORT_SOURCE!r}; also matches legacy notes prefix {IMPORT_TAG!r}) "
            "before insert"
        ),
    )
    parser.add_argument("--no-backup", action="store_true", help="Skip backup when --apply is used")
    parser.add_argument("--max-rows", type=int, default=0, help="Limit parsed inventory rows for testing (0 = no limit)")
    parser.add_argument(
        "--resolve-image-from-links",
        action="store_true",
        help=(
            "Try to resolve image URLs from linked product pages (og:image/twitter:image). "
            "Useful when sheet has product links but no direct image URLs."
        ),
    )
    parser.add_argument(
        "--image-link-timeout",
        type=float,
        default=8.0,
        help="HTTP timeout seconds when --resolve-image-from-links is enabled (default: 8.0)",
    )
    return parser.parse_args()


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_sheet_name(sheet_names: list[str], requested: str) -> str:
    target = requested.strip().casefold()
    for name in sheet_names:
        if name.strip().casefold() == target:
            return name
    raise ValueError(
        f"Sheet {requested!r} not found. Available sheets: {', '.join(sheet_names)}"
    )


def normalize_item_key(value: object) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    collapsed = re.sub(r"\s+", " ", text).strip()
    return collapsed.casefold() if collapsed else None


def canonicalize_inventory_category(value: object) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    collapsed = re.sub(r"\s+", " ", text).strip()
    lowered = collapsed.lower()
    if lowered in INFRA_CATEGORY_EXACT:
        return INFRA_CATEGORY_NAME
    if any(hint in lowered for hint in INFRA_CATEGORY_HINTS):
        return INFRA_CATEGORY_NAME
    return collapsed


def infer_category_from_item_text(*values: object) -> str | None:
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        lowered = text.lower()
        if any(hint in lowered for hint in INFRA_ITEM_HINTS):
            return INFRA_CATEGORY_NAME
    return None


def clean_url(cell) -> str | None:
    if cell is None:
        return None
    if getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "target", None):
        return str(cell.hyperlink.target).strip()
    value = clean_text(cell.value)
    if value and URL_RE.match(value):
        return value
    return None


def sanitize_unit(value: object) -> str:
    if value is None:
        return "each"
    if isinstance(value, (int, float)):
        as_float = float(value)
        return str(int(as_float)) if as_float.is_integer() else f"{as_float:g}"
    text = str(value).strip()
    return text or "each"


def parse_quantity(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return max(float(value), 0.0)

    text = str(value).strip().lower()
    if not text:
        return 0.0

    # Try expressions such as "13 packs*6 each = 78"
    left = text.split("=")[0].strip()
    match = PACK_MULT_RE.search(left)
    if match:
        try:
            return max(float(match.group(1)) * float(match.group(2)), 0.0)
        except ValueError:
            pass

    # Try cases such as "3 case(192)+75 single"
    base = CASE_PAREN_RE.search(left)
    if base:
        try:
            total = float(base.group(1)) * float(base.group(2))
            tail = left[base.end() :]
            tail_numbers = [float(raw) for raw in NUMBER_RE.findall(tail)]
            if tail_numbers:
                total += tail_numbers[0]
            return max(total, 0.0)
        except ValueError:
            pass

    numbers = [float(raw) for raw in NUMBER_RE.findall(text)]
    return max(numbers[0], 0.0) if numbers else 0.0


def normalize_location(raw_location: object) -> tuple[str | None, str | None]:
    text = clean_text(raw_location)
    if not text:
        return None, None
    match = LOCATION_CODE_RE.search(text)
    if not match:
        return None, text
    normalized = f"{match.group(1).upper()}{match.group(2)}"
    if text.strip().upper() == normalized:
        return normalized, None
    return normalized, text


def combine_notes(parts: Iterable[str | None]) -> str | None:
    cleaned: list[str] = []
    for value in parts:
        text = clean_text(value)
        if not text:
            continue
        if text not in cleaned:
            cleaned.append(text)
    if not cleaned:
        return None
    return " | ".join(cleaned)


def normalize_category(major: str | None, sub: str | None) -> str | None:
    major_clean = clean_text(major)
    sub_clean = clean_text(sub)
    combined: str | None = None
    if major_clean and sub_clean:
        if major_clean.casefold() == sub_clean.casefold():
            combined = major_clean
        else:
            combined = f"{major_clean} / {sub_clean}"
    else:
        combined = major_clean or sub_clean
    return canonicalize_inventory_category(combined)


def resolve_meta_image_url_from_html(page_url: str, html_text: str) -> str | None:
    for tag in META_TAG_RE.findall(html_text):
        attrs: dict[str, str] = {}
        for key, value in META_ATTR_RE.findall(tag):
            attrs[key.strip().lower()] = html.unescape(value.strip())
        marker = attrs.get("property", "").lower() or attrs.get("name", "").lower()
        if marker not in {"og:image", "twitter:image"}:
            continue
        content = attrs.get("content")
        if not content:
            continue
        resolved = urllib_parse.urljoin(page_url, content)
        if URL_RE.match(resolved):
            return resolved
    return None


def resolve_image_url_from_reference(
    ref_url: str | None,
    *,
    timeout_seconds: float,
    cache: dict[str, str | None],
) -> str | None:
    url = clean_text(ref_url)
    if not url or not URL_RE.match(url):
        return None
    if IMAGE_URL_RE.match(url):
        return url
    if url in cache:
        return cache[url]

    request = urllib_request.Request(
        url,
        method="GET",
        headers={
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/121.0 Safari/537.36"
            )
        },
    )
    try:
        with urllib_request.urlopen(request, timeout=max(1.0, timeout_seconds)) as response:
            content_type = str(response.headers.get("Content-Type", "")).lower()
            raw = response.read(512_000)
    except (urllib_error.HTTPError, urllib_error.URLError, TimeoutError):
        cache[url] = None
        return None

    if "image/" in content_type:
        cache[url] = url
        return url

    decoded = raw.decode("utf-8", errors="ignore")
    resolved = resolve_meta_image_url_from_html(url, decoded)
    cache[url] = resolved
    return resolved


def parse_category_overrides(workbook, category_sheet_name: str | None) -> dict[str, str]:
    requested = clean_text(category_sheet_name)
    if not requested:
        return {}
    try:
        actual_name = normalize_sheet_name(list(workbook.sheetnames), requested)
    except ValueError:
        return {}

    ws = workbook[actual_name]
    overrides: dict[str, str] = {}
    for row_idx in range(1, ws.max_row + 1):
        category = canonicalize_inventory_category(ws.cell(row_idx, 2).value)
        item_name = clean_text(ws.cell(row_idx, 3).value)
        key = normalize_item_key(item_name)
        if not key or not category:
            continue
        if key not in overrides:
            overrides[key] = category
    return overrides


def parse_inventory_rows(
    xlsx_path: Path,
    sheet_name: str,
    *,
    category_sheet_name: str | None = DEFAULT_CATEGORY_SHEET,
    max_rows: int = 0,
    resolve_image_from_links: bool = False,
    image_link_timeout: float = 8.0,
) -> list[ParsedItem]:
    workbook = load_workbook(filename=xlsx_path, read_only=False, data_only=True)
    sheet_actual = normalize_sheet_name(workbook.sheetnames, sheet_name)
    ws = workbook[sheet_actual]
    category_overrides = parse_category_overrides(workbook, category_sheet_name)
    image_cache: dict[str, str | None] = {}

    parsed: list[ParsedItem] = []
    major_category: str | None = None
    sub_category: str | None = None

    for row_idx in range(1, ws.max_row + 1):
        col_a = clean_text(ws.cell(row_idx, 1).value)
        col_b = clean_text(ws.cell(row_idx, 2).value)
        item_name = clean_text(ws.cell(row_idx, 3).value)

        if not item_name:
            if col_a and not col_b:
                major_category = col_a
                sub_category = None
            elif col_b and not col_a:
                sub_category = col_b
            elif col_a and col_b:
                major_category = col_a
                sub_category = col_b
            continue

        row_category = normalize_category(major_category, col_b or sub_category)
        desc_cell = ws.cell(row_idx, 7)
        image_cell = ws.cell(row_idx, 8)
        count_raw = ws.cell(row_idx, 9).value
        unit_raw = ws.cell(row_idx, 10).value
        location_raw = ws.cell(row_idx, 11).value
        notes_raw = ws.cell(row_idx, 12).value

        # Backup/fallback quantity and location from legacy columns.
        if count_raw in (None, ""):
            count_raw = ws.cell(row_idx, 4).value
        if location_raw in (None, ""):
            location_raw = ws.cell(row_idx, 5).value

        quantity = parse_quantity(count_raw)
        unit = sanitize_unit(unit_raw)
        location, legacy_location = normalize_location(location_raw)

        product_url = clean_url(desc_cell)
        image_ref = clean_url(image_cell)
        image_url = image_ref if image_ref and IMAGE_URL_RE.match(image_ref) else None
        order_url = product_url or (image_ref if image_ref and not IMAGE_URL_RE.match(image_ref) else None)
        if not image_url and resolve_image_from_links:
            image_url = resolve_image_url_from_reference(
                image_ref,
                timeout_seconds=image_link_timeout,
                cache=image_cache,
            )
        if not image_url and resolve_image_from_links:
            image_url = resolve_image_url_from_reference(
                product_url,
                timeout_seconds=image_link_timeout,
                cache=image_cache,
            )

        description_text = clean_text(desc_cell.value)
        item_key = normalize_item_key(item_name)
        category_override = category_overrides.get(item_key) if item_key else None
        if category_override and (not row_category or row_category.casefold() in {"items", "miscellaneous"}):
            row_category = category_override
        if row_category and row_category.casefold() in {"items", "miscellaneous"}:
            inferred = infer_category_from_item_text(item_name, description_text)
            if inferred:
                row_category = inferred
        if not row_category:
            row_category = infer_category_from_item_text(item_name, description_text)
        row_category = canonicalize_inventory_category(row_category)
        description_note = None
        if description_text and not URL_RE.match(description_text):
            description_note = f"Description: {description_text}"

        notes = combine_notes(
            [
                clean_text(notes_raw),
                description_note,
                f"Image reference: {image_ref}" if image_ref and not image_url and image_ref != order_url else None,
                f"Legacy location: {legacy_location}" if legacy_location else None,
            ]
        )

        parsed.append(
            ParsedItem(
                item_name=item_name,
                category=row_category,
                quantity=quantity,
                unit=unit,
                location=location,
                image_url=image_url,
                order_url=order_url,
                notes=notes,
                import_source=IMPORT_SOURCE,
            )
        )

        if max_rows > 0 and len(parsed) >= max_rows:
            break

    return parsed


def backup_db(db_path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = db_path.with_name(f"{db_path.name}.bak-{ts}")
    backup_path.write_bytes(db_path.read_bytes())
    return backup_path


def table_columns(cur: sqlite3.Cursor, table_name: str) -> set[str]:
    rows = cur.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {str(row[1]).strip() for row in rows}


def apply_import(cur: sqlite3.Cursor, rows: list[ParsedItem], replace_existing_import: bool) -> tuple[int, int]:
    columns = table_columns(cur, "standalone_inventory")
    supports_barcode = "barcode" in columns
    supports_image_url = "image_url" in columns
    supports_order_url = "order_url" in columns
    supports_import_source = "import_source" in columns

    deleted = 0
    if replace_existing_import:
        if supports_import_source:
            deleted = cur.execute(
                "DELETE FROM standalone_inventory WHERE lower(COALESCE(import_source, '')) = lower(?) OR notes LIKE ?",
                (IMPORT_SOURCE, f"{IMPORT_TAG}%"),
            ).rowcount
        else:
            deleted = cur.execute(
                "DELETE FROM standalone_inventory WHERE notes LIKE ?",
                (f"{IMPORT_TAG}%",),
            ).rowcount

    inserted = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in rows:
        insert_columns = [
            "item_name",
            "quantity",
            "unit",
            "category",
            "location",
        ]
        insert_values: list[object] = [
            row.item_name,
            row.quantity,
            row.unit,
            row.category,
            row.location,
        ]
        if supports_barcode:
            insert_columns.append("barcode")
            insert_values.append(None)
        if supports_image_url:
            insert_columns.append("image_url")
            insert_values.append(row.image_url)
        if supports_order_url:
            insert_columns.append("order_url")
            insert_values.append(row.order_url)
        if supports_import_source:
            insert_columns.append("import_source")
            insert_values.append(row.import_source)
        insert_columns.extend(["notes", "created_at", "updated_at"])
        insert_values.extend([row.notes, now, now])
        placeholders = ", ".join("?" for _ in insert_columns)
        cur.execute(
            f"INSERT INTO standalone_inventory({', '.join(insert_columns)}) VALUES ({placeholders})",
            tuple(insert_values),
        )
        inserted += 1

    return deleted, inserted


def summarize(rows: list[ParsedItem]) -> None:
    with_location = sum(1 for row in rows if row.location)
    with_image_url = sum(1 for row in rows if row.image_url)
    with_category = sum(1 for row in rows if row.category)
    with_qty = sum(1 for row in rows if row.quantity > 0)
    print(f"Parsed rows: {len(rows)}")
    print(f"Rows with category: {with_category}")
    print(f"Rows with quantity > 0: {with_qty}")
    print(f"Rows with normalized location: {with_location}")
    print(f"Rows with direct image URL: {with_image_url}")
    print("\nSample rows:")
    for row in rows[:12]:
        print(
            f"- item={row.item_name!r}, qty={row.quantity}, unit={row.unit!r}, "
            f"location={row.location!r}, category={row.category!r}"
        )


def main() -> None:
    args = parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(f"XLSX not found: {args.xlsx}")
    if not args.db.exists():
        raise FileNotFoundError(f"DB not found: {args.db}")

    parsed_rows = parse_inventory_rows(
        xlsx_path=args.xlsx,
        sheet_name=args.sheet,
        category_sheet_name=args.category_sheet,
        max_rows=max(0, int(args.max_rows)),
        resolve_image_from_links=bool(args.resolve_image_from_links),
        image_link_timeout=float(args.image_link_timeout),
    )
    summarize(parsed_rows)

    if not args.apply:
        print("\nDry-run only. Re-run with --apply to write changes.")
        return

    if not args.no_backup:
        backup_path = backup_db(args.db)
        print(f"Backup created: {backup_path}")

    conn = sqlite3.connect(args.db)
    try:
        cur = conn.cursor()
        deleted, inserted = apply_import(cur, parsed_rows, replace_existing_import=args.replace_existing_import)
        conn.commit()
    finally:
        conn.close()

    print(f"\nApply complete. Deleted previous imported rows: {deleted}. Inserted rows: {inserted}.")


if __name__ == "__main__":
    main()
