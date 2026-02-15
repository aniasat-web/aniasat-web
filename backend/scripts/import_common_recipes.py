from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

# Allow `python scripts/import_common_recipes.py` from backend/.
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.db import get_connection, init_db


PEOPLE_UNITS = {"people", "person"}

METADATA_TITLES = {
    "common recipes used across multiple retreats",
    "spice mixes used in the retreats",
    "number of people",
    "total yield",
    "needed for retreats",
    "multiplier needed",
    "(per retreat)",
    "(all retreats)",
}

UNIT_ALIASES = {
    "kilogram": "kg",
    "kilograms": "kg",
    "gram": "g",
    "grams": "g",
    "liter": "l",
    "liters": "l",
    "litre": "l",
    "litres": "l",
    "milliliter": "ml",
    "milliliters": "ml",
    "millilitre": "ml",
    "millilitres": "ml",
    "tablespoon": "tbsp",
    "tablespoons": "tbsp",
    "tbs": "tbsp",
    "teaspoon": "tsp",
    "teaspoons": "tsp",
    "pound": "lb",
    "pounds": "lb",
}


@dataclass
class ParsedIngredient:
    name: str
    quantity: float
    unit: str


@dataclass
class ParsedRecipe:
    name: str
    base_servings: float = 6.0
    ingredients: list[ParsedIngredient] = field(default_factory=list)
    steps: list[str] = field(default_factory=list)


@dataclass
class RegionConfig:
    title_col: int
    ingredient_name_col: int
    quantity_col: int
    unit_col: int
    step_text_col: int


def as_text(value: object) -> str | None:
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return None


def as_float(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def normalize_unit(unit: str) -> str:
    value = unit.strip().lower()
    return UNIT_ALIASES.get(value, value)


def normalize_recipe_name(name: str) -> str:
    return re.sub(r"\s+", " ", name).strip()


def is_number_like(text: str) -> bool:
    try:
        float(text)
        return True
    except ValueError:
        return False


def has_upcoming_ingredients(ws, row_idx: int, title_col: int, lookahead: int = 10) -> bool:
    max_row = ws.max_row
    for rr in range(row_idx + 1, min(max_row, row_idx + lookahead) + 1):
        maybe = as_text(ws.cell(rr, title_col).value)
        if maybe and maybe.lower() == "ingredients":
            return True
    return False


def is_recipe_title_candidate(title: str) -> bool:
    low = title.lower().strip()
    if not low:
        return False
    if low in METADATA_TITLES:
        return False
    if low.startswith("day "):
        return False
    if low == "ingredients":
        return False
    if "steps" in low:
        return False
    if is_number_like(low):
        return False
    return True


def parse_region(ws, cfg: RegionConfig) -> list[ParsedRecipe]:
    recipes: list[ParsedRecipe] = []
    current: ParsedRecipe | None = None
    mode: str | None = None

    def flush_current() -> None:
        nonlocal current, mode
        if current and current.ingredients:
            recipes.append(current)
        current = None
        mode = None

    for r in range(1, ws.max_row + 1):
        title = as_text(ws.cell(r, cfg.title_col).value)
        title_low = title.lower() if title else ""

        if title_low == "ingredients":
            if current:
                mode = "ingredients"
            continue

        if "steps" in title_low:
            if current:
                mode = "steps"
            continue

        ingredient_name = as_text(ws.cell(r, cfg.ingredient_name_col).value)
        ingredient_qty = as_float(ws.cell(r, cfg.quantity_col).value)
        ingredient_unit = as_text(ws.cell(r, cfg.unit_col).value)
        ingredient_row = bool(ingredient_name and ingredient_qty is not None and ingredient_unit)

        if current and mode == "ingredients" and ingredient_row:
            if ingredient_qty > 0:
                current.ingredients.append(
                    ParsedIngredient(
                        name=ingredient_name,
                        quantity=ingredient_qty,
                        unit=normalize_unit(ingredient_unit),
                    )
                )
            continue

        if current and mode == "steps":
            maybe_step = as_text(ws.cell(r, cfg.step_text_col).value)
            if maybe_step:
                step_low = maybe_step.lower()
                if "steps" not in step_low and step_low != "ingredients":
                    current.steps.append(maybe_step)

        if title and is_recipe_title_candidate(title) and has_upcoming_ingredients(ws, r, cfg.title_col):
            flush_current()

            base_servings = 6.0
            qty = as_float(ws.cell(r, cfg.quantity_col).value)
            unit = as_text(ws.cell(r, cfg.unit_col).value)
            if qty is not None and unit and unit.lower() in PEOPLE_UNITS:
                base_servings = qty

            current = ParsedRecipe(name=normalize_recipe_name(title), base_servings=base_servings)
            mode = None

    flush_current()
    return recipes


def dedupe_recipes(recipes: list[ParsedRecipe]) -> list[ParsedRecipe]:
    by_name: dict[str, ParsedRecipe] = {}
    for recipe in recipes:
        key = recipe.name.lower()
        existing = by_name.get(key)
        if not existing:
            by_name[key] = recipe
            continue

        existing_score = len(existing.ingredients) + len(existing.steps)
        new_score = len(recipe.ingredients) + len(recipe.steps)
        if new_score > existing_score:
            by_name[key] = recipe

    return list(by_name.values())


def upsert_ingredient(conn, ingredient_name: str) -> int:
    row = conn.execute(
        "SELECT id FROM ingredients WHERE lower(name) = lower(?)",
        (ingredient_name,),
    ).fetchone()
    if row:
        return int(row["id"])

    row = conn.execute(
        "INSERT INTO ingredients(name) VALUES (?) RETURNING id",
        (ingredient_name,),
    ).fetchone()
    return int(row["id"])


def import_recipes(recipes: list[ParsedRecipe], dry_run: bool, replace_existing: bool) -> None:
    init_db()

    created = 0
    replaced = 0
    skipped = 0

    with get_connection() as conn:
        for recipe in recipes:
            existing = conn.execute(
                "SELECT id FROM recipes WHERE lower(name) = lower(?)",
                (recipe.name,),
            ).fetchone()

            if existing and not replace_existing:
                skipped += 1
                continue

            if dry_run:
                if existing:
                    replaced += 1
                else:
                    created += 1
                continue

            if existing and replace_existing:
                conn.execute("DELETE FROM recipes WHERE id = ?", (existing["id"],))
                replaced += 1

            row = conn.execute(
                "INSERT INTO recipes(name, base_servings, notes) VALUES (?, ?, ?) RETURNING id",
                (recipe.name, recipe.base_servings, "Imported from Common Recipes sheet"),
            ).fetchone()
            recipe_id = int(row["id"])
            created += 1

            for item in recipe.ingredients:
                ingredient_id = upsert_ingredient(conn, item.name)
                conn.execute(
                    """
                    INSERT INTO recipe_ingredients(recipe_id, ingredient_id, quantity, unit, prep_notes)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (recipe_id, ingredient_id, item.quantity, item.unit, None),
                )

            for idx, step_text in enumerate(recipe.steps, start=1):
                conn.execute(
                    "INSERT INTO recipe_steps(recipe_id, step_order, instruction) VALUES (?, ?, ?)",
                    (recipe_id, idx, step_text),
                )

        if not dry_run:
            conn.commit()

    print("Import summary")
    print(f"- recipes processed: {len(recipes)}")
    print(f"- recipes created: {created}")
    print(f"- recipes replaced: {replaced}")
    print(f"- recipes skipped (exists): {skipped}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import recipes from the 'Common Recipes' worksheet into the Retreat Ops DB."
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        required=True,
        help="Path to the source .xlsx file",
    )
    parser.add_argument(
        "--sheet",
        default="Common Recipes",
        help="Worksheet name to parse (default: Common Recipes)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and report only; do not write to the DB",
    )
    parser.add_argument(
        "--replace-existing",
        action="store_true",
        help="Replace recipes with matching names if they already exist",
    )
    parser.add_argument(
        "--skip-right-region",
        action="store_true",
        help="Skip right-side recipe region (spice mixes) in the worksheet",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Workbook not found: {args.xlsx}")

    wb = load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' not found. Available: {', '.join(wb.sheetnames)}")

    ws = wb[args.sheet]

    left_region = RegionConfig(
        title_col=2,
        ingredient_name_col=3,
        quantity_col=4,
        unit_col=5,
        step_text_col=3,
    )
    right_region = RegionConfig(
        title_col=12,
        ingredient_name_col=12,
        quantity_col=13,
        unit_col=14,
        step_text_col=13,
    )

    recipes = parse_region(ws, left_region)
    if not args.skip_right_region:
        recipes.extend(parse_region(ws, right_region))

    recipes = dedupe_recipes(recipes)

    print("Parsed recipes")
    for recipe in recipes:
        print(
            f"- {recipe.name} | base_servings={recipe.base_servings:g} | "
            f"ingredients={len(recipe.ingredients)} | steps={len(recipe.steps)}"
        )

    import_recipes(recipes, dry_run=args.dry_run, replace_existing=args.replace_existing)


if __name__ == "__main__":
    main()
